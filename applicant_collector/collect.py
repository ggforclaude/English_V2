"""지원자 데이터 수집기.

로그인은 사용자가 브라우저에서 직접 수행합니다.
이 스크립트는 아이디/비밀번호를 묻지도, 저장하지도, 전송하지도 않습니다.
수집한 데이터는 이 폴더의 output/ 아래 엑셀 파일로만 저장됩니다.

실행:  python collect.py
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from playwright.sync_api import sync_playwright

BASE = Path(__file__).resolve().parent
PROFILE_DIR = BASE / ".browser_profile"
OUT_DIR = BASE / "output"

VISIT_DELAY_MS = 1200      # 상세 페이지 사이 간격 — 사이트에 부담 주지 않도록
MAX_TEXT_CHARS = 20000     # 상세 페이지에서 가져올 최대 글자 수
PAGE_TIMEOUT_MS = 30000


# --------------------------------------------------------------------------
# 브라우저에 주입되는 조작 패널
# --------------------------------------------------------------------------
PANEL_JS = r"""
(() => {
  if (window.top !== window.self) return;   // 최상위 프레임에만 패널 설치
  if (window.__collectorInstalled) return;
  window.__collectorInstalled = true;

  // 같은 출처(iframe 포함)의 모든 문서 — 사이트가 목록을 iframe에 넣는 경우 대응
  const allDocs = () => {
    const docs = [document];
    document.querySelectorAll('iframe,frame').forEach((f) => {
      try {
        if (f.contentDocument && f.contentDocument.body) docs.push(f.contentDocument);
      } catch (e) { /* 다른 출처 iframe은 접근 불가 */ }
    });
    return docs;
  };

  const ready = (fn) =>
    document.readyState === 'loading'
      ? document.addEventListener('DOMContentLoaded', fn, { once: true })
      : fn();

  ready(() => {
    if (!document.documentElement) return;

    const host = document.createElement('div');
    // 'all:initial'을 먼저 선언해 사이트 스타일을 털어낸 뒤, 위치를 !important로 못박는다.
    // (순서가 반대면 all:initial 이 position:fixed 를 static 으로 되돌려 패널이 페이지 맨 아래로 밀린다)
    host.style.cssText =
      'all:initial;' +
      'position:fixed!important;top:14px!important;right:14px!important;' +
      'left:auto!important;bottom:auto!important;' +
      'z-index:2147483647!important;display:block!important;' +
      'width:auto!important;height:auto!important;margin:0!important;padding:0!important;' +
      'visibility:visible!important;opacity:1!important;transform:none!important;';
    const root = host.attachShadow({ mode: 'open' });

    root.innerHTML = `
      <style>
        * { box-sizing: border-box; font-family: 'Malgun Gothic', system-ui, sans-serif; }
        .panel {
          width: 232px; background: #ffffff; color: #1c2030;
          border: 1px solid #d8dce8; border-radius: 12px;
          box-shadow: 0 8px 28px -8px #1c203055; overflow: hidden;
          font-size: 12.5px; line-height: 1.5;
        }
        .head {
          display: flex; align-items: center; justify-content: space-between;
          padding: 9px 12px; background: #4a55e8; color: #fff;
          font-weight: 700; font-size: 12.5px; cursor: move; user-select: none;
        }
        .head button {
          background: none; border: none; color: #fff; cursor: pointer;
          font-size: 15px; line-height: 1; padding: 0 2px; opacity: .8;
        }
        .head button:hover { opacity: 1; }
        .body { padding: 11px 12px 13px; display: flex; flex-direction: column; gap: 7px; }
        .panel.collapsed .body { display: none; }
        .status {
          background: #f2f4fb; border-radius: 7px; padding: 7px 9px;
          font-size: 11.5px; color: #4a5065; min-height: 32px;
          word-break: break-all;
        }
        button.act {
          width: 100%; text-align: left; padding: 9px 11px; cursor: pointer;
          border: 1px solid #dfe3ee; border-radius: 8px; background: #fafbfd;
          font-size: 12.5px; font-weight: 600; color: #1c2030;
        }
        button.act:hover { border-color: #4a55e8; background: #eef0fd; color: #4a55e8; }
        button.act.primary { background: #4a55e8; border-color: #4a55e8; color: #fff; }
        button.act.primary:hover { background: #3c46d1; }
        button.act:disabled { opacity: .5; cursor: not-allowed; }
        .counts {
          display: flex; justify-content: space-between;
          font-size: 11px; color: #8b90a3; padding-top: 3px;
          border-top: 1px dashed #e4e7f0;
        }
        .confirm { display: none; flex-direction: column; gap: 6px; }
        .confirm.show { display: flex; }
        .confirm .row { display: flex; gap: 6px; }
        .confirm .row button { flex: 1; }
      </style>

      <div class="panel" id="panel">
        <div class="head" id="head">
          <span>지원자 수집기</span>
          <button id="toggle" title="접기/펼치기">–</button>
        </div>
        <div class="body">
          <div class="status" id="status">로그인 후, 지원자 목록 페이지로 이동하세요.</div>

          <button class="act" id="btn-table">이 페이지의 표 수집</button>
          <button class="act" id="btn-pick">지원자 링크 지정하기</button>

          <div class="confirm" id="confirm">
            <div class="row">
              <button class="act primary" id="btn-go">수집 시작</button>
              <button class="act" id="btn-cancel">취소</button>
            </div>
          </div>

          <button class="act primary" id="btn-save">엑셀로 저장</button>

          <div class="counts">
            <span id="c-list">목록 0행</span>
            <span id="c-detail">상세 0명</span>
          </div>
        </div>
      </div>
    `;

    document.documentElement.appendChild(host);

    // 사이트가 DOM을 다시 그려 패널이 떨어져 나가면 되붙인다 (SPA 대응)
    setInterval(() => {
      if (!host.isConnected && document.documentElement) {
        document.documentElement.appendChild(host);
      }
    }, 1500);

    const $ = (id) => root.getElementById(id);
    const panel = $('panel');
    const statusEl = $('status');
    const confirmEl = $('confirm');

    // ---- 파이썬 -> 패널 ----
    window.__collectorStatus = (msg) => { statusEl.textContent = msg; };
    window.__collectorCounts = (listRows, detailCount) => {
      $('c-list').textContent = '목록 ' + listRows + '행';
      $('c-detail').textContent = '상세 ' + detailCount + '명';
    };

    const send = (obj) => {
      if (window.pyCommand) window.pyCommand(JSON.stringify(obj));
    };

    // ---- 접기 / 드래그 ----
    $('toggle').addEventListener('click', () => {
      panel.classList.toggle('collapsed');
      $('toggle').textContent = panel.classList.contains('collapsed') ? '+' : '–';
    });

    (() => {
      let dragging = false, sx = 0, sy = 0, ox = 0, oy = 0;
      $('head').addEventListener('mousedown', (e) => {
        if (e.target.id === 'toggle') return;
        dragging = true;
        sx = e.clientX; sy = e.clientY;
        const r = host.getBoundingClientRect();
        ox = r.left; oy = r.top;
        e.preventDefault();
      });
      window.addEventListener('mousemove', (e) => {
        if (!dragging) return;
        host.style.left = (ox + e.clientX - sx) + 'px';
        host.style.top = (oy + e.clientY - sy) + 'px';
        host.style.right = 'auto';
      });
      window.addEventListener('mouseup', () => { dragging = false; });
    })();

    // ---- 표 수집 ----
    $('btn-table').addEventListener('click', () => {
      const tables = [];
      allDocs().forEach((doc) => {
        doc.querySelectorAll('table').forEach((tbl) => {
          const rows = [...tbl.querySelectorAll('tr')]
            .map((tr) => [...tr.querySelectorAll('th,td')].map((c) => (c.innerText || '').trim()))
            .filter((r) => r.length > 0);
          if (rows.length >= 2) tables.push(rows);
        });
      });
      if (!tables.length) {
        statusEl.textContent = '이 페이지에서 표를 찾지 못했습니다. 링크 지정 방식을 써보세요.';
        return;
      }
      statusEl.textContent = '표를 수집하는 중…';
      send({ type: 'table', tables, url: location.href, title: document.title });
    });

    // ---- 링크 패턴 학습 ----
    const patternOf = (href, base) => {
      try {
        const u = new URL(href, base || location.href);
        if (!/^https?:$/.test(u.protocol)) return null;
        const path = u.pathname.replace(/\d+/g, '#');
        const keys = [...u.searchParams.keys()].sort().join(',');
        return u.origin + '|' + path + '|' + keys;
      } catch (e) { return null; }
    };

    const nameOf = (a) => {
      const t = (a.innerText || '').trim();
      if (t) return t.split('\n')[0].slice(0, 120);
      const tr = a.closest('tr');
      if (tr) {
        const c = tr.querySelector('td,th');
        if (c) return (c.innerText || '').trim().split('\n')[0].slice(0, 120);
      }
      return '(이름 없음)';
    };

    let picking = false;
    let matched = [];

    const stopPicking = () => {
      picking = false;
      allDocs().forEach((doc) => {
        doc.removeEventListener('click', onPick, true);
        doc.removeEventListener('keydown', onEsc, true);
        if (doc.body) doc.body.style.cursor = '';
      });
    };

    const onEsc = (e) => {
      if (e.key === 'Escape') { stopPicking(); statusEl.textContent = '취소했습니다.'; }
    };

    const onPick = (e) => {
      if (host.contains(e.target)) return;
      const a = e.target.closest && e.target.closest('a[href]');
      if (!a) return;
      e.preventDefault();
      e.stopPropagation();

      const pat = patternOf(a.getAttribute('href'), a.ownerDocument.baseURI);
      stopPicking();
      if (!pat) { statusEl.textContent = '이 링크는 사용할 수 없습니다. 다른 링크를 눌러보세요.'; return; }

      const seen = new Set();
      matched = [];
      allDocs().forEach((doc) => {
        doc.querySelectorAll('a[href]').forEach((el) => {
          if (patternOf(el.getAttribute('href'), doc.baseURI) !== pat) return;
          const url = new URL(el.getAttribute('href'), doc.baseURI).href;
          if (seen.has(url)) return;
          seen.add(url);
          matched.push({ name: nameOf(el), url });
        });
      });

      statusEl.textContent =
        '같은 형식의 링크 ' + matched.length + '개를 찾았습니다.\n' +
        '예: ' + (matched[0] ? matched[0].name : '-');
      confirmEl.classList.add('show');
    };

    $('btn-pick').addEventListener('click', () => {
      if (picking) return;
      picking = true;
      confirmEl.classList.remove('show');
      statusEl.textContent = '지원자 이름 링크를 하나 클릭하세요. (ESC = 취소)';
      allDocs().forEach((doc) => {
        if (doc.body) doc.body.style.cursor = 'crosshair';
        doc.addEventListener('click', onPick, true);
        doc.addEventListener('keydown', onEsc, true);
      });
    });

    $('btn-cancel').addEventListener('click', () => {
      confirmEl.classList.remove('show');
      matched = [];
      statusEl.textContent = '취소했습니다.';
    });

    $('btn-go').addEventListener('click', () => {
      if (!matched.length) return;
      confirmEl.classList.remove('show');
      statusEl.textContent = '수집을 준비하는 중…';
      send({ type: 'links', items: matched, sourceUrl: location.href });
      matched = [];
    });

    // ---- 저장 ----
    $('btn-save').addEventListener('click', () => {
      statusEl.textContent = '저장하는 중…';
      send({ type: 'save' });
    });
  });
})();
"""


class Store:
    """수집한 데이터를 모아두는 곳."""

    def __init__(self):
        self.list_header = []
        self.list_rows = []
        self.details = []

    def add_table(self, rows):
        """가장 행이 많은 표를 목록 데이터로 받는다."""
        header, *body = rows
        if not self.list_header:
            self.list_header = header
        # 이미 받은 표와 열 개수가 다르면 길이를 맞춘다
        width = len(self.list_header)
        for r in body:
            padded = (r + [""] * width)[:width]
            if any(c for c in padded):
                self.list_rows.append(padded)

    def add_detail(self, name, url, text):
        self.details.append(
            {
                "이름": name,
                "상세URL": url,
                "수집일시": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "내용": text,
            }
        )

    def is_empty(self):
        return not self.list_rows and not self.details

    def save(self):
        OUT_DIR.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = OUT_DIR / f"지원자수집_{stamp}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        if self.list_rows:
            ws = wb.create_sheet("목록")
            ws.append(self.list_header)
            for row in self.list_rows:
                ws.append(row)
            _autosize(ws)

        if self.details:
            ws = wb.create_sheet("상세")
            cols = ["이름", "상세URL", "수집일시", "내용"]
            ws.append(cols)
            for d in self.details:
                ws.append([d[c] for c in cols])
            _autosize(ws, wide_col=4)

        wb.save(path)
        return path


def _autosize(ws, wide_col=None):
    for i, col in enumerate(ws.columns, start=1):
        letter = col[0].column_letter
        if wide_col and i == wide_col:
            ws.column_dimensions[letter].width = 80
            continue
        longest = max((len(str(c.value or "")) for c in col[:200]), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 8), 40)


def main():
    PROFILE_DIR.mkdir(exist_ok=True)
    store = Store()
    pending = []

    print("=" * 62)
    print(" 지원자 데이터 수집기")
    print("=" * 62)
    print(" 1. 열리는 브라우저에서 직접 로그인하세요.")
    print("    (아이디/비밀번호는 이 스크립트에 저장되지 않습니다)")
    print(" 2. 지원자 목록 페이지로 이동하세요.")
    print(" 3. 화면 오른쪽 위 패널의 버튼으로 수집하세요.")
    print(" 4. 창을 닫으면 종료됩니다.")
    print("=" * 62)

    with sync_playwright() as pw:
        context = pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=False,
            viewport=None,
            args=["--start-maximized"],
        )
        context.set_default_timeout(PAGE_TIMEOUT_MS)

        def py_command(payload):
            try:
                pending.append(json.loads(payload))
            except Exception as exc:  # noqa: BLE001
                print(f"[명령 해석 실패] {exc}")
            return "ok"

        context.expose_function("pyCommand", py_command)
        context.add_init_script(PANEL_JS)

        def broadcast(fn_call):
            for p in list(context.pages):
                try:
                    p.evaluate(fn_call)
                except Exception:  # noqa: BLE001
                    pass

        def set_status(msg):
            broadcast(
                "() => window.__collectorStatus && window.__collectorStatus("
                + json.dumps(msg)
                + ")"
            )

        def set_counts():
            broadcast(
                f"() => window.__collectorCounts && window.__collectorCounts("
                f"{len(store.list_rows)}, {len(store.details)})"
            )

        page = context.pages[0] if context.pages else context.new_page()
        page.goto("about:blank")

        print("\n브라우저가 열렸습니다. 로그인 후 패널을 사용하세요.\n")

        while True:
            if not context.pages:
                break
            try:
                context.pages[0].wait_for_timeout(250)
            except Exception:  # noqa: BLE001
                break

            while pending:
                cmd = pending.pop(0)
                try:
                    _handle(cmd, context, store, set_status, set_counts)
                except Exception as exc:  # noqa: BLE001
                    print(f"[오류] {exc}")
                    set_status(f"오류: {exc}")

        # 창이 닫혔는데 저장 안 한 데이터가 있으면 자동 저장
        if not store.is_empty():
            path = store.save()
            print(f"\n창이 닫혀 자동 저장했습니다 → {path}")
        print("종료합니다.")


def _handle(cmd, context, store, set_status, set_counts):
    kind = cmd.get("type")

    if kind == "table":
        tables = cmd.get("tables") or []
        biggest = max(tables, key=len)
        before = len(store.list_rows)
        store.add_table(biggest)
        added = len(store.list_rows) - before
        print(f"[표] {added}행 수집 (표 {len(tables)}개 중 가장 큰 것)")
        set_status(f"표에서 {added}행을 수집했습니다.")
        set_counts()

    elif kind == "links":
        items = cmd.get("items") or []
        total = len(items)
        print(f"[상세] {total}명 수집 시작")
        set_status(f"수집 중… 0 / {total}")

        worker = context.new_page()
        try:
            for i, item in enumerate(items, start=1):
                name, url = item.get("name", ""), item.get("url", "")
                try:
                    worker.goto(url, wait_until="domcontentloaded")
                    worker.wait_for_timeout(600)
                    text = worker.evaluate(
                        "() => (document.body && document.body.innerText) || ''"
                    )
                    text = re.sub(r"\n{3,}", "\n\n", text).strip()[:MAX_TEXT_CHARS]
                    store.add_detail(name, url, text)
                    print(f"  {i}/{total}  {name}  ({len(text)}자)")
                except Exception as exc:  # noqa: BLE001
                    store.add_detail(name, url, f"[수집 실패] {exc}")
                    print(f"  {i}/{total}  {name}  실패: {exc}")

                set_status(f"수집 중… {i} / {total}")
                set_counts()
                worker.wait_for_timeout(VISIT_DELAY_MS)
        finally:
            try:
                worker.close()
            except Exception:  # noqa: BLE001
                pass

        set_status(f"{total}명 수집 완료. '엑셀로 저장'을 누르세요.")
        set_counts()

    elif kind == "save":
        if store.is_empty():
            set_status("아직 수집된 데이터가 없습니다.")
            return
        path = store.save()
        print(f"[저장] {path}")
        set_status(f"저장 완료: {path.name}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n중단했습니다.")
        sys.exit(0)
