"""결과 검증: 문자열이 잔고/잔액에 남아있는지 + 달러금액 열 확인"""
import openpyxl, os

BASE = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.load_workbook(os.path.join(BASE, '토스_통합거래내역.xlsx'))

# 원화 시트 검증
ws = wb['원화 거래내역']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
print('원화 헤더:', headers)

problems = []
fx_sample = []
for row in ws.iter_rows(min_row=2, values_only=True):
    잔고 = row[12]
    잔액 = row[13]
    달러 = row[14]
    거래구분 = row[1]
    if isinstance(잔고, str) or isinstance(잔액, str):
        problems.append(row)
    if 거래구분 and '환전' in str(거래구분) and len(fx_sample) < 5:
        fx_sample.append(row)

print(f'\n원화 문제 행: {len(problems)}개')
for r in problems[:5]:
    print(f'  {r}')

print(f'\n환전 샘플 (달러금액 확인):')
for r in fx_sample:
    print(f'  거래구분={r[1]}, 환율={r[3]}, 거래대금={r[5]}, 달러금액={r[14]}, 잔고={r[12]}, 잔액={r[13]}')

# 달러 시트 검증
ws2 = wb['달러 거래내역']
headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column+1)]
print('\n달러 헤더:', headers2)

problems2 = []
fx_dollar = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    krw = row[12]
    usd = row[13]
    거래구분 = row[1]
    if isinstance(krw, str) or isinstance(usd, str):
        problems2.append(row)
    if 거래구분 and '환전' in str(거래구분) and len(fx_dollar) < 5:
        fx_dollar.append(row)

print(f'\n달러 문제 행: {len(problems2)}개')
for r in problems2[:5]:
    print(f'  {r}')

print(f'\n달러 환전 샘플:')
for r in fx_dollar:
    print(f'  거래구분={r[1]}, 환율={r[3]}, 거래대금={r[5]}, 잔액KRW={r[12]}, 잔액USD={r[13]}')
