"""
신한 종합거래내역 PDF/XLS 및 토스 거래내역 PDF → Excel 변환
"""
import re
import fitz
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from bs4 import BeautifulSoup
import os

# ──────────────────────────────────────────────
# 공통 유틸
# ──────────────────────────────────────────────
def get_all_lines(path, password=None):
    doc = fitz.open(path)
    if password:
        doc.authenticate(password)
    lines = []
    for page in doc.pages():
        h = page.get_text('html')
        soup = BeautifulSoup(h, 'html.parser')
        for p in soup.find_all('p'):
            t = p.get_text().strip().replace('\xa0', ' ')
            if t:
                lines.append(t)
    return lines

def to_num(val):
    if val is None:
        return 0
    s = str(val).replace(',', '').strip()
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return val

def split_combined(field):
    """'123,456 텍스트' → (int, '텍스트')  |  '123,456' → (int, '')"""
    m = re.match(r'^([\d,]+(?:\.\d+)?)\s+(.*)', str(field).strip())
    if m:
        try:
            num = int(m.group(1).replace(',', ''))
        except ValueError:
            num = float(m.group(1).replace(',', ''))
        return num, m.group(2).strip()
    try:
        return int(str(field).replace(',', '')), ''
    except ValueError:
        return field, ''

def make_excel(wb, ws, headers, rows, col_widths=None):
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = hdr_align; c.border = border

    NUM_TYPES = (int, float)
    r_align = Alignment(horizontal='right', vertical='center')
    c_align = Alignment(horizontal='center', vertical='center')
    l_align = Alignment(vertical='center')

    for ri, row in enumerate(rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if fill:
                c.fill = fill
            c.border = border
            if isinstance(val, NUM_TYPES):
                c.alignment = r_align
                c.number_format = '#,##0' if isinstance(val, int) else '#,##0.######'
            elif ci == 1:
                c.alignment = c_align
            else:
                c.alignment = l_align

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).column_letter}1'

    if col_widths:
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = col_widths.get(h, 12)


# ══════════════════════════════════════════════
# 신한 PDF 파서 (3행 × 9컬럼 구조)
# ══════════════════════════════════════════════
LAST_HEADER_COL = '상대계좌명'
FOOTER_RE = re.compile(r'^\d{4}년 \d{2}월 \d{2}일$')
DATE_RE_SHINHAN = re.compile(r'^\d{4}-\d{2}-\d{2}$')

def filter_shinhan(lines):
    data = []
    i = 0
    while i < len(lines):
        if lines[i] == LAST_HEADER_COL:
            i += 1
            while i < len(lines):
                if FOOTER_RE.match(lines[i]) or lines[i] == LAST_HEADER_COL:
                    break
                data.append(lines[i])
                i += 1
        else:
            i += 1
    return data

def is_trade_type(s):
    return bool(s) and not s[0].isdigit() and ',' not in s and len(s) > 1

def find_row2_start(chunk):
    for i in range(2, min(len(chunk)-1, 12)):
        try:
            n = int(chunk[i])
            if 1 <= n <= 99 and is_trade_type(chunk[i+1]):
                return i
        except ValueError:
            pass
    return None

def parse_shinhan_chunk(chunk):
    row2 = find_row2_start(chunk)
    if row2 is None or row2 < 2:
        return None

    r1 = chunk[:row2]
    r2 = chunk[row2:row2+7]
    r3 = chunk[row2+7:]

    if len(r2) < 7 or len(r3) < 4:
        return None

    total_pay, counterpart = split_combined(r1[-1])
    deferred_over, channel = split_combined(r2[6])
    balance_val, acct_name = split_combined(r3[-1])
    settlement = r3[-2] if len(r3) >= 2 else '0'

    return {
        '거래일':   r1[0],
        '종목명':   r1[1] if len(r1) > 1 else '',
        '상품구분': r3[0],
        '거래구분': r2[1],
        '단가/환율': to_num(r1[2]) if len(r1) > 2 else 0,
        '수량/외화': to_num(r2[2]),
        '수수료':   to_num(r1[3]) if len(r1) > 3 else 0,
        '소득세':   to_num(r1[4]) if len(r1) > 4 else 0,
        '거래세':   to_num(r2[3]),
        '지방소득세': to_num(r2[4]),
        '신용금액': to_num(r1[5]) if len(r1) > 5 else 0,
        '신용이자': to_num(r2[5]),
        '총변제금': to_num(total_pay),
        '상대처':   counterpart,
        '미수연체료': to_num(deferred_over),
        '거래채널': channel,
        '정산금액': to_num(settlement),
        '예수금잔고': to_num(balance_val),
        '상대계좌명': acct_name,
        '거래순번': to_num(r2[0]),
    }

SHINHAN_COLS = [
    '거래일','종목명','상품구분','거래구분',
    '단가/환율','수량/외화','정산금액','수수료',
    '소득세','거래세','지방소득세','신용금액',
    '신용이자','미수연체료','총변제금','상대처',
    '예수금잔고','상대계좌명','거래채널','거래순번'
]
SHINHAN_WIDTHS = {
    '거래일':12,'종목명':22,'상품구분':14,'거래구분':16,
    '단가/환율':13,'수량/외화':10,'정산금액':14,'수수료':9,
    '소득세':9,'거래세':9,'지방소득세':9,'신용금액':10,
    '신용이자':9,'미수연체료':10,'총변제금':11,'상대처':14,
    '예수금잔고':15,'상대계좌명':14,'거래채널':22,'거래순번':8
}

def convert_shinhan(pdf_path, out_path):
    lines = get_all_lines(pdf_path)
    data = filter_shinhan(lines)
    date_idx = [i for i, l in enumerate(data) if DATE_RE_SHINHAN.match(l)]

    transactions = []
    for k, start in enumerate(date_idx):
        end = date_idx[k+1] if k+1 < len(date_idx) else len(data)
        tx = parse_shinhan_chunk(data[start:end])
        if tx:
            transactions.append([tx.get(c, '') for c in SHINHAN_COLS])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '거래내역'
    make_excel(wb, ws, SHINHAN_COLS, transactions, SHINHAN_WIDTHS)
    wb.save(out_path)
    return len(transactions)


# ══════════════════════════════════════════════
# 토스 PDF 파서 (원화 + 달러 2개 시트)
# ══════════════════════════════════════════════
DATE_RE_TOSS = re.compile(r'^\d{4}\.\d{2}\.\d{2}$')
TOSS_HEADER_COLS = {'거래일자','거래구분','종목명(종목코드)','환율','거래수량',
                    '거래대금','정산금액','단가','수수료','거래세','제세금',
                    '변제/연체합','잔고','잔액'}
TOSS_SKIP_LINE = re.compile(r'^\d+\s*/\s*\d+$')   # page numbers  "1 / 12"
USD_LINE = re.compile(r'^\(\$\s')

def split_toss_sections(lines):
    """원화 / 달러 두 섹션으로 분리"""
    won_lines, dollar_lines = [], []
    section = None
    for l in lines:
        if l == '원화 거래내역':
            section = 'won'
        elif l == '달러 거래내역':
            section = 'dollar'
        elif section == 'won':
            won_lines.append(l)
        elif section == 'dollar':
            dollar_lines.append(l)
    return won_lines, dollar_lines

def filter_toss_section(lines):
    """헤더/푸터/페이지번호 제거"""
    result = []
    for l in lines:
        if l in TOSS_HEADER_COLS:
            continue
        if TOSS_SKIP_LINE.match(l):
            continue
        if re.match(r'^\d{4}년 \d{1,2}월 \d{1,2}일', l):  # footer date
            continue
        if l == '발급일자':
            continue
        result.append(l)
    return result

# ── 원화 섹션 파서 ──────────────────────────
def parse_won_chunk(chunk, has_settlement=True):
    """
    정산금액 있을 때 (신규 PDF):
      [0] date  [1] type  [2] name  [3] qty
      [4] 거래대금  [5] 정산금액  [6] 단가  [7] 수수료
      [8] 거래세  [9] 제세금  [10] 변제
      [11] "잔고 잔액"(합산) OR [11] 잔고  [12] 잔액

    정산금액 없을 때 (구 PDF):
      [0] date  [1] type  [2] name  [3] qty
      [4] 거래대금  [5] 단가  [6] 수수료
      [7] 거래세  [8] 제세금  [9] 변제
      [10] 잔고  [11] 잔액
    """
    if len(chunk) < 10:
        return None

    date = chunk[0]
    trade_type = chunk[1]

    name = chunk[2]
    if re.match(r'^[\d,]+\.\d+$', name.replace(' ', '')):
        exchange_rate = name.replace(',', '')
        name = ''
        qty_idx = 3
    else:
        exchange_rate = '1'
        qty_idx = 3

    if qty_idx >= len(chunk):
        return None

    qty = chunk[qty_idx]
    body_start = qty_idx + 1
    n_fields = 7 if has_settlement else 6
    n_remaining = len(chunk) - body_start

    if n_remaining == n_fields + 1:
        # 마지막 항목이 "잔고 잔액" 합산
        fields = chunk[body_start:body_start + n_fields]
        balance_raw = chunk[body_start + n_fields]
        bal_str = str(balance_raw).strip()
        parts = bal_str.split()
        if len(parts) == 2:
            stock_bal_val = to_num(parts[0])
            cash_val = to_num(parts[1])
        else:
            stock_bal_val = 0
            cash_val = to_num(balance_raw)
    elif n_remaining == n_fields + 2:
        # 잔고, 잔액 분리
        fields = chunk[body_start:body_start + n_fields]
        stock_bal_val = to_num(chunk[body_start + n_fields])
        cash_val = to_num(chunk[body_start + n_fields + 1])
    else:
        # 폴백
        fields = chunk[body_start:-2] if len(chunk) > body_start + 2 else []
        stock_bal_val = to_num(chunk[-2]) if len(chunk) >= 2 else 0
        cash_val = to_num(chunk[-1])

    def safe_field(idx):
        return to_num(fields[idx]) if idx < len(fields) else 0

    if has_settlement:
        return {
            '거래일자': date,
            '거래구분': trade_type,
            '종목명':   name,
            '환율':     to_num(exchange_rate),
            '거래수량': to_num(qty),
            '거래대금': safe_field(0),
            '정산금액': safe_field(1),
            '단가':     safe_field(2),
            '수수료':   safe_field(3),
            '거래세':   safe_field(4),
            '제세금':   safe_field(5),
            '변제/연체합': safe_field(6),
            '잔고':     stock_bal_val,
            '잔액':     cash_val,
        }
    else:
        return {
            '거래일자': date,
            '거래구분': trade_type,
            '종목명':   name,
            '환율':     to_num(exchange_rate),
            '거래수량': to_num(qty),
            '거래대금': safe_field(0),
            '정산금액': 0,
            '단가':     safe_field(1),
            '수수료':   safe_field(2),
            '거래세':   safe_field(3),
            '제세금':   safe_field(4),
            '변제/연체합': safe_field(5),
            '잔고':     stock_bal_val,
            '잔액':     cash_val,
        }

# ── 달러 섹션 파서 ──────────────────────────
def filter_usd_lines(lines):
    """달러 금액 행 ($ ...) 제거 후 KRW만 유지"""
    return [l for l in lines if not USD_LINE.match(l)]

def parse_dollar_chunk(chunk, has_settlement=True):
    """
    USD lines filtered out.
    정산금액 있을 때: tail 8 items = [거래대금, 정산금액, 단가, 수수료, 제세금, 변제, 잔고, 잔액]
    정산금액 없을 때: tail 7 items = [거래대금, 단가, 수수료, 제세금, 변제, 잔고, 잔액]
    """
    if len(chunk) < 6:
        return None

    date = chunk[0]
    trade_type = chunk[1]

    # Work from back
    tail_size = 8 if has_settlement else 7
    tail = chunk[-tail_size:]
    if has_settlement:
        거래대금 = to_num(tail[0])
        정산금액 = to_num(tail[1])
        단가 = to_num(tail[2])
        수수료 = to_num(tail[3])
        제세금 = to_num(tail[4])
        변제 = to_num(tail[5])
        잔고 = tail[6]
        잔액 = to_num(tail[7])
    else:
        거래대금 = to_num(tail[0])
        정산금액 = 0
        단가 = to_num(tail[1])
        수수료 = to_num(tail[2])
        제세금 = to_num(tail[3])
        변제 = to_num(tail[4])
        잔고 = tail[5]
        잔액 = to_num(tail[6])

    try:
        잔고 = float(str(잔고).replace(',', ''))
        if 잔고 == int(잔고):
            잔고 = int(잔고)
    except (ValueError, TypeError):
        잔고 = 0

    # Front: date, type, then name/rate/qty
    front = chunk[2:-tail_size]
    name_parts = []
    exchange_rate = '1'
    qty = '0'
    i = 0
    while i < len(front):
        item = front[i]
        # ISIN in parens
        if re.match(r'^\([A-Z0-9]+\)$', item):
            name_parts.append(item[1:-1])
            i += 1
        # Exchange rate: starts with digit, has decimal point, or combined "rate qty"
        elif re.match(r'^[\d,]+\.\d+', item.replace(' ', '')):
            parts = item.split()
            exchange_rate = parts[0].replace(',', '')
            if len(parts) >= 2:
                qty = parts[1]
            i += 1
            if len(parts) < 2 and i < len(front):
                qty = front[i]
                i += 1
            break
        else:
            name_parts.append(item)
            i += 1

    name = ' '.join(name_parts)

    return {
        '거래일자': date,
        '거래구분': trade_type,
        '종목명':   name,
        '환율':     to_num(exchange_rate),
        '거래수량': to_num(qty),
        '거래대금': 거래대금,
        '정산금액': 정산금액,
        '단가':     단가,
        '수수료':   수수료,
        '제세금':   제세금,
        '변제/연체합': 변제,
        '잔고':     잔고,
        '잔액':     잔액,
    }

def convert_toss(pdf_path, out_path):
    lines = get_all_lines(pdf_path, password='840313')
    won_raw, dollar_raw = split_toss_sections(lines)

    # PDF별로 정산금액 컬럼 존재 여부 감지
    has_settlement_won = '정산금액' in won_raw
    has_settlement_dollar = '정산금액' in dollar_raw

    won_data = filter_toss_section(won_raw)
    dollar_data_raw = filter_toss_section(dollar_raw)
    dollar_data = filter_usd_lines(dollar_data_raw)

    # Parse 원화
    won_dates = [i for i, l in enumerate(won_data) if DATE_RE_TOSS.match(l)]
    won_txs = []
    for k, start in enumerate(won_dates):
        end = won_dates[k+1] if k+1 < len(won_dates) else len(won_data)
        tx = parse_won_chunk(won_data[start:end], has_settlement=has_settlement_won)
        if tx:
            won_txs.append(tx)

    # Parse 달러
    dollar_dates = [i for i, l in enumerate(dollar_data) if DATE_RE_TOSS.match(l)]
    dollar_txs = []
    for k, start in enumerate(dollar_dates):
        end = dollar_dates[k+1] if k+1 < len(dollar_dates) else len(dollar_data)
        tx = parse_dollar_chunk(dollar_data[start:end], has_settlement=has_settlement_dollar)
        if tx:
            dollar_txs.append(tx)

    TOSS_COLS_WON = ['거래일자','거래구분','종목명','환율','거래수량',
                     '거래대금','정산금액','단가','수수료','거래세','제세금',
                     '변제/연체합','잔고','잔액']
    TOSS_COLS_DOLLAR = ['거래일자','거래구분','종목명','환율','거래수량',
                        '거래대금','정산금액','단가','수수료','제세금',
                        '변제/연체합','잔고','잔액']
    TOSS_WIDTHS = {
        '거래일자':12,'거래구분':16,'종목명':28,'환율':12,'거래수량':10,
        '거래대금':15,'정산금액':15,'단가':12,'수수료':10,'거래세':10,
        '제세금':10,'변제/연체합':12,'잔고':12,'잔액':16
    }

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = '원화 거래내역'
    won_rows = [[tx.get(c, '') for c in TOSS_COLS_WON] for tx in won_txs]
    make_excel(wb, ws1, TOSS_COLS_WON, won_rows, TOSS_WIDTHS)

    ws2 = wb.create_sheet('달러 거래내역')
    dollar_rows = [[tx.get(c, '') for c in TOSS_COLS_DOLLAR] for tx in dollar_txs]
    make_excel(wb, ws2, TOSS_COLS_DOLLAR, dollar_rows, TOSS_WIDTHS)

    wb.save(out_path)
    return len(won_txs), len(dollar_txs)


# ══════════════════════════════════════════════
# 신한 XLS 파서 (HTML 형식 .xls, EUC-KR)
# ══════════════════════════════════════════════
XLS_COLS = [
    '거래일', '거래구분', '구분', '종목코드', '종목명',
    '단가', '수량', '거래대금', '수수료', '제세금',
    '신용/대출금', '신용/대출이자', '미수발생/변제', '연체료', '과표금액', '예탁금이용료',
    '변동금액', '최종금액', '거래채널', '상품구분'
]
XLS_WIDTHS = {
    '거래일': 12, '거래구분': 10, '구분': 18, '종목코드': 10, '종목명': 22,
    '단가': 13, '수량': 10, '거래대금': 14, '수수료': 9, '제세금': 9,
    '신용/대출금': 12, '신용/대출이자': 12, '미수발생/변제': 12, '연체료': 9, '과표금액': 9, '예탁금이용료': 10,
    '변동금액': 13, '최종금액': 13, '거래채널': 20, '상품구분': 8
}

def xls_to_num(s):
    """'47,400.00000' → 47400  |  '1,460.84' → 1460.84"""
    s = str(s).replace(',', '').strip()
    if not s:
        return 0
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return s

def convert_shinhan_xls(xls_path, out_path):
    with open(xls_path, 'rb') as f:
        raw = f.read()
    content = raw.decode('euc-kr', errors='replace')
    soup = BeautifulSoup(content, 'html.parser')

    data_table = next((t for t in soup.find_all('table') if t.find('th')), None)
    if not data_table:
        raise ValueError("데이터 테이블을 찾을 수 없습니다")

    data_rows = [
        [td.get_text(strip=True) for td in row.find_all('td')]
        for row in data_table.find_all('tr')
        if not row.find('th')
    ]
    data_rows = [r for r in data_rows if len(r) >= 12]

    DATE_RE_XLS = re.compile(r'^\d{4}\.\d{2}\.\d{2}$')
    transactions = []
    i = 0
    while i + 1 < len(data_rows):
        r1, r2 = data_rows[i], data_rows[i + 1]
        if DATE_RE_XLS.match(r1[0]):
            tx = {
                '거래일':        r1[0],
                '구분':          r1[1],
                '종목코드':      r1[2],
                '수량':          xls_to_num(r1[3]),
                '거래대금':      xls_to_num(r1[4]),
                '수수료':        xls_to_num(r1[5]),
                '미수발생/변제': xls_to_num(r1[6]),
                '과표금액':      xls_to_num(r1[7]),
                '연체료':        xls_to_num(r1[8]),
                '변동금액':      xls_to_num(r1[9]),
                '상품구분':      r2[0],
                '거래구분':      r2[1],
                '종목명':        r2[2],
                '단가':          xls_to_num(r2[3]),
                '신용/대출금':   xls_to_num(r2[4]),
                '제세금':        xls_to_num(r2[5]),
                '신용/대출이자': xls_to_num(r2[6]),
                '예탁금이용료':  xls_to_num(r2[7]),
                '거래채널':      r2[8].strip(),
                '최종금액':      xls_to_num(r2[9]),
            }
            transactions.append([tx.get(c, '') for c in XLS_COLS])
            i += 2
        else:
            i += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '거래내역'
    make_excel(wb, ws, XLS_COLS, transactions, XLS_WIDTHS)
    wb.save(out_path)
    return len(transactions)


# ══════════════════════════════════════════════
# 실행 (폴더 내 파일 자동 탐색)
# ══════════════════════════════════════════════
if __name__ == '__main__':
    base = os.path.dirname(os.path.abspath(__file__))

    for fname in sorted(os.listdir(base)):
        ext = os.path.splitext(fname)[1].lower()
        if ext not in ('.pdf', '.xls'):
            continue

        fpath = os.path.join(base, fname)
        out_path = os.path.join(base, os.path.splitext(fname)[0] + '.xlsx')

        if os.path.exists(out_path):
            print(f'SKIP {fname}  (출력 파일 이미 존재)')
            continue

        try:
            if ext == '.xls':
                count = convert_shinhan_xls(fpath, out_path)
                print(f'OK   {fname} → {os.path.basename(out_path)}  ({count}건)')

            elif '토스' in fname or fname.startswith('거래내역서'):
                w, d = convert_toss(fpath, out_path)
                print(f'OK   {fname} → {os.path.basename(out_path)}  (원화 {w}건 / 달러 {d}건)')

            else:
                count = convert_shinhan(fpath, out_path)
                print(f'OK   {fname} → {os.path.basename(out_path)}  ({count}건)')

        except Exception as e:
            import traceback; traceback.print_exc()
            print(f'ERR  {fname}: {e}')
