"""
토스거래1~4.pdf → 토스_통합거래내역.xlsx
- 정산금액 컬럼 유무(has_settlement)를 PDF별로 자동 감지
- 달러 섹션 환전 거래는 USD 금액을 거래대금으로 표기
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import fitz
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
PDF_FILES = [f'토스거래{i}.pdf' for i in range(1, 5)]   # 1~4
OUT_PATH = os.path.join(BASE, '토스_통합거래내역.xlsx')
PASSWORD = '840313'

DATE_RE = re.compile(r'^\d{4}\.\d{2}\.\d{2}$')
USD_LINE = re.compile(r'^\(\$\s')

# ── 공통 유틸 ──────────────────────────────────
def get_all_lines(path, password=None):
    doc = fitz.open(path)
    if password:
        doc.authenticate(password)
    lines = []
    for page in doc.pages():
        h = page.get_text('html')
        soup = BeautifulSoup(h, 'html.parser')
        for p in soup.find_all('p'):
            t = p.get_text().strip().replace('\xa0', ' ')
            if t:
                lines.append(t)
    return lines

def to_num(val):
    if val is None:
        return 0
    s = str(val).replace(',', '').strip()
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return val

def make_excel(wb, ws, headers, rows, col_widths=None):
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = hdr_align; c.border = border

    NUM_TYPES = (int, float)
    r_align = Alignment(horizontal='right',  vertical='center')
    c_align = Alignment(horizontal='center', vertical='center')
    l_align = Alignment(vertical='center')

    for ri, row in enumerate(rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if fill:
                c.fill = fill
            c.border = border
            if isinstance(val, NUM_TYPES):
                c.alignment = r_align
                c.number_format = '#,##0' if isinstance(val, int) else '#,##0.######'
            elif ci == 1:
                c.alignment = c_align
            else:
                c.alignment = l_align

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).column_letter}1'

    if col_widths:
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = col_widths.get(h, 12)


# ── 섹션 분리 & 헤더/푸터 필터 ────────────────────
HEADER_COLS = {'거래일자','거래구분','종목명(종목코드)','환율','거래수량',
               '거래대금','정산금액','단가','수수료','거래세','제세금',
               '변제/연체합','잔고','잔액'}
PAGE_NUM    = re.compile(r'^\d+\s*/\s*\d+$')
FOOTER_DATE = re.compile(r'^\d{4}년 \d{1,2}월 \d{1,2}일')
PAGE_FOOTER = re.compile(r'^잔고구분|^: 원|^금액단위$|^발급일자$|^:$')

def split_toss_sections(lines):
    won_lines, dollar_lines = [], []
    section = None
    for l in lines:
        if l == '원화 거래내역':
            section = 'won'
        elif l == '달러 거래내역':
            section = 'dollar'
        elif section == 'won':
            won_lines.append(l)
        elif section == 'dollar':
            dollar_lines.append(l)
    return won_lines, dollar_lines

def filter_section(lines):
    result = []
    for l in lines:
        if l in HEADER_COLS:
            continue
        if PAGE_NUM.match(l):
            continue
        if FOOTER_DATE.match(l):
            continue
        if PAGE_FOOTER.match(l):
            continue
        result.append(l)
    return result

def split_balance(raw):
    s = str(raw).strip()
    parts = s.split()
    if len(parts) == 2:
        try:
            return to_num(parts[0]), to_num(parts[1])
        except Exception:
            pass
    return 0, to_num(s)

def _is_numeric(s):
    s = str(s).strip()
    parts = s.split()
    for p in parts:
        try:
            float(p.replace(',', ''))
        except ValueError:
            return False
    return bool(parts)

def strip_chunk_footer(chunk):
    c = list(chunk)
    while c and not _is_numeric(c[-1]):
        c.pop()
    return c


# ══════════════════════════════════════════════
# 원화 섹션 파서
# ══════════════════════════════════════════════
def parse_won_chunk(chunk, has_settlement=True):
    chunk = strip_chunk_footer(chunk)
    if len(chunk) < 6:
        return None

    date       = chunk[0]
    trade_type = chunk[1]
    name       = chunk[2]

    if re.match(r'^[\d,]+\.\d+$', name.replace(' ', '')):
        exchange_rate = to_num(name.replace(',', ''))
        name = ''
        qty_idx = 3
    else:
        exchange_rate = 1
        qty_idx = 3

    if qty_idx >= len(chunk):
        return None

    qty        = chunk[qty_idx]
    body_start = qty_idx + 1
    body_size  = len(chunk) - body_start
    n_fields   = 7 if has_settlement else 6

    if body_size == n_fields + 1:
        fields = chunk[body_start:body_start + n_fields]
        stock_bal, cash_val = split_balance(chunk[body_start + n_fields])
    elif body_size == n_fields + 2:
        fields = chunk[body_start:body_start + n_fields]
        stock_bal = to_num(chunk[body_start + n_fields])
        cash_val  = to_num(chunk[body_start + n_fields + 1])
    else:
        fields    = chunk[body_start:-2] if len(chunk) > body_start + 2 else []
        stock_bal = to_num(chunk[-2]) if len(chunk) >= 2 else 0
        cash_val  = to_num(chunk[-1])

    def sf(i):
        return to_num(fields[i]) if i < len(fields) else 0

    # 환전 거래: KRW 거래대금 → USD 환산값을 달러금액 컬럼에
    is_fx = '환전' in str(trade_type)
    if is_fx and exchange_rate and exchange_rate != 1:
        krw_amt    = sf(0)
        dollar_amt = round(krw_amt / exchange_rate, 4) if exchange_rate else 0
    else:
        dollar_amt = 0

    if has_settlement:
        return {
            '거래일자': date, '거래구분': trade_type, '종목명': name,
            '환율': exchange_rate, '거래수량': to_num(qty),
            '거래대금': sf(0), '정산금액': sf(1), '단가': sf(2),
            '수수료': sf(3), '거래세': sf(4), '제세금': sf(5), '변제/연체합': sf(6),
            '잔고': stock_bal, '잔액': cash_val, '달러금액(USD)': dollar_amt,
        }
    else:
        return {
            '거래일자': date, '거래구분': trade_type, '종목명': name,
            '환율': exchange_rate, '거래수량': to_num(qty),
            '거래대금': sf(0), '정산금액': 0, '단가': sf(1),
            '수수료': sf(2), '거래세': sf(3), '제세금': sf(4), '변제/연체합': sf(5),
            '잔고': stock_bal, '잔액': cash_val, '달러금액(USD)': dollar_amt,
        }


# ══════════════════════════════════════════════
# 달러 섹션 파서
# ══════════════════════════════════════════════
def _parse_usd_str(s):
    """'($ 13.87)' → 13.87  |  '($ 4.00)' → 4"""
    m = re.match(r'^\(\$\s*([\d,]+\.?\d*)\)', str(s))
    if not m:
        return 0
    v = float(m.group(1).replace(',', ''))
    return int(v) if v == int(v) else v

def extract_usd_balance(chunk):
    for item in reversed(chunk):
        m = re.search(r'\$\s*([\d,]+(?:\.\d+)?)', str(item))
        if m:
            try:
                v = float(m.group(1).replace(',', ''))
                return int(v) if v == int(v) else v
            except Exception:
                pass
    return 0

def parse_dollar_chunk(chunk_raw, has_settlement=True):
    """
    chunk_raw: USD 라인 포함된 원본 청크
    환전 거래 → 거래대금에 USD 금액 표기 (원화 금액 대신)
    """
    if len(chunk_raw) < 4:
        return None

    date       = chunk_raw[0]
    trade_type = chunk_raw[1]
    usd_balance = extract_usd_balance(chunk_raw)
    is_exchange = '환전' in str(trade_type)

    if is_exchange:
        # 구조: [date, type, 환율, qty=0, KRW거래대금, ($ USD거래대금), ...]
        # 마지막은 ($ USD잔액), 그 앞이 KRW 잔액
        exchange_rate = to_num(str(chunk_raw[2]).replace(',', ''))
        usd_amount    = _parse_usd_str(chunk_raw[5]) if len(chunk_raw) > 5 else 0
        krw_balance   = to_num(chunk_raw[-2])

        return {
            '거래일자': date, '거래구분': trade_type, '종목명': '',
            '환율': exchange_rate, '거래수량': 0,
            '거래대금': usd_amount,                          # USD 금액
            '정산금액': usd_amount if has_settlement else 0,
            '단가': 0, '수수료': 0, '제세금': 0, '변제/연체합': 0,
            '잔고': 0, '잔액(KRW)': krw_balance, '잔액(USD)': usd_balance,
        }

    # 일반 거래: USD 라인 제거 후 파싱
    chunk = [l for l in chunk_raw if not USD_LINE.match(str(l))]
    if len(chunk) < 4:
        return None

    tail_size    = 8 if has_settlement else 7
    tail         = chunk[-tail_size:]
    krw_balance  = to_num(tail[tail_size - 1])

    try:
        stock_bal = float(str(tail[tail_size - 2]).replace(',', ''))
        if stock_bal == int(stock_bal):
            stock_bal = int(stock_bal)
    except Exception:
        stock_bal = 0

    front = chunk[2:-tail_size]
    name_parts    = []
    exchange_rate = 1
    qty = '0'
    i = 0
    while i < len(front):
        item = front[i]
        if re.match(r'^\([A-Z0-9]+\)$', item):
            name_parts.append(item[1:-1])
            i += 1
        elif re.match(r'^[\d,]+\.\d+', item.replace(' ', '')):
            parts = item.split()
            exchange_rate = to_num(parts[0].replace(',', ''))
            if len(parts) >= 2:
                qty = parts[1]
            i += 1
            if len(parts) < 2 and i < len(front):
                qty = front[i]
                i += 1
            break
        else:
            name_parts.append(item)
            i += 1

    name = ' '.join(name_parts)

    def tf(idx):
        return to_num(tail[idx]) if idx < len(tail) else 0

    if has_settlement:
        return {
            '거래일자': date, '거래구분': trade_type, '종목명': name,
            '환율': exchange_rate, '거래수량': to_num(qty),
            '거래대금': tf(0), '정산금액': tf(1), '단가': tf(2),
            '수수료': tf(3), '제세금': tf(4), '변제/연체합': tf(5),
            '잔고': stock_bal, '잔액(KRW)': krw_balance, '잔액(USD)': usd_balance,
        }
    else:
        return {
            '거래일자': date, '거래구분': trade_type, '종목명': name,
            '환율': exchange_rate, '거래수량': to_num(qty),
            '거래대금': tf(0), '정산금액': 0, '단가': tf(1),
            '수수료': tf(2), '제세금': tf(3), '변제/연체합': tf(4),
            '잔고': stock_bal, '잔액(KRW)': krw_balance, '잔액(USD)': usd_balance,
        }


# ══════════════════════════════════════════════
# 컬럼 정의
# ══════════════════════════════════════════════
WON_COLS = ['거래일자','거래구분','종목명','환율','거래수량',
            '거래대금','정산금액','단가','수수료','거래세','제세금',
            '변제/연체합','잔고','잔액','달러금액(USD)']
DOLLAR_COLS = ['거래일자','거래구분','종목명','환율','거래수량',
               '거래대금','정산금액','단가','수수료','제세금',
               '변제/연체합','잔고','잔액(KRW)','잔액(USD)']
WIDTHS_WON = {
    '거래일자':12,'거래구분':16,'종목명':28,'환율':12,'거래수량':10,
    '거래대금':15,'정산금액':15,'단가':12,'수수료':10,'거래세':10,
    '제세금':10,'변제/연체합':12,'잔고':12,'잔액':16,'달러금액(USD)':14
}
WIDTHS_DOLLAR = {
    '거래일자':12,'거래구분':16,'종목명':30,'환율':12,'거래수량':12,
    '거래대금':15,'정산금액':15,'단가':12,'수수료':10,'제세금':10,
    '변제/연체합':12,'잔고':12,'잔액(KRW)':16,'잔액(USD)':14
}


# ══════════════════════════════════════════════
# 메인 처리
# ══════════════════════════════════════════════
all_won    = []
all_dollar = []

for fname in PDF_FILES:
    fpath = os.path.join(BASE, fname)
    if not os.path.exists(fpath):
        print(f'SKIP {fname}')
        continue

    print(f'처리: {fname}', end='')
    lines = get_all_lines(fpath, password=PASSWORD)
    won_raw, dollar_raw = split_toss_sections(lines)

    has_settlement_won    = '정산금액' in won_raw
    has_settlement_dollar = '정산금액' in dollar_raw

    # ── 원화 ──
    won_data  = filter_section(won_raw)
    won_dates = [i for i, l in enumerate(won_data) if DATE_RE.match(l)]
    for k, start in enumerate(won_dates):
        end = won_dates[k+1] if k+1 < len(won_dates) else len(won_data)
        tx = parse_won_chunk(won_data[start:end], has_settlement=has_settlement_won)
        if tx:
            all_won.append(tx)

    # ── 달러 (USD 라인 포함 상태로 청크 분리, 파서 내부에서 처리) ──
    dollar_data  = filter_section(dollar_raw)
    dollar_dates = [i for i, l in enumerate(dollar_data) if DATE_RE.match(l)]
    for k, start in enumerate(dollar_dates):
        end = dollar_dates[k+1] if k+1 < len(dollar_dates) else len(dollar_data)
        tx = parse_dollar_chunk(dollar_data[start:end], has_settlement=has_settlement_dollar)
        if tx:
            all_dollar.append(tx)

    print(f'  원화 {len(won_dates)}건 / 달러 {len(dollar_dates)}건')

# 날짜 정렬
all_won.sort(key=lambda x: x.get('거래일자', ''))
all_dollar.sort(key=lambda x: x.get('거래일자', ''))

wb   = openpyxl.Workbook()
ws1  = wb.active
ws1.title = '원화 거래내역'
won_rows = [[tx.get(c, '') for c in WON_COLS] for tx in all_won]
make_excel(wb, ws1, WON_COLS, won_rows, WIDTHS_WON)

ws2 = wb.create_sheet('달러 거래내역')
dollar_rows = [[tx.get(c, '') for c in DOLLAR_COLS] for tx in all_dollar]
make_excel(wb, ws2, DOLLAR_COLS, dollar_rows, WIDTHS_DOLLAR)

wb.save(OUT_PATH)
print(f'\n완료: {OUT_PATH}')
print(f'  원화 합계: {len(all_won)}건')
print(f'  달러 합계: {len(all_dollar)}건')
