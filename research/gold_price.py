"""
koreagoldx.co.kr 금 시세 → Daily_index.xlsx "금 시세" 시트 업데이트
사용법:
  python gold_price.py                        # 최근 2일 업데이트 (스케줄러용)
  python gold_price.py 2024-01-01 2026-04-29  # 날짜 범위 지정 (초기 로드용)
"""
import sys
import os
import logging
import asyncio
import aiohttp
from datetime import datetime, timedelta, date
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

# ── 설정 ──────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _get_excel_path() -> str:
    """현재 주차 Excel 파일 경로 반환."""
    KST    = pytz.timezone("Asia/Seoul")
    today  = datetime.now(KST).date()
    monday = today - timedelta(days=today.weekday())
    m = monday.month
    w = (monday.day - 1) // 7 + 1
    y = monday.year
    return os.path.join(BASE_DIR, f"Daily_index_{m}m{w}w{y}.xlsx")


EXCEL_PATH = _get_excel_path()
LOG_PATH   = os.path.join(BASE_DIR, "fetch_log.txt")
KST        = pytz.timezone("Asia/Seoul")
MAX_AUTO_LOOKBACK_DAYS = 21  # 자동 모드 공백 보정 상한 (비정상적으로 큰 공백은 경고 후 제한)

API_URL = "https://koreagoldx.co.kr/api/price/chart/list"
API_HEADERS = {
    "Content-Type": "application/json",
    "Referer": "https://koreagoldx.co.kr/price/gold",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
}

SHEET_NAME = "금 시세"

# 컬럼 정의: (API 필드, 헤더명, 열 너비, 숫자형식)
COLS = [
    ("date",    "날짜",          12, None),
    ("p_pure",  "순금 매입(원/g)", 16, "#,##0"),
    ("s_pure",  "순금 매도(원/g)", 16, "#,##0"),
    ("p_18k",   "18K 매입(원/g)", 16, "#,##0"),
    ("s_18k",   "18K 매도(원/g)", 16, "#,##0"),
    ("p_14k",   "14K 매입(원/g)", 16, "#,##0"),
    ("s_14k",   "14K 매도(원/g)", 16, "#,##0"),
    ("p_white", "백금 매입(원/g)", 16, "#,##0"),
    ("s_white", "백금 매도(원/g)", 16, "#,##0"),
    ("p_silver","은 매입(원/g)",  16, "#,##0"),
    ("s_silver","은 매도(원/g)",  16, "#,##0"),
]

# ── 로거 ──────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ── 스타일 ────────────────────────────────────────────────────────────────────
_HDR_FILL = PatternFill(start_color="7B3F00", end_color="7B3F00", fill_type="solid")
_HDR_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
_DAT_FONT = Font(name="맑은 고딕", size=10)
_CENTER   = Alignment(horizontal="center", vertical="center")
_LEFT     = Alignment(horizontal="left",   vertical="center")


# ── API 호출 ──────────────────────────────────────────────────────────────────

async def fetch_gold_prices(start_date: str, end_date: str) -> list[dict]:
    """
    start_date, end_date: "YYYY-MM-DD"
    반환: 날짜 오름차순 dict 리스트
    """
    ds = start_date.replace("-", ".")
    de = end_date.replace("-", ".")
    payload = {
        "srchDt": "all",
        "type": "gold",
        "dataDateStart": ds,
        "dataDateEnd": de,
    }

    async with aiohttp.ClientSession(headers=API_HEADERS) as session:
        async with session.post(
            API_URL, json=payload, timeout=aiohttp.ClientTimeout(total=30)
        ) as resp:
            data = await resp.json(content_type=None)

    # 응답: {"list": [...]} 구조, 날짜 형식: "2026-04-29 11:30:51"
    if isinstance(data, list):
        rows = data
    elif "list" in data:
        rows = data["list"]
    else:
        rows = data.get("data", [])

    # 하루에 여러 레코드 존재 → 날짜별 첫 번째(최신) 레코드만 사용
    results: list[dict] = []
    seen: set[str] = set()
    for row in rows:
        raw_date = row.get("date", "")
        date_str = raw_date[:10]  # "2026-04-29 11:30:51" → "2026-04-29"
        if date_str in seen:
            continue
        if not (start_date <= date_str <= end_date):
            continue
        seen.add(date_str)

        def _int(key: str):
            v = row.get(key)
            try:
                return int(float(v)) if v not in (None, "", "0") else None
            except (ValueError, TypeError):
                return None

        results.append({
            "date":     date_str,
            "p_pure":   _int("p_pure"),
            "s_pure":   _int("s_pure"),
            "p_18k":    _int("p_18k"),
            "s_18k":    _int("s_18k"),
            "p_14k":    _int("p_14k"),
            "s_14k":    _int("s_14k"),
            "p_white":  _int("p_white"),
            "s_white":  _int("s_white"),
            "p_silver": _int("p_silver"),
            "s_silver": _int("s_silver"),
        })

    results.sort(key=lambda x: x["date"])
    return results


# ── 워크북 관리 ───────────────────────────────────────────────────────────────

def _setup_sheet(ws) -> None:
    for col_idx, (_, label, width, _fmt) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def _latest_existing_date(path: str) -> str | None:
    """엑셀에 이미 저장된 금 시세 중 가장 최근 날짜(YYYY-MM-DD) 반환.
    파일/시트가 없거나 잠겨 있으면 None (호출자가 fallback 처리)."""
    if not os.path.exists(path):
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return None
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[SHEET_NAME]
    dates = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
    wb.close()
    return max(str(d)[:10] for d in dates) if dates else None


def load_workbook_state(path: str):
    """기존 파일 로드 or 신규 생성. 기존 날짜 집합 반환."""
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _setup_sheet(ws)
            existing_dates: set[str] = set()
        else:
            ws = wb[SHEET_NAME]
            existing_dates = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = row[0]
                if d:
                    existing_dates.add(str(d)[:10])
        return wb, ws, existing_dates
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _setup_sheet(ws)
        return wb, ws, set()


def insert_new_prices(ws, prices: list[dict], existing_dates: set[str]) -> int:
    """신규 데이터를 헤더 바로 아래 삽입 (최신이 위)."""
    new = [p for p in prices if p["date"] not in existing_dates]
    if not new:
        return 0

    new.sort(key=lambda x: x["date"], reverse=True)
    ws.insert_rows(2, amount=len(new))

    for i, p in enumerate(new, 2):
        for col_idx, (field, _, _, fmt) in enumerate(COLS, 1):
            val = p.get(field)
            c = ws.cell(row=i, column=col_idx, value=val)
            c.font = _DAT_FONT
            c.alignment = _CENTER if col_idx == 1 else _LEFT
            if fmt and val is not None:
                c.number_format = fmt
        existing_dates.add(p["date"])

    return len(new)


# ── 진입점 ────────────────────────────────────────────────────────────────────

async def main(
    start_date: str | None = None,
    end_date: str | None = None,
    excel_path: str | None = None,
) -> None:
    """
    start_date/end_date/excel_path 를 직접 전달하면 '소급 모드'로 동작.
    미전달 시 CLI 인자 또는 자동(마지막 저장일~오늘) 모드.
    """
    _prog = (start_date is not None or excel_path is not None)
    _excel_path = excel_path or EXCEL_PATH

    if _prog and start_date and end_date:
        log.info(f"[금 시세] 소급 모드: {start_date} ~ {end_date} → {os.path.basename(_excel_path)}")
    elif not _prog and len(sys.argv) == 3:
        start_date, end_date = sys.argv[1], sys.argv[2]
        log.info(f"[금 시세] 날짜 범위: {start_date} ~ {end_date}")
    else:
        now_kst  = datetime.now(KST)
        end_date = now_kst.strftime("%Y-%m-%d")
        fallback_start = (now_kst - timedelta(days=1)).strftime("%Y-%m-%d")
        latest = _latest_existing_date(_excel_path)
        if latest:
            gap_days = (now_kst.date() - datetime.strptime(latest, "%Y-%m-%d").date()).days
            if gap_days > MAX_AUTO_LOOKBACK_DAYS:
                log.warning(
                    f"[금 시세] 마지막 저장일({latest})이 {gap_days}일 전 — "
                    f"비정상적으로 오래돼 {MAX_AUTO_LOOKBACK_DAYS}일로 제한 (수동 확인 필요)"
                )
                start_date = (now_kst - timedelta(days=MAX_AUTO_LOOKBACK_DAYS)).strftime("%Y-%m-%d")
            else:
                start_date = latest
                if gap_days > 1:
                    log.info(f"[금 시세] 마지막 저장일({latest}) 이후 {gap_days}일 공백 감지 → 자동 보정")
        else:
            start_date = fallback_start
        log.info(f"[금 시세] 자동 모드: {start_date} ~ {end_date}")

    prices = await fetch_gold_prices(start_date, end_date)
    log.info(f"수집 완료: {len(prices)}개")

    try:
        wb, ws, existing_dates = load_workbook_state(_excel_path)
    except PermissionError:
        log.warning(f"[금 시세] 엑셀 파일이 열려 있어 저장 불가: {_excel_path}")
        return

    added = insert_new_prices(ws, prices, existing_dates)

    try:
        wb.save(_excel_path)
        log.info(f"[금 시세] 저장 완료: +{added}개 신규 | 누적 {len(existing_dates)}개 | {_excel_path}")
    except PermissionError:
        log.warning(f"[금 시세] 엑셀 저장 실패 — 파일이 열려 있습니다: {_excel_path}")


if __name__ == "__main__":
    asyncio.run(main())
