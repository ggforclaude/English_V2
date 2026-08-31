"""
research/add_sheets.py
invest_2026.xlsx에 새 시트 3개를 추가합니다.
기존 시트는 변경하지 않습니다.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

FILE = os.path.join(os.path.dirname(__file__), "invest_2026.xlsx")

# ── 공통 스타일 ────────────────────────────────────────────────────────────────
def _hdr(ws, row, col, value, bg="366092", fg="FFFFFF", bold=True, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=size, name="맑은 고딕")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def _border_all(ws, min_row, min_col, max_row, max_col):
    thin = Side(style="thin", color="CCCCCC")
    for r in ws.iter_rows(min_row=min_row, min_col=min_col,
                          max_row=max_row, max_col=max_col):
        for c in r:
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def _num(ws, row, col, value=None, fmt="#,##0", bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def _label(ws, row, col, value, bold=False, indent=0, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="맑은 고딕", size=10)
    c.alignment = Alignment(vertical="center", indent=indent)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def _note(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(color="888888", italic=True, size=9, name="맑은 고딕")
    c.alignment = Alignment(vertical="center")
    return c

def _merge_title(ws, row, text, max_col, bg="1F497D"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=13, name="맑은 고딕")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


# ══════════════════════════════════════════════════════════════════════════════
# 1. 📊 순자산대시보드
# ══════════════════════════════════════════════════════════════════════════════
def make_dashboard(wb):
    name = "📊순자산대시보드"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    months = ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06",
              "2026-07","2026-08","2026-09","2026-10","2026-11","2026-12"]

    # ── 타이틀 ──────────────────────────────────────────────────────────────
    _merge_title(ws, 1, "📊 순자산 대시보드 (월별)", 15)
    _note(ws, 2, 1, "★ 파란색 셀에만 숫자를 입력하세요 (단위: 만원)")

    # ── 섹션 A: 자산 ─────────────────────────────────────────────────────────
    _merge_title(ws, 4, "▌자산 (Assets)", 15, bg="17375E")
    ws.row_dimensions[4].height = 22

    asset_rows = [
        ("국내주식 (신한/토스/한화)", "ING 시트 현재가 합산"),
        ("해외주식 (달러환산)", "ING 시트 해외종목"),
        ("연금계좌 (연금저축+IRP)", "연금&ISA 시트 평가액"),
        ("ISA 계좌", "연금&ISA 시트 평가액"),
        ("예적금·현금", "통장 잔액 합계"),
        ("아파트(분양가 기준)", "중도금 납입액 누계"),
        ("기타 부동산·실물자산", ""),
    ]
    _hdr(ws, 5, 1, "자산 항목", bg="2E75B6")
    _hdr(ws, 5, 2, "비고", bg="2E75B6")
    for ci, m in enumerate(months):
        _hdr(ws, 5, ci + 3, m, bg="2E75B6", size=9)

    for ri, (item, note) in enumerate(asset_rows):
        row = 6 + ri
        _label(ws, row, 1, item, indent=1)
        _note(ws, row, 2, note)
        for ci in range(12):
            _num(ws, row, ci + 3, bg="EEF3FA")
        ws.row_dimensions[row].height = 18

    total_asset_row = 6 + len(asset_rows)
    _label(ws, total_asset_row, 1, "자산 합계", bold=True, bg="BDD7EE")
    _note(ws, total_asset_row, 2, "SUM 자동계산")
    for ci in range(12):
        col = get_column_letter(ci + 3)
        c = ws.cell(row=total_asset_row, column=ci + 3,
                    value=f"=SUM({col}6:{col}{total_asset_row-1})")
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor="BDD7EE")
        c.font = Font(bold=True, name="맑은 고딕")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── 섹션 B: 부채 ─────────────────────────────────────────────────────────
    liab_start = total_asset_row + 2
    _merge_title(ws, liab_start, "▌부채 (Liabilities)", 15, bg="843C0C")
    ws.row_dimensions[liab_start].height = 22

    liab_rows = [
        ("전세자금 대출", "대출 시트 잔액"),
        ("약관 대출", "대출 시트 잔액"),
        ("아파트 중도금 (미납)", "26.5/11 · 27.5/11 · 28.5"),
        ("기타 부채", ""),
    ]
    _hdr(ws, liab_start + 1, 1, "부채 항목", bg="C55A11")
    _hdr(ws, liab_start + 1, 2, "비고", bg="C55A11")
    for ci, m in enumerate(months):
        _hdr(ws, liab_start + 1, ci + 3, m, bg="C55A11", size=9)

    for ri, (item, note) in enumerate(liab_rows):
        row = liab_start + 2 + ri
        _label(ws, row, 1, item, indent=1)
        _note(ws, row, 2, note)
        for ci in range(12):
            _num(ws, row, ci + 3, bg="FBE4D5")
        ws.row_dimensions[row].height = 18

    total_liab_row = liab_start + 2 + len(liab_rows)
    _label(ws, total_liab_row, 1, "부채 합계", bold=True, bg="F4B183")
    _note(ws, total_liab_row, 2, "SUM 자동계산")
    for ci in range(12):
        col = get_column_letter(ci + 3)
        c = ws.cell(row=total_liab_row, column=ci + 3,
                    value=f"=SUM({col}{liab_start+2}:{col}{total_liab_row-1})")
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor="F4B183")
        c.font = Font(bold=True, name="맑은 고딕")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── 섹션 C: 순자산 ────────────────────────────────────────────────────────
    net_row = total_liab_row + 2
    _merge_title(ws, net_row, "▌순자산 = 자산 합계 − 부채 합계", 15, bg="375623")
    ws.row_dimensions[net_row].height = 22

    for ci in range(12):
        col = get_column_letter(ci + 3)
        c = ws.cell(row=net_row + 1, column=ci + 3,
                    value=f"={col}{total_asset_row}-{col}{total_liab_row}")
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        c.font = Font(bold=True, size=11, name="맑은 고딕")
        c.alignment = Alignment(horizontal="right", vertical="center")
    _label(ws, net_row + 1, 1, "순자산", bold=True, bg="E2EFDA")
    ws.row_dimensions[net_row + 1].height = 22

    # 전월 대비 증감
    for ci in range(12):
        col = get_column_letter(ci + 3)
        if ci == 0:
            c = ws.cell(row=net_row + 2, column=ci + 3, value="")
        else:
            prev = get_column_letter(ci + 2)
            c = ws.cell(row=net_row + 2, column=ci + 3,
                        value=f"={col}{net_row+1}-{prev}{net_row+1}")
        c.number_format = '+#,##0;-#,##0;"-"'
        c.alignment = Alignment(horizontal="right", vertical="center")
    _label(ws, net_row + 2, 1, "전월 대비 증감", bg="F2F2F2")

    # ── 컬럼 너비 ─────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    for ci in range(12):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 11

    ws.freeze_panes = "C5"
    print(f"  ✓ '{name}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# 2. 💰 월별지출
# ══════════════════════════════════════════════════════════════════════════════
def make_expense(wb):
    name = "💰월별지출"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    months = ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06",
              "2026-07","2026-08","2026-09","2026-10","2026-11","2026-12"]

    _merge_title(ws, 1, "💰 월별 지출 관리표 (단위: 만원)", 15)
    _note(ws, 2, 1, "★ 숫자 입력 셀은 파란색 · 고정비는 매월 자동복사 권장")

    categories = [
        ("── 고정 지출 ──", None, "D9E1F2"),
        ("아파트 관리비", "관리비 시트", None),
        ("전기/가스/수도", "", None),
        ("통신비 (휴대폰·인터넷)", "", None),
        ("대출 이자 (전세+약관)", "대출 시트 이자 합계", None),
        ("보험료", "", None),
        ("교육비 (학원·책 등)", "", None),
        ("구독 서비스", "OTT·앱·클라우드", None),
        ("── 변동 지출 ──", None, "D9E1F2"),
        ("식비", "", None),
        ("외식·카페", "", None),
        ("교통비", "", None),
        ("의류·잡화", "", None),
        ("의료·건강", "", None),
        ("여가·여행·취미", "", None),
        ("육아·자녀 (유주·인우)", "", None),
        ("경조사·선물", "", None),
        ("기타", "", None),
        ("── 저축·투자 ──", None, "E2EFDA"),
        ("연금 납입 (ISA 포함)", "", None),
        ("주식 매수 (순유입)", "매수금 - 매도금", None),
        ("적금·예금 납입", "", None),
        ("아파트 중도금 납입", "26.5/11 · 27.5/11 · 28.5", None),
    ]

    # 헤더
    _hdr(ws, 4, 1, "카테고리", bg="2E75B6")
    _hdr(ws, 4, 2, "비고", bg="2E75B6")
    for ci, m in enumerate(months):
        _hdr(ws, 4, ci + 3, m, bg="2E75B6", size=9)
    _hdr(ws, 4, 15, "연간합계", bg="1F4E79")

    data_rows = []
    for item, note, section_bg in categories:
        row = ws.max_row + 1
        if section_bg:  # 섹션 구분 행
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=15)
            c = ws.cell(row=row, column=1, value=item)
            c.font = Font(bold=True, size=10, name="맑은 고딕")
            c.fill = PatternFill("solid", fgColor=section_bg)
            c.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[row].height = 18
        else:
            _label(ws, row, 1, item, indent=2)
            _note(ws, row, 2, note or "")
            for ci in range(12):
                _num(ws, row, ci + 3, bg="EEF3FA")
            # 연간합계 수식
            first_col = get_column_letter(3)
            last_col = get_column_letter(14)
            c = ws.cell(row=row, column=15,
                        value=f"=SUM(C{row}:N{row})")
            c.number_format = "#,##0"
            c.font = Font(bold=True, name="맑은 고딕")
            c.fill = PatternFill("solid", fgColor="DEEAF1")
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[row].height = 18
            data_rows.append(row)

    # 합계 행
    total_row = ws.max_row + 1
    _label(ws, total_row, 1, "월 지출 합계", bold=True, bg="BDD7EE")
    for ci in range(12):
        col = get_column_letter(ci + 3)
        formula_parts = [f"{col}{r}" for r in data_rows]
        c = ws.cell(row=total_row, column=ci + 3,
                    value="=" + "+".join(formula_parts))
        c.number_format = "#,##0"
        c.font = Font(bold=True, name="맑은 고딕")
        c.fill = PatternFill("solid", fgColor="BDD7EE")
        c.alignment = Alignment(horizontal="right", vertical="center")
    c = ws.cell(row=total_row, column=15,
                value=f"=SUM(C{total_row}:N{total_row})")
    c.number_format = "#,##0"
    c.font = Font(bold=True, name="맑은 고딕")
    c.fill = PatternFill("solid", fgColor="9DC3E6")
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[total_row].height = 20

    # ── 수입 입력 섹션 ──────────────────────────────────────────────────────
    income_start = total_row + 2
    _merge_title(ws, income_start, "▌월 수입 (세후)", 15, bg="375623")
    income_items = ["급여 (본인)", "배우자 급여", "임대·기타 수입", "투자 배당·이자"]
    for ri, item in enumerate(income_items):
        row = income_start + 1 + ri
        _label(ws, row, 1, item, indent=2)
        for ci in range(12):
            _num(ws, row, ci + 3, bg="E2EFDA")
        c = ws.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor="C6E0B4")
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 18

    inc_total = income_start + 1 + len(income_items)
    _label(ws, inc_total, 1, "수입 합계", bold=True, bg="C6E0B4")
    for ci in range(12):
        col = get_column_letter(ci + 3)
        c = ws.cell(row=inc_total, column=ci + 3,
                    value=f"=SUM({col}{income_start+1}:{col}{inc_total-1})")
        c.number_format = "#,##0"
        c.font = Font(bold=True, name="맑은 고딕")
        c.fill = PatternFill("solid", fgColor="C6E0B4")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # 잉여 = 수입 - 지출
    surplus_row = inc_total + 1
    _label(ws, surplus_row, 1, "잉여 (수입 - 지출)", bold=True, bg="FFF2CC")
    for ci in range(12):
        col = get_column_letter(ci + 3)
        c = ws.cell(row=surplus_row, column=ci + 3,
                    value=f"={col}{inc_total}-{col}{total_row}")
        c.number_format = '+#,##0;-#,##0;"-"'
        c.font = Font(bold=True, name="맑은 고딕")
        c.fill = PatternFill("solid", fgColor="FFF2CC")
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[surplus_row].height = 20

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    for ci in range(12):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 11
    ws.column_dimensions["O"].width = 11
    ws.freeze_panes = "C5"
    print(f"  ✓ '{name}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# 3. 📈 포트폴리오요약
# ══════════════════════════════════════════════════════════════════════════════
def make_portfolio(wb):
    name = "📈포트폴리오요약"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    _merge_title(ws, 1, "📈 포트폴리오 현황 요약 (매월 말 업데이트)", 12)
    _note(ws, 2, 1, "★ ING 시트 데이터를 월말에 이 시트에 복사해 스냅샷으로 보관하세요")

    # ── 섹션 1: 계좌별 보유현황 ───────────────────────────────────────────────
    _merge_title(ws, 4, "▌계좌별 보유 현황 (입력일: )", 12, bg="17375E")
    headers = ["종목명", "티커/코드", "계좌", "평균단가", "현재가",
               "보유수량", "매입금액", "평가금액", "손익", "수익률(%)"]
    for ci, h in enumerate(headers):
        _hdr(ws, 5, ci + 1, h, bg="2E75B6", size=10)

    accounts = [
        ("삼성전자", "005930", "신한증권"),
        ("NVDA", "NVDA", "토스증권"),
        ("카카오", "035720", "한화투자"),
        ("(직접 입력)", "", ""),
        ("(직접 입력)", "", ""),
        ("(직접 입력)", "", ""),
        ("(직접 입력)", "", ""),
        ("(직접 입력)", "", ""),
        ("(직접 입력)", "", ""),
        ("(직접 입력)", "", ""),
    ]
    for ri, (name_, ticker, acct) in enumerate(accounts):
        row = 6 + ri
        _label(ws, row, 1, name_)
        _label(ws, row, 2, ticker)
        _label(ws, row, 3, acct)
        for ci in range(3, 8):   # 평균단가~평가금액
            _num(ws, row, ci + 1, bg="EEF3FA")
        # 손익 = 평가금액 - 매입금액
        c = ws.cell(row=row, column=9,
                    value=f"=H{row}-G{row}" if ri < 3 else "=H{row}-G{row}".replace("{row}", str(row)))
        c.number_format = '+#,##0;-#,##0;"-"'
        c.alignment = Alignment(horizontal="right", vertical="center")
        # 수익률
        c2 = ws.cell(row=row, column=10,
                     value=f"=IF(G{row}>0,(H{row}-G{row})/G{row}*100,\"\")")
        c2.number_format = '+0.00%;-0.00%'
        c2.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 18

    # 소계
    sub_row = 6 + len(accounts)
    _label(ws, sub_row, 1, "합계", bold=True, bg="BDD7EE")
    c_buy = ws.cell(row=sub_row, column=7, value=f"=SUM(G6:G{sub_row-1})")
    c_buy.number_format = "#,##0"
    c_buy.font = Font(bold=True, name="맑은 고딕")
    c_buy.fill = PatternFill("solid", fgColor="BDD7EE")
    c_buy.alignment = Alignment(horizontal="right", vertical="center")
    c_cur = ws.cell(row=sub_row, column=8, value=f"=SUM(H6:H{sub_row-1})")
    c_cur.number_format = "#,##0"
    c_cur.font = Font(bold=True, name="맑은 고딕")
    c_cur.fill = PatternFill("solid", fgColor="BDD7EE")
    c_cur.alignment = Alignment(horizontal="right", vertical="center")
    c_pl = ws.cell(row=sub_row, column=9, value=f"=H{sub_row}-G{sub_row}")
    c_pl.number_format = '+#,##0;-#,##0;"-"'
    c_pl.font = Font(bold=True, name="맑은 고딕")
    c_pl.fill = PatternFill("solid", fgColor="BDD7EE")
    c_pl.alignment = Alignment(horizontal="right", vertical="center")
    c_rt = ws.cell(row=sub_row, column=10,
                   value=f"=IF(G{sub_row}>0,(H{sub_row}-G{sub_row})/G{sub_row}*100,\"\")")
    c_rt.number_format = '+0.00%;-0.00%'
    c_rt.font = Font(bold=True, name="맑은 고딕")
    c_rt.fill = PatternFill("solid", fgColor="BDD7EE")
    c_rt.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[sub_row].height = 20

    # ── 섹션 2: 계좌별 비중 요약 ────────────────────────────────────────────
    sec2 = sub_row + 2
    _merge_title(ws, sec2, "▌계좌별 비중 (자동계산)", 12, bg="17375E")
    acct_list = ["신한증권", "토스증권", "한화투자", "연금저축/IRP", "ISA"]
    _hdr(ws, sec2 + 1, 1, "계좌", bg="2E75B6")
    _hdr(ws, sec2 + 1, 2, "평가금액 (만원)", bg="2E75B6")
    _hdr(ws, sec2 + 1, 3, "비중 (%)", bg="2E75B6")
    _note(ws, sec2 + 1, 4, "← ING 시트 기준 수동 입력 또는 SUMIF 연결")
    for ri, acct in enumerate(acct_list):
        row = sec2 + 2 + ri
        _label(ws, row, 1, acct)
        _num(ws, row, 2, bg="EEF3FA")
        # 비중 = 해당 계좌 / 전체 합계
        c = ws.cell(row=row, column=3, value="")
        c.number_format = "0.0%"
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 18
    acct_total = sec2 + 2 + len(acct_list)
    _label(ws, acct_total, 1, "합계", bold=True, bg="C6E0B4")
    ct = ws.cell(row=acct_total, column=2,
                 value=f"=SUM(B{sec2+2}:B{acct_total-1})")
    ct.number_format = "#,##0"
    ct.font = Font(bold=True, name="맑은 고딕")
    ct.fill = PatternFill("solid", fgColor="C6E0B4")
    ct.alignment = Alignment(horizontal="right", vertical="center")
    # 비중 수식 채우기
    for ri in range(len(acct_list)):
        row = sec2 + 2 + ri
        c = ws.cell(row=row, column=3,
                    value=f"=IF(B{acct_total}>0,B{row}/B{acct_total},\"\")")
        c.number_format = "0.0%"
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── 섹션 3: 자산군 배분 가이드 ──────────────────────────────────────────
    sec3 = acct_total + 2
    _merge_title(ws, sec3, "▌자산배분 가이드 (목표 vs 실제)", 12, bg="375623")
    alloc_hdrs = ["자산군", "목표 비중(%)", "실제 비중(%)", "목표 금액(만)", "실제 금액(만)", "차이"]
    for ci, h in enumerate(alloc_hdrs):
        _hdr(ws, sec3 + 1, ci + 1, h, bg="548235", size=10)
    alloc_items = [
        ("국내주식", 30),
        ("해외주식", 30),
        ("연금/ISA", 20),
        ("현금·채권", 15),
        ("부동산·실물", 5),
    ]
    for ri, (item, pct) in enumerate(alloc_items):
        row = sec3 + 2 + ri
        _label(ws, row, 1, item)
        _num(ws, row, 2, value=pct)
        _num(ws, row, 3, bg="EEF3FA")
        _num(ws, row, 4, bg="F2F2F2")
        _num(ws, row, 5, bg="EEF3FA")
        c = ws.cell(row=row, column=6, value=f"=E{row}-D{row}")
        c.number_format = '+#,##0;-#,##0;"-"'
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 18

    _note(ws, sec3 + 2 + len(alloc_items), 1,
          "※ 목표 비중은 자유롭게 수정 · 실제 비중/금액은 위 계좌별 요약에서 계산 후 복사")

    # 컬럼 너비
    widths = [22, 14, 14, 10, 12, 12, 12, 12, 12, 12]
    for ci, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w
    ws.freeze_panes = "A6"
    print(f"  ✓ '📈포트폴리오요약' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# main
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"파일 열기: {FILE}")
    wb = load_workbook(FILE)
    print(f"기존 시트: {wb.sheetnames}")

    make_dashboard(wb)
    make_expense(wb)
    make_portfolio(wb)

    wb.save(FILE)
    print(f"\n✅ 저장 완료: {FILE}")
    print(f"최종 시트 목록: {wb.sheetnames}")
