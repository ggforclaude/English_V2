"""
KOSPI 업종별 + 미국 섹터 RSI 비교 분석기

파일 정책:
  - 매주 월요일 새 파일 생성: Daily_index_{월}m{주차}w{년}.xlsx
  - 대표종목1,2 : KOSPI 업종 시가총액 상위 2개 (월요일 갱신)
  - 대표종목3   : 직전 주 업종 내 주간 상승폭 1위 종목 (월요일 갱신)
  - 한국ETF     : 대표종목 1~3 편입 비중 최고 ETF (매일 갱신)

사용법:
  python sector_rsi.py          # 자동 (월요일=새 파일 생성, 그 외=일별 추가)
  python sector_rsi.py --init   # 2024-01-02부터 전체 초기 로드 (수동 실행용)
"""
import os, sys, json, logging
from datetime import datetime, timedelta, date
import pytz
import pandas as pd
import FinanceDataReader as fdr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from pykrx import stock as krx

# ── 설정 ─────────────────────────────────────────────────────────────────────
KST           = pytz.timezone("Asia/Seoul")
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
LOG_PATH      = os.path.join(BASE_DIR, "sector_rsi_log.txt")
TOPS_CACHE    = os.path.join(BASE_DIR, "sector_tops_cache.json")
RSI_PERIOD    = 14
WARMUP_DAYS   = 60
DISPLAY_START = date(2024, 1, 2)
MAX_DATA_ROWS = 3000

# ── 주차 파일명 ───────────────────────────────────────────────────────────────

def get_weekly_filepath(dt: date | None = None) -> str:
    """현재(또는 지정) 날짜 기준 주차 Excel 파일 경로 반환.
    예) 2026-05-04(월) → Daily_index_5m1w2026.xlsx
    """
    if dt is None:
        dt = datetime.now(KST).date()
    monday = dt - timedelta(days=dt.weekday())
    m = monday.month
    w = (monday.day - 1) // 7 + 1
    y = monday.year
    return os.path.join(BASE_DIR, f"Daily_index_{m}m{w}w{y}.xlsx")


# ── KOSPI 업종 지수 코드 ──────────────────────────────────────────────────────
SECTORS: dict[str, str] = {
    "1005": "음식료·담배",  "1006": "섬유·의류",   "1007": "종이·목재",
    "1008": "화학",        "1009": "제약",         "1010": "비금속",
    "1011": "금속",        "1012": "기계·장비",    "1013": "전기전자",
    "1014": "의료·정밀",   "1015": "운송장비",     "1016": "유통",
    "1017": "전기·가스",   "1018": "건설",         "1019": "운송·창고",
    "1020": "통신",        "1021": "금융",         "1024": "증권",
    "1025": "보험",        "1026": "일반서비스",   "1027": "제조",
    "1045": "부동산",      "1046": "IT서비스",     "1047": "오락·문화",
}

# ── 미국 SPDR 섹터 ETF ────────────────────────────────────────────────────────
US_SECTORS: dict[str, str] = {
    "XLK":  "기술(US)",        "XLF":  "금융(US)",
    "XLV":  "헬스케어(US)",    "XLE":  "에너지(US)",
    "XLB":  "소재(US)",        "XLI":  "산업재(US)",
    "XLY":  "경기소비재(US)",  "XLP":  "필수소비재(US)",
    "XLU":  "유틸리티(US)",    "XLC":  "통신서비스(US)",
    "XLRE": "부동산(US)",
}

# ── 한국 섹터 → 대응 미국 섹터 ───────────────────────────────────────────────
KR_TO_US: dict[str, str] = {
    "음식료·담배": "XLP", "섬유·의류":  "XLY", "종이·목재":  "XLB",
    "화학":        "XLB", "제약":        "XLV", "비금속":      "XLB",
    "금속":        "XLB", "기계·장비":  "XLI", "전기전자":    "XLK",
    "의료·정밀":  "XLV", "운송장비":   "XLI", "유통":        "XLY",
    "전기·가스":  "XLU", "건설":        "XLI", "운송·창고":  "XLI",
    "통신":        "XLC", "금융":        "XLF", "증권":        "XLF",
    "보험":        "XLF", "일반서비스": "XLI", "제조":        "XLI",
    "부동산":      "XLRE","IT서비스":   "XLK", "오락·문화":  "XLC",
}

# ── 한국 섹터 추종 ETF 후보 (ticker, 표시명) ──────────────────────────────────
# 섹터별 여러 후보를 정의. 매일 대표종목 편입 스코어로 최적 ETF 선택.
KR_ETF_CANDIDATES: dict[str, list[tuple[str, str]]] = {
    "음식료·담배": [("157450", "TIGER 음식료&생활용품")],
    "섬유·의류":   [],
    "종이·목재":   [],
    "화학":        [("117460", "KODEX 에너지화학")],
    "제약":        [("244580", "KODEX 바이오"), ("143850", "TIGER 헬스케어")],
    "비금속":      [],
    "금속":        [("139230", "KODEX 철강")],
    "기계·장비":   [],
    "전기전자":    [("091160", "KODEX 반도체")],
    "의료·정밀":   [("143850", "TIGER 헬스케어"), ("244580", "KODEX 바이오")],
    "운송장비":    [("091180", "KODEX 자동차")],
    "유통":        [],
    "전기·가스":   [],
    "건설":        [("104520", "KODEX 건설")],
    "운송·창고":   [],
    "통신":        [],
    "금융":        [("091170", "KODEX 은행")],
    "증권":        [("266410", "TIGER 증권")],
    "보험":        [],
    "일반서비스":  [],
    "제조":        [],
    "부동산":      [("329200", "TIGER 부동산인프라")],
    "IT서비스":    [],
    "오락·문화":   [("091230", "KODEX 미디어&엔터")],
}

# ── 스타일 ────────────────────────────────────────────────────────────────────
_HDR_FILL_KR  = PatternFill("solid", fgColor="1F4E79")
_HDR_FILL_US  = PatternFill("solid", fgColor="7B2C2C")
_HDR_FILL_INFO= PatternFill("solid", fgColor="375623")
_HDR_FONT     = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
_DAT_FONT     = Font(name="맑은 고딕", size=10)
_SUB_FONT     = Font(name="맑은 고딕", size=9, italic=True, color="444444")
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RED_FILL     = PatternFill("solid", fgColor="FFCCCC")
_GREEN_FILL   = PatternFill("solid", fgColor="CCFFCC")
_ORANGE_FILL  = PatternFill("solid", fgColor="FFE4B5")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ── RSI 계산 ─────────────────────────────────────────────────────────────────

def calc_rsi(close: pd.Series, period: int = RSI_PERIOD) -> pd.Series:
    """Wilder's RSI"""
    delta    = close.diff()
    gain     = delta.clip(lower=0)
    loss     = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, float("nan"))
    return (100 - (100 / (1 + rs))).round(2)


# ── 한국 업종 데이터 ──────────────────────────────────────────────────────────

def fetch_korea_rsi(start_str: str, end_str: str) -> pd.DataFrame:
    close_dict: dict[str, pd.Series] = {}
    for ticker, name in SECTORS.items():
        log.info(f"  [한국] {name} ({ticker})")
        try:
            df = krx.get_index_ohlcv(start_str, end_str, ticker)
            if df is not None and not df.empty:
                close_dict[name] = df["종가"].rename(name)
        except Exception as e:
            log.warning(f"    수집 오류: {e}")
    close_df = pd.DataFrame(close_dict)
    close_df.index = pd.to_datetime(close_df.index)
    close_df.sort_index(inplace=True)
    return close_df.apply(calc_rsi)


# ── 미국 섹터 데이터 ──────────────────────────────────────────────────────────

def fetch_us_rsi(start_str: str, end_str: str) -> pd.DataFrame:
    close_dict: dict[str, pd.Series] = {}
    for etf, label in US_SECTORS.items():
        log.info(f"  [미국] {label} ({etf})")
        try:
            df = fdr.DataReader(etf, start_str, end_str)
            if df is not None and not df.empty:
                close_dict[label] = df["Close"].rename(label)
        except Exception as e:
            log.warning(f"    수집 오류: {e}")
    close_df = pd.DataFrame(close_dict)
    close_df.index = pd.to_datetime(close_df.index)
    close_df.sort_index(inplace=True)
    return close_df.apply(calc_rsi)


# ── 날짜 통합 인덱스 ──────────────────────────────────────────────────────────

def align_dates(kr_rsi: pd.DataFrame, us_rsi: pd.DataFrame
                ) -> tuple[pd.DataFrame, pd.DataFrame]:
    common_idx = kr_rsi.index.union(us_rsi.index).sort_values()
    return kr_rsi.reindex(common_idx), us_rsi.reindex(common_idx)


# ── 대표종목 수집 (월요일) ────────────────────────────────────────────────────

def fetch_sector_caps_top2(today_str: str) -> dict[str, list[str]]:
    """업종별 시가총액 상위 2개 종목코드 반환."""
    log.info("  대표종목 1,2 (시총 top2) 수집 중...")
    try:
        caps = krx.get_market_cap(today_str, market="KOSPI")
    except Exception as e:
        log.warning(f"시가총액 로드 실패: {e}")
        return {name: ["-", "-"] for name in SECTORS.values()}

    result: dict[str, list[str]] = {}
    for ticker, name in SECTORS.items():
        try:
            constituents = krx.get_index_portfolio_deposit_file(ticker)
            top2 = caps[caps.index.isin(constituents)].nlargest(2, "시가총액")
            codes = list(top2.index[:2])
            while len(codes) < 2:
                codes.append("-")
            result[name] = codes
        except Exception as e:
            log.warning(f"  [{name}] 시총 top2 오류: {e}")
            result[name] = ["-", "-"]
    return result


def fetch_all_weekly_gainers(prev_mon_str: str, prev_fri_str: str) -> dict[str, tuple[str, str]]:
    """직전 주 업종별 최고 상승 종목 {업종명: (종목코드, 종목명)}.
    시장 전체 OHLCV를 두 날짜만 조회해 계산 (API 호출 최소화).
    """
    log.info("  대표종목3 (주간 최고 상승) 수집 중...")
    try:
        start_df = krx.get_market_ohlcv_by_ticker(prev_mon_str, market="KOSPI")
        end_df   = krx.get_market_ohlcv_by_ticker(prev_fri_str,  market="KOSPI")
    except Exception as e:
        log.warning(f"주간 시세 로드 실패: {e}")
        return {name: ("-", "-") for name in SECTORS.values()}

    # 상승폭 계산
    gains: dict[str, float] = {}
    for code in start_df.index:
        if code in end_df.index:
            try:
                s = float(start_df.loc[code, "종가"])
                e = float(end_df.loc[code, "종가"])
                if s > 0:
                    gains[code] = (e - s) / s
            except Exception:
                pass

    result: dict[str, tuple[str, str]] = {}
    for ticker, name in SECTORS.items():
        try:
            constituents = set(str(c) for c in krx.get_index_portfolio_deposit_file(ticker))
            sector_gains = {c: g for c, g in gains.items() if c in constituents}
            if sector_gains:
                best = max(sector_gains, key=sector_gains.get)
                best_name = krx.get_market_ticker_name(best)
                result[name] = (best, best_name)
            else:
                result[name] = ("-", "-")
        except Exception as e:
            log.warning(f"  [{name}] 주간 상승 종목 오류: {e}")
            result[name] = ("-", "-")

    return result


def build_monday_tops(today_str: str, prev_mon_str: str, prev_fri_str: str) -> dict[str, dict]:
    """월요일 업종별 대표종목 수집.
    Returns: {업종명: {"codes": [c1,c2,c3], "names": [n1,n2,n3]}}
    """
    top2_codes = fetch_sector_caps_top2(today_str)
    gainers    = fetch_all_weekly_gainers(prev_mon_str, prev_fri_str)

    tops: dict[str, dict] = {}
    for name in SECTORS.values():
        codes = list(top2_codes.get(name, ["-", "-"]))
        names = []
        for code in codes:
            if code == "-":
                names.append("-")
            else:
                try:
                    names.append(krx.get_market_ticker_name(code))
                except Exception:
                    names.append(code)

        g_code, g_name = gainers.get(name, ("-", "-"))
        codes.append(g_code)
        names.append(g_name)

        tops[name] = {"codes": codes, "names": names}

    return tops


# ── ETF 스코어링 (매일) ───────────────────────────────────────────────────────

def _score_etf(etf_ticker: str, stock_codes: list[str]) -> int:
    """ETF 보유 종목 중 대표종목이 몇 개 포함되는지 반환."""
    try:
        df = krx.get_etf_portfolio_deposit_file(etf_ticker)
        if df is None or df.empty:
            return 0
        holdings = set(str(c) for c in df.index)
        return sum(1 for code in stock_codes if code not in ("-", "") and code in holdings)
    except Exception:
        return 0


def build_sector_etfs(tops: dict[str, dict]) -> dict[str, str]:
    """업종별 대표종목 편입 비중 최고 ETF 선택. {업종명: "표시명(티커)"}"""
    log.info("  한국ETF 최적화 중...")
    result: dict[str, str] = {}
    for name in SECTORS.values():
        candidates = KR_ETF_CANDIDATES.get(name, [])
        if not candidates:
            result[name] = "-"
            continue

        codes = [c for c in tops.get(name, {}).get("codes", []) if c != "-"]

        if not codes or len(candidates) == 1:
            ticker, label = candidates[0]
            result[name] = f"{label}({ticker})"
            continue

        best_ticker, best_label = candidates[0]
        best_score = -1
        for ticker, label in candidates:
            score = _score_etf(ticker, codes)
            if score > best_score:
                best_score = score
                best_ticker, best_label = ticker, label

        result[name] = f"{best_label}({best_ticker})"

    return result


# ── tops 캐시 (일별 ETF 갱신용) ──────────────────────────────────────────────

def save_tops_cache(tops: dict) -> None:
    with open(TOPS_CACHE, "w", encoding="utf-8") as f:
        json.dump(tops, f, ensure_ascii=False, indent=2)


def load_tops_cache() -> dict | None:
    if os.path.exists(TOPS_CACHE):
        try:
            with open(TOPS_CACHE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return None


# ── Excel 스타일 헬퍼 ─────────────────────────────────────────────────────────

def _rsi_fill(val) -> PatternFill | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if val >= 70:  return _RED_FILL
    if val <= 30:  return _GREEN_FILL
    if val >= 60:  return _ORANGE_FILL
    return None


def _write_rsi_sheet(ws, rsi_df: pd.DataFrame, hdr_fill: PatternFill,
                     subrows: dict[str, list] | None = None) -> int:
    """RSI 데이터를 시트에 작성. 반환: 데이터 시작 행 번호."""
    ws.delete_rows(1, ws.max_row)
    sector_names = list(rsi_df.columns)
    n_sectors    = len(sector_names)

    ws.cell(1, 1, "날짜").fill = hdr_fill
    ws.cell(1, 1).font         = _HDR_FONT
    ws.cell(1, 1).alignment    = _CENTER
    ws.column_dimensions["A"].width = 12

    for ci, name in enumerate(sector_names, 2):
        c = ws.cell(1, ci, name)
        c.fill = hdr_fill; c.font = _HDR_FONT; c.alignment = _CENTER
        ws.column_dimensions[get_column_letter(ci)].width = 13

    ws.row_dimensions[1].height = 20

    cur_row = 2
    if subrows:
        _META_FILL = PatternFill("solid", fgColor="D9D9D9")
        for label, values in subrows.items():
            c = ws.cell(cur_row, 1, label)
            c.fill = _META_FILL; c.font = _SUB_FONT; c.alignment = _CENTER
            for ci, val in enumerate(values, 2):
                cell = ws.cell(cur_row, ci, val)
                cell.font      = _SUB_FONT
                cell.alignment = _CENTER
                cell.fill      = _META_FILL
            ws.row_dimensions[cur_row].height = 28
            cur_row += 1
        cur_row += 1  # 구분 빈 행

    data_start_row = cur_row
    ws.freeze_panes = f"B{data_start_row}"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_sectors + 1)}1"

    sorted_df = rsi_df.sort_index(ascending=False)
    for ri, (dt, row) in enumerate(sorted_df.iterrows(), data_start_row):
        dc = ws.cell(ri, 1, dt.date())
        dc.number_format = "YYYY-MM-DD"; dc.font = _DAT_FONT; dc.alignment = _CENTER
        for ci, name in enumerate(sector_names, 2):
            val = row[name]
            v   = None if pd.isna(val) else float(val)
            c   = ws.cell(ri, ci, v)
            c.font = _DAT_FONT; c.alignment = _CENTER; c.number_format = "0.00"
            fill = _rsi_fill(val)
            if fill: c.fill = fill
        ws.row_dimensions[ri].height = 15

    return data_start_row


def _write_info_sheet(ws, tops: dict[str, dict], sector_etfs: dict[str, str]) -> None:
    ws.delete_rows(1, ws.max_row)
    headers = ["업종명", "대표종목1", "대표종목2", "대표종목3", "한국추종ETF", "대응미국섹터"]
    widths  = [14, 16, 16, 16, 30, 16]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, ci, h)
        c.fill = _HDR_FILL_INFO; c.font = _HDR_FONT; c.alignment = _CENTER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20

    for ri, (_, name) in enumerate(SECTORS.items(), 2):
        data = tops.get(name, {})
        names = data.get("names", ["-", "-", "-"])
        etf   = sector_etfs.get(name, "-")
        us    = KR_TO_US.get(name, "-")
        n1 = names[0] if len(names) > 0 else "-"
        n2 = names[1] if len(names) > 1 else "-"
        n3 = names[2] if len(names) > 2 else "-"
        for ci, val in enumerate([name, n1, n2, n3, etf, us], 1):
            c = ws.cell(ri, ci, val)
            c.font = _DAT_FONT; c.alignment = _CENTER
        ws.row_dimensions[ri].height = 18


def _write_chart_sheet(wb, kr_cols, us_cols, kr_data_start, us_data_start) -> None:
    CHART_SHEET = "비교 차트"
    if CHART_SHEET in wb.sheetnames:
        del wb[CHART_SHEET]
    wc = wb.create_sheet(CHART_SHEET)

    ws_kr = wb["한국 RSI"]
    ws_us = wb["미국 RSI"]
    n_rows_kr = kr_data_start - 1 + MAX_DATA_ROWS
    n_rows_us = us_data_start - 1 + MAX_DATA_ROWS

    CHART_W = 13
    CHART_H = 18

    for idx, (_, kr_name) in enumerate(SECTORS.items()):
        us_etf   = KR_TO_US.get(kr_name, "")
        us_label = US_SECTORS.get(us_etf, "")

        col_offset = (idx % 2) * (CHART_W + 1) + 1
        row_offset = (idx // 2) * (CHART_H + 2) + 1
        anchor_col = get_column_letter(col_offset)
        anchor_cell= f"{anchor_col}{row_offset}"

        chart = LineChart()
        chart.title         = f"{kr_name}  vs  {us_label or us_etf}"
        chart.style         = 10
        chart.y_axis.title  = "RSI"
        chart.x_axis.title  = "날짜"
        chart.y_axis.numFmt = "0"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.width  = 14
        chart.height = 11

        if kr_name in kr_cols:
            kr_col_idx = kr_cols.index(kr_name) + 2
            kr_vals = Reference(ws_kr, min_col=kr_col_idx,
                                min_row=kr_data_start - 1, max_row=n_rows_kr)
            chart.add_data(kr_vals, titles_from_data=True)
            if chart.series:
                chart.series[-1].graphicalProperties.line.solidFill = "1F4E79"
                chart.series[-1].graphicalProperties.line.width = 20000

        if us_label and us_label in us_cols:
            us_col_idx = us_cols.index(us_label) + 2
            us_vals = Reference(ws_us, min_col=us_col_idx,
                                min_row=us_data_start - 1, max_row=n_rows_us)
            chart.add_data(us_vals, titles_from_data=True)
            if len(chart.series) >= 2:
                chart.series[-1].graphicalProperties.line.solidFill = "C00000"
                chart.series[-1].graphicalProperties.line.width = 20000

        dates = Reference(ws_kr, min_col=1, min_row=kr_data_start, max_row=n_rows_kr)
        chart.set_categories(dates)
        wc.add_chart(chart, anchor_cell)

    log.info(f"비교 차트 {len(SECTORS)}개 생성 완료")


# ── 시트 행 탐색 헬퍼 ─────────────────────────────────────────────────────────

def _find_last_data_row(ws) -> int:
    for r in range(ws.max_row, 0, -1):
        v = ws.cell(r, 1).value
        if isinstance(v, (datetime, date)):
            return r
    return 1


def _find_data_start_row(ws) -> int:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, (datetime, date)):
            return r
    return ws.max_row + 1


def _find_subrow(ws, label: str) -> int | None:
    """레이블 문자열이 있는 메타 행 번호 반환 (없으면 None)."""
    for r in range(2, min(15, ws.max_row + 1)):
        if ws.cell(r, 1).value == label:
            return r
    return None


def _append_rsi_rows(ws, new_rows: pd.DataFrame, col_names: list[str]) -> None:
    data_start = _find_data_start_row(ws)
    new_rows_desc = new_rows.sort_index(ascending=False)
    ws.insert_rows(data_start, len(new_rows_desc))
    for ri, (dt, row) in enumerate(new_rows_desc.iterrows(), data_start):
        dc = ws.cell(ri, 1, dt.date())
        dc.number_format = "YYYY-MM-DD"
        dc.font = _DAT_FONT
        dc.alignment = _CENTER
        for ci, name in enumerate(col_names, 2):
            val = row.get(name, None)
            v   = None if (val is None or pd.isna(val)) else float(val)
            c   = ws.cell(ri, ci, v)
            c.font = _DAT_FONT; c.alignment = _CENTER; c.number_format = "0.00"
            fill = _rsi_fill(val)
            if fill: c.fill = fill
        ws.row_dimensions[ri].height = 15


def _update_etf_subrow(ws, sector_etfs: dict[str, str]) -> None:
    """한국 RSI 시트의 '한국ETF' 메타 행 값을 갱신."""
    etf_row = _find_subrow(ws, "한국ETF")
    if etf_row is None:
        return
    kr_cols = list(SECTORS.values())
    _META_FILL = PatternFill("solid", fgColor="D9D9D9")
    for ci, name in enumerate(kr_cols, 2):
        cell = ws.cell(etf_row, ci)
        cell.value     = sector_etfs.get(name, "-")
        cell.font      = _SUB_FONT
        cell.alignment = _CENTER
        cell.fill      = _META_FILL


def _update_info_sheet(ws, tops: dict[str, dict], sector_etfs: dict[str, str]) -> None:
    """업종 정보 시트의 ETF 열(5번째)을 갱신."""
    for ri, (_, name) in enumerate(SECTORS.items(), 2):
        ws.cell(ri, 5).value = sector_etfs.get(name, "-")


# ── 기존 파일에서 RSI 로드 ────────────────────────────────────────────────────

def _load_existing_rsi(path: str, sheet: str) -> pd.DataFrame | None:
    if not os.path.exists(path):
        return None
    try:
        wb = openpyxl.load_workbook(path)
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        headers = [ws.cell(1, c).value for c in range(2, ws.max_column + 1)]
        data: dict[str, list] = {h: [] for h in headers if h}
        dates: list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if isinstance(row[0], (datetime, date)):
                dates.append(row[0])
                for i, h in enumerate(headers):
                    if h:
                        data[h].append(row[i + 1])
        if not dates:
            return None
        return pd.DataFrame(data, index=pd.to_datetime(dates))
    except PermissionError:
        raise  # 파일이 열려있음 → 호출자에서 처리
    except Exception as e:
        log.warning(f"기존 데이터 로드 오류 ({sheet}): {e}")
        return None


# ── 월요일: 새 주차 파일 초기화 ──────────────────────────────────────────────

def monday_init_excel(excel_path: str, today: date) -> None:
    """월요일 전용: 주차 파일 생성 + RSI 전체 로드 + 대표종목/ETF 설정."""
    log.info(f"[월요일 초기화] 파일: {os.path.basename(excel_path)}")

    # 직전 주 범위 (월~금)
    prev_fri = today - timedelta(days=3)
    prev_mon = today - timedelta(days=7)
    today_str    = today.strftime("%Y%m%d")
    prev_mon_str = prev_mon.strftime("%Y%m%d")
    prev_fri_str = prev_fri.strftime("%Y%m%d")

    # 대표종목 수집 (시총은 직전 금요일 확정 종가 기준 — 월요일 09:00 개장 전엔 당일 시총이 0)
    tops = build_monday_tops(prev_fri_str, prev_mon_str, prev_fri_str)
    save_tops_cache(tops)

    # ETF 선정
    sector_etfs = build_sector_etfs(tops)

    # RSI 데이터 수집 (DISPLAY_START 이전 WARMUP_DAYS 포함)
    warmup_start  = date(2023, 10, 1)
    fetch_start   = warmup_start.strftime("%Y%m%d")
    fetch_end     = today.strftime("%Y%m%d")
    yesterday_str = (today - timedelta(days=1)).strftime("%Y%m%d")

    log.info("한국 업종 RSI 수집...")
    kr_rsi = fetch_korea_rsi(fetch_start, yesterday_str)
    kr_rsi = kr_rsi[kr_rsi.index.date >= DISPLAY_START]

    log.info("미국 섹터 RSI 수집...")
    us_rsi = fetch_us_rsi(fetch_start, yesterday_str)
    us_rsi = us_rsi[us_rsi.index.date >= DISPLAY_START]

    kr_rsi, us_rsi = align_dates(kr_rsi, us_rsi)

    # 서브행 구성
    kr_cols = list(SECTORS.values())
    kr_subrows: dict[str, list] = {
        "대표종목1": [tops.get(n, {}).get("names", ["-","-","-"])[0] for n in kr_cols],
        "대표종목2": [tops.get(n, {}).get("names", ["-","-","-"])[1] for n in kr_cols],
        "대표종목3": [tops.get(n, {}).get("names", ["-","-","-"])[2] for n in kr_cols],
        "한국ETF":   [sector_etfs.get(n, "-") for n in kr_cols],
        "대응미국":  [KR_TO_US.get(n, "-")    for n in kr_cols],
    }

    # 워크북 생성 (기존 시트 보존, RSI 관련 시트만 재생성)
    RSI_SHEETS = ["한국 RSI", "미국 RSI", "비교 차트", "업종 정보"]
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        for s in RSI_SHEETS:
            if s in wb.sheetnames:
                del wb[s]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for s in RSI_SHEETS:
        wb.create_sheet(s)

    _write_info_sheet(wb["업종 정보"], tops, sector_etfs)
    kr_data_start = _write_rsi_sheet(wb["한국 RSI"], kr_rsi, _HDR_FILL_KR, subrows=kr_subrows)
    us_data_start = _write_rsi_sheet(wb["미국 RSI"], us_rsi, _HDR_FILL_US)
    _write_chart_sheet(wb, kr_cols, list(us_rsi.columns), kr_data_start, us_data_start)

    # 시트 순서 정렬
    desired = ["한국 RSI", "미국 RSI", "비교 차트", "업종 정보"]
    for i, s in enumerate(desired):
        if s in wb.sheetnames:
            wb.move_sheet(s, offset=wb.sheetnames.index(s) - i)
    if "한국 RSI" in wb.sheetnames:
        wb.active = wb["한국 RSI"]

    wb.save(excel_path)
    log.info(f"[월요일 초기화 완료] {os.path.basename(excel_path)}")


# ── 일별 업데이트 ─────────────────────────────────────────────────────────────

def daily_update_excel(excel_path: str, today: date) -> None:
    """RSI 데이터 추가 + 한국ETF 행 갱신 (차트·서식 보존)."""
    if not os.path.exists(excel_path):
        log.warning(f"파일 없음 → 월요일 초기화 필요: {excel_path}")
        monday_init_excel(excel_path, today)
        return

    try:
        kr_existing = _load_existing_rsi(excel_path, "한국 RSI")
        us_existing = _load_existing_rsi(excel_path, "미국 RSI")
    except PermissionError:
        log.warning(f"[접근 실패] Excel 파일이 열려 있어 RSI 업데이트를 건너뜁니다: {os.path.basename(excel_path)}")
        return

    if kr_existing is None:
        log.warning("한국 RSI 시트 없음 → 초기화 실행")
        monday_init_excel(excel_path, today)
        return

    yesterday   = today - timedelta(days=1)
    latest_kr   = kr_existing.index.max().date()
    latest_us   = (us_existing.index.max().date() if us_existing is not None else latest_kr)

    if latest_kr >= yesterday and latest_us >= yesterday:
        log.info(f"RSI 최신 상태 (한국: {latest_kr}, 미국: {latest_us})")
    else:
        fetch_from  = (pd.Timestamp(min(latest_kr, latest_us))
                       - pd.tseries.offsets.BDay(WARMUP_DAYS)).date()
        fetch_start = fetch_from.strftime("%Y%m%d")
        fetch_end   = yesterday.strftime("%Y%m%d")

        log.info(f"RSI 업데이트: {fetch_from} → {yesterday}")
        new_kr = fetch_korea_rsi(fetch_start, fetch_end)
        new_us = fetch_us_rsi(fetch_start, fetch_end)

        kr_new = new_kr[new_kr.index.date > latest_kr]
        us_new = new_us[new_us.index.date > latest_us]

        if not kr_new.empty or not us_new.empty:
            kr_new, us_new = align_dates(
                kr_new if not kr_new.empty else pd.DataFrame(index=us_new.index),
                us_new if not us_new.empty else pd.DataFrame(index=kr_new.index),
            )
            wb = openpyxl.load_workbook(excel_path)
            _append_rsi_rows(wb["한국 RSI"], kr_new, list(SECTORS.values()))
            _append_rsi_rows(wb["미국 RSI"], us_new, list(us_new.columns))
            wb.save(excel_path)
            log.info(f"RSI 추가: +{len(kr_new)}일(한국) / +{len(us_new)}일(미국)")

    # ETF 행 갱신
    tops = load_tops_cache()
    if tops is None:
        log.warning("tops 캐시 없음 — ETF 갱신 건너뜀")
        return

    log.info("한국ETF 행 갱신 중...")
    sector_etfs = build_sector_etfs(tops)

    wb = openpyxl.load_workbook(excel_path)
    if "한국 RSI" in wb.sheetnames:
        _update_etf_subrow(wb["한국 RSI"], sector_etfs)
    if "업종 정보" in wb.sheetnames:
        _update_info_sheet(wb["업종 정보"], tops, sector_etfs)
    wb.save(excel_path)
    log.info("ETF 행 갱신 완료")


# ── --init 전용 전체 초기화 ───────────────────────────────────────────────────

def bulk_init_excel(excel_path: str, today: date) -> None:
    """수동 --init: DISPLAY_START부터 전체 로드 후 파일 생성."""
    log.info(f"[전체 초기화] {os.path.basename(excel_path)}")
    today_str = today.strftime("%Y%m%d")
    prev_fri  = (today - timedelta(days=3)).strftime("%Y%m%d")
    prev_mon  = (today - timedelta(days=7)).strftime("%Y%m%d")

    tops = build_monday_tops(today_str, prev_mon, prev_fri)
    save_tops_cache(tops)
    sector_etfs = build_sector_etfs(tops)

    warmup_start  = date(2023, 10, 1)
    yesterday     = (today - timedelta(days=1)).strftime("%Y%m%d")

    log.info("한국 업종 RSI 수집...")
    kr_rsi = fetch_korea_rsi(warmup_start.strftime("%Y%m%d"), yesterday)
    kr_rsi = kr_rsi[kr_rsi.index.date >= DISPLAY_START]

    log.info("미국 섹터 RSI 수집...")
    us_rsi = fetch_us_rsi(warmup_start.strftime("%Y%m%d"), yesterday)
    us_rsi = us_rsi[us_rsi.index.date >= DISPLAY_START]

    kr_rsi, us_rsi = align_dates(kr_rsi, us_rsi)

    kr_cols = list(SECTORS.values())
    kr_subrows: dict[str, list] = {
        "대표종목1": [tops.get(n, {}).get("names", ["-","-","-"])[0] for n in kr_cols],
        "대표종목2": [tops.get(n, {}).get("names", ["-","-","-"])[1] for n in kr_cols],
        "대표종목3": [tops.get(n, {}).get("names", ["-","-","-"])[2] for n in kr_cols],
        "한국ETF":   [sector_etfs.get(n, "-") for n in kr_cols],
        "대응미국":  [KR_TO_US.get(n, "-")    for n in kr_cols],
    }

    RSI_SHEETS = ["한국 RSI", "미국 RSI", "비교 차트", "업종 정보"]
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        for s in RSI_SHEETS:
            if s in wb.sheetnames:
                del wb[s]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for s in RSI_SHEETS:
        wb.create_sheet(s)

    _write_info_sheet(wb["업종 정보"], tops, sector_etfs)
    kr_data_start = _write_rsi_sheet(wb["한국 RSI"], kr_rsi, _HDR_FILL_KR, subrows=kr_subrows)
    us_data_start = _write_rsi_sheet(wb["미국 RSI"], us_rsi, _HDR_FILL_US)
    _write_chart_sheet(wb, kr_cols, list(us_rsi.columns), kr_data_start, us_data_start)

    desired = ["한국 RSI", "미국 RSI", "비교 차트", "업종 정보"]
    for i, s in enumerate(desired):
        if s in wb.sheetnames:
            wb.move_sheet(s, offset=wb.sheetnames.index(s) - i)
    if "한국 RSI" in wb.sheetnames:
        wb.active = wb["한국 RSI"]

    wb.save(excel_path)
    log.info(f"[전체 초기화 완료] {len(kr_rsi)}일×{len(kr_rsi.columns)}업종 | {excel_path}")


# ── 진입점 ────────────────────────────────────────────────────────────────────

def main() -> None:
    now_kst    = datetime.now(KST)
    today      = now_kst.date()
    init_mode  = "--init" in sys.argv
    excel_path = get_weekly_filepath(today)

    log.info("=" * 55)
    log.info(f"sector_rsi 시작 | {today} | {os.path.basename(excel_path)}")
    log.info("=" * 55)

    if init_mode:
        bulk_init_excel(excel_path, today)
    elif today.weekday() == 0:  # 월요일
        monday_init_excel(excel_path, today)
    else:
        daily_update_excel(excel_path, today)


if __name__ == "__main__":
    main()
