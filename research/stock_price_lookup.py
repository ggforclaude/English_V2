#!/usr/bin/env python3
"""
종목명/코드 입력 -> 최근 1년 일별 종가 + 그 이전 월말 종가 Excel 생성

실행:
  python stock_price_lookup.py         (대화형 입력, 여러 종목 연속 조회 가능)
  python stock_price_lookup.py 005930  (인자로 바로 실행)

출력: research/주가조회/{종목명}_주가_{YYYYMMDD}.xlsx
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import json
from pathlib import Path
from datetime import date, datetime, timedelta

import pandas as pd
import FinanceDataReader as fdr
from pykrx import stock as krx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = Path(__file__).parent
OUT_DIR = BASE_DIR / "주가조회"
MAP_CACHE = BASE_DIR / "stock_name_map.json"
MAP_MAX_AGE_DAYS = 7

OUT_DIR.mkdir(exist_ok=True)


def _load_name_map() -> dict:
    """종목명 -> 종목코드 매핑. 캐시가 없거나 오래되면 FinanceDataReader로 재생성."""
    if MAP_CACHE.exists():
        age = datetime.now() - datetime.fromtimestamp(MAP_CACHE.stat().st_mtime)
        if age.days < MAP_MAX_AGE_DAYS:
            with open(MAP_CACHE, encoding="utf-8") as f:
                return json.load(f)

    print("종목 목록 갱신 중 (최초 1회 또는 주 1회)...")
    df = fdr.StockListing("KRX")
    name_map = dict(zip(df["Name"], df["Code"]))
    with open(MAP_CACHE, "w", encoding="utf-8") as f:
        json.dump(name_map, f, ensure_ascii=False)
    return name_map


def resolve_code(user_input: str, name_map: dict):
    """입력값 -> (종목코드, 종목명). 못 찾으면 None."""
    s = user_input.strip().lstrip("﻿")

    if s.isdigit():
        s = s.zfill(6)
        try:
            name = krx.get_market_ticker_name(s)
            if name:
                return s, name
        except Exception:
            pass
        return None

    if s in name_map:
        return name_map[s], s

    matches = [(n, c) for n, c in name_map.items() if s in n]
    if len(matches) == 1:
        return matches[0][1], matches[0][0]
    return None


def fetch_prices(code: str):
    """일별(최근 1년) / 월별(그 이전) 종가 Series + 현재가 + 기준일 반환."""
    today = date.today()
    one_year_ago = today - timedelta(days=365)

    df_daily = krx.get_market_ohlcv_by_date(
        one_year_ago.strftime("%Y%m%d"), today.strftime("%Y%m%d"), code
    )
    daily_close = df_daily["종가"]
    daily_close.index = pd.to_datetime(daily_close.index)

    monthly_end = (one_year_ago - timedelta(days=1)).strftime("%Y%m%d")
    df_old = krx.get_market_ohlcv_by_date("19900101", monthly_end, code)
    if df_old.empty:
        monthly_close = pd.Series(dtype=float)
    else:
        old_close = df_old["종가"]
        old_close.index = pd.to_datetime(old_close.index)
        monthly_close = old_close.resample("ME").last()

    if not daily_close.empty:
        current_price = float(daily_close.iloc[-1])
        current_date = daily_close.index[-1].date()
    else:
        current_price = float("nan")
        current_date = today

    return daily_close, monthly_close, current_price, current_date


def write_excel(name, code, daily, monthly, current_price, current_date) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주가"

    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=11)
    SUB_FONT = Font(name="맑은 고딕", size=10, bold=True)
    DAT_FONT = Font(name="맑은 고딕", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")

    info = [("종목명", name), ("종목코드", code), ("현재가", current_price), ("기준일", current_date)]
    for i, (label, val) in enumerate(info, start=1):
        ws.cell(i, 1, label).font = SUB_FONT
        c = ws.cell(i, 2, val)
        c.font = DAT_FONT
        if label == "현재가":
            c.number_format = "#,##0"
        elif label == "기준일":
            c.number_format = "YYYY-MM-DD"

    header_row = 6
    for col, text in ((1, "날짜"), (2, "종가"), (3, "구분")):
        c = ws.cell(header_row, col, text)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1
    for dt, val in daily.sort_index(ascending=False).items():
        ws.cell(row, 1, dt.date()).number_format = "YYYY-MM-DD"
        ws.cell(row, 1).font = DAT_FONT
        ws.cell(row, 2, float(val)).number_format = "#,##0"
        ws.cell(row, 2).font = DAT_FONT
        ws.cell(row, 3, "일별").font = DAT_FONT
        row += 1

    for dt, val in monthly.sort_index(ascending=False).items():
        ws.cell(row, 1, dt.date()).number_format = "YYYY-MM-DD"
        ws.cell(row, 1).font = DAT_FONT
        ws.cell(row, 2, float(val)).number_format = "#,##0"
        ws.cell(row, 2).font = DAT_FONT
        ws.cell(row, 3, "월말").font = DAT_FONT
        row += 1

    out_path = OUT_DIR / f"{name}_주가_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(out_path)
    return out_path


def run_one(user_input: str, name_map: dict) -> None:
    resolved = resolve_code(user_input, name_map)
    if resolved is None:
        print(f"'{user_input}' 종목을 찾을 수 없습니다. 종목명 또는 6자리 종목코드를 정확히 입력해주세요.")
        return
    code, name = resolved
    print(f"조회 중: {name} ({code}) ...")
    try:
        daily, monthly, current_price, current_date = fetch_prices(code)
    except Exception as e:
        print(f"시세 조회 오류: {e}")
        return
    out_path = write_excel(name, code, daily, monthly, current_price, current_date)
    print(f"완료: {out_path}")
    print(f"  현재가({current_date}): {current_price:,.0f}원 / 일별 {len(daily)}건, 월별 {len(monthly)}건")


def main() -> None:
    name_map = _load_name_map()

    if len(sys.argv) > 1:
        run_one(" ".join(sys.argv[1:]), name_map)
        return

    print("=" * 50)
    print(" 종목 주가 조회")
    print(" 종목명 또는 종목코드(6자리)를 입력하세요. (종료: 그냥 Enter)")
    print("=" * 50)
    while True:
        user_input = input("\n종목명/코드 > ").strip()
        if not user_input:
            print("종료합니다.")
            break
        run_one(user_input, name_map)


if __name__ == "__main__":
    main()
