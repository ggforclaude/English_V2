"""
research/create_asset_manager.py
자산관리_2026.xlsx 생성 스크립트

사용법:
  python create_asset_manager.py            # 신규 생성
  python create_asset_manager.py --update   # KRX 시세만 업데이트
"""
import sys, os, argparse
sys.stdout.reconfigure(encoding="utf-8")

import pytz
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(BASE_DIR, "자산관리_2026.xlsx")
KST         = pytz.timezone("Asia/Seoul")

# ── 계좌 · 코드 목록 ────────────────────────────────────────────────────────
ACCOUNTS      = ["신한증권", "토스증권", "한화투자", "연금저축(한화)", "IRP(한화)", "ISA(한화)", "기타"]
TRADE_TYPES   = ["매수", "매도", "배당", "권리"]
CASH_TYPES    = ["입금", "출금", "이체입금", "이체출금"]
IDEA_STATUS   = ["아이디어", "검토중", "투자결정", "보류", "기각"]
IDEA_PRIORITY = ["긴급", "높음", "보통", "낮음"]
IDEA_CATEGORY = ["가치투자", "성장주", "배당주", "이벤트드리븐", "섹터로테이션", "테마", "글로벌매크로", "기타"]
SELL_REASON   = ["목표가달성", "손절", "리밸런싱", "급전필요", "전략변경", "기타"]
REENTRY       = ["재매수검토", "관망", "기회없음", "재매수완료"]
YESNO         = ["Yes", "No", "미확인"]
MARKETS       = ["KOSPI", "KOSDAQ", "KONEX", "ETF", "해외"]

# ── 공통 색상 ────────────────────────────────────────────────────────────────
C = {
    "navy":    "17375E",  # 진한 네이비 (주 헤더)
    "blue":    "2E75B6",  # 파란 (보조 헤더)
    "lblue":   "BDD7EE",  # 연파랑 (합계행)
    "input":   "EEF3FA",  # 입력 셀
    "formula": "F2F2F2",  # 수식 셀 (읽기전용)
    "green":   "375623",  # 초록 섹션
    "lgreen":  "E2EFDA",  # 연초록
    "orange":  "C55A11",  # 주황 섹션
    "lorange": "FCE4D6",  # 연주황
    "yellow":  "FFF2CC",  # 노랑 강조
    "white":   "FFFFFF",
    "hdr_fg":  "FFFFFF",
    "gray":    "D9D9D9",
    "lgray":   "F5F5F5",
    "red":     "C00000",
}


# ── 스타일 헬퍼 ──────────────────────────────────────────────────────────────

def _f(bold=False, size=10, color="000000", italic=False, underline=None):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color,
                italic=italic, underline=underline)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(top=s, left=s, right=s, bottom=s)

def _thick_bottom(color="AAAAAA"):
    thin = Side(style="thin", color="CCCCCC")
    thick = Side(style="medium", color=color)
    return Border(top=thin, left=thin, right=thin, bottom=thick)


def hdr(ws, row, col, text, bg=C["navy"], fg=C["hdr_fg"], bold=True,
        size=10, h="center", wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _f(bold=bold, size=size, color=fg)
    c.fill      = _fill(bg)
    c.alignment = _align(h=h, wrap=wrap)
    c.border    = _border(color="888888")
    return c


def data(ws, row, col, value=None, bg=C["input"], fmt=None,
         h="left", bold=False, color="000000", formula=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _f(bold=bold, color=color)
    c.fill      = _fill(bg if not formula else C["formula"])
    c.alignment = _align(h=h)
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    return c


def title_row(ws, row, text, n_cols, bg=C["navy"], size=13):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _f(bold=True, size=size, color=C["hdr_fg"])
    c.fill      = _fill(bg)
    c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 26
    return c


def note(ws, row, col, text, color="888888"):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _f(italic=True, size=9, color=color)
    c.alignment = _align(h="left")
    return c


def dv_list(ws, formula_str, cell_range, allow_blank=True):
    dv = DataValidation(type="list", formula1=formula_str,
                        allow_blank=allow_blank, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def col_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — 투자거래내역
# ══════════════════════════════════════════════════════════════════════════════
TRADE  = "투자거래내역"
CASH   = "입출금내역"
KRX    = "KRX종목마스터"
NROWS  = 2000   # 데이터 행 수

def make_trade_log(wb):
    ws = wb.create_sheet(TRADE)

    # ── 컬럼 정의 ─────────────────────────────────────────────────────────────
    # A  B    C    D      E      F   G   H      I    J      K
    # 날 계좌 유형 종목코드 종목명 수량 단가 금액(수식) 수수료 실제금액(수식) 메모
    COLS = [
        ("날짜",         11, "#,##0",        "center"),
        ("계좌",         14, None,            "center"),
        ("유형",          8, None,            "center"),
        ("종목코드",      10, "@",             "center"),
        ("종목명",        20, None,            "left"),
        ("수량",          10, "#,##0",        "right"),
        ("단가",          12, "#,##0",        "right"),
        ("금액",          14, "#,##0",        "right"),
        ("수수료",        10, "#,##0",        "right"),
        ("실제금액",      14, "#,##0",        "right"),
        ("메모",          30, None,            "left"),
    ]

    title_row(ws, 1, "📋 투자거래내역 — 매수·매도·배당 기록", len(COLS))
    note(ws, 2, 1, "★ 노란색=입력필수 | 회색=자동계산 | 종목코드 입력 시 종목명 자동완성")

    # 헤더
    for ci, (label, width, _, align) in enumerate(COLS, 1):
        hdr(ws, 3, ci, label)
        col_w(ws, ci, width)
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}3"

    # 데이터 행
    for r in range(4, NROWS + 4):
        # 날짜
        c = ws.cell(row=r, column=1)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "YYYY-MM-DD"; c.alignment = _align("center")
        # 계좌 (입력)
        for col in [2, 3]:
            c = ws.cell(row=r, column=col)
            c.fill = _fill(C["input"]); c.border = _border()
            c.alignment = _align("center")
        # 종목코드
        c = ws.cell(row=r, column=4)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "@"; c.alignment = _align("center")
        # 종목명 — VLOOKUP
        c = ws.cell(row=r, column=5,
                    value=f"=IFERROR(VLOOKUP(D{r},'{KRX}'!$A:$B,2,FALSE),\"\")")
        c.fill = _fill(C["formula"]); c.border = _border()
        # 수량, 단가
        for col in [6, 7]:
            c = ws.cell(row=r, column=col)
            c.fill = _fill(C["input"]); c.border = _border()
            c.number_format = "#,##0"; c.alignment = _align("right")
        # 금액 = 수량 × 단가
        c = ws.cell(row=r, column=8,
                    value=f"=IF(AND(F{r}<>\"\",G{r}<>\"\"),F{r}*G{r},\"\")")
        c.fill = _fill(C["formula"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 수수료
        c = ws.cell(row=r, column=9)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 실제금액 = 매수:금액+수수료 / 매도:금액-수수료 / 배당:금액
        c = ws.cell(row=r, column=10,
                    value=(f"=IF(H{r}=\"\",\"\","
                           f"IF(C{r}=\"매수\",H{r}+IF(I{r}=\"\",0,I{r}),"
                           f"IF(C{r}=\"매도\",H{r}-IF(I{r}=\"\",0,I{r}),H{r})))"))
        c.fill = _fill(C["formula"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        c.font = _f(bold=True)
        # 메모
        c = ws.cell(row=r, column=11)
        c.fill = _fill(C["input"]); c.border = _border()

    # 데이터 유효성
    dv_list(ws, f'"{",".join(ACCOUNTS)}"', f"B4:B{NROWS+3}")
    dv_list(ws, f'"{",".join(TRADE_TYPES)}"', f"C4:C{NROWS+3}")

    print(f"  ✓ '{TRADE}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — 입출금내역
# ══════════════════════════════════════════════════════════════════════════════

def make_cash_flow(wb):
    ws = wb.create_sheet(CASH)

    COLS = [
        ("날짜",          11),
        ("계좌",          14),
        ("유형",          10),
        ("금액(원)",      16),
        ("상대계좌/출처", 18),
        ("내용",          35),
        ("메모",          25),
    ]

    title_row(ws, 1, "💰 입출금내역 — 계좌별 자금 흐름 기록", len(COLS), bg=C["green"])
    note(ws, 2, 1, "★ 계좌 간 이체: 보내는 쪽=이체출금, 받는 쪽=이체입금으로 각각 입력")

    for ci, (label, width) in enumerate(COLS, 1):
        hdr(ws, 3, ci, label, bg=C["green"])
        col_w(ws, ci, width)
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}3"

    for r in range(4, NROWS + 4):
        # 날짜
        c = ws.cell(row=r, column=1)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "YYYY-MM-DD"; c.alignment = _align("center")
        # 계좌, 유형
        for col in [2, 3]:
            c = ws.cell(row=r, column=col)
            c.fill = _fill(C["input"]); c.border = _border()
            c.alignment = _align("center")
        # 금액
        c = ws.cell(row=r, column=4)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 상대계좌/출처, 내용, 메모
        for col in [5, 6, 7]:
            c = ws.cell(row=r, column=col)
            c.fill = _fill(C["input"]); c.border = _border()

    dv_list(ws, f'"{",".join(ACCOUNTS)}"',          f"B4:B{NROWS+3}")
    dv_list(ws, f'"{",".join(CASH_TYPES)}"',        f"C4:C{NROWS+3}")
    dv_list(ws, f'"{",".join(ACCOUNTS + ["외부"])}\"', f"E4:E{NROWS+3}")

    print(f"  ✓ '{CASH}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — 계좌별현황
# ══════════════════════════════════════════════════════════════════════════════
SUMMARY = "계좌별현황"

def make_account_summary(wb):
    ws = wb.create_sheet(SUMMARY)
    N  = len(ACCOUNTS)

    # ── 섹션 1: 계좌별 자금 요약 ───────────────────────────────────────────
    NCOLS = 10
    title_row(ws, 1, "📊 계좌별 현황 — 자금 흐름 & 수익률 요약", NCOLS)
    note(ws, 2, 1, "★ '평가금액'만 수동 입력. 나머지는 투자거래내역·입출금내역 시트에서 자동 집계")
    note(ws, 2, 6, f"최종 업데이트: {datetime.now(KST).strftime('%Y-%m-%d %H:%M')}")

    # 헤더
    sub_hdrs = [
        ("계좌",          14, "center"),
        ("순입금 합계",   14, "right"),   # 입금+이체입금 - 출금 - 이체출금
        ("총 매수금액",   14, "right"),
        ("총 매도금액",   14, "right"),
        ("추정 현금잔고", 14, "right"),   # 순입금 - 매수 + 매도
        ("평가금액\n(수동입력)", 14, "right"),
        ("평가손익",      14, "right"),   # 평가금액 + 현금 - 순입금
        ("수익률",        10, "right"),
        ("메모",          22, "left"),
    ]
    for ci, (label, width, align) in enumerate(sub_hdrs, 1):
        hdr(ws, 4, ci, label, wrap=True)
        col_w(ws, ci, width)
        ws.row_dimensions[4].height = 30

    # 계좌별 행 (SUMIFS 수식)
    for ri, acct in enumerate(ACCOUNTS):
        row = 5 + ri
        # 계좌명
        c = ws.cell(row=row, column=1, value=acct)
        c.font = _f(bold=True); c.fill = _fill(C["lgray"]); c.border = _border()
        c.alignment = _align("center")

        tl = f"'{TRADE}'!"   # 투자거래내역 시트
        cl = f"'{CASH}'!"    # 입출금내역 시트
        acct_ref = f'"{acct}"'

        # B: 순입금 = 입금 + 이체입금 - 출금 - 이체출금
        c = ws.cell(row=row, column=2,
            value=(f'=SUMIFS({cl}D:D,{cl}B:B,{acct_ref},{cl}C:C,"입금")'
                   f'+SUMIFS({cl}D:D,{cl}B:B,{acct_ref},{cl}C:C,"이체입금")'
                   f'-SUMIFS({cl}D:D,{cl}B:B,{acct_ref},{cl}C:C,"출금")'
                   f'-SUMIFS({cl}D:D,{cl}B:B,{acct_ref},{cl}C:C,"이체출금")'))
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")

        # C: 총 매수금액
        c = ws.cell(row=row, column=3,
            value=f'=SUMIFS({tl}J:J,{tl}B:B,{acct_ref},{tl}C:C,"매수")')
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")

        # D: 총 매도금액
        c = ws.cell(row=row, column=4,
            value=f'=SUMIFS({tl}J:J,{tl}B:B,{acct_ref},{tl}C:C,"매도")')
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")

        # E: 추정 현금잔고 = 순입금 - 매수 + 매도
        c = ws.cell(row=row, column=5, value=f"=B{row}-C{row}+D{row}")
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")

        # F: 평가금액 (수동)
        c = ws.cell(row=row, column=6)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")

        # G: 평가손익 = (현금잔고 + 평가금액) - 순입금
        c = ws.cell(row=row, column=7,
            value=f"=IF(F{row}=\"\",\"\",(E{row}+F{row})-B{row})")
        c.number_format = '+#,##0;-#,##0;"-"'
        c.fill = _fill(C["formula"]); c.border = _border()
        c.alignment = _align("right")

        # H: 수익률
        c = ws.cell(row=row, column=8,
            value=f"=IF(OR(F{row}=\"\",B{row}=0),\"\",(E{row}+F{row}-B{row})/B{row})")
        c.number_format = "0.00%"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")

        # I: 메모
        c = ws.cell(row=row, column=9)
        c.fill = _fill(C["input"]); c.border = _border()

        ws.row_dimensions[row].height = 20

    # 합계 행
    total_row = 5 + N
    c = ws.cell(row=total_row, column=1, value="합  계")
    c.font = _f(bold=True, size=11); c.fill = _fill(C["lblue"])
    c.border = _border(); c.alignment = _align("center")

    for col in range(2, 6):
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_letter}5:{col_letter}{5+N-1})")
        c.number_format = "#,##0"; c.font = _f(bold=True)
        c.fill = _fill(C["lblue"]); c.border = _border()
        c.alignment = _align("right")
    # 평가금액 합계
    c = ws.cell(row=total_row, column=6,
                value=f"=SUMIF(F5:F{4+N},\"<>\"&\"\")")
    c.number_format = "#,##0"; c.font = _f(bold=True)
    c.fill = _fill(C["lblue"]); c.border = _border()
    c.alignment = _align("right")
    # 평가손익 합계
    c = ws.cell(row=total_row, column=7,
                value=f"=IF(F{total_row}=\"\",\"\",(E{total_row}+F{total_row})-B{total_row})")
    c.number_format = '+#,##0;-#,##0;"-"'; c.font = _f(bold=True)
    c.fill = _fill(C["lblue"]); c.border = _border()
    c.alignment = _align("right")
    # 전체 수익률
    c = ws.cell(row=total_row, column=8,
                value=f"=IF(OR(F{total_row}=\"\",B{total_row}=0),\"\",(E{total_row}+F{total_row}-B{total_row})/B{total_row})")
    c.number_format = "0.00%"; c.font = _f(bold=True)
    c.fill = _fill(C["lblue"]); c.border = _border()
    c.alignment = _align("right")
    ws.row_dimensions[total_row].height = 22

    # ── 섹션 2: 월별 거래 건수 요약 ────────────────────────────────────────
    sec2_row = total_row + 2
    title_row(ws, sec2_row, "▌월별 거래 통계 (투자거래내역 자동 집계)", NCOLS,
              bg=C["blue"], size=11)
    ws.row_dimensions[sec2_row].height = 22

    months = [f"2026-{m:02d}" for m in range(1, 13)]
    hdr(ws, sec2_row + 1, 1, "월",     bg=C["blue"])
    hdr(ws, sec2_row + 1, 2, "매수건", bg=C["blue"])
    hdr(ws, sec2_row + 1, 3, "매수액", bg=C["blue"])
    hdr(ws, sec2_row + 1, 4, "매도건", bg=C["blue"])
    hdr(ws, sec2_row + 1, 5, "매도액", bg=C["blue"])
    hdr(ws, sec2_row + 1, 6, "순매수", bg=C["blue"])

    tl = f"'{TRADE}'!"
    for mi, m in enumerate(months):
        r = sec2_row + 2 + mi
        c = ws.cell(row=r, column=1, value=m)
        c.fill = _fill(C["lgray"]); c.border = _border(); c.alignment = _align("center")
        # 매수 건수: COUNTIFS
        c = ws.cell(row=r, column=2,
            value=f'=COUNTIFS({tl}A:A,">="&DATE(YEAR("{m}-01"),MONTH("{m}-01"),1),'
                  f'{tl}A:A,"<="&EOMONTH(DATE(YEAR("{m}-01"),MONTH("{m}-01"),1),0),'
                  f'{tl}C:C,"매수")')
        c.fill = _fill(C["formula"]); c.border = _border(); c.alignment = _align("right")
        # 매수액
        c = ws.cell(row=r, column=3,
            value=f'=SUMPRODUCT(({tl}A$4:{tl}A${NROWS+3}>=DATE(YEAR("{m}-01"),MONTH("{m}-01"),1))*'
                  f'({tl}A$4:{tl}A${NROWS+3}<=EOMONTH(DATE(YEAR("{m}-01"),MONTH("{m}-01"),1),0))*'
                  f'({tl}C$4:{tl}C${NROWS+3}="매수")*{tl}J$4:{tl}J${NROWS+3})')
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")
        # 매도 건수
        c = ws.cell(row=r, column=4,
            value=f'=COUNTIFS({tl}A:A,">="&DATE(YEAR("{m}-01"),MONTH("{m}-01"),1),'
                  f'{tl}A:A,"<="&EOMONTH(DATE(YEAR("{m}-01"),MONTH("{m}-01"),1),0),'
                  f'{tl}C:C,"매도")')
        c.fill = _fill(C["formula"]); c.border = _border(); c.alignment = _align("right")
        # 매도액
        c = ws.cell(row=r, column=5,
            value=f'=SUMPRODUCT(({tl}A$4:{tl}A${NROWS+3}>=DATE(YEAR("{m}-01"),MONTH("{m}-01"),1))*'
                  f'({tl}A$4:{tl}A${NROWS+3}<=EOMONTH(DATE(YEAR("{m}-01"),MONTH("{m}-01"),1),0))*'
                  f'({tl}C$4:{tl}C${NROWS+3}="매도")*{tl}J$4:{tl}J${NROWS+3})')
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")
        # 순매수 = 매수액 - 매도액
        c = ws.cell(row=r, column=6, value=f"=C{r}-E{r}")
        c.number_format = '+#,##0;-#,##0;0'; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A4"
    print(f"  ✓ '{SUMMARY}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — 투자아이디어
# 버핏의 체크리스트 + 퀀트 스코어 기반 발굴 프레임워크
# ══════════════════════════════════════════════════════════════════════════════
IDEA = "투자아이디어"

def make_idea_tracker(wb):
    ws = wb.create_sheet(IDEA)

    # ── 사용 설명 섹션 ─────────────────────────────────────────────────────
    NCOLS = 16
    title_row(ws, 1, "💡 투자아이디어 발굴 & 체크리스트", NCOLS, bg=C["orange"])
    note(ws, 2, 1,
         "★ 아이디어 발굴 → 체크리스트 → 투자결정 3단계 프로세스 | 상태별 필터 권장")
    note(ws, 2, 9,
         "【체크리스트 기준】 버핏: 이해가능성·경쟁우위·경영진·적정가·장기보유 "
         "| 켈리공식: 승률·배당배율 기반 적정 투자비중 계산 가능")

    # ── 컬럼 헤더 ─────────────────────────────────────────────────────────
    COLS = [
        # 기본정보
        ("등록일",       10, C["orange"]),
        ("분류",          10, C["orange"]),
        ("종목코드",      10, C["orange"]),
        ("종목명\n(자동)", 16, C["orange"]),
        # 밸류에이션
        ("현재가\n(자동)", 12, "C55A11"),
        ("목표주가",      12, "C55A11"),
        ("상승여력",      10, "C55A11"),
        ("손절가",        10, "C55A11"),
        # 체크리스트 (Yes/No)
        ("사업\n이해?",    9, C["blue"]),
        ("경쟁\n우위?",    9, C["blue"]),
        ("경영진\n신뢰?",  9, C["blue"]),
        ("적정가\n이하?",  9, C["blue"]),
        ("장기\n보유?",    9, C["blue"]),
        # 관리
        ("우선순위",      10, C["navy"]),
        ("진행상태",      12, C["navy"]),
        ("핵심근거·메모", 50, C["navy"]),
    ]

    for ci, (label, width, bg) in enumerate(COLS, 1):
        hdr(ws, 3, ci, label, bg=bg, wrap=True, size=9)
        col_w(ws, ci, width)
    ws.row_dimensions[3].height = 34
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NCOLS)}3"

    IDEA_ROWS = 200
    for r in range(4, IDEA_ROWS + 4):
        # 등록일
        c = ws.cell(row=r, column=1)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "YYYY-MM-DD"; c.alignment = _align("center")
        # 분류
        c = ws.cell(row=r, column=2)
        c.fill = _fill(C["input"]); c.border = _border(); c.alignment = _align("center")
        # 종목코드
        c = ws.cell(row=r, column=3)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "@"; c.alignment = _align("center")
        # 종목명 VLOOKUP
        c = ws.cell(row=r, column=4,
                    value=f"=IFERROR(VLOOKUP(C{r},'{KRX}'!$A:$B,2,FALSE),\"\")")
        c.fill = _fill(C["formula"]); c.border = _border()
        # 현재가 VLOOKUP
        c = ws.cell(row=r, column=5,
                    value=f"=IFERROR(VLOOKUP(C{r},'{KRX}'!$A:$C,3,FALSE),\"\")")
        c.number_format = "#,##0"
        c.fill = _fill(C["formula"]); c.border = _border(); c.alignment = _align("right")
        # 목표주가
        c = ws.cell(row=r, column=6)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 상승여력 = (목표가 - 현재가) / 현재가
        c = ws.cell(row=r, column=7,
                    value=f"=IF(OR(E{r}=\"\",F{r}=\"\"),\"\",(F{r}-E{r})/E{r})")
        c.number_format = "0.0%"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")
        # 손절가
        c = ws.cell(row=r, column=8)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 체크리스트 5항목 (Yes/No/미확인)
        for col in range(9, 14):
            c = ws.cell(row=r, column=col)
            c.fill = _fill(C["input"]); c.border = _border(); c.alignment = _align("center")
        # 우선순위, 진행상태
        for col in [14, 15]:
            c = ws.cell(row=r, column=col)
            c.fill = _fill(C["input"]); c.border = _border(); c.alignment = _align("center")
        # 핵심근거
        c = ws.cell(row=r, column=16)
        c.fill = _fill(C["input"]); c.border = _border()
        ws.row_dimensions[r].height = 18

    # 데이터 유효성
    r_end = IDEA_ROWS + 3
    dv_list(ws, f'"{",".join(IDEA_CATEGORY)}"',  f"B4:B{r_end}")
    for col_letter in ["I", "J", "K", "L", "M"]:
        dv_list(ws, f'"{",".join(YESNO)}"', f"{col_letter}4:{col_letter}{r_end}")
    dv_list(ws, f'"{",".join(IDEA_PRIORITY)}"',  f"N4:N{r_end}")
    dv_list(ws, f'"{",".join(IDEA_STATUS)}"',    f"O4:O{r_end}")

    # ── 투자 프레임워크 안내 섹션 ─────────────────────────────────────────
    guide_row = IDEA_ROWS + 6
    title_row(ws, guide_row, "📖 투자 아이디어 발굴 프레임워크 (참고용)", NCOLS,
              bg=C["navy"], size=11)
    ws.row_dimensions[guide_row].height = 22

    guide = [
        ("분류",        "발굴 방법",
         "주요 지표·기준"),
        ("가치투자",    "PER<10, PBR<1, ROE>15%, 부채비율<100%인 종목 스크리닝 (KRX 정보데이터시스템·증권사 HTS)",
         "안전마진 30%이상 | 5년 실적 안정성 확인"),
        ("성장주",      "매출성장률 20%+, 영업이익률 개선 추세, TAM 대비 시장점유율 확대 여부",
         "PEG<1 (PER÷성장률) | 내러티브와 숫자 모두 확인"),
        ("배당주",      "배당수익률>3%, 5년 배당성장, FCF 대비 배당성향<60%",
         "배당컷 리스크: 부채비율·이자보상배율 체크"),
        ("이벤트드리븐","합병·분할·지배구조 변경·실적 서프라이즈 직후 과잉반응 구간 매수",
         "이벤트 완료 후 3~6개월 모멘텀 추적"),
        ("섹터로테이션", "경기 사이클(초기:금융/소재→성장:IT/산업재→후기:에너지/헬스케어→침체:유틸리티/필수소비) 확인",
         "RSI 시트 섹터 데이터 활용 | 상대강도 분석"),
        ("리스크 관리",  "개별종목 5% 이하 | 섹터 집중도 30% 이하 | 분기마다 포트폴리오 리뷰",
         "손절선: 매수가 대비 -15~20% | 목표가 달성 시 일부 실현"),
    ]

    for ri, (cat, method, metric) in enumerate(guide):
        r = guide_row + 1 + ri
        ws.row_dimensions[r].height = 30
        c = ws.cell(row=r, column=1, value=cat)
        c.font = _f(bold=(ri == 0), size=9); c.fill = _fill(C["lblue"])
        c.border = _border(); c.alignment = _align("center", wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        c2 = ws.cell(row=r, column=2, value=method)
        c2.font = _f(bold=(ri == 0), size=9); c2.fill = _fill(C["lgray"])
        c2.border = _border(); c2.alignment = _align(h="left", wrap=True)
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=NCOLS)
        c3 = ws.cell(row=r, column=10, value=metric)
        c3.font = _f(italic=(ri > 0), size=9); c3.fill = _fill(C["yellow"])
        c3.border = _border(); c3.alignment = _align(h="left", wrap=True)

    print(f"  ✓ '{IDEA}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — 매도모니터링
# 매도 후 현재가 vs 매도가 비교, 재매수 판단
# ══════════════════════════════════════════════════════════════════════════════
SELL_MON = "매도모니터링"

def make_sell_monitor(wb):
    ws = wb.create_sheet(SELL_MON)
    NCOLS = 15

    title_row(ws, 1, "🔍 매도 후 모니터링 — 재매수 타이밍 판단", NCOLS, bg="4B0082")
    note(ws, 2, 1,
         "★ 매도 후에도 해당 종목을 추적해 재진입 기회 포착 "
         "| 현재가는 KRX종목마스터 시트에서 자동 참조 (시세 업데이트 필요)")

    COLS = [
        ("종목코드",    10),
        ("종목명\n(자동)", 18),
        ("계좌",        12),
        ("매수평균가",  12),
        ("매도일",      11),
        ("매도가",      12),
        ("매도수량",    10),
        ("매도금액",    14),
        ("실현손익",    12),
        ("실현수익률",  11),
        ("매도사유",    14),
        ("현재가\n(자동)", 12),
        ("매도후등락",  11),
        ("재매수검토",  12),
        ("재매수조건·메모", 35),
    ]

    for ci, (label, width) in enumerate(COLS, 1):
        hdr(ws, 3, ci, label, bg="5B2C8E", wrap=True, size=9)
        col_w(ws, ci, width)
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NCOLS)}3"

    MON_ROWS = 300
    for r in range(4, MON_ROWS + 4):
        # 종목코드
        c = ws.cell(row=r, column=1)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "@"; c.alignment = _align("center")
        # 종목명 VLOOKUP
        c = ws.cell(row=r, column=2,
                    value=f"=IFERROR(VLOOKUP(A{r},'{KRX}'!$A:$B,2,FALSE),\"\")")
        c.fill = _fill(C["formula"]); c.border = _border()
        # 계좌
        c = ws.cell(row=r, column=3)
        c.fill = _fill(C["input"]); c.border = _border(); c.alignment = _align("center")
        # 매수평균가
        c = ws.cell(row=r, column=4)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 매도일
        c = ws.cell(row=r, column=5)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "YYYY-MM-DD"; c.alignment = _align("center")
        # 매도가
        c = ws.cell(row=r, column=6)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 매도수량
        c = ws.cell(row=r, column=7)
        c.fill = _fill(C["input"]); c.border = _border()
        c.number_format = "#,##0"; c.alignment = _align("right")
        # 매도금액 = 매도가 × 수량
        c = ws.cell(row=r, column=8,
                    value=f"=IF(AND(F{r}<>\"\",G{r}<>\"\"),F{r}*G{r},\"\")")
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")
        # 실현손익 = (매도가 - 매수평균가) × 수량
        c = ws.cell(row=r, column=9,
                    value=f"=IF(OR(D{r}=\"\",F{r}=\"\",G{r}=\"\"),\"\",(F{r}-D{r})*G{r})")
        c.number_format = '+#,##0;-#,##0;"-"'
        c.fill = _fill(C["formula"]); c.border = _border(); c.alignment = _align("right")
        # 실현수익률
        c = ws.cell(row=r, column=10,
                    value=f"=IF(OR(D{r}=\"\",F{r}=\"\"),\"\",(F{r}-D{r})/D{r})")
        c.number_format = "0.0%"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")
        # 매도사유
        c = ws.cell(row=r, column=11)
        c.fill = _fill(C["input"]); c.border = _border(); c.alignment = _align("center")
        # 현재가 VLOOKUP
        c = ws.cell(row=r, column=12,
                    value=f"=IFERROR(VLOOKUP(A{r},'{KRX}'!$A:$C,3,FALSE),\"\")")
        c.number_format = "#,##0"; c.fill = _fill(C["formula"])
        c.border = _border(); c.alignment = _align("right")
        # 매도후등락 = (현재가 - 매도가) / 매도가
        c = ws.cell(row=r, column=13,
                    value=f"=IF(OR(F{r}=\"\",L{r}=\"\"),\"\",(L{r}-F{r})/F{r})")
        c.number_format = '+0.0%;-0.0%;"-"'
        c.fill = _fill(C["formula"]); c.border = _border(); c.alignment = _align("right")
        # 재매수검토
        c = ws.cell(row=r, column=14)
        c.fill = _fill(C["input"]); c.border = _border(); c.alignment = _align("center")
        # 메모
        c = ws.cell(row=r, column=15)
        c.fill = _fill(C["input"]); c.border = _border()
        ws.row_dimensions[r].height = 18

    dv_list(ws, f'"{",".join(ACCOUNTS)}"',    f"C4:C{MON_ROWS+3}")
    dv_list(ws, f'"{",".join(SELL_REASON)}"', f"K4:K{MON_ROWS+3}")
    dv_list(ws, f'"{",".join(REENTRY)}"',     f"N4:N{MON_ROWS+3}")

    print(f"  ✓ '{SELL_MON}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 — KRX 종목 마스터
# pykrx로 전 종목 로드 + 현재가 기록
# ══════════════════════════════════════════════════════════════════════════════

def make_krx_master(wb, skip_fetch=False):
    ws = wb.create_sheet(KRX)
    NCOLS = 6

    title_row(ws, 1, f"📑 KRX 종목 마스터 — VLOOKUP 참조용 | 업데이트: {datetime.now(KST).strftime('%Y-%m-%d %H:%M')}", NCOLS)
    note(ws, 2, 1,
         "★ 현재가는 Python 스크립트 실행 시 자동 업데이트 | "
         "실시간 아님 (마지막 거래일 종가) | python create_asset_manager.py --update 로 시세 갱신")

    COLS = [
        ("종목코드", 12, "center"),
        ("종목명",   22, "left"),
        ("현재가",   12, "right"),
        ("전일대비", 10, "right"),
        ("등락률",   10, "right"),
        ("시장구분", 10, "center"),
    ]
    for ci, (label, width, _) in enumerate(COLS, 1):
        hdr(ws, 3, ci, label)
        col_w(ws, ci, width)
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NCOLS)}3"

    if skip_fetch:
        note(ws, 4, 1, "→ python create_asset_manager.py --update 실행 시 KRX 종목 데이터 채워짐")
        print(f"  ✓ '{KRX}' 시트 생성 완료 (시세 로드 생략)")
        return

    print("  KRX 전 종목 데이터 로드 중 (1~2분 소요)...")
    try:
        from pykrx import stock as krx_stock
        today = datetime.now(KST).date()

        # 최근 거래일 탐색 (주말·공휴일 건너뜀)
        trade_date = None
        for delta in range(7):
            d = today - timedelta(days=delta)
            date_str = d.strftime("%Y%m%d")
            try:
                df = krx_stock.get_market_ohlcv_by_ticker(date_str, market="ALL")
                if not df.empty:
                    trade_date = date_str
                    break
            except Exception:
                continue

        if trade_date is None:
            raise RuntimeError("최근 7일 거래일 데이터 없음")

        print(f"    기준일: {trade_date} | 종목 수: {len(df)}")

        _write_krx_rows(ws, df, trade_date, krx_stock)

    except Exception as e:
        print(f"  ⚠ KRX 데이터 로드 실패: {e}")
        note(ws, 4, 1, f"KRX 로드 실패: {e} | --update 로 재시도하세요")

    print(f"  ✓ '{KRX}' 시트 생성 완료")


def _write_krx_rows(ws, df, trade_date, krx_stock):
    """DataFrame → KRX마스터 시트에 쓰기."""
    from pykrx import stock as krx_stock

    # 시장구분 맵핑
    kospi_tickers  = set(krx_stock.get_market_ticker_list(trade_date, market="KOSPI"))
    kosdaq_tickers = set(krx_stock.get_market_ticker_list(trade_date, market="KOSDAQ"))

    _HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    _HDR_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    _DAT_FONT = Font(name="맑은 고딕", size=10)
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(top=thin, left=thin, right=thin, bottom=thin)

    price_col  = "종가" if "종가" in df.columns else ("Close" if "Close" in df.columns else df.columns[3])
    change_col = "등락률" if "등락률" in df.columns else None

    for row_idx, (ticker, row_data) in enumerate(df.iterrows(), 4):
        name = krx_stock.get_market_ticker_name(str(ticker))
        price = int(row_data[price_col]) if price_col in df.columns else 0
        chg_rate = float(row_data[change_col]) / 100 if change_col and change_col in df.columns else None

        if ticker in kospi_tickers:
            market = "KOSPI"
        elif ticker in kosdaq_tickers:
            market = "KOSDAQ"
        else:
            market = "기타"

        vals = [str(ticker), name, price, None, chg_rate, market]
        fmts = ["@", None, "#,##0", "+#,##0;-#,##0", "+0.00%;-0.00%", None]
        aligns = ["center", "left", "right", "right", "right", "center"]

        for ci, (val, fmt, al) in enumerate(zip(vals, fmts, aligns), 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font = _DAT_FONT
            c.border = bdr
            c.alignment = Alignment(horizontal=al, vertical="center")
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[row_idx].height = 16

    # 타이틀 업데이트
    ws.cell(row=1, column=1).value = (
        f"📑 KRX 종목 마스터 — VLOOKUP 참조용 | "
        f"기준일: {trade_date} | 총 {len(df)}종목"
    )


# ══════════════════════════════════════════════════════════════════════════════
# --update 모드: 기존 파일의 KRX 시세만 업데이트
# ══════════════════════════════════════════════════════════════════════════════

def update_krx_prices():
    if not os.path.exists(OUTPUT_FILE):
        print(f"파일 없음: {OUTPUT_FILE} — 먼저 python create_asset_manager.py 실행")
        return

    print(f"KRX 시세 업데이트: {OUTPUT_FILE}")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if KRX not in wb.sheetnames:
        print("KRX종목마스터 시트 없음")
        return

    ws = wb[KRX]
    # 기존 데이터 행 삭제 (헤더 3행 유지)
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)

    try:
        from pykrx import stock as krx_stock
        today = datetime.now(KST).date()
        for delta in range(7):
            d = today - timedelta(days=delta)
            date_str = d.strftime("%Y%m%d")
            try:
                df = krx_stock.get_market_ohlcv_by_ticker(date_str, market="ALL")
                if not df.empty:
                    _write_krx_rows(ws, df, date_str, krx_stock)
                    print(f"  {len(df)}종목 업데이트 완료 (기준일: {date_str})")
                    break
            except Exception:
                continue
    except Exception as e:
        print(f"업데이트 실패: {e}")

    wb.save(OUTPUT_FILE)
    print(f"저장 완료: {OUTPUT_FILE}")


# ══════════════════════════════════════════════════════════════════════════════
# main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--update",      action="store_true", help="KRX 시세만 업데이트")
    parser.add_argument("--skip-krx",   action="store_true", help="KRX 로드 생략 (빠른 생성)")
    args = parser.parse_args()

    if args.update:
        update_krx_prices()
        return

    print(f"자산관리_2026.xlsx 생성 중...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 시트 순서
    make_trade_log(wb)
    make_cash_flow(wb)
    make_account_summary(wb)
    make_idea_tracker(wb)
    make_sell_monitor(wb)
    make_krx_master(wb, skip_fetch=args.skip_krx)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 저장 완료: {OUTPUT_FILE}")
    print(f"   시트 목록: {wb.sheetnames}")
    print()
    print("【다음 단계】")
    print("  1. 파일을 열고 KRX종목마스터 시트 확인")
    print("  2. 투자거래내역·입출금내역에 데이터 입력")
    print("  3. 계좌별현황은 자동 집계됨")
    print("  4. 시세 갱신 시: python create_asset_manager.py --update")
    print("  5. 빠른 재생성(KRX 제외): python create_asset_manager.py --skip-krx")


if __name__ == "__main__":
    main()
