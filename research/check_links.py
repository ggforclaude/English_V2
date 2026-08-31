"""
Daily_index.xlsx 리서치 보고서 시트 링크 유효성 검사 및 자동 수정
- researchId와 hyperlink URL 불일치 자동 수정
- HTTP 상태 코드 샘플 점검 (날짜 범위 내 전체)
사용: python check_links.py
"""
import asyncio
import aiohttp
import openpyxl
from openpyxl.styles import Font
import os
import re

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "Daily_index.xlsx")
SHEET_NAME = "리서치 보고서"
URL_BASE   = "https://m.stock.naver.com/research/company/"
START_DATE = "2026-03-31"
END_DATE   = "2026-04-29"
CONCURRENCY = 10   # 동시 요청 수

_LNK_FONT = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")


# ── 1단계: hyperlink URL 정합성 수정 ─────────────────────────────────────────

def fix_all_hyperlinks(ws) -> int:
    """researchId(K열=col11)와 링크 URL이 맞지 않는 셀을 전부 재구성."""
    fixed = 0
    for row_num in range(2, ws.max_row + 1):
        rid = str(ws.cell(row=row_num, column=11).value or "").strip()  # col 11 = ID
        if not rid:
            continue
        link_cell = ws.cell(row=row_num, column=9)
        correct_url = f"{URL_BASE}{rid}"
        hl = link_cell.hyperlink
        current_url = (hl.target or "") if hl else ""
        if current_url == correct_url and link_cell.value == "보기":
            continue
        link_cell.value = "보기"
        link_cell.hyperlink = correct_url
        link_cell.font = _LNK_FONT
        fixed += 1
    return fixed


# ── 2단계: HTTP 상태 확인 ─────────────────────────────────────────────────────

async def _check(session, rid: str, sem: asyncio.Semaphore):
    url = f"{URL_BASE}{rid}"
    async with sem:
        try:
            async with session.get(
                url,
                allow_redirects=True,
                timeout=aiohttp.ClientTimeout(total=12),
            ) as r:
                return rid, r.status
        except Exception as e:
            return rid, f"ERR({e})"


async def check_http(ws) -> tuple[int, list]:
    """날짜 범위 내 리포트 링크 HTTP 상태 일괄 확인."""
    rids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d   = str(row[0] or "")
        rid = str(row[10] or "").strip()  # col 11 (index 10) = researchId
        if START_DATE <= d <= END_DATE and rid:
            rids.append(rid)

    if not rids:
        print(f"  날짜 범위({START_DATE}~{END_DATE})에 데이터 없음")
        return 0, []

    print(f"  HTTP 점검 대상: {len(rids)}개 링크")

    sem = asyncio.Semaphore(CONCURRENCY)
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
    }
    async with aiohttp.ClientSession(headers=headers) as session:
        results = await asyncio.gather(*[_check(session, rid, sem) for rid in rids])

    ok      = sum(1 for _, s in results if s == 200)
    broken  = [(rid, s) for rid, s in results if s != 200]
    return ok, broken


# ── 메인 ─────────────────────────────────────────────────────────────────────

async def main():
    print(f"파일: {EXCEL_PATH}\n")

    print("[1/3] 파일 로드...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"  오류: '{SHEET_NAME}' 시트 없음. 시트 목록: {wb.sheetnames}")
        return
    ws = wb[SHEET_NAME]
    total_rows = ws.max_row - 1
    print(f"  전체 행 수: {total_rows:,}개")

    print("\n[2/3] hyperlink URL 정합성 수정...")
    fixed = fix_all_hyperlinks(ws)
    if fixed:
        wb.save(EXCEL_PATH)
        print(f"  수정 완료: {fixed}개 링크 재구성 → 저장")
    else:
        print("  이상 없음 (수정 불필요)")

    print(f"\n[3/3] HTTP 상태 점검 ({START_DATE} ~ {END_DATE})...")
    ok, broken = await check_http(ws)

    print(f"\n  결과: 정상 {ok}개 / 오류 {len(broken)}개")
    if broken:
        print("  오류 링크 (최대 30개):")
        for rid, status in broken[:30]:
            print(f"    {status}  →  {URL_BASE}{rid}")
        if len(broken) > 30:
            print(f"    ... 외 {len(broken) - 30}개")
    else:
        print("  모든 링크 정상!")

    print("\n완료.")


if __name__ == "__main__":
    asyncio.run(main())
