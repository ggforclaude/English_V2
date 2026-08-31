"""
seibro 종목보유현황 → 대표종목별 ETF 후보 선정 → pykrx 편입 상세 → Excel

파이프라인:
  1. sector_tops_cache.json 에서 대표종목(최대 72개) 로드
  2. 세이브로(seibro.or.kr) 에서 각 종목 보유 ETF 상위 5개 조회 (CU당보유비중 기준)
  3. 중복 제거 → 후보 ETF 목록 (예상 100~200개)
  4. pykrx 로 후보 ETF 편입 상위 10종목 조회 (주식코드 6자리만, 채권 제외)
  5. Excel 생성:
     - Sheet "대표종목별 ETF"  : 종목별 상위 ETF 5개 및 보유비중 (편입정보없는 ETF 제외)
     - Sheet "ETF 편입 현황"   : ETF 1행, 편입1~10순위+비중, 대표종목 컬러 표시

실행:
  python etf_analysis.py           # 단독 실행 → etf_analysis.xlsx
  python etf_analysis.py --excel   # 기존 캐시로 Excel만 재생성

daily_update.py 월요일 통합:
  etf_analysis.run_weekly_update() → Daily_index 주간파일에 시트 추가
"""
import os, json, time, logging, sys, threading
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import xml.etree.ElementTree as ET
import pandas as pd
from pykrx import stock as krx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pytz

# ── 설정 ─────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
TOPS_CACHE  = os.path.join(BASE_DIR, "sector_tops_cache.json")
HOLD_CACHE  = os.path.join(BASE_DIR, "etf_holdings_cache.json")
OUTPUT_PATH = os.path.join(BASE_DIR, "etf_analysis.xlsx")  # 단독 실행 시 출력
KST         = pytz.timezone("Asia/Seoul")


def _get_weekly_excel_path() -> str:
    """주간 Daily_index 파일 경로 (sector_rsi.py 와 동일한 로직)"""
    today  = datetime.now(KST).date()
    monday = today - timedelta(days=today.weekday())
    m = monday.month
    w = (monday.day - 1) // 7 + 1
    y = monday.year
    return os.path.join(BASE_DIR, f"Daily_index_{m}m{w}w{y}.xlsx")
MAX_HOLD    = 10   # ETF당 편입종목 표시 개수
TOP_N       = 5    # 종목당 상위 ETF 개수
WORKERS     = 6    # pykrx 병렬 스레드 수

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(message)s",
                    datefmt="%Y-%m-%d %H:%M:%S")
log = logging.getLogger(__name__)

# ── seibro API ────────────────────────────────────────────────────────────────
_SEIBRO_PAGE = ("https://seibro.or.kr/websquare/control.jsp"
                "?w2xPath=/IPORTAL/user/etf/BIP_CNTS06036V.xml&menuNo=186")
_SEIBRO_API  = "https://seibro.or.kr/websquare/engine/proworks/callServletService.jsp"
_SEIBRO_HDRS = {
    "Content-Type":    "application/xml; charset=UTF-8",
    "Referer":         _SEIBRO_PAGE,
    "Origin":          "https://seibro.or.kr",
    "X-Requested-With":"XMLHttpRequest",
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}


def _seibro_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": _SEIBRO_HDRS["User-Agent"]})
    try:
        s.get(_SEIBRO_PAGE, timeout=15)
    except Exception:
        pass
    return s


def _seibro_xml(action: str, stock_code: str, std_dt: str,
                start: int = 1, end: int = TOP_N) -> bytes:
    body = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<reqParam action="{action}" task="ksd.safe.bip.cnts.etf.process.EtfMartInfoPTask">'
        '<MENU_NO value="186"/>'
        '<CMM_BTN_ABBR_NM value="total,search,openall,print,hwp,word,pdf,searchIcon,seach,"/>'
        '<W2XPATH value="/IPORTAL/user/etf/BIP_CNTS06036V.xml"/>'
        f'<VCTF_ISIN value="{stock_code}"/>'
        f'<STD_DT value="{std_dt}"/>'
        '<radio2 value="0"/>'
        f'<START_PAGE value="{start}"/>'
        f'<END_PAGE value="{end}"/>'
        '</reqParam>'
    )
    return body.encode("utf-8")


def _get(el, tag: str, attr: str = "value", default="") -> str:
    node = el.find(tag)
    return node.get(attr, default) if node is not None else default


def query_seibro_etfs(session: requests.Session, stock_code: str,
                      std_dt: str, top_n: int = TOP_N) -> list[dict]:
    """세이브로: 특정 종목을 보유한 ETF 상위 top_n개 (CU당보유비중 내림차순)"""
    try:
        r = session.post(
            _SEIBRO_API,
            data=_seibro_xml("secnHoldStatStkPList", stock_code, std_dt, 1, top_n),
            headers=_SEIBRO_HDRS,
            timeout=20,
        )
        root = ET.fromstring(r.text)
    except Exception as e:
        log.warning(f"  [seibro] {stock_code} 조회 오류: {e}")
        return []

    results = []
    for data in root.findall("data"):
        res = data.find("result")
        if res is None:
            continue
        isin_val = _get(res, "ISIN")               # ETF ISIN (예: KR7448330001)
        etf_code = isin_val[3:9] if len(isin_val) >= 9 else ""
        if not etf_code.isdigit():
            continue
        try:
            hold = float(_get(res, "HOLD_IMPO") or 0)
            aum  = int(_get(res, "NETASST_TOTAMT") or 0)
        except ValueError:
            hold, aum = 0.0, 0
        results.append({
            "ticker":    etf_code,
            "name":      _get(res, "KOR_SECN_NM"),
            "manager":   _get(res, "MNGCO_NM"),
            "aum":       aum,
            "hold_impo": hold,
        })
    return results


# ── STEP 1: 대표종목 로드 ─────────────────────────────────────────────────────

def load_rep_stocks() -> dict:
    """sector_tops_cache.json → {sector: {"codes": [...], "names": [...]}}"""
    if not os.path.exists(TOPS_CACHE):
        log.error(f"캐시 없음: {TOPS_CACHE}")
        return {}
    with open(TOPS_CACHE, encoding="utf-8") as f:
        return json.load(f)


def load_rep_for_coloring() -> dict[str, set[str]]:
    """색상 표시용: {"대표1": {codes}, "대표2": {codes}, "대표3": {codes}}"""
    tops = load_rep_stocks()
    rep: dict[str, set[str]] = {"대표1": set(), "대표2": set(), "대표3": set()}
    for d in tops.values():
        codes = d.get("codes", [])
        for i, label in enumerate(["대표1", "대표2", "대표3"]):
            if i < len(codes) and codes[i] and codes[i] != "-":
                rep[label].add(str(codes[i]))
    return rep


# ── STEP 2: seibro로 ETF 후보 목록 구성 ──────────────────────────────────────

def get_valid_std_dt(session: requests.Session) -> str:
    """세이브로에 실제 데이터가 있는 가장 최근 거래일 탐색 (최대 7일 전까지)"""
    today = datetime.now(KST).date()
    for days_back in range(1, 8):
        d = today - timedelta(days=days_back)
        if d.weekday() >= 5:
            continue
        dt_str = d.strftime("%Y%m%d")
        try:
            body = _seibro_xml("secnHoldStatStkPList", "005930", dt_str, 1, 1)
            r    = session.post(_SEIBRO_API, data=body, headers=_SEIBRO_HDRS, timeout=12)
            root = ET.fromstring(r.text)
            if len(root.findall("data")) > 0:
                return dt_str
        except Exception:
            pass
    return (today - timedelta(days=1)).strftime("%Y%m%d")


def build_candidate_list(tops: dict, std_dt: str = "") -> dict[str, dict]:
    """
    대표종목별 상위 ETF 세이브로 조회 → 유니크 ETF 딕셔너리 반환
    {ticker: {name, aum, manager, sources: [{sector, stock_code, stock_name, rank, hold_impo}]}}
    """
    session = _seibro_session()
    candidates: dict[str, dict] = {}

    all_stocks = []
    for sector, d in tops.items():
        codes = d.get("codes", [])
        names = d.get("names", [])
        for rank, (code, name) in enumerate(zip(codes, names), 1):
            if code and code != "-":
                all_stocks.append((sector, code, name, rank))

    std_dt = get_valid_std_dt(session)
    log.info(f"세이브로 기준일: {std_dt}")
    log.info(f"세이브로 조회 시작: {len(all_stocks)}개 대표종목 × 상위 {TOP_N}개 ETF")

    for i, (sector, code, name, rank) in enumerate(all_stocks, 1):
        etfs = query_seibro_etfs(session, code, std_dt, TOP_N)
        for etf in etfs:
            t = etf["ticker"]
            if t not in candidates:
                candidates[t] = {
                    "name":    etf["name"],
                    "aum":     etf["aum"],
                    "manager": etf["manager"],
                    "sources": [],
                }
            candidates[t]["sources"].append({
                "sector":     sector,
                "stock_code": code,
                "stock_name": name,
                "rank":       rank,
                "hold_impo":  etf["hold_impo"],
            })
        print(f"  [{i:2d}/{len(all_stocks)}] {sector} {name}({code}): {len(etfs)}개 ETF 발견", end="\r")
        time.sleep(0.35)

    print()
    log.info(f"세이브로 완료: 총 {len(candidates)}개 고유 ETF 후보")
    return candidates


# ── STEP 3: pykrx 편입 종목 조회 ─────────────────────────────────────────────

def fetch_holdings(ticker: str) -> list[dict] | None:
    """pykrx로 ETF 편입 종목 상위 MAX_HOLD개 반환.
    6자리 숫자 코드(주식)만 포함 → 채권·인프라 등 비주식 편입은 자동 제외.
    실패 또는 주식 편입이 없으면 None 반환.
    """
    try:
        df = krx.get_etf_portfolio_deposit_file(ticker)
    except Exception:
        return None
    if df is None or df.empty:
        return None

    df.index    = df.index.astype(str)
    name_col    = df.columns[0]
    weight_col  = df.columns[-1]
    df[weight_col] = pd.to_numeric(df[weight_col], errors="coerce")
    df = df[df[weight_col] > 0].dropna(subset=[weight_col])
    # 6자리 숫자 코드(주식)만 유지 — 채권(ISIN), 기타 제외
    df = df[df.index.str.match(r"^\d{6}$")]
    if df.empty:
        return None

    df = df.sort_values(weight_col, ascending=False).head(MAX_HOLD)
    return [
        {"code": str(idx), "name": str(row[name_col]), "weight": float(row[weight_col])}
        for idx, row in df.iterrows()
    ]


def load_hold_cache() -> dict:
    if os.path.exists(HOLD_CACHE):
        try:
            with open(HOLD_CACHE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_hold_cache(cache: dict) -> None:
    tmp = HOLD_CACHE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)
    os.replace(tmp, HOLD_CACHE)


def collect_holdings(candidates: dict[str, dict]) -> dict:
    """후보 ETF별 편입 종목 수집 (캐시 재사용)"""
    cache = load_hold_cache()
    pending = [(t, info) for t, info in candidates.items() if t not in cache]
    already = len(candidates) - len(pending)
    log.info(f"pykrx 편입 조회: 총 {len(candidates)}개 | 기수집 {already}개 | 신규 {len(pending)}개")

    if not pending:
        return cache

    done = 0

    def _worker(item):
        ticker, info = item
        h = fetch_holdings(ticker)
        return ticker, info, h

    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = {ex.submit(_worker, item): item for item in pending}
        for fut in as_completed(futs):
            ticker, info, holdings = fut.result()
            cache[ticker] = {
                "name":     info.get("name", ""),
                "aum":      info.get("aum", 0),
                "manager":  info.get("manager", ""),
                "sources":  info.get("sources", []),
                "holdings": holdings or [],
            }
            done += 1
            if done % 10 == 0 or done == len(pending):
                save_hold_cache(cache)
                log.info(f"  진행: {already + done}/{len(candidates)}")

    return cache


# ── STEP 4: Excel 생성 ────────────────────────────────────────────────────────

# 스타일 상수
_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT  = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
_DAT_FONT  = Font(name="맑은 고딕", size=10)
_CENTER    = Alignment(horizontal="center", vertical="center")
_LEFT      = Alignment(horizontal="left",   vertical="center")

_R1_NAME   = PatternFill("solid", fgColor="FFD700")   # 금  - 대표1 종목명
_R1_WGHT   = PatternFill("solid", fgColor="FFE97F")   # 연금 - 대표1 비중
_R2_NAME   = PatternFill("solid", fgColor="70AD47")   # 초록 - 대표2 종목명
_R2_WGHT   = PatternFill("solid", fgColor="A9D18E")   # 연초 - 대표2 비중
_R3_NAME   = PatternFill("solid", fgColor="2E75B6")   # 파랑 - 대표3 종목명
_R3_WGHT   = PatternFill("solid", fgColor="9DC3E6")   # 연파 - 대표3 비중


def _rep_fills(code: str, rep: dict[str, set[str]]):
    if code in rep.get("대표1", set()): return _R1_NAME, _R1_WGHT
    if code in rep.get("대표2", set()): return _R2_NAME, _R2_WGHT
    if code in rep.get("대표3", set()): return _R3_NAME, _R3_WGHT
    return None, None


def _hdr(ws, row, col, val, width=None):
    c = ws.cell(row, col, val)
    c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CENTER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def build_sheet1_summary(wb, tops: dict, cache: dict) -> None:
    """Sheet1: 대표종목별 ETF 상위 5개 및 CU당보유비중"""
    ws = wb.create_sheet("대표종목별 ETF")

    # 헤더
    base = [("섹터", 14), ("종목명", 14), ("코드", 8), ("구분", 6)]
    for col, (h, w) in enumerate(base, 1):
        _hdr(ws, 1, col, h, w)

    for i in range(1, TOP_N + 1):
        c = (i - 1) * 2 + len(base) + 1
        _hdr(ws, 1, c,     f"ETF{i}명",  22)
        _hdr(ws, 1, c + 1, f"비중{i}(%)", 8)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # 편입정보있는 ETF만 역방향 인덱스: stock_code → [(hold_impo, ticker, etf_name)]
    stock_to_etf: dict[str, list[tuple]] = {}
    for ticker, info in cache.items():
        if not info.get("holdings"):   # 편입정보없는 ETF 제외
            continue
        for src in info.get("sources", []):
            code = src["stock_code"]
            stock_to_etf.setdefault(code, []).append(
                (src["hold_impo"], ticker, info["name"])
            )
    for code in stock_to_etf:
        stock_to_etf[code].sort(reverse=True)

    row = 2
    rank_labels = ["대표1", "대표2", "대표3"]
    for sector, d in tops.items():
        codes = d.get("codes", [])
        names = d.get("names", [])
        for ri, (code, name) in enumerate(zip(codes, names)):
            if not code or code == "-":
                continue
            ws.cell(row, 1, sector).font = _DAT_FONT
            ws.cell(row, 1).alignment = _CENTER
            ws.cell(row, 2, name).font = _DAT_FONT
            ws.cell(row, 2).alignment = _LEFT
            ws.cell(row, 3, code).font = _DAT_FONT
            ws.cell(row, 3).alignment = _CENTER
            lbl = rank_labels[ri] if ri < 3 else f"대표{ri+1}"
            ws.cell(row, 4, lbl).font = _DAT_FONT
            ws.cell(row, 4).alignment = _CENTER

            etfs = stock_to_etf.get(code, [])[:TOP_N]
            for j, (hold, ticker, etf_name) in enumerate(etfs):
                c = (j) * 2 + len(base) + 1
                ws.cell(row, c,     etf_name).font = _DAT_FONT
                ws.cell(row, c).alignment = _LEFT
                wc = ws.cell(row, c + 1, hold / 100)
                wc.number_format = "0.00%"
                wc.font = _DAT_FONT
                wc.alignment = _CENTER

            ws.row_dimensions[row].height = 16
            row += 1

    log.info(f"Sheet '대표종목별 ETF' 생성: {row - 2}행")


def build_sheet2_etf(wb, cache: dict, rep: dict[str, set[str]]) -> None:
    """Sheet2: ETF 1행, 편입 1~10순위 + 비중, 대표종목 컬러"""
    ws = wb.create_sheet("ETF 편입 현황")

    now_str    = datetime.now(KST).strftime("%Y-%m-%d %H:%M")
    total_cols = 4 + MAX_HOLD * 2
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    info_c = ws.cell(1, 1,
        f"생성: {now_str}  |  출처: SEIBro + pykrx(KRX)  |  "
        "■ 금=대표1(시총1위)  ■ 초록=대표2(시총2위)  ■ 파랑=대표3(직전주최고상승)")
    info_c.font      = Font(name="맑은 고딕", size=9, italic=True, color="444444")
    info_c.alignment = _LEFT
    ws.row_dimensions[1].height = 14

    # 헤더 (행 2)
    base_h = [("ETF명", 28), ("티커", 8), ("운용사", 14), ("AUM(억)", 10)]
    for col, (h, w) in enumerate(base_h, 1):
        _hdr(ws, 2, col, h, w)
    for i in range(1, MAX_HOLD + 1):
        c = (i - 1) * 2 + len(base_h) + 1
        _hdr(ws, 2, c,     f"편입{i}순위", 18)
        _hdr(ws, 2, c + 1, f"비중{i}(%)",   8)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}2"

    # 데이터 (AUM 내림차순, 편입정보없는 ETF 제외)
    ri = 3
    for ticker, info in sorted(cache.items(),
                                key=lambda x: x[1].get("aum", 0), reverse=True):
        etf_name = info.get("name", "")
        aum      = info.get("aum", 0)
        manager  = info.get("manager", "")
        holdings = info.get("holdings", [])
        if not holdings:               # 채권혼합·인프라 등 주식 편입 없는 ETF 제외
            continue

        ws.cell(ri, 1, etf_name).font = _DAT_FONT
        ws.cell(ri, 1).alignment = _LEFT
        ws.cell(ri, 2, ticker).font   = _DAT_FONT
        ws.cell(ri, 2).alignment = _CENTER
        ws.cell(ri, 3, manager).font  = _DAT_FONT
        ws.cell(ri, 3).alignment = _LEFT
        aum_cell = ws.cell(ri, 4, round(aum / 1e8) if aum else "")
        aum_cell.font = _DAT_FONT; aum_cell.alignment = _CENTER
        if aum:
            aum_cell.number_format = "#,##0"

        for i, h in enumerate(holdings):
            c        = i * 2 + len(base_h) + 1
            code     = h.get("code", "")
            name     = h.get("name", "")
            weight   = h.get("weight", 0.0)
            nf, wf   = _rep_fills(code, rep)

            nc = ws.cell(ri, c, name)
            nc.font      = Font(name="맑은 고딕", size=10, bold=bool(nf))
            nc.alignment = _LEFT
            if nf: nc.fill = nf

            wc = ws.cell(ri, c + 1, weight / 100)
            wc.number_format = "0.00%"
            wc.font      = Font(name="맑은 고딕", size=10, bold=bool(wf))
            wc.alignment = _CENTER
            if wf: wc.fill = wf

        if not holdings:
            nc = ws.cell(ri, len(base_h) + 1, "편입정보없음")
            nc.font = Font(name="맑은 고딕", size=9, color="AAAAAA", italic=True)
            nc.alignment = _LEFT

        ws.row_dimensions[ri].height = 16
        ri += 1

    log.info(f"Sheet 'ETF 편입 현황' 생성: {ri - 3}행")


_ETF_SHEETS = ["대표종목별 ETF", "ETF 편입 현황"]


def build_excel(cache: dict, tops: dict, rep: dict[str, set[str]],
                excel_path: str = OUTPUT_PATH) -> None:
    """cache 데이터로 ETF 시트를 생성/갱신. 기존 파일이 있으면 시트만 교체."""
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    for name in _ETF_SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    build_sheet1_summary(wb, tops, cache)
    build_sheet2_etf(wb, cache, rep)
    wb.save(excel_path)
    log.info(f"저장 완료: {excel_path}")


# ── 진입점 ────────────────────────────────────────────────────────────────────

def run_weekly_update() -> None:
    """daily_update.py 에서 월요일에 호출.
    세이브로 + pykrx 전체 수집 후 주간 Daily_index 파일에 ETF 시트 추가."""
    excel_path = _get_weekly_excel_path()
    log.info(f"=== ETF 분석 (월요일 주간파일 통합): {excel_path} ===")

    tops = load_rep_stocks()
    if not tops:
        log.error("sector_tops_cache.json 없음 — sector_rsi 먼저 실행 필요")
        return

    rep        = load_rep_for_coloring()
    candidates = build_candidate_list(tops)
    cache      = collect_holdings(candidates)
    save_hold_cache(cache)

    build_excel(cache, tops, rep, excel_path=excel_path)
    log.info("=== ETF 분석 완료 ===")


def main():
    excel_only = "--excel" in sys.argv

    log.info("=" * 60)
    log.info(f"ETF 분석 {'(Excel 재생성)' if excel_only else '(전체 실행)'} → {OUTPUT_PATH}")
    log.info("=" * 60)

    tops = load_rep_stocks()
    if not tops:
        return

    rep = load_rep_for_coloring()
    total_rep = sum(len(v) for v in rep.values())
    log.info(f"대표종목: {total_rep}개 ({len(tops)}개 섹터)")

    if excel_only:
        cache = load_hold_cache()
        if not cache:
            log.error("캐시 없음 — 먼저 python etf_analysis.py 를 실행하세요")
            return
    else:
        candidates = build_candidate_list(tops)
        cache      = collect_holdings(candidates)
        save_hold_cache(cache)

    log.info(f"총 {len(cache)}개 ETF → Excel 생성 중...")
    build_excel(cache, tops, rep)
    log.info("완료")


if __name__ == "__main__":
    main()
