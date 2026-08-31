"""
네이버 증권 리서치 보고서 → Excel 업데이트
사용법:
  python fetch_research.py                        # 최근 2일 업데이트 (스케줄러용)
  python fetch_research.py 2026-04-01 2026-04-27  # 날짜 범위 지정 (초기 로드용)
  python fetch_research.py --backfill-pdf         # 현재 주차 파일 PDF 링크 소급 적용
  python fetch_research.py --backfill-pdf all     # 모든 Daily_index*.xlsx 소급 적용
  python fetch_research.py --backfill-pdf Daily_index.xlsx  # 특정 파일 소급 적용
"""
import asyncio
import aiohttp
import sys
import os
import json
import glob
import logging
from datetime import datetime, timedelta, date
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 설정 ──────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))


def _get_excel_path() -> str:
    """현재 주차 Excel 파일 경로 반환."""
    KST    = pytz.timezone("Asia/Seoul")
    today  = datetime.now(KST).date()
    monday = today - timedelta(days=today.weekday())
    m = monday.month
    w = (monday.day - 1) // 7 + 1
    y = monday.year
    return os.path.join(BASE_DIR, f"Daily_index_{m}m{w}w{y}.xlsx")


EXCEL_PATH = _get_excel_path()
LOG_PATH     = os.path.join(BASE_DIR, "fetch_log.txt")
PENDING_DIR  = os.path.join(BASE_DIR, "pending")
API_BASE    = "https://m.stock.naver.com/api/research"
API_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://m.stock.naver.com/research",
    "Accept":  "application/json",
}
KST       = pytz.timezone("Asia/Seoul")
PAGE_SIZE = 50
MAX_AUTO_LOOKBACK_DAYS = 21  # 자동 모드 공백 보정 상한 (비정상적으로 큰 공백은 경고 후 제한)

# 컬럼 정의: (API 필드명, 헤더 표시명, 열 너비)
COLS = [
    ("writeDate",     "작성일",         12),
    ("itemCode",      "종목코드",       10),
    ("itemName",      "종목명",         16),
    ("brokerName",    "증권사",         16),
    ("opinion",       "투자의견",       10),
    ("goalPrice",     "목표주가",       12),
    ("prevGoalPrice", "발간 시점 종가", 14),
    ("title",         "제목",           65),
    ("endUrl",        "링크",            6),
    ("_pdfUrl",       "PDF",             6),   # col 10: 원문 PDF 직접 다운로드
    ("researchId",    "ID",             10),   # col 11
    ("_divergence",   "괴리율",         10),   # col 12
]

# ── 로거 ──────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ── API 수집 ──────────────────────────────────────────────────────────────────

async def _fetch_page(session: aiohttp.ClientSession, page: int) -> list:
    url = f"{API_BASE}/company?page={page}&pageSize={PAGE_SIZE}"
    async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as r:
        return await r.json(content_type=None)


async def fetch_reports(start_date: str, end_date: str) -> list[dict]:
    """start_date ~ end_date 범위의 국내 종목 리서치 보고서 수집 후 상세 정보(투자의견/목표주가/PDF) 첨부."""
    all_reports: list[dict] = []

    async with aiohttp.ClientSession(headers=API_HEADERS) as session:
        page = 1
        while True:
            try:
                items = await _fetch_page(session, page)
            except Exception as e:
                log.warning(f"페이지 {page} 요청 오류: {e}")
                break

            if not items:
                break

            stop = False
            for item in items:
                d = item.get("writeDate", "")
                if d > end_date:
                    continue
                if d < start_date:
                    stop = True
                    break
                all_reports.append(item)

            print(f"  페이지 {page:3d} | 수집 {len(all_reports):5d}개", end="\r")

            if stop or len(items) < PAGE_SIZE:
                break

            page += 1
            await asyncio.sleep(0.3)

        print()
        # 리포트 상세 API에서 투자의견/목표주가/PDF URL 병렬 수집
        if all_reports:
            await _attach_details(session, all_reports)

    return all_reports


# ── 리포트 상세 정보 수집 ──────────────────────────────────────────────────────

_DETAIL_BASE  = "https://m.stock.naver.com/api/research/company/"
_DETAIL_SEM_SIZE = 20  # 동시 요청 수


async def _fetch_detail(
    session: aiohttp.ClientSession,
    rid: str | int,
    sem: asyncio.Semaphore,
) -> tuple:
    """리포트 상세 API에서 투자의견/목표주가/PDF URL 추출. 실패 시 빈 값 반환."""
    url = f"{_DETAIL_BASE}{rid}"
    async with sem:
        try:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as r:
                data = await r.json(content_type=None)
            rc = data.get("researchContent", {}) or {}
            return (
                rc.get("opinion", ""),
                rc.get("goalPrice"),
                rc.get("prevGoalPrice"),
                rc.get("attachUrl", "") or "",
            )
        except Exception as e:
            log.debug(f"상세 정보 수집 실패 (rid={rid}): {e}")
            return "", None, None, ""


async def _attach_details(
    session: aiohttp.ClientSession,
    reports: list[dict],
) -> None:
    """reports 각 항목에 opinion/goalPrice/prevGoalPrice/_pdfUrl 필드를 병렬로 채웁니다."""
    sem = asyncio.Semaphore(_DETAIL_SEM_SIZE)
    tasks = [
        _fetch_detail(session, r.get("researchId", ""), sem)
        for r in reports
    ]
    log.info(f"상세 정보 수집 중 ({len(tasks)}개)...")
    results = await asyncio.gather(*tasks)
    for report, (opinion, goal_price, prev_goal_price, pdf_url) in zip(reports, results):
        report["opinion"]       = opinion
        report["goalPrice"]      = goal_price
        report["prevGoalPrice"]  = prev_goal_price
        report["_pdfUrl"]        = pdf_url
    found = sum(1 for *_, u in results if u)
    log.info(f"상세 정보 수집 완료: {found}/{len(tasks)}개 (PDF 링크 기준)")


# ── Excel 스타일 ───────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
_DAT_FONT = Font(name="맑은 고딕", size=10)
_LNK_FONT = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
_CENTER   = Alignment(horizontal="center", vertical="center")
_LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# 가운데 정렬 컬럼: 작성일(1), 종목코드(2), 투자의견(5), 링크(9), PDF(10), ID(11), 괴리율(12)
_CENTER_COLS = {1, 2, 5, 9, 10, 11, 12}


def _setup_sheet(ws) -> None:
    """헤더 행 스타일 설정"""
    for col_idx, (_, label, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def _row_values(r: dict) -> list:
    gp = r.get("goalPrice")
    pg = r.get("prevGoalPrice")
    gp_val = int(gp) if gp else None
    pg_val = int(pg) if pg else None
    divergence = (gp_val - pg_val) / pg_val if (gp_val and pg_val) else None
    return [
        r.get("writeDate", ""),
        r.get("itemCode", ""),
        r.get("itemName", ""),
        r.get("brokerName", ""),
        r.get("opinion", ""),
        gp_val,
        pg_val,
        r.get("title", ""),
        r.get("endUrl", ""),       # col 9: 링크 (researchId로 재조합)
        r.get("_pdfUrl", ""),      # col 10: PDF (페이지 스크래핑으로 얻은 직접 URL)
        str(r.get("researchId", "")),
        divergence,
    ]


def _write_data_row(ws, row_num: int, values: list) -> None:
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        c.alignment = _CENTER if col_idx in _CENTER_COLS else _LEFT
        if col_idx == 9:  # 링크: Naver 리포트 페이지
            rid = str(values[10]) if len(values) > 10 else ""
            if rid:
                c.hyperlink = f"{_URL_BASE}{rid}"
                c.value = "보기"
                c.font = _LNK_FONT
            else:
                c.font = _DAT_FONT
        elif col_idx == 10:  # PDF: stock.pstatic.net 직접 다운로드
            pdf_url = str(val or "").strip()
            if pdf_url:
                safe_url = pdf_url.replace('"', '%22')
                c.value = f'=HYPERLINK("{safe_url}","PDF")'
                c.font = _LNK_FONT
            else:
                c.value = None
                c.font = _DAT_FONT
        elif col_idx in (6, 7):
            c.font = _DAT_FONT
            if val:
                c.number_format = "#,##0"
        elif col_idx == 12:  # 괴리율
            c.font = _DAT_FONT
            if val is not None:
                c.number_format = "0.0%"
        else:
            c.font = _DAT_FONT


# ── 워크북 관리 ───────────────────────────────────────────────────────────────

SHEET_NAME = "리서치 보고서"
_URL_BASE  = "https://m.stock.naver.com/research/company/"


def _is_new_format(ws) -> bool:
    """PDF 컬럼이 추가된 신형식인지 확인."""
    return ws.cell(row=1, column=10).value == "PDF"


def _migrate_to_new_format(ws) -> None:
    """기존 11컬럼 파일에 PDF 컬럼(col 10) 삽입 마이그레이션."""
    ws.insert_cols(10, 1)
    c = ws.cell(row=1, column=10, value="PDF")
    c.fill = _HDR_FILL
    c.font = _HDR_FONT
    c.alignment = _CENTER
    ws.column_dimensions[get_column_letter(10)].width = 6
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    _rebuild_all_hyperlinks(ws)
    log.info("마이그레이션 완료: PDF 컬럼 추가")


def _latest_existing_date(path: str) -> str | None:
    """엑셀에 이미 저장된 리서치 보고서 중 가장 최근 작성일(YYYY-MM-DD) 반환.
    파일/시트가 없거나 잠겨 있으면 None (호출자가 fallback 처리)."""
    if not os.path.exists(path):
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return None
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[SHEET_NAME]
    dates = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
    wb.close()
    return max(str(d) for d in dates) if dates else None


def load_workbook_state(path: str):
    """기존 파일 로드 or 신규 생성. 기존 researchId 집합 반환."""
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _setup_sheet(ws)
            return wb, ws, set()
        ws = wb[SHEET_NAME]
        ws.freeze_panes = "A2"
        if ws.max_row > 1 and not _is_new_format(ws):
            _migrate_to_new_format(ws)
        existing_ids: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = row[10]  # researchId: col 11 (index 10)
            if rid:
                existing_ids.add(str(rid))
        return wb, ws, existing_ids
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _setup_sheet(ws)
        return wb, ws, set()


def _rebuild_all_hyperlinks(ws) -> None:
    """col 11 (researchId) 기준으로 col 9 (링크) 하이퍼링크 전체 재구성."""
    for row_num in range(2, ws.max_row + 1):
        rid = str(ws.cell(row=row_num, column=11).value or "").strip()
        if not rid:
            continue
        link_cell = ws.cell(row=row_num, column=9)
        link_cell.hyperlink = f"{_URL_BASE}{rid}"
        link_cell.value = "보기"
        link_cell.font = _LNK_FONT
        link_cell.alignment = _CENTER


def insert_new_reports(ws, reports: list[dict], existing_ids: set[str]) -> int:
    """신규 리포트를 헤더 바로 아래 삽입 (최신이 위에 표시)."""
    new = [r for r in reports if str(r.get("researchId", "")) not in existing_ids]
    if not new:
        return 0

    new.sort(
        key=lambda x: (x.get("writeDate", ""), int(x.get("researchId", 0) or 0)),
        reverse=True,
    )

    n = len(new)
    ws.insert_rows(2, amount=n)

    for i, r in enumerate(new, 2):
        _write_data_row(ws, i, _row_values(r))
        existing_ids.add(str(r.get("researchId", "")))

    _rebuild_all_hyperlinks(ws)
    return n


# ── PDF 소급 적용 ─────────────────────────────────────────────────────────────

async def backfill_pdf_links(excel_path: str) -> int:
    """기존 파일에서 PDF 링크 없는 행을 리포트 상세 API로 채웁니다."""
    wb, ws, _ = load_workbook_state(excel_path)

    # PDF 링크 없거나 잘못된(pstatic.net 아닌) 행의 researchId 수집
    id_to_row: dict[str, int] = {}
    for row_num in range(2, ws.max_row + 1):
        rid = str(ws.cell(row=row_num, column=11).value or "").strip()
        pdf = str(ws.cell(row=row_num, column=10).value or "").strip()
        if rid and ("pstatic.net" not in pdf):
            id_to_row[rid] = row_num

    if not id_to_row:
        log.info(f"[백필] 대상 없음: {excel_path}")
        return 0

    log.info(f"[백필] {len(id_to_row)}개 리포트 PDF URL 수집 시작: {os.path.basename(excel_path)}")

    sem = asyncio.Semaphore(_DETAIL_SEM_SIZE)
    async with aiohttp.ClientSession(headers=API_HEADERS) as session:
        tasks = {
            rid: asyncio.ensure_future(_fetch_detail(session, rid, sem))
            for rid in id_to_row
        }
        total = len(tasks)
        done = 0
        results: dict[str, str] = {}
        for rid, task in tasks.items():
            *_, pdf_url = await task
            results[rid] = pdf_url
            done += 1
            if done % 100 == 0 or done == total:
                print(f"  {done}/{total} 완료", end="\r")
    print()

    updated = 0
    for rid, pdf_url in results.items():
        if pdf_url:
            row_num = id_to_row[rid]
            cell = ws.cell(row=row_num, column=10)
            safe_url = pdf_url.replace('"', '%22')
            cell.value     = f'=HYPERLINK("{safe_url}","PDF")'
            cell.font      = _LNK_FONT
            cell.alignment = _CENTER
            updated += 1

    try:
        wb.save(excel_path)
        log.info(f"[백필 완료] +{updated}개 PDF 링크 → {excel_path}")
    except PermissionError:
        log.warning(f"[백필 저장 실패] 파일이 열려 있습니다: {excel_path}")

    return updated


# ── 임시 저장 (파일 잠금 대비) ────────────────────────────────────────────────

def _load_pending() -> tuple[list[dict], list[str]]:
    if not os.path.exists(PENDING_DIR):
        return [], []
    files = sorted(glob.glob(os.path.join(PENDING_DIR, "pending_*.json")))
    all_reports: list[dict] = []
    for f in files:
        try:
            with open(f, encoding="utf-8") as fp:
                all_reports.extend(json.load(fp))
        except Exception as e:
            log.warning(f"pending 파일 읽기 오류 ({f}): {e}")
    return all_reports, files


def _save_pending(reports: list[dict]) -> str:
    os.makedirs(PENDING_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(PENDING_DIR, f"pending_{ts}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(reports, f, ensure_ascii=False)
    return path


def _delete_pending(files: list[str]) -> None:
    for f in files:
        try:
            os.remove(f)
        except Exception:
            pass


# ── 진입점 ────────────────────────────────────────────────────────────────────

async def main(
    start_date: str | None = None,
    end_date: str | None = None,
    excel_path: str | None = None,
) -> None:
    """
    start_date/end_date/excel_path 를 직접 전달하면 '소급 모드'로 동작.
    미전달 시 CLI 인자 또는 자동(어제~오늘) 모드.
    """
    _prog = (start_date is not None or excel_path is not None)

    # --backfill-pdf 모드 (CLI 전용)
    if not _prog and "--backfill-pdf" in sys.argv:
        extra = [a for a in sys.argv[1:] if a != "--backfill-pdf"]
        if "all" in extra:
            files = sorted(glob.glob(os.path.join(BASE_DIR, "Daily_index*.xlsx")))
        elif extra:
            f = extra[0]
            files = [f if os.path.isabs(f) else os.path.join(BASE_DIR, f)]
        else:
            files = [EXCEL_PATH]
        for f in files:
            if os.path.exists(f):
                await backfill_pdf_links(f)
            else:
                log.warning(f"파일 없음: {f}")
        return

    # ── 대상 Excel 파일 결정 ──────────────────────────────────────────────────
    _excel_path = excel_path or EXCEL_PATH

    # ── 날짜 범위 결정 ────────────────────────────────────────────────────────
    if _prog and start_date and end_date:
        log.info(f"[시작] 소급 모드: {start_date} ~ {end_date} → {os.path.basename(_excel_path)}")
    elif not _prog and len(sys.argv) == 3:
        start_date, end_date = sys.argv[1], sys.argv[2]
        log.info(f"[시작] 날짜 범위: {start_date} ~ {end_date}")
    else:
        now_kst  = datetime.now(KST)
        end_date = now_kst.strftime("%Y-%m-%d")
        fallback_start = (now_kst - timedelta(days=1)).strftime("%Y-%m-%d")
        latest = _latest_existing_date(_excel_path)
        if latest:
            gap_days = (now_kst.date() - datetime.strptime(latest, "%Y-%m-%d").date()).days
            if gap_days > MAX_AUTO_LOOKBACK_DAYS:
                log.warning(
                    f"[자동 모드] 마지막 저장일({latest})이 {gap_days}일 전 — "
                    f"비정상적으로 오래돼 {MAX_AUTO_LOOKBACK_DAYS}일로 제한 (수동 확인 필요)"
                )
                start_date = (now_kst - timedelta(days=MAX_AUTO_LOOKBACK_DAYS)).strftime("%Y-%m-%d")
            else:
                start_date = latest
                if gap_days > 1:
                    log.info(f"[자동 모드] 마지막 저장일({latest}) 이후 {gap_days}일 공백 감지 → 자동 보정")
        else:
            start_date = fallback_start
        log.info(f"[시작] 자동 모드: {start_date} ~ {end_date}")

    # pending: 소급 모드에서는 건너뜀 (pending은 현재 주 파일 전용)
    if _prog:
        pending_reports, pending_files = [], []
    else:
        pending_reports, pending_files = _load_pending()
        if pending_reports:
            log.info(f"임시 저장 데이터 발견: {len(pending_reports)}개 → 이번에 함께 반영")

    new_reports = await fetch_reports(start_date, end_date)
    log.info(f"수집 완료: {len(new_reports)}개")

    all_reports = pending_reports + new_reports

    try:
        wb, ws, existing_ids = load_workbook_state(_excel_path)
    except PermissionError:
        if _prog:
            log.warning(f"[소급 실패] 파일이 열려 있습니다: {_excel_path}")
        else:
            _delete_pending(pending_files)
            saved_path = _save_pending(all_reports)
            log.warning(
                f"[접근 실패] 엑셀 파일이 열려 있어 읽을 수 없습니다. "
                f"{len(all_reports)}개를 임시 저장했습니다 → {saved_path}"
            )
        return

    added = insert_new_reports(ws, all_reports, existing_ids)

    try:
        wb.save(_excel_path)
        if not _prog:
            _delete_pending(pending_files)
        log.info(f"엑셀 저장: +{added}개 신규 추가 | 누적 {len(existing_ids)}개 | {_excel_path}")
    except PermissionError:
        if _prog:
            log.warning(f"[소급 저장 실패] 파일이 열려 있습니다: {_excel_path}")
        else:
            _delete_pending(pending_files)
            saved_path = _save_pending(all_reports)
            log.warning(
                f"[저장 실패] 엑셀 파일이 열려 있습니다. "
                f"{len(all_reports)}개를 임시 저장했습니다 → {saved_path}"
            )


if __name__ == "__main__":
    asyncio.run(main())
