"""
신한증권 거래내역서 PDF → Excel 변환
Usage: python pdf_to_excel.py
"""
import sys, os, re
sys.stdout.reconfigure(encoding="utf-8")

import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(BASE_DIR, "거래내역서.xlsx")

PDF_FILES = sorted(
    f for f in os.listdir(BASE_DIR)
    if f.startswith("거래내역서_") and f.endswith(".pdf")
)

# ── 출력 컬럼 정의 ─────────────────────────────────────────────────────────────
OUT_COLS = [
    # (헤더,         너비,  서식,            정렬)
    ("날짜",         11,  "YYYY-MM-DD",    "center"),
    ("거래구분",     15,  "@",             "left"),
    ("종목명",       22,  "@",             "left"),
    ("종목코드",     15,  "@",             "center"),
    ("환율",         11,  "#,##0.00",      "right"),
    ("거래수량",     13,  "#,##0.######",  "right"),
    ("거래대금₩",   14,  "#,##0",         "right"),
    ("거래대금$",   12,  "#,##0.0000",    "right"),
    ("단가₩",       12,  "#,##0",         "right"),
    ("단가$",       11,  "#,##0.0000",    "right"),
    ("수수료₩",     11,  "#,##0",         "right"),
    ("수수료$",     11,  "#,##0.0000",    "right"),
    ("거래세",       10,  "#,##0",         "right"),
    ("제세금₩",     10,  "#,##0",         "right"),
    ("제세금$",     10,  "#,##0.0000",    "right"),
    ("변제연체합",   12,  "#,##0",         "right"),
    ("잔고",         13,  "#,##0.######",  "right"),
    ("잔액₩",       13,  "#,##0",         "right"),
    ("잔액$",       12,  "#,##0.0000",    "right"),
]

# PDF 표 헤더에 등장하는 단어들
KNOWN_HEADERS = {
    "거래일자", "거래구분", "종목명(종목코드)", "환율", "거래수량",
    "거래대금", "단가", "수수료", "거래세", "제세금", "변제/연체합", "잔고", "잔액",
}

# ── 스타일 헬퍼 ────────────────────────────────────────────────────────────────
def _f(bold=False, size=10, color="000000"):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color)

def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _aln(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v)

def _bdr(clr="CCCCCC"):
    s = Side(style="thin", color=clr)
    return Border(top=s, left=s, right=s, bottom=s)

# ── 파싱 유틸 ──────────────────────────────────────────────────────────────────
def parse_date(s: str):
    try:
        return datetime.strptime(s.strip(), "%Y.%m.%d").date()
    except Exception:
        return None

def parse_num(s: str):
    if not s:
        return None
    s = s.strip().replace(",", "")
    if s in ("", "-"):
        return None
    # USD 형식: "($X.XX)"
    m = re.search(r"\(\$?\s*([0-9.]+)\)", s)
    if m:
        return float(m.group(1))
    try:
        return float(s)
    except ValueError:
        return None

def parse_usd(s: str):
    """서브행 USD 값 파싱. "($3.99)" → 3.99. 없으면 None."""
    if not s:
        return None
    m = re.search(r"\(\$?\s*([0-9.]+)\)", s)
    return float(m.group(1)) if m else None

def parse_stock(raw: str):
    """"종목명(종목코드)" → (종목명, 종목코드)."""
    raw = raw.strip()
    m = re.match(r"^(.*?)\(([A-Za-z0-9]+)\)\s*$", raw)
    if m:
        return m.group(1).strip(), m.group(2)
    return raw, ""

# ── 컬럼 경계 계산 ─────────────────────────────────────────────────────────────
def build_boundaries(header_words: list[tuple[float, str]]) -> list[tuple[str, float, float]]:
    """(x, 헤더명) 리스트 → [(헤더명, x_lo, x_hi), ...] 반환."""
    srt = sorted(header_words, key=lambda w: w[0])
    result = []
    for i, (x, name) in enumerate(srt):
        lo = (srt[i - 1][0] + x) / 2 if i > 0 else 0
        hi = (x + srt[i + 1][0]) / 2 if i < len(srt) - 1 else 9999
        result.append((name, lo, hi))
    return result

def get_col(x: float, boundaries: list) -> str | None:
    for name, lo, hi in boundaries:
        if lo <= x < hi:
            return name
    return None

def join_words(word_list: list[tuple[float, str]]) -> str:
    if not word_list:
        return ""
    # x 정렬 후 결합 — USD 값("($" + "3.99)") 은 공백 없이, 일반 단어는 공백으로 구분
    parts = [t for _, t in sorted(word_list)]
    # "($" 로 시작하는 부분은 인접 단어와 공백 없이 합침
    result = parts[0]
    for t in parts[1:]:
        if result.endswith("($") or result.endswith("($ "):
            result += t
        else:
            result += " " + t
    return result.strip()

# ── 페이지 파싱 ────────────────────────────────────────────────────────────────
def parse_page(page) -> list[dict]:
    raw_words = page.get_text("words")  # (x0,y0,x1,y1,text,block,line,word)

    # 중복 단어 제거 (페이지 번호 N/M 반복 문제)
    seen: set = set()
    unique = []
    for w in raw_words:
        key = (round(w[0], 1), round(w[1], 1), w[4])
        if key not in seen:
            seen.add(key)
            unique.append(w)

    # y 기준 그룹핑 (1pt 단위)
    y_groups: dict[int, list] = defaultdict(list)
    for w in unique:
        y_groups[round(w[1])].append((w[0], w[4]))

    sorted_ys = sorted(y_groups.keys())

    # 섹션 레이블 위치 (원화/외화 거래내역)
    section_ys: dict[int, str] = {}
    for yr in sorted_ys:
        joined = " ".join(t for _, t in y_groups[yr])
        if "원화" in joined and "거래내역" in joined:
            section_ys[yr] = "원화"
        elif "외화" in joined and "거래내역" in joined:
            section_ys[yr] = "외화"

    # 헤더 행 탐색 → 컬럼 경계 설정
    header_sections: list[tuple] = []  # (header_y, boundaries, section_name)
    for yr in sorted_ys:
        texts = [t for _, t in y_groups[yr]]
        if "거래일자" not in texts:
            continue
        # 가장 가까운 이전 섹션 레이블
        sec_name = None
        for sy in sorted(section_ys, reverse=True):
            if sy < yr:
                sec_name = section_ys[sy]
                break
        hwords = [(x, t) for x, t in y_groups[yr] if t in KNOWN_HEADERS]
        if hwords:
            header_sections.append((yr, build_boundaries(hwords), sec_name))

    if not header_sections:
        return []

    # 각 섹션별 데이터 추출
    all_rows: list[dict] = []
    for sec_i, (h_y, boundaries, sec_name) in enumerate(header_sections):
        next_h_y = header_sections[sec_i + 1][0] if sec_i + 1 < len(header_sections) else float("inf")
        data_ys = [yr for yr in sorted_ys if h_y < yr < next_h_y]

        i = 0
        while i < len(data_ys):
            yr = data_ys[i]

            # 컬럼 할당
            col_data: dict[str, list] = defaultdict(list)
            for x, text in y_groups[yr]:
                col = get_col(x, boundaries)
                if col:
                    col_data[col].append((x, text))

            # 날짜 확인
            date_val = parse_date(join_words(col_data.get("거래일자", [])))
            if date_val is None:
                i += 1
                continue

            # 서브행 확인 (USD 값, y 차이 ≤ 15pt)
            sub_data: dict[str, list] = defaultdict(list)
            if i + 1 < len(data_ys):
                next_yr = data_ys[i + 1]
                if next_yr - yr <= 15:
                    next_col: dict[str, list] = defaultdict(list)
                    for x, text in y_groups[next_yr]:
                        col = get_col(x, boundaries)
                        if col:
                            next_col[col].append((x, text))
                    # 서브행: 날짜 없어야 함
                    if not parse_date(join_words(next_col.get("거래일자", []))):
                        sub_data = next_col
                        i += 2
                    else:
                        i += 1
                else:
                    i += 1
            else:
                i += 1

            all_rows.append(_build_row(date_val, col_data, sub_data, sec_name))

    return all_rows


def _build_row(date_val, col_data: dict, sub_data: dict, sec_name: str | None) -> dict:
    def g(col):   return join_words(col_data.get(col, []))
    def s(col):   return join_words(sub_data.get(col, []))
    def n(col):   return parse_num(g(col))
    def usd(col): return parse_usd(s(col))

    stock_name, stock_code = parse_stock(g("종목명(종목코드)"))
    환율 = n("환율")

    return {
        "날짜":      date_val,
        "거래구분":  g("거래구분").strip(),
        "종목명":    stock_name,
        "종목코드":  stock_code,
        "환율":      환율 if 환율 else None,   # 0은 공란 처리
        "거래수량":  n("거래수량"),
        "거래대금₩": n("거래대금"),
        "거래대금$": usd("거래대금"),
        "단가₩":     n("단가"),
        "단가$":     usd("단가"),
        "수수료₩":   n("수수료"),
        "수수료$":   usd("수수료"),
        "거래세":    n("거래세"),
        "제세금₩":   n("제세금"),
        "제세금$":   usd("제세금"),
        "변제연체합": n("변제/연체합"),
        "잔고":      n("잔고"),
        "잔액₩":     n("잔액"),
        "잔액$":     usd("잔액"),
        "_sec":      sec_name,
    }


# ── Excel 출력 ─────────────────────────────────────────────────────────────────
_HDR_BG = "17375E"
_ROW_ODD  = "FFFFFF"
_ROW_EVEN = "EEF3FA"

def write_excel(all_rows: list[dict], out_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래내역"

    # ── 헤더 행 ──────────────────────────────────────────────────────────────
    for ci, (label, width, fmt, align) in enumerate(OUT_COLS, 1):
        c = ws.cell(1, ci, label)
        c.font      = _f(bold=True, color="FFFFFF")
        c.fill      = _fill(_HDR_BG)
        c.alignment = _aln("center")
        c.border    = _bdr("888888")
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUT_COLS))}1"

    # ── 데이터 행 ──────────────────────────────────────────────────────────
    col_keys = [h for h, *_ in OUT_COLS]
    # 컬럼 헤더 → 내부 dict 키 매핑
    key_map = {
        "날짜": "날짜", "거래구분": "거래구분", "종목명": "종목명",
        "종목코드": "종목코드", "환율": "환율", "거래수량": "거래수량",
        "거래대금₩": "거래대금₩", "거래대금$": "거래대금$",
        "단가₩": "단가₩", "단가$": "단가$",
        "수수료₩": "수수료₩", "수수료$": "수수료$",
        "거래세": "거래세", "제세금₩": "제세금₩", "제세금$": "제세금$",
        "변제연체합": "변제연체합", "잔고": "잔고",
        "잔액₩": "잔액₩", "잔액$": "잔액$",
    }

    for row_i, row in enumerate(all_rows, 2):
        bg = _ROW_EVEN if row_i % 2 == 0 else _ROW_ODD
        for ci, (label, width, fmt, align) in enumerate(OUT_COLS, 1):
            val = row.get(key_map.get(label, label))
            c = ws.cell(row_i, ci, val)
            c.font      = _f()
            c.fill      = _fill(bg)
            c.border    = _bdr()
            c.alignment = _aln(align)
            if fmt and val is not None:
                c.number_format = fmt
        ws.row_dimensions[row_i].height = 17

    wb.save(out_path)
    print(f"✅ 저장: {out_path} (총 {len(all_rows):,}행)")


# ── 진입점 ─────────────────────────────────────────────────────────────────────
def main():
    if not PDF_FILES:
        print("거래내역서_*.pdf 파일을 찾을 수 없습니다.")
        return

    all_rows: list[dict] = []
    for pdf_file in PDF_FILES:
        path = os.path.join(BASE_DIR, pdf_file)
        doc = fitz.open(path)
        rows: list[dict] = []
        for page in doc:
            rows.extend(parse_page(page))
        doc.close()
        print(f"  {pdf_file}: {len(rows):,}행 추출")
        all_rows.extend(rows)

    # 날짜 오름차순 정렬
    all_rows.sort(key=lambda r: r.get("날짜") or datetime.min.date())

    write_excel(all_rows, OUT_FILE)

    # 거래구분 요약
    from collections import Counter
    cnt = Counter(r["거래구분"] for r in all_rows)
    print("\n  거래구분 요약:")
    for k, v in sorted(cnt.items(), key=lambda x: -x[1]):
        print(f"    {k or '(공백)':<15} {v:>5}건")


if __name__ == "__main__":
    main()
