#!/usr/bin/env python3
"""
관심종목 리포트 PDF 다운로드

Daily_index*.xlsx 에서 특정 종목의 2025년 이후 리서치 보고서를 찾아 PDF 다운로드.
Daily_index.xlsx에 관심종목 시트를 추가해 거기서 종목 목록을 읽습니다.

사용법:
  python download_stock_reports.py                  # 관심종목 시트 기반 전체 다운로드
  python download_stock_reports.py 005930           # 종목코드 직접 지정
  python download_stock_reports.py 삼성전자         # 종목명 직접 지정
  python download_stock_reports.py --setup          # 관심종목 시트 추가만
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os, re, glob, time, argparse, requests, openpyxl
from collections import Counter
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pytz

BASE_DIR       = Path(__file__).parent
MAIN_EXCEL     = BASE_DIR / "Daily_index.xlsx"
REPORT_SHEET   = "리서치 보고서"
WATCHLIST_SHEET = "관심종목"
MIN_DATE       = "2025-01-01"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
DAT_FONT = Font(name="맑은 고딕", size=10)
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left", vertical="center")
THIN     = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

DL_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://stock.pstatic.net/",
}


# ── 관심종목 시트 ─────────────────────────────────────────────────────────────

def setup_watchlist_sheet():
    """Daily_index.xlsx에 관심종목 시트 추가."""
    if not MAIN_EXCEL.exists():
        print(f"오류: {MAIN_EXCEL.name} 파일이 없습니다.")
        return False

    wb = openpyxl.load_workbook(str(MAIN_EXCEL))

    if WATCHLIST_SHEET in wb.sheetnames:
        print(f"관심종목 시트가 이미 존재합니다.")
        wb.close()
        return True

    ws = wb.create_sheet(WATCHLIST_SHEET, 0)

    cols = [("종목코드/종목명", 20), ("종목명(확인용)", 18), ("다운로드 경로", 42),
            ("마지막 업데이트", 18), ("리포트 수", 10)]

    for i, (h, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # 안내 행
    note = ws.cell(row=2, column=1,
                   value="▼ A열에 종목코드(예: 005930) 또는 종목명(예: 삼성전자) 입력 후 저장하세요")
    note.font = Font(name="맑은 고딕", size=9, color="666666", italic=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 18

    try:
        wb.save(str(MAIN_EXCEL))
        print(f"관심종목 시트 추가 완료 ({MAIN_EXCEL.name})")
        print("  → A3 셀부터 종목코드 또는 종목명 입력")
        return True
    except PermissionError:
        print(f"오류: {MAIN_EXCEL.name} 파일이 열려 있습니다. 닫고 다시 실행하세요.")
        return False


def get_watchlist():
    """관심종목 시트에서 입력된 종목 목록 반환."""
    if not MAIN_EXCEL.exists():
        return []

    wb = openpyxl.load_workbook(str(MAIN_EXCEL), data_only=True)
    if WATCHLIST_SHEET not in wb.sheetnames:
        return []

    ws = wb[WATCHLIST_SHEET]
    stocks = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        val = row[0]
        if val and str(val).strip() and not str(val).startswith("▼"):
            stocks.append(str(val).strip())
    return stocks


# ── PDF URL 추출 ───────────────────────────────────────────────────────────────

def extract_pdf_url(cell_val):
    """HYPERLINK 수식 또는 직접 URL에서 PDF 주소 추출."""
    if not cell_val:
        return None
    s = str(cell_val)
    m = re.search(r'HYPERLINK\("([^"]+)"', s)
    if m:
        return m.group(1)
    if s.startswith("http") and "pstatic.net" in s:
        return s
    return None


# ── Daily_index 파일 스캔 ─────────────────────────────────────────────────────

def scan_all_excel_files(stock_query: str, min_date: str = MIN_DATE) -> list[dict]:
    """
    모든 Daily_index*.xlsx 파일에서 종목 매칭 리포트 수집.
    종목코드 (숫자) 또는 종목명(부분 일치) 검색.
    """
    files = sorted(glob.glob(str(BASE_DIR / "Daily_index*.xlsx")))
    if not files:
        print("오류: Daily_index*.xlsx 파일 없음")
        return []

    q = stock_query.strip()
    is_code = bool(re.match(r"^\d{4,6}$", q))
    q_lower = q.lower()

    all_reports: list[dict] = []

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            # data_only=False 필수: True로 읽으면 =HYPERLINK() 수식의 URL이 소실됨
            wb = openpyxl.load_workbook(fpath, data_only=False, read_only=True)
        except Exception as e:
            print(f"  [경고] {fname} 읽기 실패: {e}")
            continue

        if REPORT_SHEET not in wb.sheetnames:
            wb.close()
            continue

        ws = wb[REPORT_SHEET]
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            date_val = str(row[0] or "").strip()
            if date_val < min_date:
                continue

            code    = str(row[1] or "").strip()
            name    = str(row[2] or "").strip()
            firm    = str(row[3] or "").strip()
            opinion = str(row[4] or "").strip()
            target  = row[5]
            prev_p  = row[6]
            title   = str(row[7] or "").strip()
            pdf_cell = row[9]   # col 10: PDF

            if is_code:
                matched = (code == q or code.lstrip("0") == q or q == code.lstrip("0"))
            else:
                matched = q_lower in name.lower()

            if not matched:
                continue

            pdf_url = extract_pdf_url(pdf_cell)

            all_reports.append({
                "date":       date_val,
                "code":       code,
                "name":       name,
                "firm":       firm,
                "opinion":    opinion,
                "target":     int(str(target).replace(",", "")) if target else None,
                "prev_price": int(str(prev_p).replace(",", "")) if prev_p else None,
                "title":      title,
                "pdf_url":    pdf_url,
                "source":     fname,
            })
            count += 1

        wb.close()
        if count > 0:
            print(f"  {fname}: {count}건")

    # 날짜 내림차순 정렬
    all_reports.sort(key=lambda x: x["date"], reverse=True)
    return all_reports


# ── PDF 다운로드 ───────────────────────────────────────────────────────────────

def sanitize(s: str) -> str:
    return re.sub(r'[<>:"/\\|?*\n\r]', '_', s).strip()


def download_pdf(url: str, save_path: Path) -> tuple[bool, str]:
    try:
        resp = requests.get(url, headers=DL_HEADERS, timeout=25, stream=True)
        if resp.status_code != 200:
            return False, f"HTTP {resp.status_code}"
        data = resp.content
        if len(data) < 500:
            return False, "응답 너무 짧음"
        save_path.parent.mkdir(parents=True, exist_ok=True)
        save_path.write_bytes(data)
        return True, f"{len(data)//1024}KB"
    except Exception as e:
        return False, str(e)[:60]


def download_reports_for_stock(stock_query: str, min_date: str = MIN_DATE) -> tuple[list, str | None]:
    print(f"\n=== [{stock_query}] 리포트 검색 ({min_date} 이후) ===")

    reports = scan_all_excel_files(stock_query, min_date)
    if not reports:
        print("매칭 리포트 없음")
        return [], None

    name_counts = Counter(r["name"] for r in reports if r["name"])
    stock_name  = name_counts.most_common(1)[0][0] if name_counts else stock_query

    total_with_pdf = sum(1 for r in reports if r["pdf_url"])
    print(f"\n총 {len(reports)}건 | PDF 있음: {total_with_pdf}건 | 종목명: {stock_name}")

    folder = BASE_DIR / sanitize(stock_name)
    folder.mkdir(parents=True, exist_ok=True)

    ok = fail = skip = 0

    for r in reports:
        if not r["pdf_url"]:
            skip += 1
            continue

        date_s  = r["date"].replace("-", "")
        fname   = f"{date_s}_{sanitize(r['firm'])}_{sanitize(r['title'])[:40]}.pdf"
        spath   = folder / fname

        if spath.exists():
            skip += 1
            continue

        success, msg = download_pdf(r["pdf_url"], spath)
        if success:
            ok += 1
            print(f"  ✓ {r['date']} {r['firm']:<12} {r['title'][:28]} ({msg})")
        else:
            fail += 1
            print(f"  ✗ {r['date']} {r['firm']:<12} {msg}")
        time.sleep(0.25)

    print(f"\n다운로드 완료: {ok}건 ✓  {fail}건 ✗  {skip}건 건너뜀")
    print(f"저장 위치: {folder}")
    return reports, stock_name


def update_watchlist_status(stock_query: str, stock_name: str, count: int):
    if not MAIN_EXCEL.exists():
        return
    try:
        wb = openpyxl.load_workbook(str(MAIN_EXCEL))
        if WATCHLIST_SHEET not in wb.sheetnames:
            return
        ws = wb[WATCHLIST_SHEET]
        now_str = datetime.now(pytz.timezone("Asia/Seoul")).strftime("%Y-%m-%d %H:%M")
        for row in ws.iter_rows(min_row=3):
            if row[0].value and str(row[0].value).strip() == stock_query:
                row[1].value = stock_name
                row[2].value = str(BASE_DIR / sanitize(stock_name))
                row[3].value = now_str
                row[4].value = count
                break
        wb.save(str(MAIN_EXCEL))
    except Exception:
        pass


# ── 진입점 ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="관심종목 리포트 PDF 다운로드")
    parser.add_argument("stock", nargs="?", help="종목코드 또는 종목명")
    parser.add_argument("--setup",     action="store_true", help="관심종목 시트 추가")
    parser.add_argument("--from-date", default=MIN_DATE,    help=f"시작일 (기본: {MIN_DATE})")
    args = parser.parse_args()

    if args.setup:
        setup_watchlist_sheet()
        return

    if args.stock:
        stocks = [args.stock]
    else:
        stocks = get_watchlist()
        if not stocks:
            if not MAIN_EXCEL.exists():
                print(f"오류: {MAIN_EXCEL.name} 파일을 찾을 수 없습니다.")
            else:
                ok = setup_watchlist_sheet()
                if ok:
                    print("\n관심종목 시트가 추가됐습니다.")
                    print("Daily_index.xlsx → 관심종목 시트에서 종목 입력 후 다시 실행하세요.")
            return

    for s in stocks:
        reports, name = download_reports_for_stock(s, args.from_date)
        if name and reports:
            update_watchlist_status(s, name, len(reports))


if __name__ == "__main__":
    main()
