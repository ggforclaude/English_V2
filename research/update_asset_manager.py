"""
research/update_asset_manager.py
자산관리_2026.xlsx — 거래내역/참고/배당금/투자현황 시트 재구성
KRX종목마스터·투자아이디어·매도모니터링은 유지
"""
import sys, os, argparse
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE     = os.path.join(BASE_DIR, "자산관리_2026.xlsx")

# ── 시트 이름 ──────────────────────────────────────────────────────────────
TRADE = "거래내역"
REF   = "참고"
DIV   = "배당금"
HOLD  = "투자현황"
KRX   = "KRX종목마스터"

# ── 데이터 행 범위 ──────────────────────────────────────────────────────────
TD_S, TD_E = 5, 2004   # 거래내역 데이터 (2000행)
DV_S, DV_E = 5, 504    # 배당금 데이터 (500행)
HD_S, HD_E = 5, 204    # 투자현황 데이터 (200행)
FR_S, FR_E = 31, 130   # 참고-수수료 테이블 (100행)

# ── 계좌 목록 ───────────────────────────────────────────────────────────────
ACCOUNTS = [
    "신한", "신한-해외", "토스", "토스-해외",
    "신한-연금저축", "미래-연금저축", "신한-ISA",
    "미래에셋", "한화", "한화-해외", "유주", "인우",
]
TRADE_TYPES = ["매수", "매도", "배당재투자"]

# 참고 시트 계좌 목록 위치: A열 4행~
ACCT_COL    = 1   # A열
ACCT_HDR    = 3   # 헤더 행
ACCT_START  = 4   # 데이터 시작 행
ACCT_END    = ACCT_START + len(ACCOUNTS) - 1  # 15행

# 수수료 테이블 기본값 (계좌, 시작일, 수수료율%, 매매세율%)
FEE_DEFAULTS = [
    ("신한",          "2024-01-01", 0.015, 0.20),
    ("신한-해외",     "2024-01-01", 0.250, 0.00),
    ("토스",          "2024-01-01", 0.015, 0.20),
    ("토스-해외",     "2024-01-01", 0.250, 0.00),
    ("신한-연금저축", "2024-01-01", 0.015, 0.00),
    ("미래-연금저축", "2024-01-01", 0.014, 0.00),
    ("신한-ISA",      "2024-01-01", 0.015, 0.00),
    ("미래에셋",      "2024-01-01", 0.014, 0.20),
    ("한화",          "2024-01-01", 0.015, 0.20),
    ("한화-해외",     "2024-01-01", 0.300, 0.00),
    ("유주",          "2024-01-01", 0.015, 0.20),
    ("인우",          "2024-01-01", 0.015, 0.20),
]

# ── 공통 스타일 ─────────────────────────────────────────────────────────────
C = {
    "navy":   "17375E", "blue":   "2E75B6", "lblue":  "BDD7EE",
    "input":  "EEF3FA", "formula":"F2F2F2", "green":  "375623",
    "lgreen": "E2EFDA", "orange": "C55A11", "yellow": "FFF2CC",
    "gray":   "D9D9D9", "lgray":  "F5F5F5", "white":  "FFFFFF",
    "lorange":"FCE4D6", "hidden": "ECECEC",  # 숨김 보조컬럼
}

def _f(bold=False, size=10, color="000000", italic=False):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _aln(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _bdr(clr="CCCCCC"):
    s = Side(style="thin", color=clr)
    return Border(top=s, left=s, right=s, bottom=s)

def hdr_cell(ws, row, col, text, bg=None, fg="FFFFFF", bold=True,
             h="center", size=10, wrap=False):
    bg = bg or C["navy"]
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _f(bold=bold, size=size, color=fg)
    c.fill      = _fill(bg)
    c.alignment = _aln(h=h, wrap=wrap)
    c.border    = _bdr("888888")
    return c

def inp(ws, row, col, value=None, fmt=None, h="left", bg=None):
    bg = bg or C["input"]
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg); c.border = _bdr(); c.alignment = _aln(h=h)
    c.font = _f()
    if fmt: c.number_format = fmt
    return c

def fml(ws, row, col, formula, fmt=None, h="right", bg=None):
    bg = bg or C["formula"]
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = _fill(bg); c.border = _bdr(); c.alignment = _aln(h=h)
    c.font = _f()
    if fmt: c.number_format = fmt
    return c

def title(ws, row, text, ncols, bg=None, size=13):
    bg = bg or C["navy"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _f(bold=True, size=size, color="FFFFFF")
    c.fill = _fill(bg); c.alignment = _aln(h="center")
    ws.row_dimensions[row].height = 26

def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _f(italic=True, size=9, color="888888")

def cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def dv(ws, formula, rng):
    v = DataValidation(type="list", formula1=formula,
                       allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(v); v.add(rng)


# ══════════════════════════════════════════════════════════════════════════════
# 참고 시트
# ══════════════════════════════════════════════════════════════════════════════
def make_ref(wb):
    ws = wb.create_sheet(REF)

    # ── 섹션1: 계좌 목록 ─────────────────────────────────────────────────
    title(ws, 1, "📋 참고 데이터 — 계좌 목록 & 수수료 테이블", 6, bg=C["green"])
    note(ws, 2, 1, "★ 계좌 추가: A열에 행 추가 → 거래내역·투자현황 드롭다운 자동 반영")

    hdr_cell(ws, ACCT_HDR, 1, "계좌명 (드롭다운 소스)", bg=C["green"])
    cw(ws, 1, 20)

    for i, acct in enumerate(ACCOUNTS):
        r = ACCT_START + i
        c = ws.cell(row=r, column=1, value=acct)
        c.fill = _fill(C["input"]); c.border = _bdr()
        c.font = _f(); c.alignment = _aln("center")
        ws.row_dimensions[r].height = 18

    # 추가 여분 행 (사용자가 계좌 추가)
    for i in range(10):
        r = ACCT_END + 1 + i
        c = ws.cell(row=r, column=1)
        c.fill = _fill(C["input"]); c.border = _bdr()

    # ── 섹션2: 수수료/세금 테이블 ─────────────────────────────────────────
    fee_hdr_row = FR_S - 1  # = 30
    note(ws, fee_hdr_row - 1, 1,
         "★ 수수료율·매매세율(%) 입력. 계좌+날짜 기준으로 최근 행이 자동 적용됨")
    for ci, (lbl, w) in enumerate([
        ("계좌",         20), ("적용시작일",   14),
        ("수수료율(%)",  14), ("매매세율(%)",  14),
        ("비고",         30)
    ], 1):
        hdr_cell(ws, fee_hdr_row, ci, lbl, bg=C["orange"])
        cw(ws, ci, w)

    for i, (acct, sdate, fee, tax) in enumerate(FEE_DEFAULTS):
        r = FR_S + i
        vals = [acct, sdate, fee, tax, "기본값 — 실제 수수료 확인 후 수정"]
        fmts = [None, "YYYY-MM-DD", "0.000", "0.000", None]
        aligns = ["center", "center", "right", "right", "left"]
        for ci, (v, fmt, al) in enumerate(zip(vals, fmts, aligns), 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = _fill(C["input"]); c.border = _bdr()
            c.font = _f(); c.alignment = _aln(al)
            if fmt: c.number_format = fmt
        ws.row_dimensions[r].height = 18

    # 여분 행
    for i in range(len(FEE_DEFAULTS), FR_E - FR_S + 1):
        r = FR_S + i
        for ci in range(1, 6):
            c = ws.cell(row=r, column=ci)
            c.fill = _fill(C["input"]); c.border = _bdr()

    ws.freeze_panes = "A4"
    print(f"  ✓ '{REF}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# 거래내역 시트
# ══════════════════════════════════════════════════════════════════════════════
def _fee_formula(r, rate_col_letter):
    """계좌·날짜 기준으로 참고 시트에서 수수료율/세금율 조회 SUMPRODUCT."""
    a, b = f"'{REF}'!$A${FR_S}:$A${FR_E}", f"'{REF}'!$B${FR_S}:$B${FR_E}"
    rc   = f"'{REF}'!${rate_col_letter}${FR_S}:${rate_col_letter}${FR_E}"
    mx   = f"MAXIFS({b},{a},C{r},{b},\"<=\"&A{r})"
    return (f"=IFERROR(SUMPRODUCT(({a}=C{r})*({b}<=A{r})"
            f"*({b}={mx})*{rc})/100,0)")

def make_trade(wb):
    ws = wb.create_sheet(TRADE)
    NCOLS = 13

    # 컬럼 레이아웃
    # A날짜 B거래구분 C계좌 D종목코드 E종목명 F단가 G수량
    # H거래금액(수식) I수수료율(수식,hidden) J수수료(수식) K세금율(수식,hidden) L세금(수식) M메모
    COL_META = [
        # (헤더, 너비, 입력여부, 정렬, 서식)
        ("날짜",         11, True,  "center", "YYYY-MM-DD"),
        ("거래구분",      9, True,  "center", None),
        ("계좌",         14, True,  "center", None),
        ("종목코드",     10, True,  "center", "@"),
        ("종목명",       20, False, "left",   None),   # VLOOKUP
        ("단가",         13, True,  "right",  "#,##0"),
        ("수량",         10, True,  "right",  "#,##0"),
        ("거래금액",     14, False, "right",  "#,##0"),
        ("수수료율",      9, False, "right",  "0.000%"),  # hidden helper
        ("수수료",       11, False, "right",  "#,##0"),
        ("세금율",        8, False, "right",  "0.000%"),  # hidden helper
        ("세금",         11, False, "right",  "#,##0"),
        ("메모",         28, True,  "left",   None),
    ]

    title(ws, 1, "📋 거래내역 — 투자 거래 기록 (매수/매도)", NCOLS)
    note(ws, 2, 1,
         "★ A~D, F~G, M열만 입력 | 나머지 자동계산 | "
         "수수료·세금율 수정: 참고 시트 수수료 테이블")
    note(ws, 2, 8,
         "※ 수수료율·세금율은 참고 시트 적용. 계좌/날짜가 없으면 0 반환")

    for ci, (lbl, w, _, al, _fmt) in enumerate(COL_META, 1):
        bg = C["navy"] if _ else C["blue"]   # 입력컬럼=진네이비, 수식=파랑
        hdr_cell(ws, 3, ci, lbl, bg=bg, wrap=False)
        cw(ws, ci, w)
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NCOLS)}3"

    # I·K열 (수수료율·세금율) 숨기기
    ws.column_dimensions["I"].hidden = True
    ws.column_dimensions["K"].hidden = True

    for r in range(TD_S, TD_E + 1):
        # A: 날짜
        inp(ws, r, 1, fmt="YYYY-MM-DD", h="center")
        # B: 거래구분
        inp(ws, r, 2, h="center")
        # C: 계좌
        inp(ws, r, 3, h="center")
        # D: 종목코드
        inp(ws, r, 4, fmt="@", h="center")
        # E: 종목명 VLOOKUP
        fml(ws, r, 5,
            f"=IFERROR(VLOOKUP(D{r},'{KRX}'!$A:$B,2,FALSE),\"\")",
            h="left")
        # F: 단가
        inp(ws, r, 6, fmt="#,##0", h="right")
        # G: 수량
        inp(ws, r, 7, fmt="#,##0", h="right")
        # H: 거래금액 = F×G
        fml(ws, r, 8,
            f"=IF(AND(F{r}<>\"\",G{r}<>\"\"),F{r}*G{r},\"\")",
            fmt="#,##0")
        # I: 수수료율 (hidden)
        fml(ws, r, 9, _fee_formula(r, "C"),
            fmt="0.000%", bg=C["hidden"])
        # J: 수수료 = H×I
        fml(ws, r, 10,
            f"=IF(H{r}=\"\",\"\",ROUND(H{r}*I{r},0))",
            fmt="#,##0")
        # K: 세금율 (hidden, 매도만)
        fml(ws, r, 11,
            f'=IF(B{r}="매도",{_fee_formula(r, "D")[1:]},0)',
            fmt="0.000%", bg=C["hidden"])
        # L: 세금 = H×K (매도만)
        fml(ws, r, 12,
            f"=IF(H{r}=\"\",\"\",ROUND(H{r}*K{r},0))",
            fmt="#,##0")
        # M: 메모
        inp(ws, r, 13)
        ws.row_dimensions[r].height = 18

    # 드롭다운: 거래구분 하드코딩, 계좌는 참고 시트 동적 참조
    dv(ws, f'"{",".join(TRADE_TYPES)}"',        f"B{TD_S}:B{TD_E}")
    dv(ws, f"'{REF}'!$A${ACCT_START}:$A${ACCT_END+10}", f"C{TD_S}:C{TD_E}")

    print(f"  ✓ '{TRADE}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# 배당금 시트
# ══════════════════════════════════════════════════════════════════════════════
def make_dividend(wb):
    ws = wb.create_sheet(DIV)
    NCOLS = 8

    COL_META = [
        ("날짜",         11, "YYYY-MM-DD", "center"),
        ("종목코드",     10, "@",           "center"),
        ("종목명",       20, None,          "left"),
        ("계좌",         14, None,          "center"),
        ("배당금총액",   14, "#,##0",       "right"),
        ("원천세율(%)",  12, "0.0",         "right"),
        ("원천세액",     12, "#,##0",       "right"),
        ("실수령액",     12, "#,##0",       "right"),
        # ── 투자현황에서는 H열(실수령액)을 SUMPRODUCT로 가져옴 ──
    ]
    NCOLS = len(COL_META) + 1  # +메모
    title(ws, 1, "💵 배당금 내역", NCOLS, bg=C["green"])
    note(ws, 2, 1,
         "★ D·E·F 입력 → G(원천세)·H(실수령) 자동계산 | 투자현황의 배당금은 H열 합산")

    for ci, (lbl, w, fmt, al) in enumerate(COL_META, 1):
        hdr_cell(ws, 3, ci, lbl, bg=C["green"])
        cw(ws, ci, w)
    hdr_cell(ws, 3, 9, "메모", bg=C["green"])
    cw(ws, 9, 25)
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I3"

    for r in range(DV_S, DV_E + 1):
        inp(ws, r, 1, fmt="YYYY-MM-DD", h="center")
        inp(ws, r, 2, fmt="@", h="center")
        # 종목명 VLOOKUP
        fml(ws, r, 3,
            f"=IFERROR(VLOOKUP(B{r},'{KRX}'!$A:$B,2,FALSE),\"\")",
            h="left")
        inp(ws, r, 4, h="center")
        inp(ws, r, 5, fmt="#,##0", h="right")   # 배당금총액
        inp(ws, r, 6, fmt="0.0", h="right")      # 원천세율(%)
        # 원천세액 = 총액 × 세율/100
        fml(ws, r, 7,
            f"=IF(OR(E{r}=\"\",F{r}=\"\"),\"\",ROUND(E{r}*F{r}/100,0))",
            fmt="#,##0")
        # 실수령액 = 총액 - 원천세
        fml(ws, r, 8,
            f"=IF(E{r}=\"\",\"\",E{r}-IF(G{r}=\"\",0,G{r}))",
            fmt="#,##0")
        inp(ws, r, 9)
        ws.row_dimensions[r].height = 18

    dv(ws, f"'{REF}'!$A${ACCT_START}:$A${ACCT_END+10}", f"D{DV_S}:D{DV_E}")
    print(f"  ✓ '{DIV}' 시트 생성 완료")


# ══════════════════════════════════════════════════════════════════════════════
# 투자현황 시트 — FIFO 로트 추적 방식
#
# 컬럼:
#   A매수일  B종목코드  C종목명(vlookup)  D계좌
#   E매수수량 F매수단가 G매수금액 H매수수수료
#   I매도일  J매도수량  K매도단가  L매도금액  M매도수수료 N매매세
#   O보유수량(수식) P보유상태
#   Q실현손익(수식) R현재가(vlookup) S미실현손익(수식)
#   T총손익(수식)   U손익률(수식)    V보유일수(수식)
#   W배당금(sumproduct)
# ══════════════════════════════════════════════════════════════════════════════

HOLD_COLS = [
    # (헤더,           너비,  서식,            정렬)
    ("매수일",          11,  "YYYY-MM-DD",   "center"),  # A
    ("종목코드",        10,  "@",            "center"),  # B
    ("종목명",          20,  None,           "left"),    # C  vlookup
    ("계좌",            14,  None,           "center"),  # D
    ("매수수량",        10,  "#,##0",        "right"),   # E
    ("매수단가",        12,  "#,##0",        "right"),   # F
    ("매수금액",        14,  "#,##0",        "right"),   # G
    ("매수수수료",      12,  "#,##0",        "right"),   # H
    ("매도일",          11,  "YYYY-MM-DD",   "center"),  # I
    ("매도수량",        10,  "#,##0",        "right"),   # J
    ("매도단가",        12,  "#,##0",        "right"),   # K
    ("매도금액",        14,  "#,##0",        "right"),   # L
    ("매도수수료",      12,  "#,##0",        "right"),   # M
    ("매매세",          10,  "#,##0",        "right"),   # N
    ("보유수량",        10,  "#,##0",        "right"),   # O  수식
    ("보유상태",        10,  None,           "center"),  # P
    ("실현손익",        14,  '+#,##0;-#,##0;"-"', "right"),  # Q  수식
    ("현재가",          12,  "#,##0",        "right"),   # R  vlookup
    ("미실현손익",      14,  '+#,##0;-#,##0;"-"', "right"),  # S  수식
    ("총  손익",        14,  '+#,##0;-#,##0;"-"', "right"),  # T  수식
    ("손익률",          10,  "+0.00%;-0.00%","right"),   # U  수식
    ("보유일수",        10,  "#,##0",        "right"),   # V  수식
    ("배당금",          11,  "#,##0",        "right"),   # W  sumproduct
]
HOLD_NCOLS = len(HOLD_COLS)  # 23

# 수식 컬럼 인덱스 (1-based)
HC = {l: i for i, (l, *_) in enumerate(HOLD_COLS, 1)}
# HC["매수일"]=1, HC["종목코드"]=2, ..., HC["배당금"]=23


def make_holdings(wb):
    """투자현황 시트 템플릿 생성 (헤더·서식만; 데이터는 refresh_holdings가 채움)."""
    ws = wb.create_sheet(HOLD)

    title(ws, 1, "📊 투자현황 — FIFO 로트별 손익 추적", HOLD_NCOLS)
    note(ws, 2, 1,
         "★ 거래내역 입력 후 python update_asset_manager.py --refresh 실행 → 자동 갱신")
    note(ws, 2, 12,
         "※ 현재가 갱신: python update_asset_manager.py --update | "
         "손익 = 실현(매도분) + 미실현(보유분)")

    # 헤더
    for ci, (lbl, w, fmt, al) in enumerate(HOLD_COLS, 1):
        static_cols = {1,2,4,5,6,7,8,9,10,11,12,13,14,16}  # 스크립트가 값으로 씀
        bg = C["navy"] if ci in static_cols else C["blue"]
        hdr_cell(ws, 3, ci, lbl, bg=bg, size=9, wrap=True)
        cw(ws, ci, w)
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(HOLD_NCOLS)}3"

    # 데이터 행: 서식만 적용 (값은 refresh_holdings가 씀)
    for r in range(HD_S, HD_E + 1):
        for ci, (_, w, fmt, al) in enumerate(HOLD_COLS, 1):
            c = ws.cell(row=r, column=ci)
            c.fill   = _fill(C["lgray"])
            c.border = _bdr()
            c.font   = _f()
            c.alignment = _aln(al)
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[r].height = 18

    print(f"  ✓ '{HOLD}' 시트 생성 완료 (데이터: --refresh 로 채움)")


# ══════════════════════════════════════════════════════════════════════════════
# FIFO 매칭 & 투자현황 갱신
# ══════════════════════════════════════════════════════════════════════════════

def refresh_holdings(wb):
    """거래내역 → FIFO 매칭 → 투자현황 재기록."""
    from collections import defaultdict

    # 수식 결과값을 읽으려면 data_only=True로 별도 로드
    read_wb  = openpyxl.load_workbook(FILE, data_only=True)
    trade_ws = read_wb[TRADE]
    hold_ws  = wb[HOLD]

    def _num(v):
        """수식 문자열이나 None 을 0으로 처리."""
        if v is None or isinstance(v, str):
            return 0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0

    # ── 1. 거래내역 읽기 ─────────────────────────────────────────────────────
    txs = []
    for r in range(TD_S, TD_E + 1):
        date_val = trade_ws.cell(r, 1).value
        t_type   = (trade_ws.cell(r, 2).value or "").strip()
        account  = (trade_ws.cell(r, 3).value or "").strip()
        code     = str(trade_ws.cell(r, 4).value or "").strip()
        price    = _num(trade_ws.cell(r, 6).value)
        qty      = _num(trade_ws.cell(r, 7).value)
        fee      = _num(trade_ws.cell(r, 10).value)  # J열 수수료 (수식 결과)
        tax      = _num(trade_ws.cell(r, 12).value)  # L열 세금 (수식 결과)

        if not date_val or not t_type or not code or qty == 0:
            continue
        txs.append(dict(date=date_val, type=t_type, account=account,
                        code=code, price=price, qty=qty, fee=fee, tax=tax))

    txs.sort(key=lambda x: (str(x["date"]), x["code"]))

    # ── 2. FIFO 매칭 ─────────────────────────────────────────────────────────
    # lots[code][account] = list of lot dicts
    lots = defaultdict(lambda: defaultdict(list))
    output = []

    for tx in txs:
        code, account = tx["code"], tx["account"]

        if tx["type"] in ("매수", "배당재투자"):
            lots[code][account].append({
                "buy_date":   tx["date"],
                "total_qty":  tx["qty"],
                "remain_qty": tx["qty"],
                "price":      tx["price"],
                "amount":     tx["price"] * tx["qty"],
                "fee":        tx["fee"],
            })

        elif tx["type"] == "매도":
            remain_sell = tx["qty"]
            for lot in lots[code][account]:
                if remain_sell <= 0 or lot["remain_qty"] <= 0:
                    continue
                matched    = min(lot["remain_qty"], remain_sell)
                buy_ratio  = matched / lot["total_qty"]
                sell_ratio = matched / tx["qty"]

                output.append({
                    "buy_date":    lot["buy_date"],
                    "code":        code,
                    "account":     account,
                    "buy_qty":     matched,
                    "buy_price":   lot["price"],
                    "buy_amount":  round(lot["price"] * matched),
                    "buy_fee":     round(lot["fee"] * buy_ratio),
                    "sell_date":   tx["date"],
                    "sell_qty":    matched,
                    "sell_price":  tx["price"],
                    "sell_amount": round(tx["price"] * matched),
                    "sell_fee":    round(tx["fee"] * sell_ratio),
                    "sell_tax":    round(tx["tax"] * sell_ratio),
                    "status":      "매도완료",
                })
                lot["remain_qty"] -= matched
                remain_sell       -= matched

    # 잔여 보유 로트
    for code, acct_map in lots.items():
        for account, lot_list in acct_map.items():
            for lot in lot_list:
                if lot["remain_qty"] <= 0:
                    continue
                remain  = lot["remain_qty"]
                ratio   = remain / lot["total_qty"]
                output.append({
                    "buy_date":    lot["buy_date"],
                    "code":        code,
                    "account":     account,
                    "buy_qty":     remain,
                    "buy_price":   lot["price"],
                    "buy_amount":  round(lot["price"] * remain),
                    "buy_fee":     round(lot["fee"] * ratio),
                    "sell_date":   None,
                    "sell_qty":    0,
                    "sell_price":  None,
                    "sell_amount": 0,
                    "sell_fee":    0,
                    "sell_tax":    0,
                    "status":      "보유중",
                })

    # 정렬: 종목코드 → 계좌 → 매수일 → 매도완료 먼저
    output.sort(key=lambda x: (
        x["code"], x["account"],
        str(x["buy_date"]),
        0 if x["status"] == "매도완료" else 1,
    ))

    # ── 3. 투자현황 시트 재기록 ──────────────────────────────────────────────
    # 기존 데이터 행 초기화
    for r in range(HD_S, HD_E + 1):
        for ci in range(1, HOLD_NCOLS + 1):
            c = hold_ws.cell(r, ci)
            c.value          = None
            c.fill           = _fill(C["lgray"])
            c.border         = _bdr()
            c.font           = _f()
            c.number_format  = HOLD_COLS[ci - 1][2] or "General"
            c.alignment      = _aln(HOLD_COLS[ci - 1][3])

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(top=thin, left=thin, right=thin, bottom=thin)

    def _wv(r, col_name, value, fmt=None, h=None):
        """값 쓰기."""
        ci = HC[col_name]
        _, _, default_fmt, default_al = HOLD_COLS[ci - 1]
        c = hold_ws.cell(r, ci, value=value)
        c.fill          = _fill(C["white"])
        c.border        = bdr
        c.font          = _f()
        c.number_format = fmt or default_fmt or "General"
        c.alignment     = _aln(h or default_al)
        return c

    def _wf(r, col_name, formula, fmt=None, h=None):
        """수식 쓰기."""
        ci = HC[col_name]
        _, _, default_fmt, default_al = HOLD_COLS[ci - 1]
        c = hold_ws.cell(r, ci, value=formula)
        c.fill          = _fill(C["formula"])
        c.border        = bdr
        c.font          = _f()
        c.number_format = fmt or default_fmt or "General"
        c.alignment     = _aln(h or default_al)
        return c

    for i, row in enumerate(output):
        if i >= (HD_E - HD_S + 1):
            print(f"  ⚠ 행 수 초과 ({len(output)}개 > {HD_E - HD_S + 1}). 뒤 항목 생략")
            break
        r = HD_S + i

        # ── 정적 값 (스크립트가 씀) ───────────────────────────────────────
        _wv(r, "매수일",     row["buy_date"])
        _wv(r, "종목코드",   row["code"])
        _wv(r, "계좌",       row["account"])
        _wv(r, "매수수량",   row["buy_qty"])
        _wv(r, "매수단가",   row["buy_price"])
        _wv(r, "매수금액",   row["buy_amount"])
        _wv(r, "매수수수료", row["buy_fee"])
        _wv(r, "매도수량",   row["sell_qty"] if row["sell_qty"] else None)
        _wv(r, "매도단가",   row["sell_price"])
        _wv(r, "매도금액",   row["sell_amount"] if row["sell_amount"] else None)
        _wv(r, "매도수수료", row["sell_fee"] if row["sell_fee"] else None)
        _wv(r, "매매세",     row["sell_tax"] if row["sell_tax"] else None)
        _wv(r, "보유상태",   row["status"])

        # 매도일
        if row["sell_date"]:
            _wv(r, "매도일", row["sell_date"])

        # ── 보유상태별 배경색 ────────────────────────────────────────────
        status_bg = C["lgreen"] if row["status"] == "보유중" else C["lorange"]
        hold_ws.cell(r, HC["보유상태"]).fill = _fill(status_bg)

        # ── 수식 컬럼 ────────────────────────────────────────────────────
        # C: 종목명
        _wf(r, "종목명",
            f"=IFERROR(VLOOKUP(B{r},'{KRX}'!$A:$B,2,FALSE),\"\")")
        # O: 보유수량 = 매수수량 - 매도수량
        _wf(r, "보유수량",
            f"=E{r}-IF(J{r}=\"\",0,J{r})")
        # Q: 실현손익 = 매도금액 - 매수금액 - 매수수수료 - 매도수수료 - 매매세
        _wf(r, "실현손익",
            f"=IF(J{r}>0,L{r}-G{r}-H{r}-IF(M{r}=\"\",0,M{r})-IF(N{r}=\"\",0,N{r}),0)")
        # R: 현재가
        _wf(r, "현재가",
            f"=IFERROR(VLOOKUP(B{r},'{KRX}'!$A:$C,3,FALSE),\"\")")
        # S: 미실현손익 = 보유수량 × 현재가 - 매수금액 - 매수수수료
        _wf(r, "미실현손익",
            f"=IF(O{r}>0,O{r}*R{r}-G{r}-H{r},0)")
        # T: 총손익
        _wf(r, "총  손익",
            f"=Q{r}+S{r}")
        # U: 손익률
        _wf(r, "손익률",
            f"=IF(G{r}>0,T{r}/G{r},\"\")")
        # V: 보유일수
        _wf(r, "보유일수",
            f"=IF(A{r}=\"\",\"\",IF(O{r}>0,TODAY()-A{r},"
            f"IF(I{r}=\"\",\"\",I{r}-A{r})))",
            fmt="#,##0")
        # W: 배당금 (종목코드+계좌 기준 합산)
        _wf(r, "배당금",
            f"=SUMPRODUCT(('{DIV}'!$B${DV_S}:$B${DV_E}=B{r})"
            f"*('{DIV}'!$D${DV_S}:$D${DV_E}=D{r})"
            f"*'{DIV}'!$H${DV_S}:$H${DV_E})")

        hold_ws.row_dimensions[r].height = 18

    total_rows = len(output)
    print(f"  ✓ '{HOLD}' 갱신 완료 — {total_rows}개 로트 (보유중: "
          f"{sum(1 for r in output if r['status']=='보유중')}, "
          f"매도완료: {sum(1 for r in output if r['status']=='매도완료')})")


# ══════════════════════════════════════════════════════════════════════════════
# main — 기존 파일 업데이트
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="자산관리_2026.xlsx 업데이트")
    parser.add_argument("--refresh", action="store_true",
                        help="거래내역 → FIFO → 투자현황 갱신")
    parser.add_argument("--update", action="store_true",
                        help="KRX 현재가만 갱신 (create_asset_manager.py --update 호출)")
    args = parser.parse_args()

    if not os.path.exists(FILE):
        print(f"파일 없음: {FILE}")
        return

    # ── --refresh: 거래내역 읽어 투자현황 재기록 ────────────────────────────
    if args.refresh:
        print(f"[--refresh] 투자현황 FIFO 갱신 시작")
        wb = openpyxl.load_workbook(FILE)
        if TRADE not in wb.sheetnames:
            print(f"  ⚠ '{TRADE}' 시트 없음. 먼저 기본 실행으로 시트를 생성하세요.")
            return
        if HOLD not in wb.sheetnames:
            print(f"  ⚠ '{HOLD}' 시트 없음. 먼저 기본 실행으로 시트를 생성하세요.")
            return
        refresh_holdings(wb)
        wb.save(FILE)
        print(f"✅ 저장 완료: {FILE}")
        return

    # ── --update: KRX 현재가 갱신은 create_asset_manager.py 위임 ─────────────
    if args.update:
        import subprocess
        script = os.path.join(BASE_DIR, "create_asset_manager.py")
        subprocess.run([sys.executable, script, "--update"], check=True)
        return

    # ── 기본 실행: 시트 재구성 ───────────────────────────────────────────────
    print(f"파일 로드: {FILE}")
    wb = openpyxl.load_workbook(FILE)
    print(f"기존 시트: {wb.sheetnames}")

    # 대체할 시트 삭제
    DROP = ["투자거래내역", "입출금내역", "계좌별현황"]
    for name in DROP:
        if name in wb.sheetnames:
            del wb[name]
            print(f"  삭제: {name}")
    # 새로 만들 시트도 미리 삭제 (재실행 시 중복 방지)
    for name in [REF, TRADE, DIV, HOLD]:
        if name in wb.sheetnames:
            del wb[name]
            print(f"  재생성 위해 삭제: {name}")

    # 새 시트 생성
    make_ref(wb)
    make_trade(wb)
    make_dividend(wb)
    make_holdings(wb)

    # 시트 순서 재정렬
    TARGET_ORDER = [
        TRADE, DIV, REF, HOLD,
        "투자아이디어", "매도모니터링", KRX,
        "📊순자산대시보드", "💰월별지출", "📈포트폴리오요약",
    ]
    current = wb.sheetnames
    ordered = [s for s in TARGET_ORDER if s in current]
    rest    = [s for s in current if s not in TARGET_ORDER]
    for i, name in enumerate(ordered + rest):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(FILE)
    print(f"\n✅ 저장 완료: {FILE}")
    print(f"   최종 시트: {wb.sheetnames}")

if __name__ == "__main__":
    main()
