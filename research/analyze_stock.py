#!/usr/bin/env python3
"""
종목 심층 분석 보고서 생성 (PDF 다운로드 + 분석 통합)

사용법:
  python analyze_stock.py                  # 관심종목 시트의 모든 종목 처리
  python analyze_stock.py 005930           # 특정 종목코드
  python analyze_stock.py 삼성전자         # 특정 종목명

선택적 환경변수:
  ANTHROPIC_API_KEY         → PDF AI 분석  (pip install anthropic)
  NAVER_CLIENT_ID/SECRET    → 경제신문 뉴스 검색
  DART_API_KEY              → DART 공시 조회 (opendart.fss.or.kr 무료 발급)

출력: research/{종목명}_분석_{YYYYMMDD}.xlsx
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os, re, glob, time, json, argparse, requests, openpyxl
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import pytz

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

BASE_DIR      = Path(__file__).parent
MAIN_EXCEL    = BASE_DIR / "Daily_index.xlsx"
REPORT_SHEET  = "리서치 보고서"
WATCHLIST_SHEET = "관심종목"
MIN_DATE      = "2025-01-01"
KST           = pytz.timezone("Asia/Seoul")

# ── api_keys.txt 파일에서 키 로드 (환경변수보다 우선) ─────────────────────────
def _load_api_keys():
    """research/api_keys.txt 에서 키=값 형태로 읽어 os.environ에 주입."""
    keys_file = BASE_DIR / "api_keys.txt"
    if not keys_file.exists():
        return
    with open(keys_file, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                k = k.strip(); v = v.strip().strip('"').strip("'")
                if k and v:
                    os.environ.setdefault(k, v)

_load_api_keys()

TOP5_DOMAINS = {
    "mk.co.kr":       "매일경제",
    "hankyung.com":   "한국경제",
    "heraldcorp.com": "헤럴드경제",
    "sedaily.com":    "서울경제",
    "edaily.co.kr":   "이데일리",
}

DL_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://stock.pstatic.net/",
}

# ── 스타일 ────────────────────────────────────────────────────────────────────

S = {
    "hdr_dark":  PatternFill("solid", fgColor="1F4E79"),
    "hdr_mid":   PatternFill("solid", fgColor="2E75B6"),
    "hdr_light": PatternFill("solid", fgColor="D6E4F0"),
    "buy":       PatternFill("solid", fgColor="E8F5E9"),
    "sell":      PatternFill("solid", fgColor="FFEBEE"),
    "hold":      PatternFill("solid", fgColor="FFF9C4"),
    "alt_row":   PatternFill("solid", fgColor="F5F9FF"),
    "white":     PatternFill("solid", fgColor="FFFFFF"),
    "gold":      PatternFill("solid", fgColor="FFF3CD"),
    "thin": Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    ),
    "f_white_bold": Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10),
    "f_dark":       Font(color="1F4E79", bold=True, name="맑은 고딕", size=10),
    "f_normal":     Font(name="맑은 고딕", size=10),
    "f_small":      Font(name="맑은 고딕", size=9),
    "f_link":       Font(name="맑은 고딕", size=10, color="0563C1", underline="single"),
    "center":  Alignment(horizontal="center", vertical="center"),
    "left":    Alignment(horizontal="left",   vertical="center"),
    "wrap":    Alignment(horizontal="left",   vertical="top", wrap_text=True),
}


def _hdr(ws, row, col, val, style="dark", w=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = S[f"hdr_{style}"]
    c.font = S["f_white_bold"] if style != "light" else S["f_dark"]
    c.alignment = S["center"]
    c.border = S["thin"]
    if w:
        ws.column_dimensions[get_column_letter(col)].width = w
    return c


def _cell(ws, row, col, val, fill=None, font=None, align="left", fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill or S["white"]
    c.font = font or S["f_normal"]
    c.alignment = S[align] if isinstance(align, str) else align
    c.border = S["thin"]
    if fmt:
        c.number_format = fmt
    return c


def _op_fill(op: str):
    op = (op or "").upper()
    if "매수" in op or "BUY" in op or "아웃퍼폼" in op:
        return S["buy"]
    if "매도" in op or "SELL" in op or "언더퍼폼" in op:
        return S["sell"]
    return S["hold"]


# ── 관심종목 시트 읽기 ─────────────────────────────────────────────────────────

def get_watchlist() -> list[str]:
    if not MAIN_EXCEL.exists():
        return []
    wb = openpyxl.load_workbook(str(MAIN_EXCEL), data_only=True, read_only=True)
    if WATCHLIST_SHEET not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[WATCHLIST_SHEET]
    stocks = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        val = row[0] if row else None
        if val and str(val).strip() and not str(val).startswith("▼"):
            stocks.append(str(val).strip())
    wb.close()
    return stocks


# ── PDF URL 추출 ───────────────────────────────────────────────────────────────

def extract_pdf_url(cell_val) -> str | None:
    """=HYPERLINK("url","PDF") 수식 또는 직접 URL에서 PDF 주소 추출."""
    if not cell_val:
        return None
    s = str(cell_val)
    # 수식 형태: =HYPERLINK("https://...", "PDF")
    m = re.search(r'HYPERLINK\("([^"]+)"', s)
    if m:
        return m.group(1)
    # 직접 URL
    if s.startswith("http") and "pstatic.net" in s:
        return s
    return None


# ── Daily_index 파일 스캔 ─────────────────────────────────────────────────────

def scan_all_excel_files(stock_query: str, min_date: str = MIN_DATE) -> list[dict]:
    """
    모든 Daily_index*.xlsx에서 종목 매칭 리포트 수집.
    data_only=False 필수: True로 읽으면 =HYPERLINK() 수식 URL이 소실됨.
    """
    files = sorted(glob.glob(str(BASE_DIR / "Daily_index*.xlsx")))
    if not files:
        print("  오류: Daily_index*.xlsx 파일 없음")
        return []

    q = stock_query.strip()
    is_code = bool(re.match(r"^\d{4,6}$", q))
    q_lower = q.lower()

    all_reports: list[dict] = []

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=False, read_only=True)
        except Exception as e:
            print(f"  [경고] {fname}: {e}")
            continue

        if REPORT_SHEET not in wb.sheetnames:
            wb.close()
            continue

        ws = wb[REPORT_SHEET]
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            date_val = str(row[0] or "").strip()
            if date_val < min_date:
                continue

            code    = str(row[1] or "").strip()
            name    = str(row[2] or "").strip()
            firm    = str(row[3] or "").strip()
            opinion = str(row[4] or "").strip()
            target  = row[5]
            prev_p  = row[6]
            title   = str(row[7] or "").strip()
            pdf_cell = row[9]   # col 10: PDF (수식 문자열 그대로)

            if is_code:
                matched = (code == q or code.lstrip("0") == q)
            else:
                matched = q_lower in name.lower()
            if not matched:
                continue

            pdf_url = extract_pdf_url(pdf_cell)

            all_reports.append({
                "date":      date_val,
                "code":      code,
                "name":      name,
                "firm":      firm,
                "opinion":   opinion,
                "target":    int(str(target).replace(",", "")) if target else None,
                "prev_price": int(str(prev_p).replace(",", "")) if prev_p else None,
                "title":     title,
                "pdf_url":   pdf_url,
                "source":    fname,
                "_key":      f"{date_val}_{firm}",
            })
            count += 1

        wb.close()
        if count > 0:
            print(f"    {fname}: {count}건")

    all_reports.sort(key=lambda x: x["date"], reverse=True)
    return all_reports


# ── PDF 다운로드 ───────────────────────────────────────────────────────────────

def sanitize(s: str) -> str:
    return re.sub(r'[<>:"/\\|?*\n\r\t]', '_', str(s)).strip("_. ")


def download_pdfs(reports: list[dict], stock_name: str) -> Path:
    """리포트 목록에서 PDF 다운로드. 저장 폴더 경로 반환."""
    folder = BASE_DIR / sanitize(stock_name)
    folder.mkdir(parents=True, exist_ok=True)

    with_url = [r for r in reports if r.get("pdf_url")]
    print(f"  PDF 링크 있음: {len(with_url)}/{len(reports)}건")

    ok = fail = skip = 0
    for r in with_url:
        date_s = r["date"].replace("-", "")
        fname  = f"{date_s}_{sanitize(r['firm'])}_{sanitize(r['title'])[:40]}.pdf"
        spath  = folder / fname

        if spath.exists():
            skip += 1
            continue

        try:
            resp = requests.get(r["pdf_url"], headers=DL_HEADERS, timeout=25)
            if resp.status_code == 200 and len(resp.content) > 500:
                spath.write_bytes(resp.content)
                ok += 1
                print(f"    ✓ {r['date']} {r['firm']:<12} ({len(resp.content)//1024}KB)")
            else:
                fail += 1
                print(f"    ✗ {r['date']} {r['firm']:<12} HTTP {resp.status_code}")
        except Exception as e:
            fail += 1
            print(f"    ✗ {r['date']} {r['firm']:<12} {str(e)[:50]}")
        time.sleep(0.25)

    print(f"  다운로드: {ok}건 ✓  {fail}건 ✗  {skip}건 이미 존재")
    return folder


# ── PDF 텍스트 추출 ───────────────────────────────────────────────────────────

def extract_pdf_text(pdf_path: Path) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            parts = [p.extract_text() for p in pdf.pages[:8] if p.extract_text()]
        return "\n".join(parts)[:8000]
    except ImportError:
        return ""
    except Exception as e:
        return f"[추출 실패: {e}]"


# ── AI 분석 ───────────────────────────────────────────────────────────────────

_AI_PROMPT = """다음은 증권사 리서치 보고서입니다. JSON만 반환하세요 (설명 없이):
{
  "bull_reasons": ["긍정/매수 근거 (최대 5개, 간결하게)"],
  "bear_risks":   ["리스크/부정 요인 (최대 3개)"],
  "catalysts":    ["주요 촉매/이벤트 (최대 3개)"],
  "summary":      "핵심 투자 논리 1-2문장",
  "topic_tags":   ["핵심 키워드 2-3개"]
}

보고서:
"""

def analyze_with_ai(text: str, firm: str) -> dict | None:
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key or not text or text.startswith("[추출 실패"):
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=key)
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=800,
            messages=[{"role": "user", "content": _AI_PROMPT + text[:6000]}],
        )
        raw = msg.content[0].text.strip()
        m = re.search(r"\{[\s\S]+\}", raw)
        if m:
            return json.loads(m.group(0))
    except Exception as e:
        print(f"    [AI 오류 - {firm}]: {e}")
    return None


# ── 현재가 조회 ───────────────────────────────────────────────────────────────

def get_current_price(code: str) -> tuple[int | None, str]:
    clean = re.sub(r"^A", "", code)
    url = f"https://m.stock.naver.com/api/stock/{clean}/basic"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer":    "https://m.stock.naver.com/",
    }
    try:
        data = requests.get(url, headers=headers, timeout=8).json()
        price = None
        for f in ("closePrice", "currentPrice", "stockEndPrice"):
            v = data.get(f, "")
            if v:
                price = int(str(v).replace(",", ""))
                break
        return price, str(data.get("fluctuationsRatio", ""))
    except Exception:
        return None, ""


# ── 이전 분석 파일 로드 ───────────────────────────────────────────────────────

def load_previous_analysis(stock_name: str) -> tuple[dict, list, list]:
    """
    가장 최근 {종목명}_분석_*.xlsx 에서 기존 데이터 로드.
    반환: (analyses_dict, old_news_list, old_dart_list)
    """
    pattern = str(BASE_DIR / f"{sanitize(stock_name)}_분석_*.xlsx")
    files   = sorted(glob.glob(pattern), reverse=True)
    if not files:
        return {}, [], []

    prev_path = Path(files[0])
    print(f"  이전 파일 발견: {prev_path.name} → 기존 데이터 재사용")

    analyses: dict    = {}
    old_news: list    = []
    old_dart: list    = []

    try:
        wb = openpyxl.load_workbook(str(prev_path), data_only=True, read_only=False)

        # ── 리포트분석 시트에서 AI 결과 복원 ──
        if "리포트분석" in wb.sheetnames:
            ws = wb["리포트분석"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                date = str(row[0] or "").strip()
                firm = str(row[1] or "").strip()
                if not date or not firm or not re.match(r"\d{4}-\d{2}-\d{2}", date):
                    continue  # 차트 섹션 헤더 등 스킵
                summary = str(row[8]  or "").strip()
                bull    = str(row[9]  or "").strip()
                bear    = str(row[10] or "").strip()
                cats    = str(row[11] or "").strip()
                tags    = str(row[12] or "").strip()
                if summary or bull:
                    def _to_list(s):
                        return [x.lstrip("•· ").strip() for x in s.split("\n")
                                if x.lstrip("•· ").strip()]
                    analyses[f"{date}_{firm}"] = {
                        "summary":      summary,
                        "bull_reasons": _to_list(bull),
                        "bear_risks":   _to_list(bear),
                        "catalysts":    _to_list(cats),
                        "topic_tags":   [t.strip() for t in tags.split(",") if t.strip()],
                    }

        # ── 뉴스 시트에서 기존 기사 복원 ──
        if "뉴스" in wb.sheetnames:
            ws = wb["뉴스"]
            for row_cells in ws.iter_rows(min_row=2):
                vals = [c.value for c in row_cells]
                if not vals or not vals[0]:
                    continue
                link = ""
                if len(row_cells) >= 5 and row_cells[4].hyperlink:
                    link = row_cells[4].hyperlink.target or ""
                old_news.append({
                    "date":   str(vals[0] or "").strip(),
                    "source": str(vals[1] or "").strip(),
                    "title":  str(vals[2] or "").strip(),
                    "desc":   str(vals[3] or "").strip(),
                    "link":   link,
                })

        # ── 공시 시트에서 기존 공시 복원 ──
        if "공시" in wb.sheetnames:
            ws = wb["공시"]
            for row_cells in ws.iter_rows(min_row=2):
                vals = [c.value for c in row_cells]
                if not vals or not vals[0]:
                    continue
                rcept_no = ""
                if len(row_cells) >= 4 and row_cells[3].hyperlink:
                    href = row_cells[3].hyperlink.target or ""
                    m = re.search(r"rcpNo=(\w+)", href)
                    if m:
                        rcept_no = m.group(1)
                old_dart.append({
                    "date":     str(vals[0] or "").strip(),
                    "type":     str(vals[1] or "").strip(),
                    "filer":    str(vals[2] or "").strip(),
                    "rcept_no": rcept_no,
                })

        wb.close()
    except Exception as e:
        print(f"  [경고] 이전 파일 로드 중 오류: {e}")

    print(f"  → AI분석 {len(analyses)}건 / 뉴스 {len(old_news)}건 / 공시 {len(old_dart)}건 재사용")
    return analyses, old_news, old_dart


# ── 뉴스 검색 ─────────────────────────────────────────────────────────────────

def fetch_naver_news(query: str, max_items: int = 80,
                     since_date: str | None = None) -> list[dict]:
    """since_date 이후 기사만 수집. None이면 전체."""
    cid  = os.environ.get("NAVER_CLIENT_ID", "")
    csec = os.environ.get("NAVER_CLIENT_SECRET", "")
    if not cid or not csec:
        print("  [뉴스 건너뜀] NAVER_CLIENT_ID/SECRET 미설정")
        return []

    headers   = {"X-Naver-Client-Id": cid, "X-Naver-Client-Secret": csec}
    collected = []

    for start in range(1, max_items + 1, 25):
        params = {"query": query, "display": 25, "start": start, "sort": "date"}
        stop   = False
        try:
            items = requests.get(
                "https://openapi.naver.com/v1/search/news.json",
                headers=headers, params=params, timeout=10
            ).json().get("items", [])
            if not items:
                break
            for it in items:
                link   = it.get("originallink") or it.get("link", "")
                source = next((n for d, n in TOP5_DOMAINS.items() if d in link), None)
                if not source:
                    continue
                title = re.sub(r"<[^>]+>", "", it.get("title", ""))
                desc  = re.sub(r"<[^>]+>", "", it.get("description", ""))
                if len(desc) < 50:
                    continue
                try:
                    from email.utils import parsedate_to_datetime
                    date_str = parsedate_to_datetime(it.get("pubDate","")).strftime("%Y-%m-%d")
                except Exception:
                    date_str = it.get("pubDate","")[:10]
                if date_str < "2025-01-01":
                    stop = True; break
                if since_date and date_str <= since_date:
                    stop = True; break
                collected.append({"date": date_str, "source": source,
                                   "title": title, "desc": desc[:300], "link": link})
        except Exception as e:
            print(f"  [뉴스 오류] {e}")
            break
        if stop:
            break
        time.sleep(0.2)

    collected.sort(key=lambda x: x["date"], reverse=True)
    return collected


# ── 뉴스 AI 필터링·중요도 선별 ────────────────────────────────────────────────

# 키워드 기반 사전 제거 패턴 (AI 호출 없이 빠르게)
_NEWS_DROP_PATTERNS = [
    re.compile(r"(증권|리서치|투자|운용).{0,10}(목표주가|투자의견).{0,15}(상향|하향|유지|제시|변경)", re.I),
    re.compile(r"(목표주가|투자의견).{0,15}(증권|리서치|투자|운용)", re.I),
    re.compile(r"(네이밍\s*스폰서|타이틀\s*스폰서|스포츠\s*후원|후원\s*협약)", re.I),
    re.compile(r"(임직원\s*(봉사|기부)|사회공헌|자원봉사|헌혈)", re.I),
    re.compile(r"(창립\s*\d+주년|기념식|시무식|종무식)", re.I),
]


def _keyword_filter_news(news_list: list) -> list:
    result = []
    for n in news_list:
        text = n["title"] + " " + n.get("desc", "")
        if any(p.search(text) for p in _NEWS_DROP_PATTERNS):
            continue
        result.append(n)
    return result


def filter_rank_news_ai(news_list: list, stock_name: str, max_output: int = 30) -> list:
    """키워드 사전 필터 → AI로 투자 중요도 순 선별."""
    if not news_list:
        return []

    # 1단계: 키워드 필터
    filtered = _keyword_filter_news(news_list)
    removed = len(news_list) - len(filtered)
    if removed:
        print(f"  키워드 필터: {removed}건 제거 → {len(filtered)}건 남음")

    if not filtered:
        return []

    # 2단계: AI 중요도 선별 (API 키 있을 때만)
    if not os.environ.get("ANTHROPIC_API_KEY") or len(filtered) <= max_output:
        return filtered[:max_output]

    client = anthropic.Anthropic()
    items_text = "\n".join(
        f"[{i}] {n['date']} {n['title']} | {n.get('desc','')[:80]}"
        for i, n in enumerate(filtered)
    )
    prompt = f"""{stock_name} 관련 뉴스 {len(filtered)}건입니다. 투자자 관점에서 중요도 순으로 최대 {max_output}건을 선별하세요.

뉴스:
{items_text}

중요 기준: 실적·매출·이익 변화, 신규 수주·계약, M&A·투자, 원가·공급망, 규제·소송, 신제품·기술, 주요 인사
불필요: 단순 홍보, 행사 참가, 임직원 현황, 스포츠/문화 관련

중요도 순으로 선택한 인덱스만 JSON으로 반환: {{"selected": [0, 3, 7, ...]}}"""

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=400,
            messages=[{"role": "user", "content": prompt}]
        )
        m = re.search(r'\{[^}]+\}', msg.content[0].text, re.DOTALL)
        if m:
            selected = json.loads(m.group()).get("selected", [])
            ranked = [filtered[i] for i in selected if 0 <= i < len(filtered)]
            if ranked:
                print(f"  AI 선별: {len(filtered)}건 → {len(ranked)}건")
                return ranked
    except Exception as e:
        print(f"  [AI 뉴스 선별 오류] {e}")

    return filtered[:max_output]


# ── DART 공시 ─────────────────────────────────────────────────────────────────

# 투자자 관점 중요 공시 키워드
IMPORTANT_DART_KEYWORDS = [
    "실적", "배당", "증자", "감자", "합병", "분할", "주주총회",
    "자기주식", "임원", "계약", "풍문", "조회공시", "사업보고서",
    "반기보고서", "분기보고서", "영업양수", "투자", "기술이전",
    "소송", "제재", "전환사채", "교환사채", "신주인수권", "공개매수",
    "최대주주", "주요사항", "잠정실적", "공급계약", "업무협약",
]


def fetch_dart_disclosures(code: str, since_date: str | None = None) -> list[dict]:
    """since_date 다음날 이후 공시만 조회. None이면 2025-01-01부터."""
    key = os.environ.get("DART_API_KEY", "")
    if not key:
        print("  [공시 건너뜀] DART_API_KEY 미설정 (opendart.fss.or.kr 무료 발급)")
        return []
    try:
        clean = re.sub(r"^A", "", code.strip())
        corp_resp = requests.get(
            "https://opendart.fss.or.kr/api/company.json",
            params={"crtfc_key": key, "stock_code": clean}, timeout=10
        ).json()
        if corp_resp.get("status") != "000":
            print(f"  [공시] 기업코드 조회 실패 ({clean}): {corp_resp.get('message', corp_resp.get('status',''))}")
            return []
        corp_code = corp_resp.get("corp_code", "")
        if not corp_code:
            print(f"  [공시] 기업코드 없음 ({clean})")
            return []

        if since_date:
            from datetime import timedelta
            bgn = (datetime.strptime(since_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y%m%d")
        else:
            bgn = "20250101"

        today = datetime.now(KST).strftime("%Y%m%d")
        if bgn > today:
            print("  공시: 새로운 항목 없음")
            return []

        list_resp = requests.get(
            "https://opendart.fss.or.kr/api/list.json",
            params={"crtfc_key": key, "corp_code": corp_code,
                    "bgn_de": bgn, "end_de": today,
                    "page_count": 100, "sort": "date", "sort_mth": "desc"},
            timeout=15
        ).json()
        if list_resp.get("status") not in ("000", "013"):
            print(f"  [공시] 목록 조회 실패: {list_resp.get('message', list_resp.get('status',''))}")
            return []

        all_items = list_resp.get("list", [])
        # 투자자 관점 중요 공시만 필터링
        result = []
        for it in all_items:
            rpt = it.get("report_nm", "")
            if any(kw in rpt for kw in IMPORTANT_DART_KEYWORDS):
                result.append({
                    "date":     it.get("rcept_dt", "")[:10],
                    "type":     rpt,
                    "filer":    it.get("flr_nm", ""),
                    "rcept_no": it.get("rcept_no", ""),
                })
        print(f"  전체 {len(all_items)}건 중 중요 공시 {len(result)}건 선별")
        return result
    except Exception as e:
        print(f"  [공시 오류] {e}")
        return []


# ── Excel 생성 ────────────────────────────────────────────────────────────────

def write_summary_sheet(wb, stock_name, code, history, current_price, change_rate):
    ws = wb.active
    ws.title = "요약"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.sheet_view.showGridLines = False

    for col, w in enumerate([20, 28, 28, 28, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"  {stock_name} ({code})  —  종목 분석 보고서"
    t.font  = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
    t.fill  = S["hdr_dark"]
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    ws["A2"].value = f"  분석일: {datetime.now(KST).strftime('%Y-%m-%d %H:%M')} KST  |  기간: {MIN_DATE} 이후"
    ws["A2"].font  = Font(name="맑은 고딕", size=9, color="666666", italic=True)
    ws["A2"].fill  = PatternFill("solid", fgColor="EEF4FB")
    ws.row_dimensions[2].height = 18

    _hdr(ws, 4, 1, "항목", "dark"); _hdr(ws, 4, 2, "값", "dark")
    _hdr(ws, 4, 3, "항목", "dark"); _hdr(ws, 4, 4, "값", "dark")

    left_data = [
        ("현재가",   f"{current_price:,}원" if current_price else "조회 실패"),
        ("등락률",   f"{change_rate}%" if change_rate else "-"),
        ("종목코드", code),
    ]
    right_data = []
    if history:
        targets  = [r["target"] for r in history if r["target"]]
        opinions = [r["opinion"] for r in history if r["opinion"]]
        consensus = Counter(opinions).most_common(1)[0][0] if opinions else "-"
        med_tgt   = sorted(targets)[len(targets)//2] if targets else None
        latest    = max(history, key=lambda x: x["date"])
        right_data = [
            ("컨센서스",     consensus),
            ("중위 목표주가", f"{med_tgt:,}원" if med_tgt else "-"),
            ("최근 리포트",  f"{latest['date']} {latest['firm']}"),
            ("리포트 수",    f"{len(history)}건"),
        ]

    for i, (k, v) in enumerate(left_data, 5):
        _cell(ws, i, 1, k, fill=S["hdr_light"], font=S["f_dark"], align="center")
        c = _cell(ws, i, 2, v, fill=S["buy"] if k == "현재가" else S["white"], align="center")
        if k == "현재가":
            c.hyperlink = f"https://finance.naver.com/item/main.naver?code={code}"
            c.font = Font(name="맑은 고딕", size=10, bold=True,
                          color="FFFFFF", underline="single")
    for i, (k, v) in enumerate(right_data, 5):
        _cell(ws, i, 3, k, fill=S["hdr_light"], font=S["f_dark"], align="center")
        _cell(ws, i, 4, v, fill=S["alt_row"] if i % 2 == 0 else S["white"], align="center")

    start = max(5 + len(left_data), 5 + len(right_data)) + 2
    ws.merge_cells(f"A{start}:E{start}")
    c = ws.cell(row=start, column=1, value="  최근 리포트 목록 (최대 15건)")
    c.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    c.fill = S["hdr_mid"]
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start].height = 24

    for col, (h, w) in enumerate([("날짜",12),("증권사",18),("투자의견",12),
                                   ("목표주가",12),("제목",45)], 1):
        _hdr(ws, start+1, col, h, "light")
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, r in enumerate(sorted(history, key=lambda x: x["date"], reverse=True)[:15], start+2):
        _cell(ws, i, 1, r["date"],   align="center")
        _cell(ws, i, 2, r["firm"])
        _cell(ws, i, 3, r["opinion"], fill=_op_fill(r["opinion"]), align="center")
        _cell(ws, i, 4, r["target"],  align="center", fmt="#,##0")
        _cell(ws, i, 5, r["title"])


# ── 뉴스+공시 AI 배치 분석 ────────────────────────────────────────────────────

def analyze_news_dart_timeline(items: list[dict], stock_name: str) -> dict[int, dict]:
    """
    뉴스+공시 항목 배치 AI 분석 (API 1회 호출).
    반환: {index: {"sentiment": "긍정/부정/중립", "implication": "한 문장"}}
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or not items:
        return {}

    lines = []
    for i, it in enumerate(items[:60]):
        line = f"{i}. [{it['date']}] [{it['type']}] {it['source']}: {it['title']}"
        if it.get("desc"):
            line += f"  ({it['desc'][:80]})"
        lines.append(line)

    prompt = f"""다음은 {stock_name} 관련 뉴스·공시 목록입니다.
각 항목의 투자 시사점을 한 문장으로 분석하고 JSON 배열만 반환하세요 (설명 없이):
[{{"id":0,"sentiment":"긍정","implication":"시사점"}},...]
sentiment는 "긍정" "부정" "중립" 중 하나.

{chr(10).join(lines)}"""

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        m = re.search(r'\[[\s\S]+\]', raw)
        if m:
            arr = json.loads(m.group(0))
            return {entry["id"]: entry for entry in arr if "id" in entry}
    except Exception as e:
        print(f"    [AI 뉴스·공시 분석 오류]: {e}")
    return {}


# ── 통합 리포트분석 시트 (리포트목록 + 투자논리 + 목표주가) ─────────────────

def write_report_analysis_sheet(wb, history: list, analyses: dict):
    ws = wb.create_sheet("리포트분석")
    ws.sheet_properties.tabColor = "2E75B6"
    ws.sheet_view.showGridLines = False

    # ── 헤더 ──
    HDR_COLS = [
        ("날짜",12), ("증권사",16), ("투자의견",11), ("목표주가",11),
        ("이전목표",11), ("변동",9), ("변동률",9), ("제목",42),
        ("AI 요약",48), ("긍정 논리",44), ("리스크",38), ("촉매",32), ("키워드",22),
    ]
    for col, (h, w) in enumerate(HDR_COLS, 1):
        _hdr(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── 데이터 행 ──
    sorted_h    = sorted(history, key=lambda x: x["date"], reverse=True)
    prev_by_firm: dict[str, int] = {}
    chart_data: list[tuple] = []   # (date, target) for chart

    for i, r in enumerate(sorted_h, 2):
        firm = r["firm"]; tgt = r["target"]
        prev = prev_by_firm.get(firm)
        chg  = (tgt - prev) if (tgt and prev) else None
        chgr = chg / prev    if (chg is not None and prev) else None
        if tgt:
            prev_by_firm[firm] = tgt
            chart_data.append((r["date"], tgt))

        rf  = S["alt_row"] if i % 2 == 0 else S["white"]
        ai  = analyses.get(f"{r['date']}_{r['firm']}") or {}
        chg_f = S["buy"] if (chg and chg > 0) else (S["sell"] if (chg and chg < 0) else rf)

        bull = ("• " + "\n• ".join(ai.get("bull_reasons", []))) if ai else ""
        bear = ("• " + "\n• ".join(ai.get("bear_risks",   []))) if ai else ""
        cats = ("• " + "\n• ".join(ai.get("catalysts",    []))) if ai else ""
        tags = ", ".join(ai.get("topic_tags", []))

        _cell(ws, i,  1, r["date"],    align="center")
        _cell(ws, i,  2, firm,         fill=rf)
        _cell(ws, i,  3, r["opinion"], fill=_op_fill(r["opinion"]), align="center")
        _cell(ws, i,  4, tgt,          fill=rf,    align="center", fmt="#,##0")
        _cell(ws, i,  5, prev,         fill=rf,    align="center", fmt="#,##0")
        _cell(ws, i,  6, f"{'+' if chg and chg>0 else ''}{chg:,}" if chg is not None else "-",
              fill=chg_f, align="center")
        _cell(ws, i,  7, chgr,         fill=chg_f, align="center",
              fmt="+0.0%;-0.0%" if chgr is not None else None)
        _cell(ws, i,  8, r["title"],   fill=rf)
        c9  = _cell(ws, i,  9, ai.get("summary",""), fill=rf); c9.alignment  = S["wrap"]
        c10 = _cell(ws, i, 10, bull,  fill=rf); c10.alignment = S["wrap"]
        c11 = _cell(ws, i, 11, bear,  fill=rf); c11.alignment = S["wrap"]
        c12 = _cell(ws, i, 12, cats,  fill=rf); c12.alignment = S["wrap"]
        _cell(ws, i, 13, tags, fill=S["gold"] if tags else rf)

        lines = max(bull.count("\n"), bear.count("\n"), cats.count("\n"),
                    (ai.get("summary","")).count("\n"), 0) + 1
        ws.row_dimensions[i].height = max(18, lines * 15)

    # ── 목표주가 차트 (시트 하단) ──
    if len(chart_data) >= 3:
        chart_data.sort(key=lambda x: x[0])
        chart_row_start = len(sorted_h) + 4

        ws.merge_cells(f"A{chart_row_start}:M{chart_row_start}")
        c = ws.cell(row=chart_row_start, column=1, value="  목표주가 변동 추이")
        c.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        c.fill = S["hdr_mid"]
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[chart_row_start].height = 22

        tbl_row = chart_row_start + 1
        _hdr(ws, tbl_row, 1, "날짜", "light")
        _hdr(ws, tbl_row, 2, "목표주가", "light")
        for j, (dt, tp) in enumerate(chart_data, tbl_row + 1):
            _cell(ws, j, 1, dt, align="center")
            _cell(ws, j, 2, tp, align="center", fmt="#,##0")

        try:
            n = len(chart_data)
            chart = LineChart()
            chart.title = "목표주가 추이"; chart.style = 10
            chart.y_axis.title = "원"; chart.x_axis.title = "날짜"
            chart.width = 28; chart.height = 14
            chart.add_data(Reference(ws, min_col=2, min_row=tbl_row,
                                     max_row=tbl_row + n), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=tbl_row + 1,
                                           max_row=tbl_row + n))
            ws.add_chart(chart, f"D{chart_row_start}")
        except Exception:
            pass


# ── 뉴스·공시 시계열 시트 ────────────────────────────────────────────────────

def write_news_dart_timeline_sheet(wb, news: list, disclosures: list,
                                   stock_name: str, ai_timeline: dict):
    ws = wb.create_sheet("뉴스·공시_시계열")
    ws.sheet_properties.tabColor = "7030A0"
    ws.sheet_view.showGridLines = False

    # 뉴스 + 공시 통합
    items: list[dict] = []
    for n in news:
        items.append({"date": n["date"], "type": "뉴스", "source": n["source"],
                      "title": n["title"], "desc": n.get("desc",""), "link": n.get("link","")})
    for d in disclosures:
        rcept = d.get("rcept_no","")
        items.append({"date": d["date"], "type": "공시", "source": d.get("filer",""),
                      "title": d["type"], "desc": "",
                      "link": f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept}" if rcept else ""})

    if not items:
        ws["A1"].value = "뉴스·공시 데이터 없음 — api_keys.txt에 NAVER / DART_API_KEY 확인"
        ws["A1"].font  = Font(name="맑은 고딕", size=10, color="888888", italic=True)
        return

    items.sort(key=lambda x: x["date"], reverse=True)

    HDR_COLS = [("날짜",12),("유형",8),("출처",16),("제목",50),
                ("내용",55),("AI 시사점",42),("링크",8)]
    for col, (h, w) in enumerate(HDR_COLS, 1):
        _hdr(ws, 1, col, h, "mid")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    NEWS_FILL  = PatternFill("solid", fgColor="E3F2FD")   # 연파랑
    DART_FILL  = PatternFill("solid", fgColor="FFF3E0")   # 연주황
    POS_FILL   = PatternFill("solid", fgColor="E8F5E9")
    NEG_FILL   = PatternFill("solid", fgColor="FFEBEE")
    NEU_FILL   = PatternFill("solid", fgColor="F5F5F5")

    src_colors = {"매일경제":"BBDEFB","한국경제":"C8E6C9",
                  "헤럴드경제":"FFF9C4","서울경제":"E1BEE7","이데일리":"FFCCBC"}

    for i, it in enumerate(items, 2):
        is_news = it["type"] == "뉴스"
        base_f  = NEWS_FILL if is_news else DART_FILL
        src_hex = src_colors.get(it["source"], "") if is_news else ""
        src_f   = PatternFill("solid", fgColor=src_hex) if src_hex else base_f

        ai_entry = ai_timeline.get(i - 2, {})
        sentiment = ai_entry.get("sentiment", "")
        implication = ai_entry.get("implication", "")
        ai_fill = (POS_FILL if sentiment == "긍정" else
                   NEG_FILL if sentiment == "부정" else NEU_FILL) if sentiment else base_f

        type_font = Font(name="맑은 고딕", size=9, bold=True,
                         color="1565C0" if is_news else "E65100")

        _cell(ws, i, 1, it["date"],    fill=base_f, align="center")
        c2 = _cell(ws, i, 2, it["type"], fill=base_f, align="center")
        c2.font = type_font
        _cell(ws, i, 3, it["source"],  fill=src_f,  align="center")
        _cell(ws, i, 4, it["title"],   fill=base_f)
        c5 = _cell(ws, i, 5, it["desc"], fill=base_f)
        c5.alignment = S["wrap"]
        c6 = _cell(ws, i, 6, implication, fill=ai_fill)
        c6.alignment = S["wrap"]
        lc = _cell(ws, i, 7, "보기" if it["link"] else "", fill=base_f, align="center")
        if it["link"]:
            lc.hyperlink = it["link"]
            lc.font = S["f_link"]
        ws.row_dimensions[i].height = 32 if it["desc"] or implication else 18


def write_news_sheet(wb, news):
    ws = wb.create_sheet("뉴스")
    ws.sheet_properties.tabColor = "833C11"
    ws.sheet_view.showGridLines = False

    if not news:
        ws["A1"].value = "NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 설정 시 뉴스 검색 활성화"
        ws["A1"].font  = Font(name="맑은 고딕", size=10, color="888888", italic=True)
        ws["A2"].value = "  발급: https://developers.naver.com/apps"
        ws["A2"].font  = Font(name="맑은 고딕", size=9, color="0563C1")
        return

    headers = [("날짜",12),("신문사",16),("제목",55),("요약",70),("링크",8)]
    for col, (h, w) in enumerate(headers, 1):
        _hdr(ws, 1, col, h, "mid"); ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    src_colors = {"매일경제": "E3F2FD", "한국경제": "E8F5E9",
                  "헤럴드경제": "FFF8E1", "서울경제": "F3E5F5", "이데일리": "FCE4EC"}

    for i, n in enumerate(news, 2):
        sf = PatternFill("solid", fgColor=src_colors.get(n["source"], "FFFFFF"))
        rf = S["alt_row"] if i % 2 == 0 else S["white"]
        _cell(ws, i, 1, n["date"],   align="center")
        _cell(ws, i, 2, n["source"], fill=sf, align="center")
        _cell(ws, i, 3, n["title"],  fill=rf)
        c = _cell(ws, i, 4, n["desc"], fill=rf); c.alignment = S["wrap"]
        lc = _cell(ws, i, 5, "링크" if n["link"] else "", fill=rf, align="center")
        if n["link"]:
            lc.hyperlink = n["link"]; lc.font = S["f_link"]
        ws.row_dimensions[i].height = 32


def write_dart_sheet(wb, disclosures):
    ws = wb.create_sheet("공시")
    ws.sheet_properties.tabColor = "BE4B48"
    ws.sheet_view.showGridLines = False

    if not disclosures:
        ws["A1"].value = "DART_API_KEY 설정 시 공시 조회 활성화 (opendart.fss.or.kr 무료 발급)"
        ws["A1"].font  = Font(name="맑은 고딕", size=10, color="888888", italic=True)
        return

    headers = [("접수일",12),("공시 유형",50),("제출인",20),("DART",8)]
    for col, (h, w) in enumerate(headers, 1):
        _hdr(ws, 1, col, h, "mid"); ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    for i, d in enumerate(disclosures, 2):
        rf = S["alt_row"] if i % 2 == 0 else S["white"]
        _cell(ws, i, 1, d["date"],  align="center")
        _cell(ws, i, 2, d["type"],  fill=rf)
        _cell(ws, i, 3, d["filer"], fill=rf)
        rc = d.get("rcept_no", "")
        lc = _cell(ws, i, 4, "보기" if rc else "", fill=rf, align="center")
        if rc:
            lc.hyperlink = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rc}"
            lc.font = S["f_link"]
        ws.row_dimensions[i].height = 18


# ── 단일 종목 전체 처리 ───────────────────────────────────────────────────────

def _pdf_key(pdf) -> str:
    """PDF 파일명에서 'YYYY-MM-DD_증권사' 형태의 analyses 키 추출."""
    parts = Path(pdf).stem.split("_", 2)
    date  = (f"{parts[0][:4]}-{parts[0][4:6]}-{parts[0][6:8]}"
             if len(parts) > 0 and len(parts[0]) == 8 else "")
    firm  = parts[1] if len(parts) > 1 else ""
    return f"{date}_{firm}"


def run_for_stock(stock_q: str, min_date: str, quick: bool = False) -> bool:
    mode_label = "빠른 다운로드" if quick else "AI 심층 분석"
    bar = "=" * 55
    print(f"\n{bar}")
    print(f"  [{mode_label}] {stock_q}")
    print(f"{bar}")

    # 1) 리포트 검색 + PDF 다운로드
    total_steps = 3 if quick else 6
    print(f"\n[1/{total_steps}] 리포트 검색 및 PDF 다운로드...")
    reports = scan_all_excel_files(stock_q, min_date)
    if not reports:
        print(f"  오류: '{stock_q}' 관련 리포트를 찾을 수 없습니다.")
        return False

    name_cnt   = Counter(r["name"] for r in reports if r["name"])
    stock_name = name_cnt.most_common(1)[0][0]
    stock_code = reports[0]["code"]
    clean_code = re.sub(r"^A", "", stock_code)
    print(f"  종목: {stock_name} ({clean_code}) | 총 {len(reports)}건")

    download_pdfs(reports, stock_name)

    # 2) 현재가
    print(f"\n[2/{total_steps}] 현재가 조회...")
    price, change_rate = get_current_price(clean_code)
    print(f"  {stock_name}: {price:,}원 ({change_rate}%)" if price else "  조회 실패")

    # ── 빠른 모드: 목록 Excel만 생성 후 종료 ──────────────────────────────────
    if quick:
        print(f"\n[3/{total_steps}] 리포트 목록 파일 생성...")
        today_s = datetime.now(KST).strftime("%Y%m%d")
        output  = BASE_DIR / f"{sanitize(stock_name)}_리포트목록_{today_s}.xlsx"
        wb = openpyxl.Workbook()
        write_summary_sheet(wb, stock_name, clean_code, reports, price, change_rate)
        write_report_analysis_sheet(wb, reports, {})
        wb.save(str(output))
        print(f"\n완료 → {output.name}")
        return True

    # ── 심층 분석 모드 ─────────────────────────────────────────────────────────

    # 이전 분석 파일에서 기존 데이터 로드 (증분 분석)
    prev_analyses, old_news, old_dart = load_previous_analysis(stock_name)

    # 3) PDF AI 분석 (새 PDF만 분석, 기존 결과 재사용)
    print(f"\n[3/{total_steps}] PDF AI 분석...")
    has_ai = bool(os.environ.get("ANTHROPIC_API_KEY"))
    folder = BASE_DIR / sanitize(stock_name)
    analyses: dict = dict(prev_analyses)

    new_pdf_count = 0
    if folder.exists():
        pdfs    = sorted(folder.glob("*.pdf"), reverse=True)
        skipped = sum(1 for p in pdfs if _pdf_key(p) in analyses)
        print(f"  PDF {len(pdfs)}개 | 재사용 {skipped}건 | "
              f"AI: {'활성' if has_ai else '비활성 (ANTHROPIC_API_KEY 미설정)'}")
        for pdf in pdfs:
            key = _pdf_key(pdf)
            if key in analyses:
                continue
            parts    = pdf.stem.split("_", 2)
            pdf_date = f"{parts[0][:4]}-{parts[0][4:6]}-{parts[0][6:8]}" if len(parts[0]) == 8 else ""
            pdf_firm = parts[1] if len(parts) > 1 else ""
            if has_ai:
                text   = extract_pdf_text(pdf)
                result = analyze_with_ai(text, pdf_firm)
                if result:
                    analyses[key] = result
                    new_pdf_count += 1
                    print(f"    ✓ {pdf_date} {pdf_firm}: {result.get('summary','')[:45]}")
    print(f"  신규 분석: {new_pdf_count}건")

    # 4) 뉴스 (이전 파일 이후 신규 기사만 수집 후 병합)
    since_news = max((n["date"] for n in old_news if n.get("date")), default=None)
    print(f"\n[4/{total_steps}] 뉴스 검색 "
          f"({'전체' if not since_news else since_news + ' 이후 신규'})...")
    new_news = fetch_naver_news(stock_name, since_date=since_news)
    # 신규 기사 키워드+AI 필터링 (기존 기사는 이미 필터링됨)
    new_news = filter_rank_news_ai(new_news, stock_name)
    seen_news:  set = set()
    dedup_news: list = []
    for n in (new_news + old_news):
        k = (n["date"], n["title"][:60])
        if k not in seen_news:
            seen_news.add(k)
            dedup_news.append(n)
    dedup_news.sort(key=lambda x: x["date"], reverse=True)
    print(f"  신규 {len(new_news)}건 + 이전 {len(old_news)}건 = 합계 {len(dedup_news)}건")

    # 5) DART 공시 (이전 파일 이후 신규 공시만 수집 후 병합)
    since_dart = max((d["date"] for d in old_dart if d.get("date")), default=None)
    print(f"\n[5/{total_steps}] DART 공시 조회 "
          f"({'전체' if not since_dart else since_dart + ' 이후 신규'})...")
    new_dart = fetch_dart_disclosures(clean_code, since_date=since_dart)
    seen_dart:  set = set()
    dedup_dart: list = []
    for d in (new_dart + old_dart):
        k = d.get("rcept_no") or f"{d['date']}_{d['type']}"
        if k not in seen_dart:
            seen_dart.add(k)
            dedup_dart.append(d)
    dedup_dart.sort(key=lambda x: x["date"], reverse=True)
    print(f"  신규 {len(new_dart)}건 + 이전 {len(old_dart)}건 = 합계 {len(dedup_dart)}건")

    # 6) 뉴스+공시 AI 시계열 분석 (배치 1회 호출)
    print(f"\n[6/6] 뉴스·공시 시계열 AI 분석...")
    timeline_items: list[dict] = []
    for n in dedup_news:
        timeline_items.append({"date": n["date"], "type": "뉴스",
                                "source": n["source"], "title": n["title"],
                                "desc": n.get("desc", ""), "link": n.get("link", "")})
    for d in dedup_dart:
        timeline_items.append({"date": d["date"], "type": "공시",
                                "source": d.get("filer", ""), "title": d["type"],
                                "desc": "", "link": ""})
    timeline_items.sort(key=lambda x: x["date"], reverse=True)

    ai_timeline: dict = {}
    if timeline_items and os.environ.get("ANTHROPIC_API_KEY"):
        ai_timeline = analyze_news_dart_timeline(timeline_items, stock_name)
        print(f"  시사점 분석: {len(ai_timeline)}건")
    else:
        print("  건너뜀 (항목 없음 또는 AI 키 미설정)")

    # Excel 생성
    today_s = datetime.now(KST).strftime("%Y%m%d")
    output  = BASE_DIR / f"{sanitize(stock_name)}_분석_{today_s}.xlsx"

    wb = openpyxl.Workbook()
    write_summary_sheet(wb, stock_name, clean_code, reports, price, change_rate)
    write_report_analysis_sheet(wb, reports, analyses)
    write_news_dart_timeline_sheet(wb, dedup_news, dedup_dart, stock_name, ai_timeline)
    write_news_sheet(wb, dedup_news)
    write_dart_sheet(wb, dedup_dart)
    wb.save(str(output))

    print(f"\n완료 → {output.name}")
    print(f"  (5개 시트: 요약 / 리포트분석 / 뉴스·공시_시계열 / 뉴스 / 공시)")
    return True


# ── 현재가 새로고침 ───────────────────────────────────────────────────────────

def refresh_price(stock_q: str) -> bool:
    """최근 분석 Excel 파일의 현재가·등락률을 실시간으로 업데이트."""
    # 종목명 해석: Daily_index에서 코드 및 이름 확인
    reports = scan_all_excel_files(stock_q, "2025-01-01")
    if not reports:
        print(f"  '{stock_q}' 관련 리포트를 찾을 수 없습니다.")
        return False

    name_cnt   = Counter(r["name"] for r in reports if r["name"])
    stock_name = name_cnt.most_common(1)[0][0]
    clean_code = re.sub(r"^A", "", reports[0]["code"])

    # 가장 최근 분석 파일 찾기
    pattern = str(BASE_DIR / f"{sanitize(stock_name)}_분석_*.xlsx")
    files   = sorted(glob.glob(pattern), reverse=True)
    if not files:
        print(f"  분석 파일 없음: {sanitize(stock_name)}_분석_*.xlsx")
        print("  먼저 종목별분석_dailyindex관심종목.bat 으로 분석을 실행하세요.")
        return False

    target = Path(files[0])
    print(f"  파일: {target.name}")

    # 현재가 조회
    price, change_rate = get_current_price(clean_code)
    if not price:
        print("  현재가 조회 실패")
        return False

    print(f"  {stock_name} ({clean_code}): {price:,}원  ({change_rate}%)")

    # Excel 업데이트
    try:
        wb = openpyxl.load_workbook(str(target))
        ws = wb["요약"]

        # B5 = 현재가, B6 = 등락률 (write_summary_sheet 구조 기준)
        ws["B5"].value = f"{price:,}원"
        ws["B6"].value = f"{change_rate}%"

        # A2 타임스탬프 업데이트
        now_str = datetime.now(KST).strftime("%Y-%m-%d %H:%M")
        old_a2  = str(ws["A2"].value or "")
        ws["A2"].value = re.sub(r"분석일: [\d\- :]+KST", f"분석일: {now_str} KST", old_a2) \
                         if "분석일:" in old_a2 else old_a2

        wb.save(str(target))
        print(f"  업데이트 완료 → {target.name}")
        return True
    except PermissionError:
        print(f"  오류: {target.name} 파일이 열려 있습니다. 닫고 다시 실행하세요.")
        return False
    except Exception as e:
        print(f"  Excel 업데이트 오류: {e}")
        return False


# ── 진입점 ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="종목 분석 보고서 생성")
    parser.add_argument("stock",          nargs="?", help="종목코드 또는 종목명 (없으면 관심종목 시트 사용)")
    parser.add_argument("--from-date",    default=MIN_DATE, dest="from_date")
    parser.add_argument("--quick",        action="store_true",
                        help="빠른 모드: PDF 다운로드 + 리포트 목록만 생성 (AI 분석 없음)")
    parser.add_argument("--refresh-price", action="store_true", dest="refresh_price",
                        help="현재가만 새로고침 (기존 분석 파일 업데이트)")
    args = parser.parse_args()

    if args.refresh_price:
        stocks = [args.stock] if args.stock else get_watchlist()
        if not stocks:
            print("종목을 지정하거나 관심종목 시트에 입력하세요.")
            return
        print(f"=== 현재가 새로고침 ({len(stocks)}개) ===")
        for s in stocks:
            refresh_price(s)
        return

    if args.stock:
        stocks = [args.stock]
    else:
        stocks = get_watchlist()
        if not stocks:
            if not MAIN_EXCEL.exists():
                print(f"오류: {MAIN_EXCEL.name} 파일 없음")
            else:
                print("관심종목 시트에 종목이 없습니다.")
                print(f"Daily_index.xlsx → 관심종목 시트 → A3 셀부터 종목코드/종목명 입력")
            return

    mode = "빠른 다운로드" if args.quick else "AI 심층 분석"
    print(f"=== 종목 {mode} 시작 ({len(stocks)}개) ===")
    for s in stocks:
        run_for_stock(s, args.from_date, quick=args.quick)


if __name__ == "__main__":
    main()
