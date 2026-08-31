# -*- coding: utf-8 -*-
"""
지각 해시(pHash)로 리사이즈된 중복 사진 탐지
- pHash 64비트 + numpy 벡터 연산으로 빠른 Hamming 거리 비교
- 해시를 JSON에 캐시해서 재실행 시 즉시 처리
"""
import os, sys, json, time
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

SCAN_DIR   = r'D:\OneDrive\사진'
OUT_EXCEL  = r'D:\OneDrive\1. NA\cluade\visual_duplicates.xlsx'
CACHE_FILE = r'D:\OneDrive\1. NA\cluade\phash_cache.json'

PHOTO_EXTS = {'.jpg','.jpeg','.png','.gif','.bmp',
              '.heic','.heif','.tiff','.tif','.webp'}

# Hamming 거리 임계값 (64비트 기준)
# 0~3: 거의 동일(리사이즈)  4~8: 유사  9+: 다른 사진
HASH_THRESHOLD = 8


# ── 해시 계산 (서브프로세스용 최상위 함수) ──────────────────────────
def _compute_hash(path):
    try:
        from PIL import Image
        import imagehash
        img = Image.open(path).convert('RGB')
        h = str(imagehash.phash(img, hash_size=8))  # 64비트
        w, ht = img.size
        return path, h, w, ht
    except Exception:
        return path, None, 0, 0


def collect_files():
    files = []
    for root, dirs, names in os.walk(SCAN_DIR):
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        for n in names:
            if Path(n).suffix.lower() in PHOTO_EXTS:
                files.append(os.path.join(root, n))
    return files


def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False)


def compute_hashes(files, cache):
    need = [f for f in files if f not in cache]
    total = len(need)

    if total == 0:
        print(f"  캐시 적중: {len(files):,}개 — 해시 계산 생략")
        return cache

    print(f"  계산 필요: {total:,}개  |  캐시 적중: {len(files)-total:,}개")
    workers = max(1, multiprocessing.cpu_count() - 1)
    print(f"  병렬 처리: {workers}코어")

    done = 0
    start = time.time()
    batch_size = 2000

    for b_start in range(0, total, batch_size):
        batch = need[b_start:b_start + batch_size]
        with ProcessPoolExecutor(max_workers=workers) as ex:
            for path, h, w, ht in ex.map(_compute_hash, batch, chunksize=50):
                if h:
                    cache[path] = {'hash': h, 'w': w, 'h': ht}
                done += 1
                if done % 1000 == 0:
                    elapsed = time.time() - start
                    eta = elapsed / done * (total - done)
                    print(f"  {done:,}/{total:,}  ({done*100//total}%)"
                          f"  남은시간 약 {int(eta//60)}분 {int(eta%60)}초")
        save_cache(cache)
        print(f"  중간 저장 ({b_start + len(batch):,}/{total:,})")

    return cache


def group_duplicates(files, cache):
    """numpy 벡터 연산으로 전체 쌍 Hamming 거리 비교"""
    import numpy as np

    valid_files = [f for f in files if f in cache]
    hashes = [cache[f]['hash'] for f in valid_files]
    widths  = [cache[f]['w']    for f in valid_files]
    heights = [cache[f]['h']    for f in valid_files]
    N = len(valid_files)
    print(f"  유효 해시: {N:,}개")

    # 64비트 정수 배열로 변환
    hash_ints = np.array([int(h, 16) for h in hashes], dtype=np.uint64)

    # Hamming 거리 함수 (비트 카운트)
    def hamming_batch(a_int, b_arr):
        """a_int(스칼라) vs b_arr(배열) — bit count of XOR"""
        xor = np.bitwise_xor(a_int, b_arr)
        # 비트 카운트: 비트 트릭
        c = xor - ((xor >> np.uint64(1)) & np.uint64(0x5555555555555555))
        c = (c & np.uint64(0x3333333333333333)) + ((c >> np.uint64(2)) & np.uint64(0x3333333333333333))
        c = (c + (c >> np.uint64(4))) & np.uint64(0x0F0F0F0F0F0F0F0F)
        return ((c * np.uint64(0x0101010101010101)) >> np.uint64(56)).astype(np.int32)

    assigned = np.zeros(N, dtype=bool)
    groups = []
    thr = np.int32(HASH_THRESHOLD)

    for i in range(N):
        if assigned[i]:
            continue
        dists = hamming_batch(hash_ints[i], hash_ints[i+1:])
        matches = np.where(dists <= thr)[0] + (i + 1)
        if len(matches) > 0:
            group_idx = [i] + matches.tolist()
            for idx in group_idx:
                assigned[idx] = True
            groups.append(group_idx)

        if i % 2000 == 0 and i > 0:
            print(f"  비교 진행: {i:,}/{N:,}  (그룹 {len(groups)}개 발견)")

    print(f"  유사 이미지 그룹: {len(groups)}개")
    return [
        [(valid_files[idx], hashes[idx], widths[idx], heights[idx])
         for idx in g]
        for g in groups
    ]


def build_excel(groups):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    wb = Workbook()

    # ── 요약 시트 ──
    ws_sum = wb.active
    ws_sum.title = '요약'
    summary = [
        ('시각적 중복 사진 분석 결과', ''),
        ('', ''),
        ('분석 일시',  datetime.datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('스캔 경로',  SCAN_DIR),
        ('유사 그룹 수', len(groups)),
        ('해당 파일 총 수', sum(len(g) for g in groups)),
        ('판정 기준 (Hamming 거리)', f'<= {HASH_THRESHOLD}  (0=완전동일, 높을수록 더 다름)'),
        ('', ''),
        ('* 해상도가 가장 큰 파일 = 원본 추정', ''),
        ('* 그 외 = 리사이즈 추정 (빨간색)', ''),
    ]
    ws_sum.column_dimensions['A'].width = 32
    ws_sum.column_dimensions['B'].width = 50
    for r, (a, b) in enumerate(summary, 1):
        ws_sum.cell(r, 1, a).font = Font(bold=(r == 1), size=(14 if r == 1 else 10),
                                          color=('1F4E79' if r == 1 else '000000'))
        ws_sum.cell(r, 2, b)

    # ── 목록 시트 ──
    ws = wb.create_sheet('유사이미지목록')
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill  = PatternFill('solid', fgColor='1F4E79')
    hfont  = Font(bold=True, color='FFFFFF', size=10)
    gfills = [PatternFill('solid', fgColor='D6E4F0'),
              PatternFill('solid', fgColor='FFFFFF')]

    headers = ['그룹', '비고', '파일명', '해상도', '크기(KB)', 'pHash', '폴더 경로', '전체 경로']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    groups_sorted = sorted(groups, key=lambda g: -len(g))
    row = 2
    for g_idx, group in enumerate(groups_sorted, 1):
        fill = gfills[g_idx % 2]
        # 해상도 큰 순으로 정렬
        group_s = sorted(group, key=lambda x: -(x[2] * x[3]))
        for rank, (fpath, fhash, w, h) in enumerate(group_s):
            try:
                sz = round(os.path.getsize(fpath) / 1024, 1)
            except OSError:
                sz = 0
            note = '원본(최고해상도)' if rank == 0 else '리사이즈 추정'
            vals = [g_idx, note, os.path.basename(fpath),
                    f'{w} x {h}', sz, fhash,
                    os.path.dirname(fpath), fpath]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row, c, v)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical='center',
                                           wrap_text=(c in (7, 8)))
                if c == 2 and rank > 0:
                    cell.font = Font(color='C00000')
            row += 1

    widths = [7, 16, 35, 14, 10, 18, 50, 65]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    wb.save(OUT_EXCEL)
    print(f'Excel 저장: {OUT_EXCEL}')


def main():
    print('=== 시각적 중복 사진 탐지 (pHash) ===\n')

    print('1단계: 파일 목록 수집...')
    files = collect_files()
    print(f'  {len(files):,}개\n')

    print('2단계: 지각 해시 계산 (최초 실행 시 30~60분 소요)...')
    cache = load_cache()
    cache = compute_hashes(files, cache)
    print()

    print('3단계: 유사 이미지 그룹화 (numpy 벡터 연산)...')
    groups = group_duplicates(files, cache)
    print()

    print('4단계: Excel 생성...')
    build_excel(groups)

    print(f'\n완료! 유사 그룹: {len(groups):,}개  |  해당 파일: {sum(len(g) for g in groups):,}개')


if __name__ == '__main__':
    main()
