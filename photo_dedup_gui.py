# -*- coding: utf-8 -*-
"""
중복 사진 GUI 뷰어 & 삭제 도구
visual_duplicates.xlsx 또는 duplicate_photos.xlsx 에서 그룹을 읽어
썸네일 미리보기와 파일 정보를 나란히 보여주고 선택 삭제합니다.

실행: python photo_dedup_gui.py
"""
import os
import sys
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from collections import defaultdict

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

try:
    from PIL import Image, ImageTk, ImageDraw
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import openpyxl
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

LOG_PATH      = r'D:\OneDrive\1. NA\cluade\deleted_log.txt'
PROGRESS_PATH = r'D:\OneDrive\1. NA\cluade\dedup_progress.json'
THUMB_W   = 200
THUMB_H   = 165
CARD_W    = 230

DEFAULT_EXCEL_PATHS = [
    r'D:\OneDrive\1. NA\cluade\visual_duplicates.xlsx',
    r'D:\OneDrive\1. NA\cluade\duplicate_photos.xlsx',
]


# ── Excel 파싱 ────────────────────────────────────────────────────────────
def load_groups_from_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheet_priority = ['유사이미지목록', '완전중복(해시동일)']
    ws = None
    for name in sheet_priority:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if '요약' not in name:
                ws = wb[name]
                break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h else '' for h in rows[0]]

    def find_col(*keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return None

    col_group  = find_col('그룹')
    col_path   = find_col('전체 경로', '전체경로')
    col_name   = find_col('파일명')
    col_size   = find_col('크기')
    col_res    = find_col('해상도')
    col_note   = find_col('비고')
    col_folder = find_col('폴더 경로', '폴더경로')

    if col_path is None:
        raise ValueError("'전체 경로' 컬럼을 찾을 수 없습니다.")

    groups_dict = defaultdict(list)
    for row in rows[1:]:
        if not any(row):
            continue
        group_id = row[col_group] if col_group is not None else 1
        path     = str(row[col_path]).strip() if row[col_path] else ''
        if not path or path == 'None':
            continue
        name   = str(row[col_name]).strip()   if col_name   is not None and row[col_name]   else os.path.basename(path)
        size   = row[col_size]                 if col_size   is not None else 0
        res    = str(row[col_res]).strip()     if col_res    is not None and row[col_res]    else ''
        note   = str(row[col_note]).strip()    if col_note   is not None and row[col_note]   else ''
        folder = str(row[col_folder]).strip()  if col_folder is not None and row[col_folder] else os.path.dirname(path)
        groups_dict[group_id].append({
            'path': path, 'name': name, 'size': size,
            'res': res, 'note': note, 'folder': folder,
        })

    return [v for v in groups_dict.values() if len(v) >= 2]


# ── 썸네일 ────────────────────────────────────────────────────────────────
def _placeholder(w, h, text='미리보기 없음'):
    img = Image.new('RGB', (w, h), (220, 220, 225))
    draw = ImageDraw.Draw(img)
    draw.rectangle([2, 2, w - 3, h - 3], outline=(180, 180, 190), width=1)
    tw, th = w // 2, h // 2
    draw.text((tw, th - 8), text[:12], fill=(150, 150, 155))
    return img

def load_thumbnail(path):
    if not PIL_OK:
        return None
    try:
        if path.lower().endswith(('.heic', '.heif')):
            try:
                import pillow_heif
                pillow_heif.register_heif_opener()
            except ImportError:
                return _placeholder(THUMB_W, THUMB_H, 'HEIC\n미지원')
        img = Image.open(path).convert('RGB')
        img.thumbnail((THUMB_W, THUMB_H), Image.LANCZOS)
        canvas = Image.new('RGB', (THUMB_W, THUMB_H), (240, 240, 243))
        x = (THUMB_W - img.width) // 2
        y = (THUMB_H - img.height) // 2
        canvas.paste(img, (x, y))
        return canvas
    except Exception:
        return _placeholder(THUMB_W, THUMB_H, os.path.splitext(path)[1].upper())


# ── GUI ───────────────────────────────────────────────────────────────────
class DuplicatePhotoViewer:
    CLR_BG       = '#F0F2F5'
    CLR_HEADER   = '#1B3A5C'
    CLR_NAV      = '#E8EDF2'
    CLR_ORIG     = '#E8F4FB'
    CLR_DUP      = '#FFF5F5'
    CLR_BORDER_O = '#2E86C1'
    CLR_BORDER_D = '#E74C3C'
    CLR_BTN_NAV  = '#2E86C1'
    CLR_BTN_DEL  = '#C0392B'
    CLR_BTN_SKIP = '#7F8C8D'

    def __init__(self, root, groups, source_file):
        self.root        = root
        self.groups      = groups
        self.source_file = source_file
        self.idx         = 0
        self.check_vars  = []
        self.thumb_imgs  = []
        self.deleted_n   = 0
        self.log_lines   = []
        self.processed   = set()

        root.title(f'중복 사진 뷰어  —  {os.path.basename(source_file)}')
        root.configure(bg=self.CLR_BG)
        root.geometry('1280x820')
        root.minsize(900, 640)

        self._build_ui()
        start_idx = self._load_progress()
        self._show(start_idx)
        root.protocol('WM_DELETE_WINDOW', self._on_close)

    # ── UI 구성 ────────────────────────────────────────────────────────────
    def _build_ui(self):
        # 헤더
        hdr = tk.Frame(self.root, bg=self.CLR_HEADER, height=48)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Label(hdr, text='중복 사진 뷰어', bg=self.CLR_HEADER, fg='white',
                 font=('맑은 고딕', 14, 'bold')).pack(side=tk.LEFT, padx=18, pady=8)
        self._lbl_prog = tk.Label(hdr, text='', bg=self.CLR_HEADER, fg='#A9CCE3',
                                   font=('맑은 고딕', 10))
        self._lbl_prog.pack(side=tk.LEFT, padx=6)
        self._lbl_stat = tk.Label(hdr, text='', bg=self.CLR_HEADER, fg='#A9CCE3',
                                   font=('맑은 고딕', 9))
        self._lbl_stat.pack(side=tk.RIGHT, padx=18)

        # 네비게이션 바
        nav = tk.Frame(self.root, bg=self.CLR_NAV, height=46)
        nav.pack(fill=tk.X)
        nav.pack_propagate(False)

        self._btn_prev = self._nav_btn(nav, '◀  이전', self._prev)
        self._btn_prev.pack(side=tk.LEFT, padx=(10, 4), pady=7)

        self._btn_next = self._nav_btn(nav, '다음  ▶', self._next)
        self._btn_next.pack(side=tk.LEFT, padx=4, pady=7)

        self._lbl_group = tk.Label(nav, text='', bg=self.CLR_NAV, fg='#1B3A5C',
                                    font=('맑은 고딕', 11, 'bold'))
        self._lbl_group.pack(side=tk.LEFT, padx=14)

        tk.Button(nav, text='건너뛰기', command=self._skip,
                  bg=self.CLR_BTN_SKIP, fg='white',
                  font=('맑은 고딕', 10), relief=tk.FLAT, padx=10,
                  cursor='hand2').pack(side=tk.RIGHT, padx=(4, 10), pady=7)

        tk.Button(nav, text='선택한 파일 삭제', command=self._delete,
                  bg=self.CLR_BTN_DEL, fg='white',
                  font=('맑은 고딕', 11, 'bold'), relief=tk.FLAT, padx=14,
                  cursor='hand2').pack(side=tk.RIGHT, padx=4, pady=7)

        # 원본 제외 전체 선택
        tk.Button(nav, text='원본 제외 전체 선택', command=self._select_dups,
                  bg='#E67E22', fg='white',
                  font=('맑은 고딕', 9), relief=tk.FLAT, padx=8,
                  cursor='hand2').pack(side=tk.RIGHT, padx=4, pady=7)

        tk.Button(nav, text='전체 해제', command=self._deselect_all,
                  bg='#95A5A6', fg='white',
                  font=('맑은 고딕', 9), relief=tk.FLAT, padx=8,
                  cursor='hand2').pack(side=tk.RIGHT, padx=4, pady=7)

        # 스크롤 가능한 콘텐츠 영역
        wrap = tk.Frame(self.root, bg=self.CLR_BG)
        wrap.pack(fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(wrap, orient=tk.VERTICAL)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb = ttk.Scrollbar(wrap, orient=tk.HORIZONTAL)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        self._canvas = tk.Canvas(wrap, bg=self.CLR_BG,
                                  yscrollcommand=vsb.set,
                                  xscrollcommand=hsb.set,
                                  highlightthickness=0)
        self._canvas.pack(fill=tk.BOTH, expand=True)
        vsb.config(command=self._canvas.yview)
        hsb.config(command=self._canvas.xview)

        self._inner = tk.Frame(self._canvas, bg=self.CLR_BG)
        self._win_id = self._canvas.create_window((0, 0), window=self._inner, anchor='nw')

        self._inner.bind('<Configure>',
                         lambda e: self._canvas.configure(
                             scrollregion=self._canvas.bbox('all')))
        self._canvas.bind_all('<MouseWheel>',
                              lambda e: self._canvas.yview_scroll(
                                  int(-1 * e.delta / 120), 'units'))

        # 상태바
        sb = tk.Frame(self.root, bg='#2C3E50', height=26)
        sb.pack(fill=tk.X, side=tk.BOTTOM)
        sb.pack_propagate(False)
        self._lbl_hint = tk.Label(sb,
            text='← → 방향키로 이동  |  Delete 키로 삭제  |  Space: 원본 제외 전체 선택',
            bg='#2C3E50', fg='#95A5A6', font=('맑은 고딕', 8))
        self._lbl_hint.pack(side=tk.LEFT, padx=10)

        # 키보드
        self.root.bind('<Left>',   lambda e: self._prev())
        self.root.bind('<Right>',  lambda e: self._next())
        self.root.bind('<Delete>', lambda e: self._delete())
        self.root.bind('<space>',  lambda e: self._select_dups())

    def _nav_btn(self, parent, text, cmd):
        return tk.Button(parent, text=text, command=cmd,
                         bg=self.CLR_BTN_NAV, fg='white',
                         font=('맑은 고딕', 10, 'bold'),
                         relief=tk.FLAT, padx=12, cursor='hand2')

    # ── 그룹 표시 ──────────────────────────────────────────────────────────
    def _show(self, idx):
        if not self.groups or not (0 <= idx < len(self.groups)):
            return
        self.idx = idx
        group = self.groups[idx]

        for w in self._inner.winfo_children():
            w.destroy()
        self.check_vars = []
        self.thumb_imgs = []
        self._canvas.yview_moveto(0)

        done = len(self.processed)
        self._lbl_prog.config(text=f'그룹  {idx + 1} / {len(self.groups)}')
        self._lbl_stat.config(
            text=f'처리: {done}그룹  |  삭제: {self.deleted_n}개')
        self._lbl_group.config(
            text=f'그룹 {idx + 1}  —  {len(group)}개 파일' +
                 ('  ✓ 처리됨' if idx in self.processed else ''))
        self._btn_prev.config(state=tk.NORMAL if idx > 0            else tk.DISABLED)
        self._btn_next.config(state=tk.NORMAL if idx < len(self.groups)-1 else tk.DISABLED)

        bg = '#F0FFF4' if idx in self.processed else self.CLR_BG
        self._inner.config(bg=bg)
        self._canvas.config(bg=bg)

        row_frame = tk.Frame(self._inner, bg=bg)
        row_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)

        for i, info in enumerate(group):
            self._make_card(row_frame, i, info, bg)

    def _make_card(self, parent, i, info, outer_bg):
        path     = info['path']
        exists   = os.path.exists(path)
        note     = info.get('note', '')
        is_orig  = (i == 0) or ('원본' in note and '리사이즈' not in note)

        border   = self.CLR_BORDER_O if is_orig else self.CLR_BORDER_D
        card_bg  = self.CLR_ORIG if is_orig else (self.CLR_DUP if exists else '#F8F8F8')

        outer = tk.Frame(parent, bg=border)
        outer.pack(side=tk.LEFT, padx=8, pady=4, anchor='n')
        card = tk.Frame(outer, bg=card_bg, padx=10, pady=8, width=CARD_W)
        card.pack(padx=2, pady=2)

        # 비고 배지
        badge_text  = note if note else ('원본' if is_orig else f'파일 {i+1}')
        badge_fg    = '#1B6FA8' if is_orig else '#C0392B'
        badge_bg    = '#D6EAF8' if is_orig else '#FADBD8'
        tk.Label(card, text=badge_text, bg=badge_bg, fg=badge_fg,
                 font=('맑은 고딕', 9, 'bold'), padx=6, pady=2).pack(pady=(0, 6))

        # 썸네일
        if PIL_OK and exists:
            img = load_thumbnail(path)
            if img:
                photo = ImageTk.PhotoImage(img)
                self.thumb_imgs.append(photo)
                lbl = tk.Label(card, image=photo, bg=card_bg,
                               relief=tk.GROOVE, bd=1)
                lbl.pack()
            else:
                tk.Label(card, text='로드 실패', bg=card_bg, fg='#95A5A6',
                         width=24, height=8).pack()
        elif not exists:
            tk.Label(card, text='[파일 없음]', bg='#F8F8F8', fg='#BDC3C7',
                     width=24, height=8, relief=tk.GROOVE, bd=1).pack()
        else:
            tk.Label(card, text='Pillow 미설치\n미리보기 불가',
                     bg=card_bg, fg='#95A5A6',
                     width=24, height=8).pack()

        # 파일명
        fname = info['name']
        disp  = fname if len(fname) <= 26 else fname[:12] + '…' + fname[-12:]
        tk.Label(card, text=disp, bg=card_bg, fg='#1A1A2E',
                 font=('맑은 고딕', 9, 'bold'),
                 wraplength=CARD_W - 20).pack(pady=(6, 2))

        # 해상도
        res = info.get('res', '')
        if res and res != 'None':
            tk.Label(card, text=f'해상도: {res}', bg=card_bg, fg='#2980B9',
                     font=('맑은 고딕', 8)).pack()

        # 크기
        size = info.get('size', 0)
        if size:
            try:
                kb = float(size)
                s  = f'{kb/1024:.1f} MB' if kb >= 1024 else f'{kb:.0f} KB'
            except (TypeError, ValueError):
                s = str(size)
            tk.Label(card, text=f'크기: {s}', bg=card_bg, fg='#27AE60',
                     font=('맑은 고딕', 8)).pack()

        # 경로 (말줄임)
        folder = info.get('folder', os.path.dirname(path))
        parts  = folder.replace('\\', '/').split('/')
        if len(parts) > 4:
            short = '/'.join(parts[:2]) + '/…/' + '/'.join(parts[-2:])
        else:
            short = folder
        path_lbl = tk.Label(card, text=short, bg=card_bg, fg='#7F8C8D',
                             font=('맑은 고딕', 7), wraplength=CARD_W - 20,
                             justify=tk.LEFT, cursor='hand2')
        path_lbl.pack(pady=(3, 0))
        path_lbl.bind('<Button-1>', lambda e, p=folder: self._open_folder(p))
        path_lbl.bind('<Enter>', lambda e, lbl=path_lbl, bg=card_bg: lbl.config(fg='#2980B9'))
        path_lbl.bind('<Leave>', lambda e, lbl=path_lbl: lbl.config(fg='#7F8C8D'))

        # 파일 없음 경고
        if not exists:
            tk.Label(card, text='파일 없음 (이미 삭제됨?)',
                     bg='#FFF3CD', fg='#856404',
                     font=('맑은 고딕', 7, 'bold'), padx=4, pady=1).pack(pady=3)

        # 체크박스
        var = tk.BooleanVar(value=(not is_orig and exists))
        self.check_vars.append((var, info))

        chk_text  = '삭제 선택' if not is_orig else '원본 (보존)'
        chk_fg    = '#C0392B'   if not is_orig else '#1B6FA8'
        chk_state = tk.NORMAL   if exists      else tk.DISABLED
        tk.Checkbutton(card, text=chk_text, variable=var,
                       bg=card_bg, fg=chk_fg,
                       font=('맑은 고딕', 10, 'bold'),
                       activebackground=card_bg,
                       state=chk_state, cursor='hand2').pack(pady=(8, 4))

    # ── 조작 ───────────────────────────────────────────────────────────────
    def _prev(self):
        if self.idx > 0:
            self._show(self.idx - 1)

    def _next(self):
        if self.idx < len(self.groups) - 1:
            self._show(self.idx + 1)

    def _skip(self):
        self._next()

    def _select_dups(self):
        for i, (var, info) in enumerate(self.check_vars):
            is_orig = i == 0 or '원본' in info.get('note', '')
            if not is_orig and os.path.exists(info['path']):
                var.set(True)

    def _deselect_all(self):
        for var, _ in self.check_vars:
            var.set(False)

    def _open_folder(self, path):
        try:
            import subprocess
            subprocess.Popen(f'explorer /select,"{path}"')
        except Exception:
            pass

    def _delete(self):
        targets = [(v, info) for v, info in self.check_vars
                   if v.get() and os.path.exists(info['path'])]
        if not targets:
            messagebox.showinfo('알림', '삭제할 파일을 선택해 주세요.')
            return

        all_sel = len(targets) == len(self.check_vars)
        if all_sel:
            if not messagebox.askyesno('경고',
                    f'그룹 내 모든 파일 {len(targets)}개가 선택되었습니다.\n'
                    f'정말 모두 삭제하시겠습니까?', icon='warning'):
                return

        preview = '\n'.join(f'  • {info["name"]}' for _, info in targets[:8])
        if len(targets) > 8:
            preview += f'\n  ... 외 {len(targets)-8}개'
        if not messagebox.askyesno('삭제 확인',
                f'{len(targets)}개 파일을 영구 삭제합니다.\n\n{preview}\n\n계속하시겠습니까?',
                icon='warning'):
            return

        ok, errors = 0, []
        for _, info in targets:
            try:
                os.remove(info['path'])
                ok += 1
                self.deleted_n += 1
                self.log_lines.append(
                    f"[삭제] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  {info['path']}")
            except OSError as e:
                errors.append(f"{info['name']}: {e}")
                self.log_lines.append(
                    f"[오류] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  {info['path']}  ({e})")

        self._save_log()
        self.processed.add(self.idx)
        self._save_progress()

        msg = f'{ok}개 파일 삭제 완료.'
        if errors:
            msg += f'\n\n오류 {len(errors)}개:\n' + '\n'.join(errors[:5])
        messagebox.showinfo('완료', msg)

        self._show(self.idx)
        if self.idx < len(self.groups) - 1:
            self._next()

    def _save_log(self):
        if not self.log_lines:
            return
        try:
            os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
            with open(LOG_PATH, 'a', encoding='utf-8') as f:
                f.write(f"\n=== {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
                for line in self.log_lines:
                    f.write(line + '\n')
            self.log_lines.clear()
        except Exception as e:
            print(f'로그 저장 실패: {e}')

    def _save_progress(self):
        data = {
            'source_file': self.source_file,
            'current_idx': self.idx,
            'processed':   sorted(self.processed),
            'deleted_n':   self.deleted_n,
            'saved_at':    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        try:
            with open(PROGRESS_PATH, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f'진행 저장 실패: {e}')

    def _load_progress(self):
        if not os.path.exists(PROGRESS_PATH):
            return 0
        try:
            with open(PROGRESS_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if data.get('source_file') != self.source_file:
                return 0
            self.processed = set(data.get('processed', []))
            self.deleted_n = data.get('deleted_n', 0)
            saved_idx      = data.get('current_idx', 0)
            saved_at       = data.get('saved_at', '')
            if self.processed or saved_idx > 0:
                resume = messagebox.askyesno(
                    '이어서 시작',
                    f'이전에 저장된 진행 내역이 있습니다.\n\n'
                    f'  저장 시각: {saved_at}\n'
                    f'  마지막 위치: 그룹 {saved_idx + 1} / {len(self.groups)}\n'
                    f'  처리 완료: {len(self.processed)}그룹  |  삭제: {self.deleted_n}개\n\n'
                    f'이어서 시작하시겠습니까?\n(아니오 선택 시 처음부터 시작)',
                )
                if resume:
                    return min(saved_idx, len(self.groups) - 1)
                else:
                    self.processed = set()
                    self.deleted_n = 0
                    return 0
        except Exception as e:
            print(f'진행 불러오기 실패: {e}')
        return 0

    def _on_close(self):
        self._save_progress()
        self.root.destroy()


# ── 진입점 ────────────────────────────────────────────────────────────────
def main():
    if not XLSX_OK:
        tk.Tk().withdraw()
        messagebox.showerror('오류',
            'openpyxl이 설치되어 있지 않습니다.\n\n'
            '터미널에서 다음을 실행하세요:\n  pip install openpyxl')
        return

    # Excel 파일 자동 감지
    excel_path = next((p for p in DEFAULT_EXCEL_PATHS if os.path.exists(p)), None)
    if not excel_path:
        root_tmp = tk.Tk()
        root_tmp.withdraw()
        excel_path = filedialog.askopenfilename(
            title='중복 사진 Excel 파일 선택',
            filetypes=[('Excel 파일', '*.xlsx'), ('모든 파일', '*.*')],
            initialdir=os.path.dirname(DEFAULT_EXCEL_PATHS[0]))
        root_tmp.destroy()
        if not excel_path:
            return

    root_tmp = tk.Tk()
    root_tmp.withdraw()

    try:
        groups = load_groups_from_excel(excel_path)
    except Exception as e:
        messagebox.showerror('오류', f'Excel 로드 실패:\n{e}')
        root_tmp.destroy()
        return

    if not groups:
        messagebox.showinfo('알림', 'Excel에 중복 그룹이 없습니다.')
        root_tmp.destroy()
        return

    root_tmp.destroy()

    root = tk.Tk()
    DuplicatePhotoViewer(root, groups, excel_path)
    root.mainloop()


if __name__ == '__main__':
    main()
