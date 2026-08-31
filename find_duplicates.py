import os
import hashlib
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

PHOTO_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.heic', '.heif',
              '.tiff', '.tif', '.webp', '.raw', '.cr2', '.nef', '.arw',
              '.mp4', '.mov', '.avi', '.3gp', '.mkv'}

SCAN_ROOT = 'D:/OneDrive'
OUTPUT_PATH = 'D:/OneDrive/1. NA/cluade/duplicate_photos.xlsx'


def md5_hash(filepath, chunk_size=65536):
    h = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()
    except (PermissionError, OSError):
        return None


def scan_files():
    print("1단계: 파일 목록 수집 중...")
    size_map = defaultdict(list)
    total = 0
    for root, dirs, files in os.walk(SCAN_ROOT):
        dirs[:] = [d for d in dirs if not d.startswith('.') and d != 'desktop.ini']
        for fname in files:
            ext = os.path.splitext(fname.lower())[1]
            if ext in PHOTO_EXTS:
                fpath = os.path.join(root, fname)
                try:
                    size = os.path.getsize(fpath)
                    size_map[size].append(fpath)
                    total += 1
                except OSError:
                    pass
    print(f"  → 총 {total:,}개 파일 수집 완료")
    return size_map


def find_hash_duplicates(size_map):
    print("2단계: 같은 크기 파일 해시 비교 중...")
    candidates = {sz: paths for sz, paths in size_map.items() if len(paths) > 1}
    candidate_count = sum(len(v) for v in candidates.values())
    print(f"  → 같은 크기 파일 그룹: {len(candidates)}그룹, {candidate_count:,}개")

    hash_map = defaultdict(list)
    processed = 0
    for size, paths in candidates.items():
        for fpath in paths:
            h = md5_hash(fpath)
            if h:
                hash_map[(size, h)].append(fpath)
            processed += 1
            if processed % 500 == 0:
                print(f"  진행: {processed:,}/{candidate_count:,}...")

    duplicates = {k: v for k, v in hash_map.items() if len(v) > 1}
    print(f"  → 실제 중복 그룹: {len(duplicates)}그룹")
    return duplicates


def find_name_duplicates(size_map):
    """같은 파일명이지만 다른 크기/경로인 파일들도 찾기"""
    name_map = defaultdict(list)
    for size, paths in size_map.items():
        for fpath in paths:
            fname = os.path.basename(fpath).lower()
            name_map[fname].append((fpath, size))

    name_dups = {name: entries for name, entries in name_map.items() if len(entries) > 1}
    return name_dups


def build_excel(hash_dups, name_dups, size_map):
    print("3단계: Excel 파일 생성 중...")
    wb = Workbook()

    # ── Sheet 1: 완전 중복 (해시 동일) ──
    ws1 = wb.active
    ws1.title = "완전중복(해시동일)"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    group_fills = [PatternFill("solid", fgColor="D6E4F0"), PatternFill("solid", fgColor="FFFFFF")]
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers1 = ["그룹", "파일명", "크기 (KB)", "확장자", "폴더 경로", "전체 경로", "비고"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = 2
    total_dup_files = 0
    for g_idx, ((size, _), paths) in enumerate(sorted(hash_dups.items(), key=lambda x: -len(x[1])), 1):
        fill = group_fills[g_idx % 2]
        for i, fpath in enumerate(sorted(paths)):
            fname = os.path.basename(fpath)
            ext = os.path.splitext(fname.lower())[1]
            folder = os.path.dirname(fpath)
            size_kb = round(size / 1024, 1)
            note = "원본(첫번째)" if i == 0 else "중복"
            for col, val in enumerate([g_idx, fname, size_kb, ext, folder, fpath, note], 1):
                cell = ws1.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=(col in (5, 6)))
                if col == 7 and i > 0:
                    cell.font = Font(color="C00000", bold=True)
            row += 1
            total_dup_files += 1

    # Column widths
    col_widths1 = [8, 35, 12, 10, 55, 70, 12]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = "A2"

    # ── Sheet 2: 같은 파일명 (다른 경로/크기) ──
    ws2 = wb.create_sheet("같은파일명(다른경로)")
    for col, h in enumerate(["그룹", "파일명", "크기 (KB)", "확장자", "폴더 경로", "전체 경로"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row2 = 2
    name_only_dups = {}
    for fname_lower, entries in name_dups.items():
        # 해시 중복에 이미 포함되지 않은 그룹만 (경로가 다른 것들)
        paths_set = set(e[0] for e in entries)
        if len(set(os.path.dirname(e[0]) for e in entries)) > 1:
            name_only_dups[fname_lower] = entries

    for g_idx, (fname_lower, entries) in enumerate(sorted(name_only_dups.items()), 1):
        fill = group_fills[g_idx % 2]
        for fpath, size in sorted(entries, key=lambda x: x[0]):
            fname = os.path.basename(fpath)
            ext = os.path.splitext(fname.lower())[1]
            folder = os.path.dirname(fpath)
            size_kb = round(size / 1024, 1)
            for col, val in enumerate([g_idx, fname, size_kb, ext, folder, fpath], 1):
                cell = ws2.cell(row=row2, column=col, value=val)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=(col in (5, 6)))
            row2 += 1

    col_widths2 = [8, 35, 12, 10, 55, 70]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 요약 ──
    ws3 = wb.create_sheet("요약", 0)
    ws3.sheet_view.showGridLines = False
    summary_data = [
        ("분석 일시", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("스캔 경로", SCAN_ROOT),
        ("", ""),
        ("총 사진/동영상 파일 수", sum(len(v) for v in size_map.values())),
        ("", ""),
        ("── 완전 중복 (해시 동일) ──", ""),
        ("중복 그룹 수", len(hash_dups)),
        ("중복 파일 총 수", total_dup_files),
        ("삭제 가능한 파일 수 (중복본)", total_dup_files - len(hash_dups)),
        ("", ""),
        ("── 같은 파일명 (다른 경로) ──", ""),
        ("같은 이름 그룹 수", len(name_only_dups)),
        ("해당 파일 총 수", sum(len(v) for v in name_only_dups.values())),
    ]

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 40

    title_cell = ws3.cell(row=1, column=1, value="📊 중복 사진 분석 결과")
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[1].height = 30

    for r, (label, value) in enumerate(summary_data, 3):
        lc = ws3.cell(row=r, column=1, value=label)
        vc = ws3.cell(row=r, column=2, value=value)
        if label.startswith("──"):
            lc.font = Font(bold=True, color="1F4E79", size=10)
        elif label:
            lc.font = Font(size=10)
            vc.font = Font(bold=True, size=10)
            if isinstance(value, int):
                vc.number_format = '#,##0'

    wb.save(OUTPUT_PATH)
    print(f"\n✅ Excel 저장 완료: {OUTPUT_PATH}")
    print(f"   완전 중복 그룹: {len(hash_dups)}개 / 파일: {total_dup_files:,}개")
    print(f"   삭제 가능 중복본: {total_dup_files - len(hash_dups):,}개")
    print(f"   같은 파일명 그룹: {len(name_only_dups)}개")


if __name__ == '__main__':
    size_map = scan_files()
    hash_dups = find_hash_duplicates(size_map)
    name_dups = find_name_duplicates(size_map)
    build_excel(hash_dups, name_dups, size_map)
