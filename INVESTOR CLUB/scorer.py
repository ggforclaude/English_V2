#!/usr/bin/env python3
"""
DP Investor Club 2기 지원서 1차 필터링 자동 채점 시스템
사용법: python scorer.py 지원서1.docx 지원서2.pdf --output 채점결과.xlsx
"""

import os
import sys
import io
import json
import time
import argparse
from pathlib import Path
import anthropic
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Windows 콘솔 한글/특수문자 출력
if sys.stdout and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)
if sys.stderr and hasattr(sys.stderr, "buffer"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace", line_buffering=True)

# ──────────────────────────────────────────────
# 1. 문서 텍스트 추출
# ──────────────────────────────────────────────

def extract_text_docx(file_path: str) -> str:
    """DOCX → 텍스트. win32com으로 한글 인코딩 안전 처리."""
    import io, sys
    import win32com.client, pythoncom

    pythoncom.CoInitialize()
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    try:
        doc = word.Documents.Open(os.path.abspath(file_path))
        text = doc.Content.Text
        doc.Close(False)
    finally:
        word.Quit()
    return text


def extract_text_pdf(file_path: str) -> str:
    """PDF → 텍스트."""
    import pdfplumber
    pages = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
    return "\n".join(pages)


def extract_text(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    if ext in (".docx", ".doc"):
        return extract_text_docx(file_path)
    elif ext == ".pdf":
        return extract_text_pdf(file_path)
    else:
        raise ValueError(f"지원하지 않는 형식: {ext}  (docx/doc/pdf만 가능)")


# ──────────────────────────────────────────────
# 2. Claude로 구조화 데이터 추출
# ──────────────────────────────────────────────

EXTRACT_PROMPT = """다음은 DP Investor Club 2기 참가 지원서의 전체 텍스트입니다.
아래 항목들을 JSON 형식으로 추출해주세요. 없는 항목은 기본값으로 채우세요.

지원서 텍스트:
{text}

추출할 JSON 스키마:
{{
  "name": "성명 (문자열, 없으면 '미상')",
  "overseas_experience": {{
    "regions": ["지역 목록 (없으면 빈 배열)"],
    "total_months": "총 개월 수 (정수, 없으면 0)"
  }},
  "ai_proficiency": "상/중/하 중 하나 (없으면 '하')",
  "ai_experience_text": "AI 활용 경험 내용 텍스트 (없으면 빈문자열)",
  "ai_experience_chars": "AI 활용 경험 글자수 (정수)",
  "education": {{
    "school": "학교명",
    "major": "전공",
    "status": "재학/휴학/졸업 중 하나"
  }},
  "certifications": [
    {{"name": "자격증 또는 어학 자격명", "score": "점수 또는 등급 (문자열)", "date": "취득날짜"}}
  ],
  "activities": {{
    "internships": [
      {{
        "org": "기관/단체명",
        "period": "기간 (예: 2024.06~2024.08)",
        "months": "기간을 개월 수로 환산한 정수",
        "content": "활동 내용",
        "is_finance": "금융/투자 관련 여부 (true/false)"
      }}
    ],
    "clubs": [
      {{
        "org": "단체명",
        "period": "기간",
        "months": "기간을 개월 수로 환산한 정수",
        "role": "직책 (회장/팀장/운영진/일반 중 하나 또는 구체 직책명)",
        "content": "활동 내용"
      }}
    ]
  }},
  "essays": {{
    "essay1": "자소서 1번 전체 텍스트",
    "essay1_chars": "글자수 (정수)",
    "essay2": "자소서 2번 전체 텍스트",
    "essay2_chars": "글자수 (정수)",
    "essay3": "자소서 3번 전체 텍스트",
    "essay3_chars": "글자수 (정수)"
  }},
  "career": {{
    "desired_job": "희망 직군",
    "career_plan": "커리어 계획 내용"
  }},
  "program_participation": {{
    "dropout_risk_text": "이탈 가능성 기재 내용 원문",
    "dropout_risk_level": "없음/낮음/중간/높음 중 하나"
  }},
  "portfolio_submitted": "포트폴리오/증빙서류 제출 여부 (true/false, 지원서 내 언급 기준)"
}}

JSON만 출력하세요. 마크다운 코드블록 없이 순수 JSON만."""


def _api_call_with_retry(fn, retries=3):
    for attempt in range(retries):
        try:
            return fn()
        except anthropic.RateLimitError:
            wait = 60 * (attempt + 1)
            print(f"  [대기] 분당 토큰 한도 초과 → {wait}초 대기 중...")
            time.sleep(wait)
    return fn()


def _sanitize_for_json(text: str) -> str:
    """에세이 텍스트 내 ASCII 큰따옴표를 작은따옴표로 변환.
    이스케이프 없이 JSON 문자열에 포함되면 파싱이 깨지므로 사전 치환."""
    return text.replace('"', "'").replace('“', "'").replace('”', "'")


def extract_structured_data(text: str, client: anthropic.Anthropic) -> dict:
    safe_text = _sanitize_for_json(text[:8000])

    def call():
        return client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=8192,
            messages=[{"role": "user", "content": EXTRACT_PROMPT.format(text=safe_text)}],
        )
    resp = _api_call_with_retry(call)
    raw = resp.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.rsplit("```", 1)[0]
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"  [경고] 구조화 파싱 실패: {e}")
        return _empty_data()


def _empty_data() -> dict:
    return {
        "name": "미상", "overseas_experience": {"regions": [], "total_months": 0},
        "ai_proficiency": "하", "ai_experience_text": "", "ai_experience_chars": 0,
        "education": {"school": "", "major": "", "status": ""},
        "certifications": [],
        "activities": {"internships": [], "clubs": []},
        "essays": {"essay1": "", "essay1_chars": 0, "essay2": "", "essay2_chars": 0,
                   "essay3": "", "essay3_chars": 0},
        "career": {"desired_job": "", "career_plan": ""},
        "program_participation": {"dropout_risk_text": "", "dropout_risk_level": "없음"},
        "portfolio_submitted": False,
    }


# ──────────────────────────────────────────────
# 3. 정량 채점 (40점)
# ──────────────────────────────────────────────

def score_quantitative(data: dict) -> dict:
    """정량 항목 채점.
    - 해외/어학: 최대 10점
    - 대외활동:  최대 10점
    - 금융자격:  최대 6점 (가장 높은 자격 하나만 반영)
    합계 최대 26점. AI 활용 일치도(10점)는 정성 채점에서 Claude가 평가.
    """

    # ① 해외 경험 (절반 축소) + 어학 자격 (기존 유지) → 합산 최대 10점
    months = data.get("overseas_experience", {}).get("total_months", 0)
    if months >= 12:  ov = 5
    elif months >= 6: ov = 3
    elif months >= 3: ov = 2
    elif months >= 1: ov = 1
    else:             ov = 0

    lang = 0
    for c in data.get("certifications", []):
        n = c.get("name", "").upper()
        s = c.get("score", "")
        if any(x in n for x in ("TOEFL", "토플")):
            try:
                v = int(s); lang = max(lang, 5 if v >= 100 else 3 if v >= 80 else 1)
            except ValueError: pass
        elif any(x in n for x in ("TOEIC", "토익")):
            try:
                v = int(s); lang = max(lang, 5 if v >= 950 else 3 if v >= 850 else 1)
            except ValueError: pass
        elif any(x in n for x in ("IELTS", "아이엘츠")):
            try:
                v = float(s); lang = max(lang, 5 if v >= 7.0 else 3 if v >= 6.0 else 1)
            except ValueError: pass
        elif any(x in n for x in ("OPIC", "OPIc", "오픽")):
            g = s.upper()
            lang = max(lang, 5 if g == "AL" else 3 if g == "IH" else 1)
    overseas_score = min(ov + lang, 10)

    # ② 대외활동 (최대 10점)
    act = 0
    for intern in data.get("activities", {}).get("internships", []):
        act += 4 if intern.get("is_finance") else 2
        if intern.get("months", 0) >= 6:
            act += 1
    completed_clubs = 0
    for club in data.get("activities", {}).get("clubs", []):
        if completed_clubs < 4:
            act += 1
            completed_clubs += 1
        if club.get("months", 0) >= 6:
            act += 1
    activity_score = min(act, 10)

    # ③ 금융 자격증 (최대 6점, 가장 높은 자격 하나만)
    CERT_TIERS = [
        # (점수, 키워드 목록)
        (6, ["CFA레벨3", "CFA Level 3", "CFA lv3", "CFA lv.3",
             "FRM", "공인회계사", "CPA", "감정평가사", "세무사"]),
        (5, ["AICPA", "CFA레벨2", "CFA Level 2", "CFA lv2", "CFA lv.2"]),
        (3, ["재무설계사", "CFP", "CFA레벨1", "CFA Level 1", "CFA lv1", "CFA lv.1",
             "금융투자분석사", "신용위험분석사", "투자자산운용사",
             "신용분석사", "재경관리사"]),
        (1, ["증권투자자문인력", "파생투자자문인력", "펀드투자자문인력",
             "세무회계", "회계관리"]),
    ]
    # CFA 단독 표기 처리 (레벨 미상 → 레벨1 취급)
    cert_score = 0
    cert_name_for_display = "-"
    for c in data.get("certifications", []):
        n = c.get("name", "")
        for pts, keywords in CERT_TIERS:
            if any(kw.lower() in n.lower() for kw in keywords):
                if pts > cert_score:
                    cert_score = pts
                    cert_name_for_display = n
                break
        else:
            # CFA 단독 (레벨 명시 없음) → 레벨1 취급(3점)
            if "CFA" in n.upper() and cert_score < 3:
                cert_score = 3
                cert_name_for_display = n
    cert_score = min(cert_score, 6)

    return {
        "overseas": overseas_score,
        "activity": activity_score,
        "cert": cert_score,
        "cert_name": cert_name_for_display,
        "total": overseas_score + activity_score + cert_score,
    }


# ──────────────────────────────────────────────
# 4. 정성 채점 (60점) — Claude API
# ──────────────────────────────────────────────

QUALITY_PROMPT = """다음은 DP Investor Club 2기 지원자의 전체 지원 정보입니다.
아래 6가지 항목을 지정된 만점 기준으로 채점하고, 각 항목마다 핵심 근거를 제시하세요.

[AI 자가 평가 등급]: {ai_level}
[AI 활용 경험 기술 내용]: {ai_experience}

[자소서 1번 — AI 전환 시대 강점/역량]
{essay1}

[자소서 2번 — 투자 커리어 노력/인사이트]
{essay2}

[자소서 3번 — 투자 분야 선택 및 전망]
{essay3}

[대외활동 역할 및 내용 (인턴, 학회, 동아리 등)]
{activities}

[희망 직군 / 커리어 계획]
{career}

[프로그램 참여 관련 기재 내용]
{participation}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
채점 기준 (총 74점 — 각 항목 채점 후 JSON 반환)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. AI 활용 일치도 (만점 10점)
   핵심: 자가 평가 등급(상/중/하)과 실제 경험 내용이 얼마나 일치하는가
   - 자가 평가와 경험이 일치: 상→8~10점, 중→6~7점, 하→4~5점
   - 자가 평가가 경험보다 과대평가(상이라 했는데 경험은 중 수준): 각 등급에서 -3~4점 감점
   - 자가 평가가 낮더라도 향상 계획이 구체적으로 서술된 경우: +1~2점 가점
   - 경험 내용이 없거나 매우 단편적이면: 3점 이하
   근거에 반드시 포함: 자가평가 등급, 경험에서 드러난 실제 수준, 과대/과소 평가 여부, 향상 계획 언급 여부

2. 성실성/책임감 (만점 14점)
   - 지속적 노력, 완수 경험, 책임감 있는 행동이 구체적 수치/결과와 함께 드러나는가
   - 14점: 탁월 / 10점: 우수 / 7점: 보통 / 4점: 미흡 / 0점: 근거 없음
   근거에 반드시 포함: 핵심 키워드(예: "매일", "완주", "책임"), 구체적 에피소드 요약

3. 리더십 (만점 14점)
   - 주도적 기획·실행, 팀/조직을 이끈 구체적 성과가 있는가
   - 직책명이 아닌 행동+결과 중심으로 판단
   - 14점: 탁월 / 10점: 우수 / 7점: 보통 / 4점: 미흡 / 0점: 근거 없음
   근거에 반드시 포함: 리더 역할을 수행한 단체/활동명, 핵심 행동 키워드

4. 투자 전문성/관심도 (만점 14점)
   - 전문 용어를 맥락에 맞게 사용하고 본인 견해가 있는가
   - 학회/인턴 경험 외에도 개인 종목 분석, 보고서 작성, 독립적 투자 연구 경험도 반영
   - 14점: 탁월 / 10점: 우수 / 7점: 보통 / 4점: 미흡 / 0점: 근거 없음
   근거에 반드시 포함: 사용된 투자 전문 용어, 언급된 개인 분석/연구 경험 요약

5. 성장 가능성 (만점 12점)
   - 현재 역량보다 미래 발전 가능성, 자기 인식, 학습 의지가 드러나는가
   - 현재 부족하더라도 명확한 성장 계획과 실행 의지가 있으면 높은 점수
   - 12점: 탁월 / 9점: 우수 / 6점: 보통 / 3점: 미흡 / 0점: 근거 없음
   근거에 반드시 포함: 성장 계획 관련 키워드, 자기 인식 수준 평가

6. 프로그램 참여 의지/열정 (만점 10점)
   - 이 프로그램을 선택한 이유가 구체적인가, 참여 후 목표가 명확한가
   - 단순 스펙 쌓기 목적이 아닌 진정성 있는 동기가 느껴지는가
   - 10점: 탁월 / 7점: 우수 / 5점: 보통 / 3점: 미흡 / 0점: 근거 없음
   근거에 반드시 포함: 지원 동기 핵심 내용 요약, 참여 목표 구체성 평가

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
다음 JSON 형식으로만 출력하세요 (마크다운 없이 순수 JSON):
{{
  "ai_consistency": {{"score": 0~10, "reason": "자가평가 등급, 실제 수준, 일치 여부, 향상계획 포함 2~3문장"}},
  "diligence":      {{"score": 0~14, "reason": "핵심 키워드, 에피소드 요약 포함 2~3문장"}},
  "leadership":     {{"score": 0~14, "reason": "활동명, 행동 키워드 포함 2~3문장"}},
  "investment":     {{"score": 0~14, "reason": "전문 용어, 개인 분석 경험 포함 2~3문장"}},
  "growth":         {{"score": 0~12, "reason": "성장 계획, 자기 인식 수준 포함 2~3문장"}},
  "commitment":     {{"score": 0~10,  "reason": "지원 동기, 참여 목표 포함 2~3문장"}}
}}"""


def score_qualitative(data: dict, client: anthropic.Anthropic) -> dict:
    essays    = data.get("essays", {})
    activities = data.get("activities", {})
    career    = data.get("career", {})
    pp        = data.get("program_participation", {})

    activity_lines = []
    for i in activities.get("internships", []):
        activity_lines.append(f"인턴 ({i.get('org', '')}, {i.get('period', '')}): {i.get('content', '')}")
    for c in activities.get("clubs", []):
        activity_lines.append(
            f"학회/동아리 ({c.get('org', '')}, 역할:{c.get('role', '')}, {c.get('period', '')}): {c.get('content', '')}"
        )

    prompt = QUALITY_PROMPT.format(
        ai_level=data.get("ai_proficiency", "미기재"),
        ai_experience=data.get("ai_experience_text", "미기재") or "미기재",
        essay1=essays.get("essay1", "미작성") or "미작성",
        essay2=essays.get("essay2", "미작성") or "미작성",
        essay3=essays.get("essay3", "미작성") or "미작성",
        activities="\n".join(activity_lines) or "없음",
        career=f"{career.get('desired_job', '')} / {career.get('career_plan', '')}",
        participation=pp.get("dropout_risk_text", "미기재") or "미기재",
    )

    def call():
        return client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )
    resp = _api_call_with_retry(call)
    raw = resp.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.rsplit("```", 1)[0]
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        empty = {"score": 0, "reason": "파싱 실패"}
        return {k: dict(empty) for k in
                ("ai_consistency", "diligence", "leadership", "investment", "growth", "commitment")}


# ──────────────────────────────────────────────
# 5. 감점 계산
# ──────────────────────────────────────────────

ESSAY_MAX_CHARS = {
    "essay1": 800,
    "essay2": 800,
    "essay3": 1000,
}

def _essay_deduction(chars: int, max_chars: int, label: str) -> tuple[int, str]:
    """자소서 분량 비율에 따른 감점 계산."""
    if max_chars == 0:
        return 0, ""
    ratio = chars / max_chars
    if ratio > 1.01:
        return -1, f"{label} 초과 작성 (-1)"
    elif ratio >= 0.70:
        return 0, ""
    elif ratio >= 0.50:
        return -2, f"{label} 분량 부족 50~69% (-2)"
    else:
        return -5, f"{label} 분량 심각 부족 50% 미만 (-5)"


def calc_deductions(data: dict) -> dict:
    points = 0
    reasons = []

    # 프로그램 이탈 가능성
    level = data.get("program_participation", {}).get("dropout_risk_level", "없음")
    if level == "높음":
        points -= 10; reasons.append("이탈 가능성 높음 (-10)")
    elif level == "중간":
        points -= 5;  reasons.append("이탈 가능성 중간 (-5)")

    # 자소서 분량 — 최대 글자수 대비 비율
    e = data.get("essays", {})
    for key, max_c in ESSAY_MAX_CHARS.items():
        actual = e.get(f"{key}_chars", 0)
        label  = {"essay1": "자소서1번", "essay2": "자소서2번", "essay3": "자소서3번"}[key]
        pts, msg = _essay_deduction(actual, max_c, label)
        if pts:
            points += pts
            reasons.append(msg)

    # 서류 미제출
    if not data.get("portfolio_submitted", False):
        points -= 3; reasons.append("포트폴리오/증빙 미제출 (-3)")

    return {"total": points, "reasons": ", ".join(reasons) if reasons else "없음"}


# ──────────────────────────────────────────────
# 6. 합산 및 합불 판정
# ──────────────────────────────────────────────

def summarize(q: dict, ql: dict, ded: dict) -> dict:
    qual_total = (
        ql["ai_consistency"]["score"]
        + ql["diligence"]["score"]
        + ql["leadership"]["score"]
        + ql["investment"]["score"]
        + ql["growth"]["score"]
        + ql["commitment"]["score"]
    )
    total = q["total"] + qual_total + ded["total"]
    # 성실성(14점 만점) 또는 리더십(14점 만점) 중 하나가 7점 이상이어야 통과
    diligence_ok = ql["diligence"]["score"] >= 7
    leadership_ok = ql["leadership"]["score"] >= 7
    passed = total >= 60 and (diligence_ok or leadership_ok)
    return {"total": total, "passed": passed}


# ──────────────────────────────────────────────
# 7. 엑셀 출력
# ──────────────────────────────────────────────

HEADERS = [
    "순위", "성명", "총점", "결과",
    # 정량 26점 + AI일치도 10점
    "해외/어학(10)", "대외활동(10)", "금융자격(6)", "AI일치도(10)",
    # 정성 64점
    "성실/책임감(14)", "리더십(14)", "투자전문성(14)", "성장가능성(12)", "참여의지(10)",
    # 감점
    "감점", "감점 사유",
    # 상세 참고
    "AI자가평가", "해외경험(월)", "보유자격증",
    "자소서1(자)", "자소서2(자)", "자소서3(자)", "이탈가능성",
    # 근거 (키워드/활동 요약 포함)
    "AI일치도 근거", "성실성 근거", "리더십 근거",
    "투자전문성 근거", "성장가능성 근거", "참여의지 근거",
]

COL_WIDTHS = [
    6, 10, 8, 8,          # 순위~결과
    12, 12, 10, 12,       # 정량 4항목
    13, 13, 13, 12, 11,   # 정성 5항목
    8, 30,                # 감점
    10, 10, 15,           # AI자가평가, 해외, 자격증
    10, 10, 10, 12,       # 자소서 3개 + 이탈가능성
    42, 42, 42, 42, 42, 42,  # 근거 6개
]

COLOR_HEADER = "1F4E79"
COLOR_PASS   = "C6EFCE"
COLOR_FAIL   = "FFC7CE"
COLOR_SCORE  = "FFF2CC"


def build_excel(results: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1차 필터링 결과"

    hdr_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    pass_fill = PatternFill("solid", fgColor=COLOR_PASS)
    fail_fill = PatternFill("solid", fgColor=COLOR_FAIL)
    score_fill = PatternFill("solid", fgColor=COLOR_SCORE)
    bold = Font(bold=True, size=10)
    normal = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # 헤더 행
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center

    ws.row_dimensions[1].height = 32

    # 데이터 행
    for rank, res in enumerate(results, 1):
        data = res["data"]
        q    = res["quantitative"]
        ql   = res["qualitative"]
        ded  = res["deductions"]
        sm   = res["summary"]

        row_vals = [
            rank,
            data.get("name", "미상"),
            sm["total"],
            "[통과]" if sm["passed"] else "[탈락]",
            # 정량
            q["overseas"], q["activity"], q["cert"],
            ql["ai_consistency"]["score"],
            # 정성
            ql["diligence"]["score"], ql["leadership"]["score"],
            ql["investment"]["score"], ql["growth"]["score"], ql["commitment"]["score"],
            # 감점
            ded["total"], ded["reasons"],
            # 상세
            data.get("ai_proficiency", "-"),
            data.get("overseas_experience", {}).get("total_months", 0),
            q.get("cert_name", "-"),
            data.get("essays", {}).get("essay1_chars", 0),
            data.get("essays", {}).get("essay2_chars", 0),
            data.get("essays", {}).get("essay3_chars", 0),
            data.get("program_participation", {}).get("dropout_risk_level", "-"),
            # 근거
            ql["ai_consistency"]["reason"],
            ql["diligence"]["reason"],
            ql["leadership"]["reason"],
            ql["investment"]["reason"],
            ql["growth"]["reason"],
            ql["commitment"]["reason"],
        ]

        r = rank + 1
        REASON_START_COL = 22  # "AI일치도 근거" 컬럼 번호
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = bold if col in (2, 3) else normal
            c.alignment = left if col >= REASON_START_COL else center

        # 결과 칸 색상
        result_cell = ws.cell(row=r, column=4)
        result_cell.fill = pass_fill if sm["passed"] else fail_fill

        # 총점 색상
        ws.cell(row=r, column=3).fill = score_fill

        ws.row_dimensions[r].height = 60

    # 열 너비
    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    # ── 평가기준 시트 (항상 최신 버전으로 재생성) ──
    if "평가기준" in wb.sheetnames:
        del wb["평가기준"]
    _build_criteria_sheet(wb)

    wb.save(output_path)


def _build_criteria_sheet(wb: openpyxl.Workbook):
    wc = wb.create_sheet("평가기준")

    # 스타일 정의
    c_navy   = "1F4E79"
    c_blue   = "BDD7EE"
    c_green  = "C6EFCE"
    c_red    = "FFC7CE"
    c_yellow = "FFF2CC"
    c_gray   = "F2F2F2"

    def hd1(row, col, val, bg=c_navy, fg="FFFFFF", span=None):
        cell = wc.cell(row=row, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(color=fg, bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if span:
            wc.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + span - 1)
        return cell

    def row_data(row, col, label, value, note="", bg=None):
        lc = wc.cell(row=row, column=col, value=label)
        lc.font = Font(bold=True, size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bg:
            lc.fill = PatternFill("solid", fgColor=bg)

        vc = wc.cell(row=row, column=col + 1, value=value)
        vc.font = Font(size=10)
        vc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if bg:
            vc.fill = PatternFill("solid", fgColor=bg)

        if note:
            nc = wc.cell(row=row, column=col + 2, value=note)
            nc.font = Font(size=9, italic=True, color="595959")
            nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    r = 1

    # ── 제목 ──
    hd1(r, 1, "DP Investor Club 2기  1차 필터링 평가기준표 (v2)", c_navy, span=4)
    wc.row_dimensions[r].height = 28
    r += 1

    cell = wc.cell(row=r, column=1,
        value="총점 100점  |  1차 통과 기준: 60점 이상 + (성실/책임감 >= 7점  OR  리더십 >= 7점)")
    cell.font = Font(size=10, italic=True)
    wc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    wc.row_dimensions[r].height = 18
    r += 2

    # ── STEP 1: Pass/Fail ──
    hd1(r, 1, "STEP 1 — 기본 자격 확인", c_navy, span=4)
    wc.row_dimensions[r].height = 22
    r += 1
    for label, val, note in [
        ("필수 항목 기입", "학력, 이메일, 희망직군 공란 없음", ""),
    ]:
        row_data(r, 1, label, val, note, bg=c_gray)
        wc.row_dimensions[r].height = 18
        r += 1
    r += 1

    # ── STEP 2: 정량 ──
    hd1(r, 1, "STEP 2 — 정량 점수 (26점)  /  자동 채점", c_blue, fg="1F4E79", span=4)
    wc.row_dimensions[r].height = 22
    r += 1

    sections_q = [
        ("① 해외 경험 + 어학 자격 (최대 10점)", [
            ("해외 경험 1년 이상",   "5점", "해외경험 최대 5점 + 어학 최대 5점 합산"),
            ("해외 경험 6개월~1년",  "3점", ""),
            ("해외 경험 3~6개월",    "2점", ""),
            ("해외 경험 1~3개월",    "1점", ""),
            ("TOEFL 100+ / OPIc AL / IELTS 7.0+ / TOEIC 950+", "+5점", ""),
            ("TOEFL 80~99 / OPIc IH / IELTS 6.0+ / TOEIC 850+", "+3점", ""),
            ("그 외 어학 자격 보유",  "+1점", ""),
        ]),
        ("② 대외활동 이력 (최대 10점)", [
            ("인턴 경험 — 금융/투자 관련",     "4점", ""),
            ("인턴 경험 — 일반",               "2점", ""),
            ("인턴 6개월 이상 시 추가",        "+1점", ""),
            ("학회/동아리 완료 1개당",         "+1점 (최대 4점)", ""),
            ("학회/동아리 6개월 이상 시 추가", "+1점/개", ""),
        ]),
        ("③ 금융 자격증 (최대 6점, 가장 높은 자격 하나만 반영)", [
            ("6점 — CFA레벨3, FRM, 공인회계사(CPA), 감정평가사, 세무사", "6점", ""),
            ("5점 — AICPA, CFA레벨2",                                    "5점", ""),
            ("3점 — 재무설계사(CFP), CFA레벨1, 금융투자분석사, 신용위험분석사, 투자자산운용사, 신용분석사, 재경관리사", "3점", ""),
            ("1점 — 증권/파생/펀드투자자문인력, 세무회계, 회계관리",     "1점", ""),
        ]),
    ]

    for sec_title, items in sections_q:
        hd1(r, 1, sec_title, c_yellow, fg="1F4E79", span=4)
        wc.row_dimensions[r].height = 20
        r += 1
        for label, val, note in items:
            row_data(r, 1, label, val, note)
            wc.row_dimensions[r].height = 17
            r += 1
        r += 1

    # ── STEP 3: 정성 ──
    hd1(r, 1, "STEP 3 — 정성 점수 (74점)  /  Claude AI 분석", c_blue, fg="1F4E79", span=4)
    wc.row_dimensions[r].height = 22
    r += 1

    sections_ql = [
        ("④ AI 활용 일치도 (최대 10점)", [
            ("분석 대상", "자가 평가 등급(상/중/하) + AI 활용 경험 기술 내용", ""),
            ("자가평가와 경험 일치 — 상 수준", "8~10점", ""),
            ("자가평가와 경험 일치 — 중 수준",  "6~7점",  ""),
            ("자가평가와 경험 일치 — 하 수준",  "4~5점",  ""),
            ("자가평가 과대 (상 선택, 경험은 중)", "-3~4점 감점", ""),
            ("향상 계획 구체적으로 기술 시",   "+1~2점 가점", ""),
            ("경험 내용 없거나 단편적",         "3점 이하", ""),
        ]),
        ("⑤ 성실성 / 책임감 (최대 14점)  ★ 핵심 인재상", [
            ("분석 대상", "자소서 1,2번 + 대외활동 내용", ""),
            ("14점 (탁월)", "지속적 노력, 완수 경험이 수치/결과와 함께 구체 서술됨", ""),
            ("10점 (우수)", "성실성 근거가 명확하나 결과 수치 부족", ""),
            (" 7점 (보통)", "성실성 언급은 있으나 단편적", ""),
            (" 4점 (미흡)", "간접적 암시만 있음", ""),
            (" 0점",        "근거 없음 / 자소서 미작성", ""),
        ]),
        ("⑥ 리더십 (최대 14점)  ★ 핵심 인재상", [
            ("분석 대상", "대외활동 역할 + 자소서 전체", ""),
            ("14점 (탁월)", "주도적 기획, 실행, 성과가 수치와 함께 서술됨", ""),
            ("10점 (우수)", "리더 역할과 행동이 명확하나 성과 서술 미흡", ""),
            (" 7점 (보통)", "직책만 있고 주도적 행동 묘사 부족", ""),
            (" 4점 (미흡)", "리더십 관련 간접 언급만", ""),
            (" 0점",        "근거 없음 / 자소서 미작성", ""),
        ]),
        ("⑦ 투자 전문성 / 관심도 (최대 14점)", [
            ("분석 대상", "자소서 3번 + 희망 커리어 + 개인 분석/연구 경험", ""),
            ("14점 (탁월)", "전문 용어를 맥락에 맞게 사용, 개인 분석 경험 + 본인 견해 명확", ""),
            ("10점 (우수)", "전문성 있으나 개인 견해 깊이 부족", ""),
            (" 7점 (보통)", "기본 관심 표현, 전문 용어 1~2개 사용", ""),
            (" 4점 (미흡)", "일반적 내용만", ""),
            (" 0점",        "근거 없음 / 자소서 미작성", ""),
        ]),
        ("⑧ 성장 가능성 (최대 12점)", [
            ("분석 대상", "자소서 전체 + 커리어 계획", ""),
            ("12점 (탁월)", "현재 역량 대비 명확한 성장 계획, 자기 인식, 실행 의지 서술됨", ""),
            (" 9점 (우수)", "성장 의지는 있으나 계획 구체성 부족", ""),
            (" 6점 (보통)", "막연한 성장 언급", ""),
            (" 3점 (미흡)", "성장 관련 내용 거의 없음", ""),
            (" 0점",        "근거 없음", ""),
        ]),
        ("⑨ 프로그램 참여 의지 / 열정 (최대 10점)", [
            ("분석 대상", "자소서 전체 + 프로그램 참여 관련 기재 내용", ""),
            ("10점 (탁월)", "구체적 지원 동기, 프로그램 목표 명확, 진정성 느껴짐", ""),
            (" 7점 (우수)", "동기는 있으나 프로그램 특화 내용 부족", ""),
            (" 5점 (보통)", "일반적 동기 언급", ""),
            (" 3점 (미흡)", "스펙 쌓기 목적이 강하게 느껴짐", ""),
            (" 0점",        "근거 없음", ""),
        ]),
    ]

    for sec_title, items in sections_ql:
        hd1(r, 1, sec_title, c_yellow, fg="1F4E79", span=4)
        wc.row_dimensions[r].height = 20
        r += 1
        for label, val, note in items:
            row_data(r, 1, label, val, note)
            wc.row_dimensions[r].height = 17
            r += 1
        r += 1

    # ── STEP 4: 감점 ──
    hd1(r, 1, "STEP 4 — 감점 항목", c_red, fg="FFFFFF", span=4)
    wc.row_dimensions[r].height = 22
    r += 1
    for label, val, note in [
        ("프로그램 이탈 가능성 '높음' 기재",               "-10점", "이탈 가능성란 원문 기준"),
        ("프로그램 이탈 가능성 '중간' 기재",               " -5점", ""),
        ("자소서 각 항목 최대 글자수 101% 초과",           " -1점/항목", "1번 800자, 2번 800자, 3번 1000자 기준"),
        ("자소서 각 항목 50~69% 작성",                    " -2점/항목", ""),
        ("자소서 각 항목 50% 미만 작성 (심각 부족)",       " -5점/항목", ""),
        ("포트폴리오/증빙 서류 미제출",                    " -3점", "지원서 내 언급 기준"),
    ]:
        row_data(r, 1, label, val, note, bg="FFF2F2")
        wc.row_dimensions[r].height = 17
        r += 1
    r += 1

    # ── 통과 기준 요약 ──
    hd1(r, 1, "1차 통과 기준 요약", c_green, fg="1F4E79", span=4)
    wc.row_dimensions[r].height = 22
    r += 1
    for label, val, note in [
        ("총점",     "60점 이상", "정량 26점 + AI일치도 10점 + 정성 64점 = 100점 만점"),
        ("필수 조건", "성실/책임감(5번) >= 7점  또는  리더십(6번) >= 7점",
                      "둘 다 7점 미만이면 총점 무관 탈락"),
    ]:
        row_data(r, 1, label, val, note, bg=c_green)
        wc.row_dimensions[r].height = 18
        r += 1

    # ── 열 너비 ──
    wc.column_dimensions["A"].width = 50
    wc.column_dimensions["B"].width = 30
    wc.column_dimensions["C"].width = 36
    wc.column_dimensions["D"].width = 5


# ──────────────────────────────────────────────
# 8. 단일 파일 처리
# ──────────────────────────────────────────────

def process_file(file_path: str, client: anthropic.Anthropic,
                 supporting_docs: list | None = None) -> dict:
    name = Path(file_path).name
    print(f"\n[처리] {name}")

    print("  ① 텍스트 추출 중...")
    text = extract_text(file_path)

    if supporting_docs:
        supp_parts = []
        for sp in supporting_docs:
            try:
                st = extract_text(sp)
                if st.strip():
                    supp_parts.append(f"[증빙자료: {Path(sp).name}]\n{st}")
            except Exception as e:
                print(f"  [경고] 증빙자료 추출 실패 ({Path(sp).name}): {e}")
        if supp_parts:
            text = text + "\n\n" + "\n\n".join(supp_parts)
            print(f"  → 증빙자료 {len(supp_parts)}건 텍스트 합산")

    print("  ② 항목 구조화 중 (Claude)...")
    data = extract_structured_data(text, client)
    print(f"     → 지원자: {data.get('name', '미상')}")

    print("  ③ 정량 채점 중...")
    q = score_quantitative(data)

    print("  ④ 정성 채점 중 (Claude)...")
    ql = score_qualitative(data, client)

    ded = calc_deductions(data)
    sm  = summarize(q, ql, ded)

    total_label = f"{sm['total']}점 -> {'통과' if sm['passed'] else '탈락'}"
    print(f"  완료: {total_label}")

    return {"file": file_path, "data": data, "quantitative": q,
            "qualitative": ql, "deductions": ded, "summary": sm}


# ──────────────────────────────────────────────
# 9. 체크포인트 (이어하기)
# ──────────────────────────────────────────────

def _ckpt_path(output_path: str) -> Path:
    p = Path(output_path)
    return p.parent / (p.stem + "_checkpoint.json")


def _load_checkpoint(ckpt: Path) -> dict:
    """저장된 체크포인트를 {절대경로: result} 딕셔너리로 반환."""
    if not ckpt.exists():
        return {}
    with open(ckpt, encoding="utf-8") as f:
        data = json.load(f)
    return {item["file"]: item for item in data.get("results", [])}


def _save_checkpoint(ckpt: Path, results: list):
    with open(ckpt, "w", encoding="utf-8") as f:
        json.dump({"results": results}, f, ensure_ascii=False, indent=2)


def _is_valid_result(result: dict) -> bool:
    """이름이 추출됐거나 정성 점수 중 하나라도 0점 초과면 유효한 결과로 판단."""
    data = result.get("data", {})
    ql   = result.get("qualitative", {})
    name_ok = data.get("name", "미상") not in ("미상", "", None)
    any_score = any(
        ql.get(k, {}).get("score", 0) > 0
        for k in ("diligence", "leadership", "investment",
                  "growth", "commitment", "ai_consistency")
    )
    return name_ok or any_score


# ──────────────────────────────────────────────
# 10. 진입점
# ──────────────────────────────────────────────

_APP_KEYWORDS = ("지원서", "참가", "application", "지원")

def _find_application_form(subfolder: Path) -> tuple:
    """하위 폴더에서 지원서 본문과 증빙자료를 구분.
    반환: (main_form: Path | None, supporting: list[Path])
    """
    SUPPORTED = {".pdf", ".docx", ".doc"}
    files = sorted(f for f in subfolder.iterdir()
                   if f.is_file() and f.suffix.lower() in SUPPORTED)
    if not files:
        return None, []

    main_form = None
    for f in files:
        if any(kw in f.name.lower() for kw in _APP_KEYWORDS):
            main_form = f
            break

    if main_form is None:
        main_form = files[0]

    return main_form, [f for f in files if f != main_form]

def _load_env():
    """프로젝트 루트 및 상위 폴더의 .env 파일을 자동 로드."""
    search_dirs = [
        Path(__file__).parent,
        Path(__file__).parent.parent,
        Path(__file__).parent.parent / "invest_telegram",
    ]
    for d in search_dirs:
        env_path = d / ".env"
        if env_path.exists():
            with open(env_path, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if "=" in line and not line.startswith("#"):
                        k, v = line.split("=", 1)
                        os.environ.setdefault(k.strip(), v.strip())


def main():
    _load_env()

    parser = argparse.ArgumentParser(
        description="DP Investor Club 지원서 자동 채점",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("files", nargs="+",
                        help="지원서 파일(PDF/DOCX) 또는 폴더 경로 (여러 개 가능)\n"
                             "  예) python scorer.py 지원서폴더/\n"
                             "  예) python scorer.py a.docx b.pdf 지원서폴더/")
    parser.add_argument("--output", "-o", default="채점결과.xlsx",
                        help="출력 엑셀 파일명 (기본값: 채점결과.xlsx)")
    parser.add_argument("--api-key", default=None,
                        help="Anthropic API 키 (없으면 환경변수 ANTHROPIC_API_KEY 사용)")
    parser.add_argument("--reset", action="store_true",
                        help="체크포인트를 무시하고 처음부터 다시 채점")
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("오류: ANTHROPIC_API_KEY 환경변수 또는 --api-key 옵션이 필요합니다.")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # 출력 경로 확정
    out = args.output if args.output.endswith(".xlsx") else args.output + ".xlsx"
    ckpt = _ckpt_path(out)

    # 파일 목록 수집 → list[tuple[str, list[str]]]  (주지원서, [증빙자료...])
    SUPPORTED = {".pdf", ".docx", ".doc"}
    file_list: list = []
    for entry in args.files:
        p = Path(entry)
        if p.is_dir():
            subdirs = sorted(d for d in p.iterdir() if d.is_dir())
            if subdirs:
                # 개인별 하위 폴더 구조
                print(f"[폴더] {entry} -> 개인별 하위 폴더 {len(subdirs)}개 발견")
                for subdir in subdirs:
                    main_form, supporting = _find_application_form(subdir)
                    if main_form is None:
                        print(f"  [건너뜀] {subdir.name} — 지원서 파일 없음")
                        continue
                    sc = len(supporting)
                    print(f"  {subdir.name}: 지원서={main_form.name}" +
                          (f", 증빙{sc}건" if sc else ""))
                    file_list.append((str(main_form.resolve()),
                                      [str(s.resolve()) for s in supporting]))
            else:
                # 기존 flat 구조 (파일만 있는 폴더)
                found = sorted(f for f in p.iterdir()
                               if f.is_file() and f.suffix.lower() in SUPPORTED)
                if not found:
                    print(f"[경고] 폴더 안에 PDF/DOCX 파일이 없습니다: {entry}")
                else:
                    print(f"[폴더] {entry} -> {len(found)}개 파일 발견")
                file_list.extend((str(f.resolve()), []) for f in found)
        elif p.is_file():
            if p.suffix.lower() not in SUPPORTED:
                print(f"[건너뜀] 지원하지 않는 형식: {entry}")
            else:
                file_list.append((str(p.resolve()), []))
        else:
            print(f"[건너뜀] 경로 없음: {entry}")

    if not file_list:
        print("\n처리할 파일이 없습니다.")
        sys.exit(1)

    total_count = len(file_list)

    # 체크포인트 로드
    if args.reset and ckpt.exists():
        ckpt.unlink()
        print("[체크포인트 초기화] 처음부터 다시 채점합니다.")

    checkpoint = _load_checkpoint(ckpt)
    results = list(checkpoint.values())

    # 이미 완료된 파일 제외
    pending = [(fp, supp) for fp, supp in file_list if fp not in checkpoint]
    if checkpoint:
        print(f"[이어하기] 이미 완료: {len(checkpoint)}명 / 남은 파일: {len(pending)}명")

    print(f"\n총 {total_count}명 중 {len(pending)}명 채점 시작\n")

    failed = []  # 에러 또는 불완전 결과 파일

    for i, (fp, supp) in enumerate(pending, 1):
        label = f"({len(checkpoint) + i}/{total_count})"
        print(f"\n[처리 {label}] {Path(fp).name}")
        try:
            result = process_file(fp, client, supporting_docs=supp)
            if _is_valid_result(result):
                results.append(result)
                checkpoint[fp] = result
                _save_checkpoint(ckpt, results)
            else:
                print(f"  [주의] 데이터 추출 불완전 (이름: {result['data'].get('name','미상')}) -> 재시도 목록 추가")
                failed.append((fp, supp))
        except Exception as e:
            print(f"  [오류] {e} -> 재시도 목록 추가")
            failed.append((fp, supp))

    # ── 실패 파일 자동 재시도 (1회) ──
    if failed:
        print(f"\n{'='*55}")
        print(f"  [재시도] {len(failed)}개 파일 재처리 중 (30초 대기 후 시작)...")
        time.sleep(30)
        still_failed = []
        for i, (fp, supp) in enumerate(failed, 1):
            print(f"\n[재시도 {i}/{len(failed)}] {Path(fp).name}")
            try:
                result = process_file(fp, client, supporting_docs=supp)
                if _is_valid_result(result):
                    results.append(result)
                    checkpoint[fp] = result
                    _save_checkpoint(ckpt, results)
                    print(f"  재시도 성공")
                else:
                    print(f"  재시도 후에도 불완전 -> 수동 확인 필요")
                    still_failed.append((fp, supp))
            except Exception as e:
                print(f"  재시도 실패: {e}")
                still_failed.append((fp, supp))
        failed = still_failed

    if not results:
        print("\n처리된 파일이 없습니다.")
        sys.exit(1)

    # 총점 내림차순 정렬
    results.sort(key=lambda r: r["summary"]["total"], reverse=True)

    # 엑셀 저장
    build_excel(results, out)

    # 성공 시 체크포인트 삭제
    if not failed and ckpt.exists():
        ckpt.unlink()

    # 터미널 요약
    passed = [r for r in results if r["summary"]["passed"]]
    print(f"\n{'='*55}")
    print(f"  채점 완료: {len(results)}명  |  통과: {len(passed)}명  |  탈락: {len(results)-len(passed)}명")
    if failed:
        print(f"  [주의] 최종 실패 {len(failed)}개 — 수동 확인 필요:")
        for fp, _ in failed:
            print(f"    - {Path(fp).name}")
    print(f"  결과 파일: {out}")
    print(f"{'='*55}")
    print("\n  순위  성명         총점   결과")
    print("  " + "-"*35)
    for idx, r in enumerate(results, 1):
        sm   = r["summary"]
        name = r["data"].get("name", "미상")
        flag = "통과" if sm["passed"] else "탈락"
        print(f"  {idx:2}위   {name:<10} {sm['total']:3}점   {flag}")


if __name__ == "__main__":
    main()
