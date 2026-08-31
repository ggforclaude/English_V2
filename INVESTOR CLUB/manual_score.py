#!/usr/bin/env python3
"""
이수정·정하림 수동 채점 스크립트
JSON 파싱 실패 우회: 자소서 원문을 JSON에 포함하지 않고 별도 텍스트로 Claude에 전송
"""

import os, sys, io, json, time
from pathlib import Path
import anthropic, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

if sys.stdout and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)
if sys.stderr and hasattr(sys.stderr, "buffer"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace", line_buffering=True)

def _load_env():
    for d in [Path(__file__).parent, Path(__file__).parent.parent,
              Path(__file__).parent.parent / "invest_telegram"]:
        env_path = d / ".env"
        if env_path.exists():
            with open(env_path, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if "=" in line and not line.startswith("#"):
                        k, v = line.split("=", 1)
                        os.environ.setdefault(k.strip(), v.strip())

def extract_text_docx(file_path):
    import win32com.client, pythoncom
    pythoncom.CoInitialize()
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    try:
        doc = word.Documents.Open(os.path.abspath(file_path))
        text = doc.Content.Text
        doc.Close(False)
    finally:
        word.Quit()
    return text

def extract_text_pdf(file_path):
    import pdfplumber
    pages = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
    return "\n".join(pages)

def extract_text(file_path):
    ext = Path(file_path).suffix.lower()
    if ext in (".docx", ".doc"):
        return extract_text_docx(file_path)
    elif ext == ".pdf":
        return extract_text_pdf(file_path)
    raise ValueError(f"지원하지 않는 형식: {ext}")

# ── 구조화 프롬프트 (자소서 원문 제외, 글자수만) ──
EXTRACT_PROMPT_SAFE = """다음은 DP Investor Club 2기 참가 지원서 텍스트입니다.
아래 항목을 JSON으로 추출하세요. 없으면 기본값을 사용하세요.

주의: essay1/essay2/essay3 필드에 원문 텍스트를 넣지 말고 글자수(정수)만 넣으세요.

지원서 텍스트:
{text}

JSON 스키마:
{{
  "name": "성명 (없으면 미상)",
  "overseas_experience": {{"regions": [], "total_months": 0}},
  "ai_proficiency": "상/중/하",
  "ai_experience_text": "AI 활용 경험 요약 (2~3줄, 따옴표 사용 금지)",
  "education": {{"school": "학교명", "major": "전공", "status": "재학/휴학/졸업"}},
  "certifications": [{{"name": "자격증명", "score": "점수", "date": "날짜"}}],
  "activities": {{
    "internships": [{{"org": "기관명", "period": "기간", "months": 0, "content": "활동요약(따옴표금지)", "is_finance": true}}],
    "clubs": [{{"org": "단체명", "period": "기간", "months": 0, "role": "직책", "content": "활동요약(따옴표금지)"}}]
  }},
  "essay1_chars": 0,
  "essay2_chars": 0,
  "essay3_chars": 0,
  "career": {{"desired_job": "희망직군", "career_plan": "커리어계획요약(따옴표금지)"}},
  "program_participation": {{"dropout_risk_text": "이탈가능성기재내용", "dropout_risk_level": "없음/낮음/중간/높음"}},
  "portfolio_submitted": false
}}

JSON만 출력 (마크다운 없이):"""

QUALITY_PROMPT = """다음은 DP Investor Club 2기 지원자 전체 지원서 원문입니다.
6가지 항목을 채점하고 JSON으로만 반환하세요.

지원서 원문:
{text}

채점 기준:
1. AI 활용 일치도 (0~10점): 자가평가 등급과 실제 경험 일치도
2. 성실성/책임감 (0~14점): 지속적 노력, 완수 경험, 구체적 수치/결과
3. 리더십 (0~14점): 주도적 기획·실행, 팀 이끈 구체적 성과
4. 투자 전문성 (0~14점): 전문 용어 맥락 사용, 개인 분석/연구 경험
5. 성장 가능성 (0~12점): 미래 발전 계획, 자기 인식, 학습 의지
6. 참여 의지 (0~10점): 구체적 지원 동기, 프로그램 목표 명확성

JSON만 출력:
{{
  "ai_consistency": {{"score": 0, "reason": "근거 2문장"}},
  "diligence":      {{"score": 0, "reason": "근거 2문장"}},
  "leadership":     {{"score": 0, "reason": "근거 2문장"}},
  "investment":     {{"score": 0, "reason": "근거 2문장"}},
  "growth":         {{"score": 0, "reason": "근거 2문장"}},
  "commitment":     {{"score": 0, "reason": "근거 2문장"}}
}}"""

def _api_retry(fn, retries=3):
    for attempt in range(retries):
        try:
            return fn()
        except anthropic.RateLimitError:
            wait = 60 * (attempt + 1)
            print(f"  [대기] {wait}초 대기...")
            time.sleep(wait)
    return fn()

def _fix_json(raw):
    """간단한 JSON 수리: 제어문자 제거, 마지막 } 이후 잘라내기"""
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.rsplit("```", 1)[0]
    # 제어문자 제거
    raw = "".join(c for c in raw if ord(c) >= 32 or c in "\n\r\t")
    # 마지막 } 위치로 자르기
    last_brace = raw.rfind("}")
    if last_brace != -1:
        raw = raw[:last_brace+1]
    return raw

def extract_structured(text, client):
    def call():
        return client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4096,
            messages=[{"role": "user", "content": EXTRACT_PROMPT_SAFE.format(text=text[:8000])}],
        )
    resp = _api_retry(call)
    raw = _fix_json(resp.content[0].text)
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"  [경고] 구조화 파싱 실패: {e}")
        print(f"  [원문 일부] {raw[:300]}")
        return None

def score_qualitative(text, client):
    def call():
        return client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2000,
            messages=[{"role": "user", "content": QUALITY_PROMPT.format(text=text[:8000])}],
        )
    resp = _api_retry(call)
    raw = _fix_json(resp.content[0].text)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        empty = {"score": 0, "reason": "파싱 실패"}
        return {k: dict(empty) for k in
                ("ai_consistency", "diligence", "leadership", "investment", "growth", "commitment")}

def score_quantitative(d):
    months = d.get("overseas_experience", {}).get("total_months", 0)
    if months >= 12:   ov = 5
    elif months >= 6:  ov = 3
    elif months >= 3:  ov = 2
    elif months >= 1:  ov = 1
    else:              ov = 0

    lang = 0
    for c in d.get("certifications", []):
        n = c.get("name", "").upper()
        s = c.get("score", "")
        if any(x in n for x in ("TOEFL","토플")):
            try: v=int(s); lang=max(lang, 5 if v>=100 else 3 if v>=80 else 1)
            except: pass
        elif any(x in n for x in ("TOEIC","토익")):
            try: v=int(s); lang=max(lang, 5 if v>=950 else 3 if v>=850 else 1)
            except: pass
        elif any(x in n for x in ("IELTS","아이엘츠")):
            try: v=float(s); lang=max(lang, 5 if v>=7.0 else 3 if v>=6.0 else 1)
            except: pass
        elif any(x in n for x in ("OPIC","OPIc","오픽")):
            g=s.upper(); lang=max(lang, 5 if g=="AL" else 3 if g=="IH" else 1)
    overseas_score = min(ov + lang, 10)

    act = 0
    for intern in d.get("activities", {}).get("internships", []):
        act += 4 if intern.get("is_finance") else 2
        if intern.get("months", 0) >= 6: act += 1
    done_clubs = 0
    for club in d.get("activities", {}).get("clubs", []):
        if done_clubs < 4: act += 1; done_clubs += 1
        if club.get("months", 0) >= 6: act += 1
    activity_score = min(act, 10)

    CERT_TIERS = [
        (6, ["CFA레벨3","CFA Level 3","CFA lv3","FRM","공인회계사","CPA","감정평가사","세무사"]),
        (5, ["AICPA","CFA레벨2","CFA Level 2","CFA lv2"]),
        (3, ["재무설계사","CFP","CFA레벨1","CFA Level 1","CFA lv1",
             "금융투자분석사","신용위험분석사","투자자산운용사","신용분석사","재경관리사"]),
        (1, ["증권투자자문인력","파생투자자문인력","펀드투자자문인력","세무회계","회계관리"]),
    ]
    cert_score = 0; cert_name = "-"
    for c in d.get("certifications", []):
        n = c.get("name", "")
        for pts, keywords in CERT_TIERS:
            if any(kw.lower() in n.lower() for kw in keywords):
                if pts > cert_score: cert_score = pts; cert_name = n
                break
        else:
            if "CFA" in n.upper() and cert_score < 3: cert_score = 3; cert_name = n
    cert_score = min(cert_score, 6)

    return {"overseas": overseas_score, "activity": activity_score,
            "cert": cert_score, "cert_name": cert_name,
            "total": overseas_score + activity_score + cert_score}

def calc_deductions(d):
    points = 0; reasons = []
    level = d.get("program_participation", {}).get("dropout_risk_level", "없음")
    if level == "높음": points -= 10; reasons.append("이탈 가능성 높음 (-10)")
    elif level == "중간": points -= 5; reasons.append("이탈 가능성 중간 (-5)")

    MAXC = {"essay1_chars": 800, "essay2_chars": 800, "essay3_chars": 1000}
    LABELS = {"essay1_chars": "자소서1번", "essay2_chars": "자소서2번", "essay3_chars": "자소서3번"}
    for key, max_c in MAXC.items():
        actual = d.get(key, 0)
        ratio = actual / max_c if max_c else 0
        if ratio > 1.01: points -= 1; reasons.append(f"{LABELS[key]} 초과 (-1)")
        elif ratio >= 0.70: pass
        elif ratio >= 0.50: points -= 2; reasons.append(f"{LABELS[key]} 부족 50~69% (-2)")
        else: points -= 5; reasons.append(f"{LABELS[key]} 심각 부족 (-5)")

    if not d.get("portfolio_submitted", False):
        points -= 3; reasons.append("포트폴리오 미제출 (-3)")

    return {"total": points, "reasons": ", ".join(reasons) if reasons else "없음"}

def process(file_path, supporting_docs, client):
    name = Path(file_path).name
    print(f"\n[처리] {name}")

    print("  ① 텍스트 추출 중...")
    text = extract_text(file_path)
    for sp in supporting_docs:
        try:
            st = extract_text(sp)
            if st.strip():
                text += f"\n\n[증빙: {Path(sp).name}]\n{st}"
                print(f"     + 증빙자료: {Path(sp).name}")
        except Exception as e:
            print(f"  [경고] {Path(sp).name}: {e}")

    print("  ② 항목 구조화 중 (Claude)...")
    d = extract_structured(text, client)
    if d is None:
        print("  [실패] 구조화 실패")
        return None
    # essay 글자수 보정 (구 스키마 호환)
    essays = d.get("essays", {})
    if essays:
        d["essay1_chars"] = essays.get("essay1_chars", d.get("essay1_chars", 0))
        d["essay2_chars"] = essays.get("essay2_chars", d.get("essay2_chars", 0))
        d["essay3_chars"] = essays.get("essay3_chars", d.get("essay3_chars", 0))
    print(f"     → 지원자: {d.get('name','미상')}")

    print("  ③ 정량 채점 중...")
    q = score_quantitative(d)

    print("  ④ 정성 채점 중 (Claude)...")
    ql = score_qualitative(text, client)

    ded = calc_deductions(d)
    qual_total = sum(ql[k]["score"] for k in ("ai_consistency","diligence","leadership","investment","growth","commitment"))
    total = q["total"] + qual_total + ded["total"]
    diligence_ok = ql["diligence"]["score"] >= 7
    leadership_ok = ql["leadership"]["score"] >= 7
    passed = total >= 60 and (diligence_ok or leadership_ok)
    sm = {"total": total, "passed": passed}

    print(f"  완료: {total}점 -> {'통과' if passed else '탈락'}")
    return {"file": file_path, "data": d, "quantitative": q,
            "qualitative": ql, "deductions": ded, "summary": sm}

def append_to_v2(new_results, v2_path="2기_채점결과_v2.xlsx"):
    COLOR_HEADER = "1F4E79"; COLOR_PASS = "C6EFCE"; COLOR_FAIL = "FFC7CE"; COLOR_SCORE = "FFF2CC"
    COL_WIDTHS = [6,10,8,8,10,12,12,12,14,14,14,8,30,10,12,10,10,10,12,45,45,45,45,45,45,45,45,45]

    wb = openpyxl.load_workbook(v2_path)
    ws = wb.active

    # 기존 데이터 읽기 (미상 제외)
    existing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row) and row[1] != "미상":
            existing.append(list(row))

    # 새 결과 → 행 변환
    def result_to_row(res):
        d = res["data"]; q = res["quantitative"]; ql = res["qualitative"]
        ded = res["deductions"]; sm = res["summary"]
        return [
            0, d.get("name","미상"), sm["total"],
            "[통과]" if sm["passed"] else "[탈락]",
            q["overseas"], q["activity"], q["cert"],
            ql["ai_consistency"]["score"],
            ql["diligence"]["score"], ql["leadership"]["score"],
            ql["investment"]["score"], ql["growth"]["score"], ql["commitment"]["score"],
            ded["total"], ded["reasons"],
            d.get("ai_proficiency","-"),
            d.get("overseas_experience",{}).get("total_months",0),
            q.get("cert_name","-"),
            d.get("essay1_chars",0), d.get("essay2_chars",0), d.get("essay3_chars",0),
            d.get("program_participation",{}).get("dropout_risk_level","-"),
            ql["ai_consistency"]["reason"], ql["diligence"]["reason"],
            ql["leadership"]["reason"], ql["investment"]["reason"],
            ql["growth"]["reason"], ql["commitment"]["reason"],
        ]

    merged = existing + [result_to_row(r) for r in new_results]
    merged.sort(key=lambda r: (r[2] if r[2] is not None else -999), reverse=True)

    # 헤더 유지, 데이터만 재작성
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title = "1차 필터링 결과"

    hdr_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    pass_fill = PatternFill("solid", fgColor=COLOR_PASS)
    fail_fill = PatternFill("solid", fgColor=COLOR_FAIL)
    score_fill = PatternFill("solid", fgColor=COLOR_SCORE)
    bold = Font(bold=True, size=10); normal = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center
    ws2.row_dimensions[1].height = 32

    for rank, row in enumerate(merged, 1):
        r = rank + 1; row[0] = rank
        for col, val in enumerate(row, 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.font = bold if col in (2,3) else normal
            c.alignment = left if col >= 23 else center
        ws2.cell(row=r, column=4).fill = pass_fill if row[3] == "[통과]" else fail_fill
        ws2.cell(row=r, column=3).fill = score_fill
        ws2.row_dimensions[r].height = 60

    for col, w in enumerate(COL_WIDTHS[:len(headers)], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    wb2.save(v2_path)
    passed = sum(1 for r in merged if r[3] == "[통과]")
    print(f"\n병합 완료: 총 {len(merged)}명  |  통과 {passed}명  |  탈락 {len(merged)-passed}명")
    print(f"저장: {v2_path}")
    print("\n  순위  성명         총점   결과")
    print("  " + "-"*35)
    for r in merged:
        flag = "통과" if r[3] == "[통과]" else "탈락"
        print(f"  {r[0]:2}위   {str(r[1]):<10} {r[2]:3}점   {flag}")

def main():
    _load_env()
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("오류: ANTHROPIC_API_KEY 필요"); sys.exit(1)
    client = anthropic.Anthropic(api_key=api_key)

    targets = [
        ("지원서/2939. 송희경/송희경_DP Investor Club 2기 참가지원서.pdf", []),
    ]

    print(f"총 {len(targets)}명 수동 채점 시작\n")
    results = []
    for main_file, supp_files in targets:
        # 존재하는 증빙 파일만
        supp = [s for s in supp_files if Path(s).exists()]
        result = process(main_file, supp, client)
        if result:
            results.append(result)

    if not results:
        print("채점된 결과 없음"); sys.exit(1)

    append_to_v2(results)

if __name__ == "__main__":
    main()
