"""재처리_12명.xlsx 결과를 2기_채점결과.xlsx에 병합하고 순위 재정렬"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

MAIN = "2기_채점결과.xlsx"
PATCH = "재처리_12명.xlsx"
OUT   = "2기_채점결과.xlsx"

COLOR_HEADER = "1F4E79"
COLOR_PASS   = "C6EFCE"
COLOR_FAIL   = "FFC7CE"
COLOR_SCORE  = "FFF2CC"

def read_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(list(row))
    return rows

main_rows  = read_rows(MAIN)
patch_rows = read_rows(PATCH)

# 미상(-12점) 행 제거 후 패치 행 추가
filtered = [r for r in main_rows if r[1] != "미상"]
merged   = filtered + patch_rows

# 총점(col index 2) 기준 내림차순 정렬
merged.sort(key=lambda r: (r[2] if r[2] is not None else -999), reverse=True)

# 원본 헤더 읽기
wb_src = openpyxl.load_workbook(MAIN)
ws_src = wb_src.active
headers = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]

# 새 워크북 작성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "1차 필터링 결과"

hdr_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
hdr_font  = Font(color="FFFFFF", bold=True, size=10)
pass_fill = PatternFill("solid", fgColor=COLOR_PASS)
fail_fill = PatternFill("solid", fgColor=COLOR_FAIL)
score_fill= PatternFill("solid", fgColor=COLOR_SCORE)
bold   = Font(bold=True, size=10)
normal = Font(size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

COL_WIDTHS = [6,10,8,8,10,12,12,12,14,14,14,8,30,10,12,10,10,10,12,45,45,45]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = center
ws.row_dimensions[1].height = 32

for rank, row in enumerate(merged, 1):
    r = rank + 1
    row[0] = rank  # 순위 재부여
    for col, val in enumerate(row, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = bold if col in (2, 3) else normal
        c.alignment = left if col >= 20 else center
    ws.cell(row=r, column=4).fill = pass_fill if row[3] == "[통과]" else fail_fill
    ws.cell(row=r, column=3).fill = score_fill
    ws.row_dimensions[r].height = 60

for col, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A2"
wb.save(OUT)

passed = sum(1 for r in merged if r[3] == "[통과]")
print(f"병합 완료: 총 {len(merged)}명  |  통과 {passed}명  |  탈락 {len(merged)-passed}명")
print(f"저장: {OUT}")
print()
print("  순위  성명         총점   결과")
print("  " + "-"*35)
for r in merged:
    flag = "통과" if r[3] == "[통과]" else "탈락"
    print(f"  {r[0]:2}위   {str(r[1]):<10} {r[2]:3}점   {flag}")
